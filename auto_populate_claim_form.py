"""
auto_populate_claim_form.py
============================
customer_2603.csv (또는 customer.csv) 데이터를 읽어
기존 Claim Report Form Excel 템플릿에 자동으로 데이터를 채웁니다.

필드 매핑:
  - Store Name  ← ship_to_name
  - Sold-to     ← sold_to (코드) + sold_to_name
  - Ship-to     ← ship_to (코드) + ship_to_name + address

사용법:
  python auto_populate_claim_form.py --ship_to 724363
  python auto_populate_claim_form.py --ship_to 724363 731942 100142
  python auto_populate_claim_form.py --sample 3
  python auto_populate_claim_form.py --all
"""

import os
import sys
import argparse
import pandas as pd
import shutil
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────────
BASE_DIR = r"E:\01. work\2025\Data_Anal_Website"
RAWDATA_DIR = os.path.join(BASE_DIR, "rawdata")
TEMPLATE_PATH = os.path.join(RAWDATA_DIR, "Claim report form_v2(1).xlsx")
UNLOCK_DIR = os.path.join(RAWDATA_DIR, "unlock")
OUTPUT_DIR = os.path.join(RAWDATA_DIR, "output")

# Linux / 다른 환경 자동 전환
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if not os.path.exists(BASE_DIR):
    BASE_DIR = _SCRIPT_DIR
    RAWDATA_DIR = os.path.join(BASE_DIR, "rawdata")
    TEMPLATE_PATH = os.path.join(RAWDATA_DIR, "Claim report form_v2(1).xlsx")
    UNLOCK_DIR = os.path.join(RAWDATA_DIR, "unlock")
    OUTPUT_DIR = os.path.join(RAWDATA_DIR, "output")

# customer_2603.csv 우선 (unlock폴더 → 스크립트 루트 순), 없으면 customer.csv
CUSTOMER_CSV = os.path.join(UNLOCK_DIR, "customer_2603.csv")
if not os.path.exists(CUSTOMER_CSV):
    CUSTOMER_CSV = os.path.join(_SCRIPT_DIR, "customer_2603.csv")
if not os.path.exists(CUSTOMER_CSV):
    CUSTOMER_CSV = os.path.join(UNLOCK_DIR, "customer.csv")


# ─────────────────────────────────────────────
# 데이터 로드
# ─────────────────────────────────────────────
def load_customer_master() -> pd.DataFrame:
    print(f"데이터 파일: {CUSTOMER_CSV}")
    for enc in ["utf-8", "latin1", "cp1252", "utf-16"]:
        try:
            df = pd.read_csv(CUSTOMER_CSV, encoding=enc)
            break
        except (UnicodeDecodeError, Exception):
            continue
    else:
        raise ValueError(f"파일을 읽을 수 없습니다: {CUSTOMER_CSV}")

    df.columns = df.columns.str.strip()
    print(f"컬럼 목록: {df.columns.tolist()}")

    # 컬럼명 정규화: 대소문자/공백/특수문자 무관하게 매핑
    col_map = {}
    for col in df.columns:
        key = col.lower().replace(" ", "_").replace("-", "_")
        col_map[key] = col

    # 필요한 컬럼 자동 탐지 (이미 사용한 컬럼 재사용 방지)
    used_cols: set = set()
    rename = {}
    for target in ["sold_to", "sold_to_name", "ship_to", "ship_to_name", "address"]:
        if target in df.columns:
            used_cols.add(target)
            continue
        # 정확한 키 매칭 우선
        if target in col_map and col_map[target] not in used_cols:
            rename[col_map[target]] = target
            used_cols.add(col_map[target])
            continue
        # 부분 매칭 (이미 사용한 컬럼 제외)
        for key, orig in col_map.items():
            if orig in used_cols:
                continue
            if target in key or key in target:
                rename[orig] = target
                used_cols.add(orig)
                break

    if rename:
        print(f"컬럼 이름 변환: {rename}")
        df = df.rename(columns=rename)

    # sold_to, ship_to를 문자열로 변환
    for col in ["sold_to", "ship_to"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    dedup_col = "ship_to" if "ship_to" in df.columns else df.columns[0]
    return df.drop_duplicates(subset=[dedup_col], keep="first")


# ─────────────────────────────────────────────
# 템플릿 셀 자동 탐지
# ─────────────────────────────────────────────
def find_label_row(ws, label: str):
    """레이블 텍스트가 있는 행/열 반환."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and label.lower() in str(cell.value).lower():
                return cell.row, cell.column
    return None, None


def detect_value_columns(ws):
    """
    Store Name / Sold-to / Ship-to 레이블 셀을 찾아
    값을 입력할 (행, 열) 딕셔너리 반환.
    값 열은 레이블 열 + 3 (오른쪽으로 이동) 또는 탐지 실패 시 None.
    """
    mapping = {}
    for key, keywords in {
        "store_name": ["store name"],
        "sold_to":    ["sold-to", "sold to"],
        "ship_to":    ["ship-to", "ship to"],
    }.items():
        for kw in keywords:
            r, c = find_label_row(ws, kw)
            if r:
                mapping[key] = (r, c)
                break
    return mapping


# ─────────────────────────────────────────────
# 폼 채우기 (템플릿 기반)
# ─────────────────────────────────────────────
def safe_write(ws, row: int, col: int, value):
    """
    병합 셀이라도 안전하게 값을 기입.
    MergedCell인 경우 해당 병합 범위의 최상단 왼쪽(master) 셀에 씁니다.
    """
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # 이 셀이 속한 병합 범위 찾기
        for merged in ws.merged_cells.ranges:
            if (merged.min_row <= row <= merged.max_row and
                    merged.min_col <= col <= merged.max_col):
                ws.cell(row=merged.min_row, column=merged.min_col, value=value)
                return
    cell.value = value


def fill_form(customer: dict, output_path: str):
    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    mapping = detect_value_columns(ws)

    def write_right(key, fallback_row, fallback_col, value, row_offset=0):
        """레이블 오른쪽에 값 기입. 탐지 실패 시 fallback 셀 사용."""
        if key in mapping:
            r, c = mapping[key]
            safe_write(ws, r + row_offset, c + 1, value)
        else:
            safe_write(ws, fallback_row + row_offset, fallback_col, value)

    # ── Store Name ← ship_to_name ──────────────────
    write_right("store_name", 5, 5, customer["ship_to_name"])

    # ── Sold-to ← sold_to + sold_to_name ───────────
    write_right("sold_to", 8, 5, f"{customer['sold_to']}  {customer['sold_to_name']}")

    # ── Ship-to ← ship_to / ship_to_name / address ─
    write_right("ship_to", 11, 5, customer["ship_to"],      row_offset=0)
    write_right("ship_to", 12, 5, customer["ship_to_name"], row_offset=1)
    write_right("ship_to", 13, 5, customer["address"],      row_offset=2)

    wb.save(output_path)
    print(f"  ✔ {os.path.basename(output_path)}")


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Claim Report Form 자동 입력")
    parser.add_argument("--ship_to", nargs="+", help="ship_to 코드 지정 (예: 724363 731942)")
    parser.add_argument("--all", action="store_true", help="전체 고객 파일 생성")
    parser.add_argument("--sample", type=int, default=3, help="샘플 N개 생성 (기본 3)")
    args = parser.parse_args()

    if not os.path.exists(TEMPLATE_PATH):
        print(f"[ERROR] 템플릿 파일 없음: {TEMPLATE_PATH}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df = load_customer_master()

    if args.all:
        targets = df
    elif args.ship_to:
        targets = df[df["ship_to"].isin([str(s) for s in args.ship_to])]
        if targets.empty:
            print(f"[ERROR] 해당 ship_to를 찾을 수 없습니다: {args.ship_to}")
            sys.exit(1)
    else:
        targets = df.head(args.sample)

    for _, row in targets.iterrows():
        customer = {
            "sold_to":      str(row.get("sold_to", "")).strip(),
            "sold_to_name": str(row.get("sold_to_name", "")).strip(),
            "ship_to":      str(row.get("ship_to", "")).strip(),
            "ship_to_name": str(row.get("ship_to_name", "")).strip(),
            "address":      str(row.get("address", "")).strip(),
        }
        safe = customer["ship_to_name"].replace("/", "_").replace("\\", "_")[:40]
        out = os.path.join(OUTPUT_DIR, f"Claim_Form_{customer['ship_to']}_{safe}.xlsx")
        fill_form(customer, out)

    print(f"\n총 {len(targets)}개 파일 생성 완료 → {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
