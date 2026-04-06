"""
price_compare.py
----------------
Reads the latest Tempe_*.csv and creates a brand price comparison Excel file.

Brand abbreviations (from comparison table):
  MC=Michelin, BS=Bridgestone, CT=Continental, GY=Goodyear,
  KH=Kumho, FK=Falken, HK=Hankook, LF=Laufenn, DL=Dunlop, YO=Yokohama
"""
import csv
import glob
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Brand mapping ─────────────────────────────────────────────────────────────
BRANDS = {
    "MC": "Michelin",
    "BS": "Bridgestone",
    "CT": "Continental",
    "GY": "Goodyear",
    "KH": "Kumho",
    "FK": "Falken",
    "HK": "Hankook",
    "LF": "Laufenn",
    "DL": "Dunlop",
    "YO": "Yokohama",
}

# ── Competitor pattern table ───────────────────────────────────────────────────
# Maps (Line, Category, KumhoPattern) → brand → list of keyword patterns in description
# Keyword matching is case-insensitive substring search.
COMPETITOR_PATTERNS = {
    ("Car", "Touring", "K435"): {
        "MC": ["Energy XM2", "XM2+", "ENERGY XM2"],
        "BS": ["EP150", "EP300", "Ecopia EP"],
        "CT": ["UC7", "TC6", "ContiEcoContact", "EcoContact 6"],
        "GY": ["Assurance TripleMax", "Assurance Trio", "AssuranceMax"],
        "HK": ["Kinergy Eco 2", "K435", "H308", "KINERGY ECO"],
        "LF": ["G Fit AS", "LH41", "G-Fit"],
        "DL": ["SP Touring", "Enasave EC300"],
        "YO": ["BluEarth AE01", "BluEarth-ES", "AE01"],
        "FK": ["SN110", "Sincera SN110"],
    },
    ("Car", "GT", "K135"): {
        "MC": ["Primacy 4", "Primacy4", "PRIMACY"],
        "BS": ["RE004", "RE003", "Turanza T005", "T005"],
        "CT": ["UC6", "ContiUltraContact", "UltraContact"],
        "GY": ["EfficientGrip Perf", "EfficientGrip 2", "EFFICIENTGRIP"],
        "HK": ["Kinergy GT H436", "K435 GT", "KINERGY GT"],
        "LF": ["LH01", "S Fit AS"],
        "DL": ["SP Sport Maxx", "SP Sport LM705"],
        "YO": ["BluEarth GT AE51", "AE51", "BLUEARTH-GT"],
        "FK": ["FK520", "AZENIS FK520"],
    },
    ("Car", "Perf", "K127"): {
        "MC": ["Pilot Sport 4", "PS4", "PILOT SPORT 4"],
        "BS": ["RE003", "Potenza RE003", "POTENTO"],
        "CT": ["SportContact 6", "SportContact 5P", "CSC5P"],
        "GY": ["Eagle F1 Asym 5", "Eagle F1 Asym 3", "F1 ASYMMETRIC"],
        "HK": ["Ventus S1 evo3", "Ventus S1 evo2", "K127", "K137"],
        "LF": ["LH01 Sport", "Sport LH01"],
        "DL": ["SP Sport Maxx 050", "MAXX050"],
        "YO": ["ADVAN Sport V107", "ADVAN Sport V105", "V105"],
        "FK": ["FK510", "AZENIS FK510"],
    },
    ("Car", "Perf", "K137"): {
        "MC": ["Pilot Sport 4S", "PS4S", "PILOT SPORT 4S"],
        "BS": ["RE71RS", "RE050A", "Potento RE050"],
        "CT": ["SportContact 7", "SportContact 6", "CSC6"],
        "GY": ["Eagle F1 SuperSport", "F1 SuperSport"],
        "HK": ["Ventus S1 evo3 K127", "K137"],
        "DL": ["SP Sport Maxx Race", "MAXX RACE"],
        "YO": ["ADVAN Neova AD09", "AD09"],
        "FK": ["FK520", "Azenis FK520"],
    },
    ("SUV", "On-Road", "RA33"): {
        "MC": ["Primacy SUV", "Primacy 4 SUV", "PRIMACY SUV"],
        "BS": ["Dueler H/L 400", "Dueler H/L 33", "DUELER HL"],
        "CT": ["ContiCrossContact LX", "CrossContact LX2"],
        "GY": ["EfficientGrip SUV", "EfficientGrip 2 SUV"],
        "HK": ["Dynapro HP2 RA33", "RA33", "DYNAPRO HP"],
        "LF": ["Dynapro HP2 RA33", "RA33"],
        "DL": ["Grandtrek PT3", "Grandtrek PT5"],
        "YO": ["Geolandar CV G058", "G058"],
        "FK": ["Wildpeak H/T02A", "HT02A"],
    },
    ("SUV", "Premium", "RA43"): {
        "MC": ["Latitude Tour HP", "Latitude Sport 3", "LATITUDE"],
        "BS": ["Dueler H/P Sport", "Dueler Sport", "DUELER SPORT"],
        "CT": ["CrossContact RX", "ContiCrossContact RX"],
        "GY": ["Eagle F1 Asym SUV", "Eagle F1 SUV"],
        "HK": ["Dynapro HP2 RA43", "RA43", "DYNAPRO HP2"],
        "LF": ["RA43", "Dynapro RA43"],
        "DL": ["Grandtrek PT3A", "SP Sport Maxx"],
        "YO": ["Geolandar G055", "G055"],
        "FK": ["Wildpeak H/T01", "HT01"],
    },
    ("SUV", "Sports", "K127A"): {
        "MC": ["Pilot Sport 4 SUV", "Pilot Sport SUV", "PS4 SUV"],
        "BS": ["Dueler Sport AS", "Alenza Sport"],
        "CT": ["SportContact 6 SUV"],
        "GY": ["Eagle F1 Asym 5 SUV", "F1 ASYM SUV"],
        "HK": ["Ventus S1 evo3 SUV", "K127A"],
        "DL": ["SP Sport Maxx 050 SUV"],
        "YO": ["ADVAN Sport SUV V107", "V107 SUV"],
        "FK": ["FK520 SUV", "AZENIS FK520 SUV"],
    },
    ("PUP", "AT", "RF12"): {
        "BS": ["Dueler A/T 697", "Dueler A/T 001", "A/T 001", "AT 001"],
        "GY": ["Wrangler AT Adventure", "AT Adventure", "WRANGLER AT"],
        "HK": ["Dynapro AT2 RF11", "Dynapro AT RF12", "RF11", "RF12"],
        "DL": ["Grandtrek AT3", "Grandtrek AT5"],
        "YO": ["Geolandar A/T G015", "G015"],
        "FK": ["Wildpeak A/T3W", "AT3W", "WILDPEAK AT"],
        "MC": ["Latitude Cross", "LTX Force"],
        "CT": ["ContiCrossContact AT"],
    },
    ("PUP", "MT", "RT05"): {
        "BS": ["Dueler M/T 674", "MT 674"],
        "GY": ["Wrangler MT/R", "Wrangler Duratrac", "DURATRAC"],
        "HK": ["Dynapro MT2 RT05", "RT05", "MT2"],
        "DL": ["Grandtrek MT2"],
        "YO": ["Geolandar M/T G003", "G003"],
        "FK": ["Wildpeak M/T", "MT01", "WILDPEAK MT"],
        "MC": ["Latitude Trial", "BF Goodrich MT"],
        "CT": ["CrossContact M/T"],
    },
    ("Van", "Van", "RA18"): {
        "MC": ["Agilis +", "Agilis 3", "AGILIS"],
        "BS": ["R660", "Duravis R660"],
        "CT": ["VancoCamper", "VanContact 100", "VANCONTACT"],
        "GY": ["Cargo Vector 2", "Cargo Marathon 2", "CARGO"],
        "HK": ["Radial RA18", "RA18"],
        "DL": ["SP VAN01", "SPVAN"],
        "YO": ["BluEarth Van RY55", "RY55"],
        "FK": ["VAN01", "Linam Van"],
    },
}

# ── Size → category mapping ────────────────────────────────────────────────────
SIZE_CATEGORY = {
    "175/65R14": ("Car",  "Touring", "K435"),
    "185/65R14": ("Car",  "Touring", "K435"),
    "195/65R15": ("Car",  "Touring", "K435"),
    "205/65R15": ("Car",  "Touring", "K435"),
    "205/65R16": ("Car",  "Touring", "K435"),
    "215/65R16": ("Car",  "Touring", "K435"),
    "225/65R17": ("SUV",  "On-Road", "RA33"),
    "235/65R16": ("SUV",  "On-Road", "RA33"),
    "215/45R17": ("Car",  "GT",      "K135"),
    "225/45R17": ("Car",  "GT",      "K135"),
    "235/45R17": ("Car",  "GT",      "K135"),
    "225/55R18": ("SUV",  "Premium", "RA43"),
    "225/60R18": ("SUV",  "On-Road", "RA33"),
    "225/40R18": ("Car",  "Perf",    "K127"),
    "225/55R19": ("SUV",  "Premium", "RA43"),
    "225/45R19": ("SUV",  "Sports",  "K127A"),
    "235/45R19": ("Car",  "Perf",    "K127"),
    "255/45R20": ("SUV",  "Sports",  "K127A"),
    "245/35R20": ("Car",  "Perf",    "K127"),
    "265/40R21": ("SUV",  "Sports",  "K127A"),
    "245/70R16": ("PUP",  "AT",      "RF12"),
    "265/70R16": ("PUP",  "AT",      "RF12"),
    "265/60R18": ("PUP",  "AT",      "RF12"),
    "265/65R17": ("PUP",  "AT",      "RF12"),
    "265/75R16": ("PUP",  "MT",      "RT05"),
    "185R14":    ("Van",  "Van",     "RA18"),
    "195R15":    ("Van",  "Van",     "RA18"),
}

# Header colours per brand
BRAND_COLOURS = {
    "MC":  "C8102E",  # Michelin red
    "BS":  "E31837",  # Bridgestone red
    "CT":  "FFA500",  # Continental orange
    "GY":  "003087",  # Goodyear blue
    "KH":  "00539B",  # Kumho blue
    "FK":  "E8001A",  # Falken red
    "HK":  "FF6600",  # Hankook orange
    "LF":  "FF6600",  # Laufenn (Hankook sub-brand)
    "DL":  "E4002B",  # Dunlop red
    "YO":  "003DA5",  # Yokohama blue
}

ROW_FILLS = {
    "MC":  "FADADD", "BS":  "FFE0E0", "CT":  "FFF3CD",
    "GY":  "D6EAF8", "KH":  "D4E6F1", "FK":  "FCE4D6",
    "HK":  "FFF0E0", "LF":  "FDEBD0", "DL":  "FADBD8",
    "YO":  "D4E6F1",
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg="1F4E79", fg="FFFFFF", bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = PatternFill("solid", fgColor=bg)
    c.font   = Font(color=fg, bold=bold, size=11)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = thin()
    return c

def data_cell(ws, row, col, value, bg="FFFFFF", align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

def abbr_for(brand_name):
    for abbr, name in BRANDS.items():
        if name.lower() == brand_name.strip().lower():
            return abbr
    return None

def autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 55)

def normalise_size(raw):
    """Extract tyre size like '175/65R14' or '185R14' from a description string."""
    m = re.search(r'(\d{3}/\d{2}R\d{2}|\d{3}R\d{2})', raw, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return raw.strip().split()[0].upper() if raw.strip() else ""

def match_competitor(desc, keywords):
    """Return True if any keyword appears in desc (case-insensitive)."""
    desc_l = desc.lower()
    return any(kw.lower() in desc_l for kw in keywords)

# ── Load CSV ──────────────────────────────────────────────────────────────────
def load_csv(path):
    rows = []
    with open(path, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows

# ── Sheet 1: Comparison summary (one row per size, brand columns) ────────────
def sheet_summary(wb, all_rows):
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "C4"

    brand_abbrs = list(BRANDS.keys())
    n_brands = len(brand_abbrs)

    # ── Build lookup: (size_norm, abbr) → best matched (desc, cost, price) ───
    # Use COMPETITOR_PATTERNS keywords to find best match per brand per size
    raw_lookup = {}   # (size_norm, abbr) → list of (desc, cost_f, price_f)
    for r in all_rows:
        brand = r.get("brand", "").strip()
        desc  = r.get("DESCRIPTION", "").strip()
        cost  = r.get("COST",  "").strip()
        price = r.get("PRICE", "").strip()
        abbr  = abbr_for(brand) or ("KH" if brand.lower() == "kumho" else None)
        if not abbr or not desc:
            continue
        size_norm = normalise_size(desc)
        cost_f  = float(cost)  if cost  else None
        price_f = float(price) if price else None
        raw_lookup.setdefault((size_norm, abbr), []).append((desc, cost_f, price_f))

    def best_match(size, abbr):
        candidates = raw_lookup.get((size, abbr), [])
        cat_info = SIZE_CATEGORY.get(size)
        if cat_info:
            keywords = COMPETITOR_PATTERNS.get(cat_info, {}).get(abbr, [])
            matched = [x for x in candidates if keywords and match_competitor(x[0], keywords)]
            pool = matched if matched else candidates
        else:
            pool = candidates
        if not pool:
            return None, None, None
        best = sorted(pool, key=lambda x: (x[2] is None, x[2] or 0))[0]
        return best  # (desc, cost_f, price_f)

    # ── Title row ─────────────────────────────────────────────────────────────
    total_cols = 3 + n_brands * 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1, value="All Sizes — Brand Price Comparison  (COST / PRICE)")
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: brand headers ──────────────────────────────────────────────────
    hdr_cell(ws, 2, 1, "Size",    bg="2E4057")
    hdr_cell(ws, 2, 2, "Pattern", bg="2E4057")
    hdr_cell(ws, 2, 3, "Category", bg="2E4057")
    col = 4
    for abbr in brand_abbrs:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        hdr_cell(ws, 2, col, f"{abbr}  {BRANDS[abbr]}",
                 bg=BRAND_COLOURS.get(abbr, "555555"))
        ws.cell(row=2, column=col+1).fill = PatternFill(
            "solid", fgColor=BRAND_COLOURS.get(abbr, "555555"))
        col += 2

    # ── Row 3: sub-headers ────────────────────────────────────────────────────
    hdr_cell(ws, 3, 1, "Size",     bg="334155", fg="CCCCCC", bold=False)
    hdr_cell(ws, 3, 2, "Pattern",  bg="334155", fg="CCCCCC", bold=False)
    hdr_cell(ws, 3, 3, "Category", bg="334155", fg="CCCCCC", bold=False)
    col = 4
    for _ in brand_abbrs:
        hdr_cell(ws, 3, col,   "Cost",  bg="334155", fg="AAAAAA", bold=False)
        hdr_cell(ws, 3, col+1, "Price", bg="334155", fg="AAAAAA", bold=False)
        col += 2
    ws.row_dimensions[3].height = 18

    # ── Data rows: one row per size ───────────────────────────────────────────
    sizes_ordered = list(SIZE_CATEGORY.keys())
    csv_sizes = sorted({normalise_size(r.get("DESCRIPTION", "")) for r in all_rows})
    for s in csv_sizes:
        if s not in sizes_ordered:
            sizes_ordered.append(s)

    row_num = 4
    prev_line = None
    for size in sizes_ordered:
        cat_info = SIZE_CATEGORY.get(size)
        if cat_info:
            line, category, kumho_pat = cat_info
        else:
            line, category, kumho_pat = ("?", "?", "?")

        # Light separator line between Line groups (Car / SUV / PUP / Van)
        if prev_line and line != prev_line:
            for ci in range(1, total_cols + 1):
                ws.cell(row=row_num, column=ci).fill = PatternFill("solid", fgColor="D9D9D9")
            row_num += 1
        prev_line = line

        bg = "F7F9FC" if row_num % 2 == 0 else "FFFFFF"

        data_cell(ws, row_num, 1, size,       bg=bg, align="center")
        data_cell(ws, row_num, 2, kumho_pat,  bg=bg, align="center")
        data_cell(ws, row_num, 3, f"{line} / {category}", bg=bg, align="center")

        col = 4
        for abbr in brand_abbrs:
            brand_bg = ROW_FILLS.get(abbr, "FFFFFF")
            desc, cost_val, price_val = best_match(size, abbr)
            if desc:
                data_cell(ws, row_num, col,   cost_val,  bg=brand_bg, align="right", num_fmt="$#,##0.00")
                data_cell(ws, row_num, col+1, price_val, bg=brand_bg, align="right", num_fmt="$#,##0.00")
            else:
                data_cell(ws, row_num, col,   "-", bg="F0F0F0", align="center")
                data_cell(ws, row_num, col+1, "-", bg="F0F0F0", align="center")
            col += 2

        row_num += 1

    autofit(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18


# ── Sheet 2: Detail (all brands, sorted by price) ────────────────────────────
def sheet_detail(wb, all_rows):
    ws = wb.create_sheet("All Products")
    ws.freeze_panes = "A3"

    hdr_cell(ws, 1, 1, "Brand Abbr", bg="2E4057")
    hdr_cell(ws, 1, 2, "Brand",       bg="2E4057")
    hdr_cell(ws, 1, 3, "Description", bg="2E4057")
    hdr_cell(ws, 1, 4, "Cost ($)",    bg="2E4057")
    hdr_cell(ws, 1, 5, "Price ($)",   bg="2E4057")
    hdr_cell(ws, 1, 6, "Margin ($)",  bg="2E4057")
    hdr_cell(ws, 1, 7, "Margin %",    bg="2E4057")

    known_rows = []
    other_rows = []
    for r in all_rows:
        brand = r.get("brand", "").strip()
        abbr  = abbr_for(brand)
        if abbr:
            known_rows.append((abbr, r))
        else:
            other_rows.append(("--", r))

    known_rows.sort(key=lambda x: (x[0], float(x[1].get("PRICE","0") or 0)))

    all_out = known_rows + other_rows
    for i, (abbr, r) in enumerate(all_out):
        rn     = i + 2
        brand  = r.get("brand", "").strip()
        desc   = r.get("DESCRIPTION", "").strip()
        cost   = r.get("COST",  "").strip()
        price  = r.get("PRICE", "").strip()
        bg     = ROW_FILLS.get(abbr, "F9F9F9") if abbr != "--" else "F9F9F9"

        cost_f  = float(cost)  if cost  else None
        price_f = float(price) if price else None
        margin  = round(price_f - cost_f, 2)           if (cost_f and price_f) else None
        margin_pct = round(margin / price_f * 100, 1)  if (margin and price_f) else None

        data_cell(ws, rn, 1, abbr if abbr != "--" else "other", bg=bg, align="center")
        data_cell(ws, rn, 2, brand, bg=bg)
        data_cell(ws, rn, 3, desc,  bg=bg)
        data_cell(ws, rn, 4, cost_f,      bg=bg, align="right", num_fmt="$#,##0.00")
        data_cell(ws, rn, 5, price_f,     bg=bg, align="right", num_fmt="$#,##0.00")
        data_cell(ws, rn, 6, margin,      bg=bg, align="right", num_fmt="$#,##0.00")
        data_cell(ws, rn, 7, margin_pct,  bg=bg, align="right", num_fmt='0.0"%"')

    # Conditional colour scale on Price column
    last = len(all_out) + 1
    ws.conditional_formatting.add(
        f"E2:E{last}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B")
    )
    autofit(ws)
    ws.column_dimensions["C"].width = 50


# ── Sheet 3: Competitor Match ─────────────────────────────────────────────────
def sheet_competitor_match(wb, all_rows):
    """
    For each (size, category, kumho_pattern) group, show one row per brand
    listing the best-matching product found in the CSV (cost + price).
    Layout:
      Line | Category | Kumho Pattern | Size | Brand | Competitor Model | Matched Product | Cost | Price | Margin%
    """
    ws = wb.create_sheet("Competitor Match")
    ws.freeze_panes = "A3"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "Line", "Category", "Kumho\nPattern", "Size",
        "Brand", "Competitor\nModel Keywords",
        "Matched Product", "Cost ($)", "Price ($)", "Margin %",
    ]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h, bg="1F4E79")
    ws.row_dimensions[1].height = 30

    # ── Build lookup: (size_norm, brand_abbr) → list of (desc, cost, price) ──
    lookup = {}   # key: (size_norm, abbr)  value: list of (desc, cost_f, price_f)
    for r in all_rows:
        brand = r.get("brand", "").strip()
        desc  = r.get("DESCRIPTION", "").strip()
        cost  = r.get("COST",  "").strip()
        price = r.get("PRICE", "").strip()
        abbr  = abbr_for(brand)
        if not abbr or not desc:
            continue
        size_norm = normalise_size(desc)
        cost_f  = float(cost)  if cost  else None
        price_f = float(price) if price else None
        lookup.setdefault((size_norm, abbr), []).append((desc, cost_f, price_f))

    # Also build KH (Kumho) entries separately
    for r in all_rows:
        brand = r.get("brand", "").strip()
        desc  = r.get("DESCRIPTION", "").strip()
        cost  = r.get("COST",  "").strip()
        price = r.get("PRICE", "").strip()
        if brand.lower() != "kumho" or not desc:
            continue
        size_norm = normalise_size(desc)
        cost_f  = float(cost)  if cost  else None
        price_f = float(price) if price else None
        lookup.setdefault((size_norm, "KH"), []).append((desc, cost_f, price_f))

    # ── Write rows ────────────────────────────────────────────────────────────
    # Order: iterate sizes in SIZE_CATEGORY order, then each brand
    sizes_ordered = list(SIZE_CATEGORY.keys())
    # Also pick up sizes from CSV not in SIZE_CATEGORY
    csv_sizes = sorted({normalise_size(r.get("DESCRIPTION","")) for r in all_rows})
    for s in csv_sizes:
        if s not in sizes_ordered:
            sizes_ordered.append(s)

    row_num = 2
    prev_group = None
    for size in sizes_ordered:
        cat_info = SIZE_CATEGORY.get(size)  # (line, category, kumho_pattern)
        if cat_info:
            line, category, kumho_pat = cat_info
        else:
            line, category, kumho_pat = ("?", "?", "?")

        group_key = (line, category, kumho_pat, size)
        comp_dict = COMPETITOR_PATTERNS.get((line, category, kumho_pat), {})

        # Collect brands to show: KH first, then others from BRANDS
        brands_to_show = ["KH"] + [a for a in BRANDS if a != "KH"]

        group_started = False
        for abbr in brands_to_show:
            brand_name = BRANDS.get(abbr, abbr)
            keywords   = comp_dict.get(abbr, [])

            # Find matching products in CSV
            candidates = lookup.get((size, abbr), [])
            if keywords:
                matched = [(d, c, p) for d, c, p in candidates
                           if match_competitor(d, keywords)]
            else:
                matched = candidates  # no keyword filter → show all

            # Pick best match: prefer one with both cost and price, then cheapest price
            if matched:
                matched_sorted = sorted(
                    matched,
                    key=lambda x: (x[2] is None, x[2] or 0)
                )
                best_desc, best_cost, best_price = matched_sorted[0]
            else:
                best_desc, best_cost, best_price = ("—", None, None)

            margin_pct = None
            if best_cost and best_price and best_price > 0:
                margin_pct = round((best_price - best_cost) / best_price * 100, 1)

            kw_display = " / ".join(keywords) if keywords else "—"
            bg = ROW_FILLS.get(abbr, "F9F9F9")

            # Alternate row shade for readability between groups
            if abbr == "KH":
                bg = "D4E6F1"  # Kumho highlight

            data_cell(ws, row_num, 1,  line,        bg=bg, align="center")
            data_cell(ws, row_num, 2,  category,    bg=bg, align="center")
            data_cell(ws, row_num, 3,  kumho_pat,   bg=bg, align="center")
            data_cell(ws, row_num, 4,  size,        bg=bg, align="center")
            data_cell(ws, row_num, 5,  f"{abbr} – {brand_name}", bg=bg)
            data_cell(ws, row_num, 6,  kw_display,  bg=bg)
            data_cell(ws, row_num, 7,  best_desc,   bg=bg)
            data_cell(ws, row_num, 8,  best_cost,   bg=bg, align="right", num_fmt="$#,##0.00")
            data_cell(ws, row_num, 9,  best_price,  bg=bg, align="right", num_fmt="$#,##0.00")
            data_cell(ws, row_num, 10, margin_pct,  bg=bg, align="right", num_fmt='0.0"%"')

            row_num += 1
            group_started = True

        # Blank separator row between size groups
        if group_started:
            for ci in range(1, 11):
                ws.cell(row=row_num, column=ci).fill = PatternFill("solid", fgColor="E8E8E8")
            row_num += 1

    # Column widths
    autofit(ws)
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 55


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    csv_files = sorted(glob.glob("Tempe_*.csv"), reverse=True)
    if not csv_files:
        print("ERROR: No Tempe_*.csv file found in current directory.")
        return

    csv_path = csv_files[0]
    print(f"Reading: {csv_path}")
    rows = load_csv(csv_path)
    print(f"  {len(rows)} rows loaded")

    wb = Workbook()
    wb.remove(wb.active)          # remove default sheet
    sheet_summary(wb, rows)
    sheet_detail(wb, rows)
    sheet_competitor_match(wb, rows)

    from datetime import datetime
    ts  = datetime.now().strftime("%H%M%S")
    out = csv_path.replace(".csv", f"_comparison_{ts}.xlsx")
    wb.save(out)
    print(f"Saved:  {out}")
    os.startfile(out)             # auto-open in Excel (Windows)


if __name__ == "__main__":
    main()
