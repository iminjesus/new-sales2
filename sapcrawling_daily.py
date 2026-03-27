import os
import time
import glob
import csv
import win32com.client
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

"""
ZSDR24030 Daily Sales Export
- Billing date: 1st of current month ~ yesterday's business day
- Output: D:\Data-Anal website\rawdata\unlock\zsdr24030_daily.csv

Pre-req:
- SAP GUI must be OPEN and LOGGED IN
- Packages: pywin32, openpyxl
"""

# ================= CONFIG =================
TCODE = "ZSDR24030"

SAP_EXPORT_DIR  = r"C:\temp"
SAP_EXPORT_GLOB = "EXPORT_*.xlsx"

OUT_CSV = r"D:\Data-Anal website\rawdata\unlock\zsdr24030_daily.csv"

DELETE_XLSX_AFTER_CONVERT = False
MIN_CSV_SIZE = 200  # bytes


# ================= DATE HELPERS =================
def last_business_day(ref: date) -> date:
    """ref 기준 전날부터 거슬러 올라가며 첫 번째 평일 반환"""
    d = ref - timedelta(days=1)
    while d.weekday() >= 5:   # 5=Sat, 6=Sun
        d -= timedelta(days=1)
    return d

def month_start(ref: date) -> date:
    return ref.replace(day=1)

def sap_date(d: date) -> str:
    """SAP 날짜 입력 형식: DD.MM.YYYY"""
    return d.strftime("%d.%m.%Y")


# ================= HELPERS =================
def wait(sec=0.3):
    time.sleep(sec)

def ensure_out_dir(path):
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)

def exists(session, wid):
    try:
        session.findById(wid)
        return True
    except:
        return False

def close_popups(session, max_steps=6):
    for _ in range(max_steps):
        if not exists(session, "wnd[1]"):
            return
        for wid in ("wnd[1]/tbar[0]/btn[0]",
                    "wnd[1]/usr/btnSPOP-OPTION1"):
            try:
                session.findById(wid).press()
                wait(0.25)
                break
            except:
                pass
        else:
            try:
                session.findById("wnd[1]").sendVKey(0)
                wait(0.25)
            except:
                return

def start_tcode(session, tcode):
    session.findById("wnd[0]/tbar[0]/okcd").Text = f"/n{tcode}"
    session.findById("wnd[0]").sendVKey(0)
    wait(1.5)
    close_popups(session, 4)

def set_billing_date(session, date_from: str, date_to: str):
    """
    Billing date LOW/HIGH 필드 설정.
    ZSDR24030 화면의 실제 ID에 맞게 fallback 순서로 시도.
    """
    # LOW (from)
    for wid in (
        "wnd[0]/usr/ctxtS_FKDAT-LOW",
        "wnd[0]/usr/ctxtFKDAT-LOW",
        "wnd[0]/usr/txtS_FKDAT-LOW",
        "wnd[0]/usr/txtFKDAT-LOW",
    ):
        if exists(session, wid):
            try:
                session.findById(wid).Text = date_from
                break
            except:
                pass

    # HIGH (to)
    for wid in (
        "wnd[0]/usr/ctxtS_FKDAT-HIGH",
        "wnd[0]/usr/ctxtFKDAT-HIGH",
        "wnd[0]/usr/txtS_FKDAT-HIGH",
        "wnd[0]/usr/txtFKDAT-HIGH",
    ):
        if exists(session, wid):
            try:
                session.findById(wid).Text = date_to
                break
            except:
                pass

    session.findById("wnd[0]").sendVKey(0)
    wait(0.5)

def newest_export_xlsx(after_ts: float) -> str:
    pattern = os.path.join(SAP_EXPORT_DIR, SAP_EXPORT_GLOB)
    candidates = []
    for p in glob.glob(pattern):
        try:
            mtime = os.path.getmtime(p)
            if mtime >= after_ts:
                candidates.append((mtime, p))
        except:
            pass
    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][1]

def xlsx_to_csv(xlsx_path: str, csv_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    ensure_out_dir(csv_path)

    def norm(v):
        return "" if v is None else str(v).strip()

    def is_number_like(s):
        if not s:
            return False
        try:
            float(s.replace(",", ""))
            return True
        except:
            return False

    def is_trailing_summary_row(values):
        non_empty = [v for v in values if v != ""]
        if len(non_empty) == 0:
            return True
        if len(non_empty) <= 2 and all(is_number_like(v) for v in non_empty):
            return True
        return False

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([norm(v) for v in row])

    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    while rows and is_trailing_summary_row(rows[-1]):
        rows.pop()

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow(r)

def file_ok(path, min_size):
    try:
        return os.path.exists(path) and os.path.getsize(path) >= min_size
    except:
        return False


# ---------- Export trigger strategies ----------
def try_export_via_alv_toolbar(session) -> bool:
    grid_ids = [
        "wnd[0]/usr/cntlGRID1/shellcont/shell",
        "wnd[0]/usr/cntlCONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlALV_CONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlGRID1/shellcont/shellcont/shell",
    ]
    for gid in grid_ids:
        if not exists(session, gid):
            continue
        try:
            grid = session.findById(gid)
            opened = False
            for btn in ("&MB_EXPORT", "EXPORT", "&EXPORT"):
                try:
                    grid.pressToolbarContextButton(btn)
                    wait(0.4)
                    opened = True
                    break
                except:
                    continue
            if not opened:
                continue
            for item in ("&XXL", "XXL", "&SPREADSHEET", "SPREADSHEET", "&PC", "PC"):
                try:
                    grid.selectContextMenuItem(item)
                    wait(0.8)
                    close_popups(session, 4)
                    return True
                except:
                    continue
        except:
            continue
    return False

def find_menu_path_by_text(session, texts):
    try:
        mbar = session.findById("wnd[0]/mbar")
    except:
        return None
    level = []
    try:
        for i in range(mbar.Children.Count):
            level.append(mbar.Children(i))
    except:
        return None
    last_hit = None
    for t in texts:
        hit = None
        for it in level:
            try:
                txt = (getattr(it, "Text", "") or "").replace("&", "").strip().lower()
                if txt == t.lower():
                    hit = it
                    break
            except:
                continue
        if hit is None:
            return None
        last_hit = hit
        nxt = []
        try:
            for i in range(hit.Children.Count):
                nxt.append(hit.Children(i))
        except:
            nxt = []
        level = nxt
    return last_hit

def try_export_via_menu(session) -> bool:
    for path in (["List", "Export", "Spreadsheet"],
                 ["List", "Export", "Local File"],
                 ["List", "Export", "Spreadsheet..."]):
        item = find_menu_path_by_text(session, path)
        if item is not None:
            try:
                item.select()
                wait(0.8)
                close_popups(session, 6)
                return True
            except:
                pass
    fallbacks = [
        "wnd[0]/mbar/menu[0]/menu[3]/menu[1]",
        "wnd[0]/mbar/menu[0]/menu[11]/menu[1]",
        "wnd[0]/mbar/menu[0]/menu[4]/menu[1]",
        "wnd[0]/mbar/menu[0]/menu[3]/menu[2]",
    ]
    for wid in fallbacks:
        if exists(session, wid):
            try:
                session.findById(wid).select()
                wait(0.8)
                close_popups(session, 6)
                return True
            except:
                pass
    return False

def trigger_export(session):
    if try_export_via_alv_toolbar(session):
        return True
    if try_export_via_menu(session):
        return True
    return False


# ================= MAIN =================
def main():
    today     = date.today()
    d_from    = month_start(today)
    d_to      = last_business_day(today)

    date_from = sap_date(d_from)
    date_to   = sap_date(d_to)

    print(f"Billing date: {date_from} ~ {date_to}")

    ensure_out_dir(OUT_CSV)
    if os.path.exists(OUT_CSV):
        try:
            os.remove(OUT_CSV)
        except:
            pass

    export_start_ts = time.time()

    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    app = SapGuiAuto.GetScriptingEngine

    if app.Children.Count == 0:
        raise RuntimeError("No SAP connection found. Open SAP GUI and log in first.")
    conn = app.Children(0)
    if conn.Children.Count == 0:
        raise RuntimeError("No SAP session found.")
    session = conn.Children(0)

    try:
        session.findById("wnd[0]").maximize()
    except:
        pass

    start_tcode(session, TCODE)
    set_billing_date(session, date_from, date_to)

    # Execute (F8)
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    wait(2.0)
    close_popups(session, 4)

    ok = trigger_export(session)
    if not ok:
        raise RuntimeError("Could not trigger export.")

    xlsx_path = ""
    for _ in range(40):
        xlsx_path = newest_export_xlsx(export_start_ts)
        if xlsx_path and os.path.exists(xlsx_path):
            s1 = os.path.getsize(xlsx_path)
            wait(0.8)
            s2 = os.path.getsize(xlsx_path)
            if s2 >= s1 and s2 > 500:
                break
        wait(0.6)

    if not xlsx_path:
        raise RuntimeError(f"Could not find SAP export XLSX in {SAP_EXPORT_DIR}")

    print("Detected XLSX:", xlsx_path)
    xlsx_to_csv(xlsx_path, OUT_CSV)

    if not file_ok(OUT_CSV, MIN_CSV_SIZE):
        raise RuntimeError("CSV conversion failed or CSV is too small.")

    print("CSV saved:", OUT_CSV)

    if DELETE_XLSX_AFTER_CONVERT:
        try:
            os.remove(xlsx_path)
        except Exception as e:
            print("[WARN] Could not delete XLSX:", e)

if __name__ == "__main__":
    main()
