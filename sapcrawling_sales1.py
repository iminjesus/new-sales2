import os
import time
import glob
import csv
import win32com.client
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

"""
ZSDR24030 Sales Export → Local CSV
- Sales Organization: 4200
- Billing date: 1st of current month ~ last working day
- Output: <repo>/rawdata/unlock/sales_thismonth.csv  (next to this script)

Pre-req:
- SAP GUI must be OPEN and LOGGED IN
- Packages: pywin32, openpyxl
"""

# ================= CONFIG =================
TCODE          = "ZSDR24030"
SALES_ORG      = "4200"

SAP_EXPORT_DIR  = r"C:\temp"
SAP_EXPORT_GLOB = "EXPORT_*.xlsx"

# Resolve relative to this script so the same code runs on every checkout
# regardless of the absolute path the repo was cloned to.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_CSV  = os.path.join(BASE_DIR, "rawdata", "unlock", "sales_thismonth.csv")

DELETE_XLSX_AFTER_CONVERT = False
MIN_CSV_SIZE = 200  # bytes

# SAP user-profile date format varies per operator (System → User Profile
# → Own Data → Defaults → Date format).  Order matters: the first accepted
# format wins.  If SAP rejects one with a status-bar warning, main() will
# reset the selection screen and try the next.
DATE_FORMATS = [
    "%d.%m.%Y",   # 01.08.2026  (SAP default 1 — Jayden's profile)
    "%Y.%m.%d",   # 2026.08.01  (SAP default 4 — this user's profile)
    "%m/%d/%Y",   # 08/01/2026  (SAP default 2 — US)
    "%d/%m/%Y",   # 01/08/2026  (SAP default 6 — UK/AU)
    "%Y-%m-%d",   # 2026-08-01  (ISO / SAP default 6)
]

# ALV Grid saved layout to auto-apply after F8 (Settings → Layout → Choose).
# Set to "" or None to skip auto-layout and use the SAP default.
LAYOUT_NAME = "JaydenSQL"


# ================= DATE HELPERS =================
def last_business_day(ref: date) -> date:
    d = ref - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def first_business_day(ref: date) -> date:
    d = ref.replace(day=1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

def month_start(ref: date) -> date:
    return ref.replace(day=1)

def sap_date(d: date) -> str:
    return d.strftime("%d.%m.%Y")


# ================= HELPERS =================
def wait(sec=0.3):
    time.sleep(sec)

def ensure_out_dir(path):
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)

def exists(session, wid):
    try:
        session.findById(wid)
        return True
    except:
        return False

def close_popups(session, max_steps=6):
    for _ in range(max_steps):
        if not exists(session, "wnd[1]"):
            return
        for wid in ("wnd[1]/tbar[0]/btn[0]",
                    "wnd[1]/usr/btnSPOP-OPTION1"):
            try:
                session.findById(wid).press()
                wait(0.25)
                break
            except:
                pass
        else:
            try:
                session.findById("wnd[1]").sendVKey(0)
                wait(0.25)
            except:
                return

def start_tcode(session, tcode):
    session.findById("wnd[0]/tbar[0]/okcd").Text = f"/n{tcode}"
    session.findById("wnd[0]").sendVKey(0)
    wait(1.5)
    close_popups(session, 4)


def sbar_text(session) -> str:
    """Return the SAP status-bar text ('Enter data in the format XX/YY/ZZZZ',
    'No data was selected', etc.) or empty string if unavailable."""
    try:
        return (session.findById("wnd[0]/sbar").Text or "").strip()
    except:
        return ""

def sbar_is_error(session) -> bool:
    """True when the status bar carries an error/warning (MessageType E/W/A).
    Empty MessageType means no message or informational only."""
    try:
        return (session.findById("wnd[0]/sbar").MessageType or "").upper() in ("E", "W", "A")
    except:
        return False

def looks_like_date_format_error(msg: str) -> bool:
    """The SAP messages that indicate 'wrong date format' vary by language:
    'Enter data in the format XX/YY/ZZZZ', 'Please enter a valid date',
    'Ungültiges Datum', etc.  Match on the common keywords."""
    m = (msg or "").lower()
    return any(k in m for k in ("format", "date", "invalid", "enter", "gültig", "valid"))

def set_selection_fields(session, date_from: str, date_to: str):
    """Sales Organization 4200 + Billing Date 설정"""

    # Sales Organization
    for wid in (
        "wnd[0]/usr/ctxtS_VKORG-LOW",
        "wnd[0]/usr/ctxtP_VKORG",
        "wnd[0]/usr/txtS_VKORG-LOW",
    ):
        if exists(session, wid):
            try:
                session.findById(wid).Text = SALES_ORG
                print(f"  Sales Org set via {wid}")
                break
            except:
                pass

    # Billing Date FROM
    for wid in (
        "wnd[0]/usr/ctxtS_FKDAT-LOW",
        "wnd[0]/usr/ctxtP_FKDAT-LOW",
        "wnd[0]/usr/txtS_FKDAT-LOW",
        "wnd[0]/usr/txtP_FKDAT-LOW",
    ):
        if exists(session, wid):
            try:
                session.findById(wid).Text = date_from
                print(f"  Billing Date FROM set via {wid}")
                break
            except:
                pass

    # Billing Date TO
    for wid in (
        "wnd[0]/usr/ctxtS_FKDAT-HIGH",
        "wnd[0]/usr/ctxtP_FKDAT-HIGH",
        "wnd[0]/usr/txtS_FKDAT-HIGH",
        "wnd[0]/usr/txtP_FKDAT-HIGH",
    ):
        if exists(session, wid):
            try:
                session.findById(wid).Text = date_to
                print(f"  Billing Date TO   set via {wid}")
                break
            except:
                pass

    wait(0.3)

def newest_export_xlsx(after_ts: float) -> str:
    pattern = os.path.join(SAP_EXPORT_DIR, SAP_EXPORT_GLOB)
    candidates = []
    for p in glob.glob(pattern):
        try:
            mtime = os.path.getmtime(p)
            if mtime >= after_ts:
                candidates.append((mtime, p))
        except:
            pass
    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][1]

def xlsx_to_csv(xlsx_path: str, csv_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    ensure_out_dir(csv_path)

    def norm(v):
        return "" if v is None else str(v).strip()

    def is_number_like(s):
        if not s:
            return False
        try:
            float(s.replace(",", ""))
            return True
        except:
            return False

    def is_trailing_summary_row(values):
        non_empty = [v for v in values if v != ""]
        if len(non_empty) == 0:
            return True
        if len(non_empty) <= 2 and all(is_number_like(v) for v in non_empty):
            return True
        return False

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([norm(v) for v in row])

    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    while rows and is_trailing_summary_row(rows[-1]):
        rows.pop()

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow(r)

def file_ok(path, min_size):
    try:
        return os.path.exists(path) and os.path.getsize(path) >= min_size
    except:
        return False


# ---------- Export strategies ----------
def try_export_via_alv_toolbar(session) -> bool:
    grid_ids = [
        "wnd[0]/usr/cntlGRID1/shellcont/shell",
        "wnd[0]/usr/cntlCONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlALV_CONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlGRID1/shellcont/shellcont/shell",
    ]
    for gid in grid_ids:
        if not exists(session, gid):
            continue
        try:
            grid = session.findById(gid)
            opened = False
            for btn in ("&MB_EXPORT", "EXPORT", "&EXPORT"):
                try:
                    grid.pressToolbarContextButton(btn)
                    wait(0.4)
                    opened = True
                    break
                except:
                    continue
            if not opened:
                continue
            for item in ("&XXL", "XXL", "&SPREADSHEET", "SPREADSHEET", "&PC", "PC"):
                try:
                    grid.selectContextMenuItem(item)
                    wait(0.8)
                    close_popups(session, 4)
                    return True
                except:
                    continue
        except:
            continue
    return False

def find_menu_path_by_text(session, texts):
    try:
        mbar = session.findById("wnd[0]/mbar")
    except:
        return None
    level = []
    try:
        for i in range(mbar.Children.Count):
            level.append(mbar.Children(i))
    except:
        return None
    last_hit = None
    for t in texts:
        hit = None
        for it in level:
            try:
                txt = (getattr(it, "Text", "") or "").replace("&", "").strip().lower()
                if txt == t.lower():
                    hit = it
                    break
            except:
                continue
        if hit is None:
            return None
        last_hit = hit
        nxt = []
        try:
            for i in range(hit.Children.Count):
                nxt.append(hit.Children(i))
        except:
            nxt = []
        level = nxt
    return last_hit

def try_export_via_menu(session) -> bool:
    for path in (["List", "Export", "Spreadsheet"],
                 ["List", "Export", "Local File"],
                 ["List", "Export", "Spreadsheet..."]):
        item = find_menu_path_by_text(session, path)
        if item is not None:
            try:
                item.select()
                wait(0.8)
                close_popups(session, 6)
                return True
            except:
                pass
    fallbacks = [
        "wnd[0]/mbar/menu[0]/menu[3]/menu[1]",
        "wnd[0]/mbar/menu[0]/menu[11]/menu[1]",
        "wnd[0]/mbar/menu[0]/menu[4]/menu[1]",
        "wnd[0]/mbar/menu[0]/menu[3]/menu[2]",
    ]
    for wid in fallbacks:
        if exists(session, wid):
            try:
                session.findById(wid).select()
                wait(0.8)
                close_popups(session, 6)
                return True
            except:
                pass
    return False

def trigger_export(session):
    if try_export_via_alv_toolbar(session):
        return True
    if try_export_via_menu(session):
        return True
    return False


# ---------- Layout auto-apply ----------
def try_set_layout_on_selection(session, layout_name: str) -> bool:
    """Some reports (including ZSDR24030 in many customer configs) expose
    a 'Display variant' parameter on the selection screen — filling it
    means F8 opens the results with the layout already applied, no popup.
    Silently returns False if no such field exists on the current screen."""
    if not layout_name:
        return False
    for wid in (
        "wnd[0]/usr/ctxtP_VARI",
        "wnd[0]/usr/ctxtP_LAYOUT",
        "wnd[0]/usr/ctxtP_VARIANT",
        "wnd[0]/usr/ctxtLIS_LAYOUT",
        "wnd[0]/usr/ctxtP_ALV",
    ):
        if exists(session, wid):
            try:
                session.findById(wid).Text = layout_name
                print(f"  Layout '{layout_name}' set on selection screen via {wid}")
                return True
            except:
                pass
    return False

def apply_layout_via_dialog(session, layout_name: str) -> bool:
    """After the report has executed and the ALV grid is showing, open
    the 'Choose Layout' dialog and select the row whose name matches
    layout_name.  Tries three strategies in order:
      1. Grid toolbar button (&LOAD / &VARIANT).
      2. Ctrl+F9 shortcut (Settings → Layout → Choose).
      3. Menu-path walk ('Settings' → 'Layout' → 'Choose').
    Then in the popup, tries to find the layout by scanning the ALV
    list or by typing into the search box.  Returns False on any
    failure — caller may proceed with the default layout."""
    if not layout_name:
        return False

    # ---- 1. Open the 'Choose Layout' dialog --------------------------------
    opened = False
    grid_ids = [
        "wnd[0]/usr/cntlGRID1/shellcont/shell",
        "wnd[0]/usr/cntlCONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlALV_CONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlGRID1/shellcont/shellcont/shell",
    ]
    for gid in grid_ids:
        if not exists(session, gid):
            continue
        for btn in ("&LOAD", "&VARIANT", "&MB_VARIANT"):
            try:
                session.findById(gid).pressToolbarButton(btn)
                wait(0.6)
                opened = True
                break
            except:
                continue
        if opened:
            break
    if not opened:
        try:
            session.findById("wnd[0]").sendVKey(33)   # Ctrl+F9
            wait(0.6)
            opened = True
        except:
            pass
    if not opened:
        item = find_menu_path_by_text(session, ["Settings", "Layout", "Choose..."])
        if item is None:
            item = find_menu_path_by_text(session, ["Settings", "Layout", "Choose"])
        if item is not None:
            try:
                item.select(); wait(0.6); opened = True
            except:
                pass

    if not opened or not exists(session, "wnd[1]"):
        return False

    # ---- 2. Find the row for layout_name in the popup ----------------------
    lname = layout_name.strip().lower()

    # (a) ALV grid inside popup — scan rows/cols for the name
    popup_grid_ids = [
        "wnd[1]/usr/cntlG_ALV_LAYOUT/shellcont/shell",
        "wnd[1]/usr/cntlCONTAINER1/shellcont/shell",
        "wnd[1]/usr/cntlSHELL/shellcont/shell",
    ]
    for gid in popup_grid_ids:
        if not exists(session, gid):
            continue
        try:
            grid = session.findById(gid)
            row_cnt = int(getattr(grid, "RowCount", 0) or 0)
            for r in range(row_cnt):
                for col in ("VARIANT", "TEXT", "LTEXT", "NAME"):
                    try:
                        val = grid.getCellValue(r, col)
                        if val and val.strip().lower() == lname:
                            grid.setCurrentCell(r, col)
                            grid.doubleClickCurrentCell()
                            wait(0.5)
                            print(f"  Layout '{layout_name}' applied via popup grid")
                            return True
                    except:
                        continue
        except:
            continue

    # (b) Search box in popup — type name + Enter
    for wid in ("wnd[1]/usr/ctxt%%DYN001-LOW",
                "wnd[1]/usr/txtRSVAR-VARIANT",
                "wnd[1]/usr/txtV-LOW"):
        if exists(session, wid):
            try:
                session.findById(wid).Text = layout_name
                session.findById("wnd[1]/tbar[0]/btn[0]").press()  # Enter
                wait(0.5)
                print(f"  Layout '{layout_name}' applied via popup search")
                return True
            except:
                continue

    # Give up — close popup gracefully so caller can continue
    try: session.findById("wnd[1]/tbar[0]/btn[12]").press()   # Cancel
    except: pass
    return False


# ================= MAIN =================
def main():
    today   = date.today()

    # 이달 첫 번째 비즈니스 데이 당일 또는 그 이전이면 전달 데이터로 다운로드
    if today <= first_business_day(today):
        last_day_prev = month_start(today) - timedelta(days=1)
        d_from = month_start(last_day_prev)
        d_to   = last_business_day(month_start(today))
        print("[INFO] 이달 첫 번째 비즈니스 데이 이전 → 전달 데이터 범위로 다운로드")
    else:
        d_from = month_start(today)
        d_to   = last_business_day(today)

    print(f"Output file  : {OUT_CSV}")

    ensure_out_dir(OUT_CSV)
    if os.path.exists(OUT_CSV):
        try:
            os.remove(OUT_CSV)
        except:
            pass

    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    app = SapGuiAuto.GetScriptingEngine

    if app.Children.Count == 0:
        raise RuntimeError("No SAP connection found. Open SAP GUI and log in first.")
    conn = app.Children(0)
    if conn.Children.Count == 0:
        raise RuntimeError("No SAP session found.")
    session = conn.Children(0)

    try:
        session.findById("wnd[0]").maximize()
    except:
        pass

    # ---- Loop over date formats until SAP accepts one --------------------
    # The user profile date format varies (Jayden's = DD.MM.YYYY, this
    # operator's = YYYY.MM.DD, US ops = MM/DD/YYYY).  We reset the
    # selection screen and retry each format, checking the status bar for
    # 'Enter data in the format XX/YY/ZZZZ' style warnings before giving up.
    export_start_ts = None
    accepted_fmt    = None
    for i, fmt in enumerate(DATE_FORMATS, 1):
        date_from = d_from.strftime(fmt)
        date_to   = d_to.strftime(fmt)
        print(f"[Attempt {i}/{len(DATE_FORMATS)}] date format = {fmt} → {date_from} ~ {date_to}")

        start_tcode(session, TCODE)
        set_selection_fields(session, date_from, date_to)
        try_set_layout_on_selection(session, LAYOUT_NAME)

        export_start_ts = time.time()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()   # F8
        wait(2.0)
        close_popups(session, 4)

        msg = sbar_text(session)
        if sbar_is_error(session) and looks_like_date_format_error(msg):
            print(f"  ✗ SAP rejected: {msg}")
            continue

        # No date-format complaint — assume F8 worked.  If the report
        # returned zero rows, trigger_export below will just fail cleanly.
        accepted_fmt = fmt
        if msg:
            print(f"  ✓ format {fmt} accepted (status: {msg})")
        else:
            print(f"  ✓ format {fmt} accepted")
        break

    if not accepted_fmt:
        raise RuntimeError(
            "None of the candidate date formats were accepted by SAP.  "
            "Check System → User Profile → Own Data → Defaults → Date format "
            "and add that format to DATE_FORMATS at the top of this script."
        )

    # ---- Auto-apply the saved layout (JaydenSQL) -------------------------
    if apply_layout_via_dialog(session, LAYOUT_NAME):
        # give ALV a moment to redraw with the new columns
        wait(0.6)
    else:
        print(f"  [WARN] Could not auto-apply layout '{LAYOUT_NAME}' — using SAP default")

    ok = trigger_export(session)
    if not ok:
        raise RuntimeError("Could not trigger export.")

    xlsx_path = ""
    for _ in range(40):
        xlsx_path = newest_export_xlsx(export_start_ts)
        if xlsx_path and os.path.exists(xlsx_path):
            s1 = os.path.getsize(xlsx_path)
            wait(0.8)
            s2 = os.path.getsize(xlsx_path)
            if s2 >= s1 and s2 > 500:
                break
        wait(0.6)

    if not xlsx_path:
        raise RuntimeError(f"Could not find SAP export XLSX in {SAP_EXPORT_DIR}")

    print("Detected XLSX:", xlsx_path)
    xlsx_to_csv(xlsx_path, OUT_CSV)

    if not file_ok(OUT_CSV, MIN_CSV_SIZE):
        raise RuntimeError("CSV conversion failed or file too small.")

    print("CSV saved:", OUT_CSV)

    if DELETE_XLSX_AFTER_CONVERT:
        try:
            os.remove(xlsx_path)
        except Exception as e:
            print("[WARN] Could not delete XLSX:", e)

if __name__ == "__main__":
    main()
