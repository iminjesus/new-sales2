"""
create_sample_claim_forms.py
============================
customer_2603.csv (또는 customer.csv) 데이터를 읽어
Claim Report Form Excel 파일을 자동으로 생성합니다.

필드 매핑:
  - Store Name  ← ship_to_name
  - Sold-to     ← sold_to (코드) + sold_to_name
  - Ship-to     ← ship_to (코드) + ship_to_name + address

실행: python create_sample_claim_forms.py
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ─────────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
UNLOCK_DIR = os.path.join(SCRIPT_DIR, "rawdata", "unlock")

# customer_2603.csv 우선 사용, 없으면 customer.csv 사용
CUSTOMER_CSV = os.path.join(UNLOCK_DIR, "customer_2603.csv")
if not os.path.exists(CUSTOMER_CSV):
    CUSTOMER_CSV = os.path.join(UNLOCK_DIR, "customer.csv")

OUTPUT_DIR = os.path.join(SCRIPT_DIR, "rawdata", "output")


# ─────────────────────────────────────────────
# 스타일 헬퍼
# ─────────────────────────────────────────────
def thin():
    return Side(style="thin")

def no_side():
    return Side(style=None)

def border(top=False, bottom=False, left=False, right=False, dashed_bottom=False):
    b = Side(style="dashed") if dashed_bottom else (thin() if bottom else no_side())
    return Border(
        top=thin() if top else no_side(),
        bottom=b,
        left=thin() if left else no_side(),
        right=thin() if right else no_side(),
    )

def box():
    return Border(top=thin(), bottom=thin(), left=thin(), right=thin())

def font(size=9, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def align(h="left", v="top", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─────────────────────────────────────────────
# 컬럼 너비 / 행 높이 설정
# 스크린샷 기준: A~AP 정도까지 넓게 사용
# ─────────────────────────────────────────────
def setup_sheet(ws):
    # A~D: 레이블 영역, E~Z: 값 영역 (넓게)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14   # 레이블 열
    for col in "CDEFGHIJKLMNOPQRSTUVWXYZ":
        ws.column_dimensions[col].width = 3
    # 오른쪽 박스 (Manufacturer 영역)
    import openpyxl.utils as utils
    for i in range(27, 45):  # AA ~ AR
        ws.column_dimensions[utils.get_column_letter(i)].width = 3

    for r in range(1, 50):
        ws.row_dimensions[r].height = 16


# ─────────────────────────────────────────────
# 왼쪽 정보 박스 테두리 그리기
# cols: 2(B) ~ 28(AB) / rows: 3~16
# ─────────────────────────────────────────────
def draw_left_box(ws, r1=3, r2=16, c1=2, c2=28):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            top    = (r == r1)
            bottom = (r == r2)
            left   = (c == c1)
            right  = (c == c2)
            if top or bottom or left or right:
                ws.cell(row=r, column=c).border = border(
                    top=top, bottom=bottom, left=left, right=right
                )


# ─────────────────────────────────────────────
# 오른쪽 Manufacturer 박스
# cols: 30(AD) ~ 44(AR)
# ─────────────────────────────────────────────
def draw_right_box(ws):
    # 섹션별 행 범위
    sections = [
        (3, 6),   # Manufacturer / Logo
        (7, 9),   # Manufacturer Handling / Claim no.
        (10, 12), # Inspector Name / Mobile
        (13, 16), # Tyre Owner / Mobile
    ]
    c1, c2 = 30, 44
    for r1, r2 in sections:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                top    = (r == r1)
                bottom = (r == r2)
                left   = (c == c1)
                right  = (c == c2)
                if top or bottom or left or right:
                    ws.cell(row=r, column=c).border = border(
                        top=top, bottom=bottom, left=left, right=right
                    )


# ─────────────────────────────────────────────
# Claim Form 작성
# ─────────────────────────────────────────────
def build_claim_form(wb: Workbook, customer: dict):
    ws = wb.active
    ws.title = "Claim Report"
    setup_sheet(ws)

    # ── 제목 ─────────────────────────────────────────
    ws.merge_cells("B1:AB1")
    ws["B1"] = "Claim report form"
    ws["B1"].font = Font(name="Arial", size=16, bold=True)
    ws["B1"].alignment = align(h="left", v="center")
    ws.row_dimensions[1].height = 28

    # ── 왼쪽 외곽선 박스 (B3:AB16) ────────────────────
    draw_left_box(ws)

    # ── Store Name (← ship_to_name) ───────────────────
    ws.merge_cells("B4:D4")
    ws["B4"] = "Store Name :"
    ws["B4"].font = font(size=9)
    ws["B4"].alignment = align(v="center")

    ws.merge_cells("E4:AB4")
    ws["E4"] = customer["ship_to_name"]
    ws["E4"].font = font(size=9)
    ws["E4"].alignment = align(v="center")

    # Store Name 아래 dashed 구분선
    for c in range(2, 29):
        cell = ws.cell(row=5, column=c)
        cell.border = Border(bottom=Side(style="dashed"))

    # ── Sold-to (← sold_to 코드 + sold_to_name) ───────
    ws.merge_cells("B7:D7")
    ws["B7"] = "Sold-to :"
    ws["B7"].font = font(size=9)
    ws["B7"].alignment = align(v="center")

    ws.merge_cells("E7:AB7")
    ws["E7"] = f"{customer['sold_to']}  {customer['sold_to_name']}"
    ws["E7"].font = font(size=9)
    ws["E7"].alignment = align(v="center")

    # Sold-to 아래 구분선
    for c in range(2, 29):
        cell = ws.cell(row=8, column=c)
        cell.border = Border(bottom=Side(style="dashed"))

    # ── Ship-to (← ship_to 코드 / ship_to_name / address) ─
    ws.merge_cells("B10:D10")
    ws["B10"] = "Ship-to :"
    ws["B10"].font = font(size=9)
    ws["B10"].alignment = align(v="top")

    # ship_to ID
    ws.merge_cells("E10:AB10")
    ws["E10"] = customer["ship_to"]
    ws["E10"].font = font(size=9)
    ws["E10"].alignment = align(v="center")

    # ship_to_name
    ws.merge_cells("E11:AB11")
    ws["E11"] = customer["ship_to_name"]
    ws["E11"].font = font(size=9)
    ws["E11"].alignment = align(v="center")

    # address
    ws.merge_cells("E12:AB12")
    ws["E12"] = customer["address"]
    ws["E12"].font = font(size=9)
    ws["E12"].alignment = align(v="center")

    # Ship-to 아래 dashed 구분선
    for c in range(2, 29):
        cell = ws.cell(row=13, column=c)
        cell.border = Border(bottom=Side(style="dashed"))

    # ── 오른쪽 Manufacturer 박스 ──────────────────────
    draw_right_box(ws)

    ws.merge_cells("AD3:AR3")
    ws["AD3"] = "Manufacturer"
    ws["AD3"].font = font(size=8)
    ws["AD3"].alignment = align(v="top")

    ws.merge_cells("AD7:AR7")
    ws["AD7"] = "Manufacturer Handling"
    ws["AD7"].font = font(size=8)

    ws.merge_cells("AD8:AR8")
    ws["AD8"] = "Claim no."
    ws["AD8"].font = font(size=8)

    ws.merge_cells("AD10:AR10")
    ws["AD10"] = "Inspector Name :"
    ws["AD10"].font = font(size=8)

    ws.merge_cells("AD11:AR11")
    ws["AD11"] = "Mobile Phone no. :"
    ws["AD11"].font = font(size=8)

    ws.merge_cells("AD13:AR13")
    ws["AD13"] = "Tyre Owner Name :"
    ws["AD13"].font = font(size=8)

    ws.merge_cells("AD14:AR14")
    ws["AD14"] = "Mobile Phone No."
    ws["AD14"].font = font(size=8)

    # ── Tyre Purchase Date ────────────────────────────
    ws.merge_cells("B18:D18")
    ws["B18"] = "Tyre Purchase Date:"
    ws["B18"].font = font(bold=True)

    ws.merge_cells("F18:P18")
    ws["F18"].border = box()

    # ── Return Tires Information ──────────────────────
    ws.merge_cells("B20:AB20")
    ws["B20"] = "Return Tires Information"
    ws["B20"].font = font(bold=True)

    # 헤더 행
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    headers = [
        ("Tire Size\nDescription",      "B22", "F23"),
        ("Pattern Name\n(or Product Name)", "G22", "K23"),
        ("DOT Number",                  "L22", "P23"),
        ("Serial Number\n(Truck & BUS Only)", "Q22", "U23"),
        ("Remain Skid\nDepth (mm)",     "V22", "AB23"),
    ]
    for text, start, end in headers:
        ws.merge_cells(f"{start}:{end}")
        cell = ws[start]
        cell.value = text
        cell.font = font(size=8)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = box()

    # 데이터 행 (ex, 1, 2, 3)
    row_labels = ["ex", "1", "2", "3"]
    col_ranges = [("B", "F"), ("G", "K"), ("L", "P"), ("Q", "U"), ("V", "AB")]
    for i, lbl in enumerate(row_labels):
        r = 24 + i * 2
        ws.cell(row=r, column=2).value = lbl
        ws.cell(row=r, column=2).font = font(size=8)
        for sc, ec in col_ranges:
            sc_n = _col_num(sc)
            ec_n = _col_num(ec)
            ws.merge_cells(start_row=r, start_column=sc_n, end_row=r+1, end_column=ec_n)
            ws.cell(row=r, column=sc_n).border = box()

    # Description 영역
    desc_row = 24 + len(row_labels) * 2 + 1
    ws.merge_cells(start_row=desc_row, start_column=2, end_row=desc_row+3, end_column=28)
    ws.cell(row=desc_row, column=2).value = "Description (Claim reason)"
    ws.cell(row=desc_row, column=2).font = font(size=8)
    ws.cell(row=desc_row, column=2).border = box()
    ws.cell(row=desc_row, column=2).alignment = align(v="top")


def _col_num(col_letter: str) -> int:
    """컬럼 문자 → 번호 (A=1, Z=26, AA=27 …)"""
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(col_letter)


# ─────────────────────────────────────────────
# 메인 실행
# ─────────────────────────────────────────────
def load_customers(n: int = None) -> pd.DataFrame:
    print(f"데이터 파일: {CUSTOMER_CSV}")
    df = pd.read_csv(
        CUSTOMER_CSV,
        dtype={"sold_to": str, "ship_to": str},
        encoding="latin1",
    )
    df.columns = df.columns.str.strip()
    # ship_to 기준으로 파일 생성 (각 ship_to별 1개 파일)
    df = df.drop_duplicates(subset=["ship_to"], keep="first")
    return df.head(n) if n else df


def create_sample_files(n: int = 3):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df = load_customers(n)
    created = []

    for _, row in df.iterrows():
        customer = {
            "sold_to":      str(row.get("sold_to", "")).strip(),
            "sold_to_name": str(row.get("sold_to_name", "")).strip(),
            "ship_to":      str(row.get("ship_to", "")).strip(),
            "ship_to_name": str(row.get("ship_to_name", "")).strip(),
            "address":      str(row.get("address", "")).strip(),
        }

        wb = Workbook()
        build_claim_form(wb, customer)

        safe_name = customer["ship_to_name"].replace("/", "_").replace("\\", "_")[:40]
        filename = f"Claim_Form_{customer['ship_to']}_{safe_name}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        wb.save(output_path)
        created.append(output_path)
        print(f"  ✔ {filename}")

    print(f"\n총 {len(created)}개 파일 생성 완료 → {OUTPUT_DIR}")
    return created


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Claim Report Form 샘플 생성")
    parser.add_argument("--n", type=int, default=3, help="생성할 파일 수 (기본 3)")
    parser.add_argument("--all", action="store_true", help="전체 고객 파일 생성")
    args = parser.parse_args()

    print("Claim Report Form 생성 중...")
    create_sample_files(n=None if args.all else args.n)
