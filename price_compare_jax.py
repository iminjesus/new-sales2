"""
price_compare_jax.py
--------------------
Reads latest Tempe_*.csv + BobJane_*.csv + JAX_*.csv
and produces a three-way comparison Excel.

Sheets:
  1. Summary      — T.Product | T.Cost | T.Price | BJ Product | BJ Price | BJ Disc | JAX Product | JAX Price | JAX Disc
  2. BJ Products  — all Bob Jane products
  3. JAX Products — all JAX products
  4. Match Detail — full side-by-side with diff columns
"""
import csv, glob, os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BRANDS = {
    "MC": "Michelin",    "BS": "Bridgestone", "CT": "Continental",
    "GY": "Goodyear",   "KH": "Kumho",       "FK": "Falken",
    "HK": "Hankook",    "LF": "Laufenn",     "DL": "Dunlop",
    "YO": "Yokohama",
}

COMPETITOR_PATTERNS = {
    ("Car", "Touring", "K435"): {
        "MC": ["Energy XM2","XM2+"],
        "BS": ["EP150","EP300","Ecopia EP"],
        "CT": ["UC7","TC6","EcoContact 6"],
        "GY": ["Assurance TripleMax","Assurance Trio"],
        "HK": ["Kinergy Eco 2","K435","KINERGY ECO"],
        "LF": ["G Fit AS","LH41","G-Fit"],
        "DL": ["SP Touring","Enasave EC300"],
        "YO": ["BluEarth AE01","BluEarth-ES","AE01"],
        "FK": ["SN110","Sincera SN110"],
    },
    ("Car", "GT", "K135"): {
        "MC": ["Primacy 4","PRIMACY"],
        "BS": ["RE004","RE003","Turanza T005","T005"],
        "CT": ["UC6","UltraContact"],
        "GY": ["EfficientGrip Perf","EfficientGrip 2"],
        "HK": ["Kinergy GT H436","KINERGY GT"],
        "LF": ["LH01","S Fit AS"],
        "DL": ["SP Sport Maxx","SP Sport LM705"],
        "YO": ["BluEarth GT AE51","AE51"],
        "FK": ["FK520","AZENIS FK520"],
    },
    ("Car", "Perf", "K127"): {
        "MC": ["Pilot Sport 4","PS4"],
        "BS": ["RE003","Potenza RE003"],
        "CT": ["SportContact 6","SportContact 5P"],
        "GY": ["Eagle F1 Asym 5","Eagle F1 Asym 3"],
        "HK": ["Ventus S1 evo3","Ventus S1 evo2","K127"],
        "DL": ["SP Sport Maxx 050"],
        "YO": ["ADVAN Sport V107","ADVAN Sport V105"],
        "FK": ["FK510","AZENIS FK510"],
    },
    ("Car", "Perf", "K137"): {
        "MC": ["Pilot Sport 4S","PS4S"],
        "BS": ["RE71RS","RE050A"],
        "CT": ["SportContact 7","CSC6"],
        "GY": ["Eagle F1 SuperSport"],
        "HK": ["K137"],
        "DL": ["SP Sport Maxx Race"],
        "YO": ["ADVAN Neova AD09","AD09"],
        "FK": ["FK520","Azenis FK520"],
    },
    ("SUV", "On-Road", "RA33"): {
        "MC": ["Primacy SUV","Primacy 4 SUV"],
        "BS": ["Dueler H/L 400","Dueler H/L 33"],
        "CT": ["CrossContact LX2"],
        "GY": ["EfficientGrip SUV"],
        "HK": ["Dynapro HP2 RA33","RA33"],
        "DL": ["Grandtrek PT3","Grandtrek PT5"],
        "YO": ["Geolandar CV G058","G058"],
        "FK": ["Wildpeak H/T02A","HT02A"],
    },
    ("SUV", "Premium", "RA43"): {
        "MC": ["Latitude Tour HP","Latitude Sport 3"],
        "BS": ["Dueler H/P Sport","Dueler Sport"],
        "CT": ["CrossContact RX"],
        "GY": ["Eagle F1 Asym SUV"],
        "HK": ["Dynapro HP2 RA43","RA43"],
        "DL": ["Grandtrek PT3A"],
        "YO": ["Geolandar G055","G055"],
        "FK": ["Wildpeak H/T01","HT01"],
    },
    ("SUV", "Sports", "K127A"): {
        "MC": ["Pilot Sport 4 SUV","PS4 SUV"],
        "BS": ["Dueler Sport AS","Alenza Sport"],
        "GY": ["Eagle F1 Asym 5 SUV"],
        "HK": ["Ventus S1 evo3 SUV","K127A"],
        "DL": ["SP Sport Maxx 050 SUV"],
        "YO": ["ADVAN Sport SUV V107"],
        "FK": ["FK520 SUV"],
    },
    ("PUP", "AT", "RF12"): {
        "BS": ["Dueler A/T 697","Dueler A/T 001"],
        "GY": ["Wrangler AT Adventure"],
        "HK": ["Dynapro AT2 RF11","RF12"],
        "DL": ["Grandtrek AT3","Grandtrek AT5"],
        "YO": ["Geolandar A/T G015","G015"],
        "FK": ["Wildpeak A/T3W","AT3W"],
        "MC": ["Latitude Cross"],
        "CT": ["ContiCrossContact AT"],
    },
    ("PUP", "MT", "RT05"): {
        "BS": ["Dueler M/T 674"],
        "GY": ["Wrangler Duratrac"],
        "HK": ["Dynapro MT2 RT05","RT05"],
        "DL": ["Grandtrek MT2"],
        "YO": ["Geolandar M/T G003","G003"],
        "FK": ["Wildpeak M/T","MT01"],
    },
    ("Van", "Van", "RA18"): {
        "MC": ["Agilis +","Agilis 3"],
        "BS": ["R660","Duravis R660"],
        "CT": ["VanContact 100"],
        "GY": ["Cargo Vector 2"],
        "HK": ["Radial RA18","RA18"],
        "DL": ["SP VAN01"],
        "YO": ["BluEarth Van RY55","RY55"],
        "FK": ["VAN01","Linam Van"],
    },
}

SIZE_CATEGORY = {
    "175/65R14": ("Car","Touring","K435"),  "185/65R14": ("Car","Touring","K435"),
    "195/65R15": ("Car","Touring","K435"),  "205/65R15": ("Car","Touring","K435"),
    "205/65R16": ("Car","Touring","K435"),  "215/65R16": ("Car","Touring","K435"),
    "225/65R17": ("SUV","On-Road","RA33"),  "235/65R16": ("SUV","On-Road","RA33"),
    "215/45R17": ("Car","GT","K135"),       "225/45R17": ("Car","GT","K135"),
    "235/45R17": ("Car","GT","K135"),       "225/55R18": ("SUV","Premium","RA43"),
    "225/60R18": ("SUV","On-Road","RA33"),  "225/40R18": ("Car","Perf","K127"),
    "225/55R19": ("SUV","Premium","RA43"),  "225/45R19": ("SUV","Sports","K127A"),
    "235/45R19": ("Car","Perf","K127"),     "255/45R20": ("SUV","Sports","K127A"),
    "245/35R20": ("Car","Perf","K127"),     "265/40R21": ("SUV","Sports","K127A"),
    "245/70R16": ("PUP","AT","RF12"),       "265/70R16": ("PUP","AT","RF12"),
    "265/60R18": ("PUP","AT","RF12"),       "265/65R17": ("PUP","AT","RF12"),
    "265/75R16": ("PUP","MT","RT05"),       "185R14":    ("Van","Van","RA18"),
    "195R15":    ("Van","Van","RA18"),
}

BRAND_COLOURS = {
    "MC":"E53935",   # vivid red
    "BS":"1976D2",   # strong blue
    "CT":"F57C00",   # orange
    "GY":"303F9F",   # indigo
    "KH":"00897B",   # teal
    "FK":"C2185B",   # deep pink
    "HK":"FF6600",   # bright orange  (solid line)
    "LF":"2E7D32",   # forest green   (solid line)
    "DL":"6D4C41",   # brown
    "YO":"7B1FA2",   # purple
}
ROW_FILLS = {
    "MC":"FADADD","BS":"FFE0E0","CT":"FFF3CD","GY":"D6EAF8",
    "KH":"D4E6F1","FK":"FCE4D6","HK":"FFF0E0","LF":"FDEBD0",
    "DL":"FADBD8","YO":"D4E6F1",
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg="1F4E79", fg="FFFFFF", bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color=fg, bold=bold, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin()
    return c

def dc(ws, row, col, value, bg="FFFFFF", align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

def autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 55)

def abbr_for(brand_name):
    for abbr, name in BRANDS.items():
        if name.lower() == brand_name.strip().lower():
            return abbr
    return None

def normalise_size(raw):
    m = re.search(r'(\d{3}/\d{2}R\d{2}|\d{3}R\d{2})', raw, re.IGNORECASE)
    return m.group(1).upper() if m else raw.strip().split()[0].upper() if raw.strip() else ""

def kw_match(desc, keywords):
    d = desc.lower()
    return any(k.lower() in d for k in keywords)

def load_csv(path):
    with open(path, encoding="latin-1") as f:
        rows = list(csv.DictReader(f))
    # normalise header keys to uppercase so lookups work regardless of case
    return [{k.upper(): v for k, v in r.items()} for r in rows]

# ── Chunk 2: Series extraction + lookup builders + matchers ───────────────────

_LEAD_RE = re.compile(
    r'^(C\s+)?'
    r'(\d{3}/\d{2}R\d{2,3}C?\s+)?'
    r'(\d{3}R\d{2,3}\s+)?'
    r'(\d{2,3}(/\d{2,3})?[A-Za-z]{1,2}\s+)?'
    r'(XL\s+)?',
    re.IGNORECASE,
)

def extract_series(desc):
    if not desc:
        return ""
    s = _LEAD_RE.sub("", desc.strip())
    for w in s.split():
        if len(w) <= 1:
            continue
        if re.match(r'^[\d/]+[A-Za-z]{0,2}$', w):
            continue
        return w.upper()
    return ""

def build_tempe_lookup(rows):
    lk = {}
    for r in rows:
        brand = (r.get("BRAND") or r.get("brand","")).strip()
        abbr  = abbr_for(brand)
        if not abbr:
            continue
        size  = r.get("SIZE","").strip() or normalise_size(r.get("DESCRIPTION",""))
        desc  = r.get("DESCRIPTION","").strip()
        cost  = r.get("COST","").strip()
        price = r.get("PRICE","").strip()
        lk.setdefault((size, abbr), []).append((
            desc,
            float(cost)  if cost  else None,
            float(price) if price else None,
        ))
    return lk

def build_bj_lookup(rows):
    lk = {}
    for r in rows:
        brand = (r.get("BRAND") or r.get("brand","")).strip()
        abbr  = abbr_for(brand)
        if not abbr:
            continue
        size  = r.get("SIZE","").strip() or normalise_size(r.get("DESCRIPTION",""))
        desc  = r.get("DESCRIPTION","").strip()
        price = r.get("PRICE","").strip()
        disc  = r.get("DISC_PRICE","").strip()
        lk.setdefault((size, abbr), []).append((
            desc,
            float(price) if price else None,
            float(disc)  if disc  else None,
            r.get("PROMO","").strip(),
            r.get("SAVE_TEXT","").strip(),
        ))
    return lk

build_jax_lookup = build_bj_lookup   # JAX CSV has same columns (SAVE_TEXT absent → "")

def best_tempe(size, abbr, t_lk):
    cands = t_lk.get((size, abbr), [])
    cat   = SIZE_CATEGORY.get(size)
    if cat:
        kws = COMPETITOR_PATTERNS.get(cat, {}).get(abbr, [])
        pool = [x for x in cands if kw_match(x[0], kws)] if kws else cands
    else:
        pool = cands
    if not pool:
        return None, None, None
    best = sorted(pool, key=lambda x: (x[2] is None, x[2] or 0))[0]
    return best[0], best[1], best[2]

def _best_competitor(size, abbr, lk, t_desc):
    cands = lk.get((size, abbr), [])
    if not cands:
        return None, None, None, "", ""
    if not t_desc:
        # No Tempe reference — return cheapest available product
        best = sorted(cands, key=lambda x: (x[1] is None, x[1] or 0))[0]
        return best
    s = _LEAD_RE.sub("", t_desc.strip())
    for raw_w in s.split():
        w = re.sub(r'[().,+]', '', raw_w)
        if len(w) <= 1:
            continue
        if re.match(r'^[\d/]+[A-Za-z]{0,2}$', w):
            continue
        if re.match(r'^\d+[A-Za-z]{1,2}$', w):
            continue
        if w.upper() in ('XL','OWT','RWL','OBL','TL','TT','FR','MFS',
                         'DOT','BMW','AUDI','BENZ','DEMO','EV','RC'):
            continue
        matched = [x for x in cands if w.lower() in x[0].lower()]
        if matched:
            return sorted(matched, key=lambda x: (x[1] is None, x[1] or 0))[0]
    return None, None, None, "", ""

def best_bj(size, abbr, bj_lk, t_desc=None):
    return _best_competitor(size, abbr, bj_lk, t_desc)

def best_jax(size, abbr, jax_lk, t_desc=None):
    return _best_competitor(size, abbr, jax_lk, t_desc)

# Tempe WebOrder CSV columns: SIZE, brand, DESCRIPTION, "Sell in Price", "Sell out Price"
def build_tw_lookup(rows):
    lk = {}
    for r in rows:
        brand = r.get("brand","").strip()
        abbr  = abbr_for(brand)
        if not abbr:
            continue
        size  = r.get("SIZE","").strip() or normalise_size(r.get("DESCRIPTION",""))
        desc  = r.get("DESCRIPTION","").strip()
        # Support both old column names (COST/PRICE) and new (Sell in Price / Sell out Price)
        cost_raw  = (r.get("Sell in Price","")  or r.get("COST","")).strip()
        price_raw = (r.get("Sell out Price","") or r.get("PRICE","")).strip()
        try:
            cost_val  = float(cost_raw)  if cost_raw  else None
            price_val = float(price_raw) if price_raw else None
        except ValueError:
            continue
        # Keep row if either sell-in or sell-out is positive
        if (not cost_val or cost_val <= 0) and (not price_val or price_val <= 0):
            continue
        lk.setdefault((size, abbr), []).append((desc, cost_val, price_val))
    return lk

def best_tw(size, abbr, tw_lk):
    """Returns (desc, sell_in_price, sell_out_price)."""
    return best_tempe(size, abbr, tw_lk)

def best_twi(size, abbr, tw_lk):
    """Returns (desc, None, sell_in_price) — for the Sell In line."""
    items = tw_lk.get((size, abbr))
    if not items:
        return "", None, None
    desc, cost_val, _price_val = items[0]
    return desc, None, cost_val

# ── Chunk 3: Sheet 1 — Summary (9 cols/brand) ─────────────────────────────────

def sheet_summary(wb, t_rows, bj_rows, jax_rows):
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "D4"

    t_lk   = build_tempe_lookup(t_rows)
    bj_lk  = build_bj_lookup(bj_rows)
    jax_lk = build_jax_lookup(jax_rows)

    brand_abbrs  = list(BRANDS.keys())
    N            = len(brand_abbrs)
    COLS_PER_BRAND = 9   # T.Product|T.Cost|T.Price|BJ Product|BJ Price|BJ Disc|JAX Product|JAX Price|JAX Disc
    total_cols   = 3 + N * COLS_PER_BRAND

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1,
                value="Tempe vs Bob Jane vs JAX — Brand Price Comparison  "
                      "(T.Product|T.Cost|T.Price | BJ Product|BJ Price|BJ Disc | JAX Product|JAX Price|JAX Disc)")
    t.font      = Font(bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    hdr_cell(ws, 2, 1, "Size",     bg="2E4057")
    hdr_cell(ws, 2, 2, "Pattern",  bg="2E4057")
    hdr_cell(ws, 2, 3, "Category", bg="2E4057")
    col = 4
    for abbr in brand_abbrs:
        ws.merge_cells(start_row=2, start_column=col,
                       end_row=2, end_column=col + COLS_PER_BRAND - 1)
        hdr_cell(ws, 2, col, f"{abbr}  {BRANDS[abbr]}",
                 bg=BRAND_COLOURS.get(abbr, "555555"))
        for ci in range(1, COLS_PER_BRAND):
            ws.cell(row=2, column=col+ci).fill = PatternFill(
                "solid", fgColor=BRAND_COLOURS.get(abbr, "555555"))
        col += COLS_PER_BRAND

    hdr_cell(ws, 3, 1, "Size",     bg="334155", fg="CCCCCC", bold=False)
    hdr_cell(ws, 3, 2, "Pattern",  bg="334155", fg="CCCCCC", bold=False)
    hdr_cell(ws, 3, 3, "Category", bg="334155", fg="CCCCCC", bold=False)
    col = 4
    SUB = ["T.Product","T.Cost","T.Price","BJ Product","BJ Price","BJ Disc",
           "JAX Product","JAX Price","JAX Disc"]
    for _ in brand_abbrs:
        for sh in SUB:
            hdr_cell(ws, 3, col, sh, bg="334155", fg="AAAAAA", bold=False)
            col += 1

    sizes_ordered = list(SIZE_CATEGORY.keys())
    csv_sizes = sorted({r.get("SIZE","").strip() or normalise_size(r.get("DESCRIPTION",""))
                        for r in t_rows})
    for s in csv_sizes:
        if s not in sizes_ordered:
            sizes_ordered.append(s)

    row_num   = 4
    prev_line = None
    for size in sizes_ordered:
        cat_info = SIZE_CATEGORY.get(size)
        line, category, kumho_pat = cat_info if cat_info else ("?","?","?")

        if prev_line and line != prev_line:
            for ci in range(1, total_cols + 1):
                ws.cell(row=row_num, column=ci).fill = PatternFill("solid", fgColor="D9D9D9")
            row_num += 1
        prev_line = line

        bg = "F7F9FC" if row_num % 2 == 0 else "FFFFFF"
        dc(ws, row_num, 1, size,                 bg=bg, align="center")
        dc(ws, row_num, 2, kumho_pat,            bg=bg, align="center")
        dc(ws, row_num, 3, f"{line}/{category}", bg=bg, align="center")

        col = 4
        for abbr in brand_abbrs:
            bbg = ROW_FILLS.get(abbr, "FFFFFF")

            t_desc,   t_cost,   t_price  = best_tempe(size, abbr, t_lk)
            bj_desc,  bj_price, bj_disc, _, _ = best_bj(size, abbr, bj_lk, t_desc)
            jax_desc, jax_price,jax_disc, _, _ = best_jax(size, abbr, jax_lk, t_desc)

            if t_desc or bj_desc or jax_desc:
                dc(ws, row_num, col,   t_desc   or "—", bg=bbg)
                dc(ws, row_num, col+1, t_cost,           bg=bbg, align="right", num_fmt="$#,##0.00")
                dc(ws, row_num, col+2, t_price,          bg=bbg, align="right", num_fmt="$#,##0.00")
                dc(ws, row_num, col+3, bj_desc  or "—", bg=bbg)
                dc(ws, row_num, col+4, bj_price,         bg=bbg, align="right", num_fmt="$#,##0.00")
                dc(ws, row_num, col+5, bj_disc,          bg=bbg, align="right", num_fmt="$#,##0.00")
                dc(ws, row_num, col+6, jax_desc or "—", bg=bbg)
                dc(ws, row_num, col+7, jax_price,        bg=bbg, align="right", num_fmt="$#,##0.00")
                dc(ws, row_num, col+8, jax_disc,         bg=bbg, align="right", num_fmt="$#,##0.00")
            else:
                for ci in range(COLS_PER_BRAND):
                    dc(ws, row_num, col+ci, "—", bg="F0F0F0", align="center")
            col += COLS_PER_BRAND

        row_num += 1

    autofit(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    col = 4
    for _ in brand_abbrs:
        ws.column_dimensions[get_column_letter(col)].width   = 32
        ws.column_dimensions[get_column_letter(col+3)].width = 32
        ws.column_dimensions[get_column_letter(col+6)].width = 32
        col += COLS_PER_BRAND

# ── Chunk 4: Sheet 2 — BJ Products  /  Sheet 3 — JAX Products ────────────────

def _sheet_products(wb, rows, sheet_name, header_bg, has_save_text=True):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"

    headers = ["SIZE","Brand","Description","Price","Disc Price","Diff","Promo"]
    if has_save_text:
        headers.append("Save Text")
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h, bg=header_bg)

    sorted_rows = sorted(rows, key=lambda r: (
        r.get("SIZE",""), r.get("brand",""),
        float(r.get("PRICE","0") or 0)
    ))
    for i, r in enumerate(sorted_rows):
        rn    = i + 2
        price = r.get("PRICE","").strip()
        disc  = r.get("DISC_PRICE","").strip()
        abbr  = abbr_for(r.get("brand","").strip())
        bg    = ROW_FILLS.get(abbr, "F9F9F9") if abbr else "F9F9F9"

        price_f = float(price) if price else None
        disc_f  = float(disc)  if disc  else None
        diff    = round(disc_f - price_f, 2) if (price_f and disc_f) else None

        dc(ws, rn, 1, r.get("SIZE",""),        bg=bg, align="center")
        dc(ws, rn, 2, r.get("brand",""),       bg=bg)
        dc(ws, rn, 3, r.get("DESCRIPTION",""), bg=bg)
        dc(ws, rn, 4, price_f, bg=bg, align="right", num_fmt="$#,##0.00")
        dc(ws, rn, 5, disc_f,  bg=bg, align="right", num_fmt="$#,##0.00")
        dc(ws, rn, 6, diff,    bg=bg, align="right", num_fmt="$#,##0.00")
        dc(ws, rn, 7, r.get("PROMO",""), bg=bg)
        if has_save_text:
            dc(ws, rn, 8, r.get("SAVE_TEXT",""), bg=bg)

    last = len(sorted_rows) + 1
    ws.conditional_formatting.add(
        f"D2:D{last}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B")
    )
    autofit(ws)
    ws.column_dimensions["C"].width = 42

def sheet_bj_products(wb, bj_rows):
    _sheet_products(wb, bj_rows, "BJ Products",  "1F5C2E", has_save_text=True)

def sheet_jax_products(wb, jax_rows):
    _sheet_products(wb, jax_rows, "JAX Products", "1A4A6E", has_save_text=False)

# ── Chunk 5: Sheet 4 — Match Detail + main() ─────────────────────────────────

def sheet_match(wb, t_rows, bj_rows, jax_rows):
    ws = wb.create_sheet("Match Detail")
    ws.freeze_panes = "A3"

    headers = [
        "Size","Line","Category","Pattern",
        "Brand","T.Product","T.Cost","T.Price",
        "BJ Product","BJ Price","BJ Disc","BJ Promo",
        "BJ vs T.Price","BJ Disc vs T.Price",
        "JAX Product","JAX Price","JAX Disc","JAX Promo",
        "JAX vs T.Price","JAX Disc vs T.Price",
    ]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h, bg="1F4E79")
    ws.row_dimensions[1].height = 28

    t_lk   = build_tempe_lookup(t_rows)
    bj_lk  = build_bj_lookup(bj_rows)
    jax_lk = build_jax_lookup(jax_rows)

    sizes_ordered = list(SIZE_CATEGORY.keys())
    csv_sizes = sorted({r.get("SIZE","").strip() or normalise_size(r.get("DESCRIPTION",""))
                        for r in t_rows})
    for s in csv_sizes:
        if s not in sizes_ordered:
            sizes_ordered.append(s)

    row_num = 2
    for size in sizes_ordered:
        cat_info = SIZE_CATEGORY.get(size)
        line, category, kumho_pat = cat_info if cat_info else ("?","?","?")

        for abbr in BRANDS:
            bg = ROW_FILLS.get(abbr, "F9F9F9")

            t_desc,   t_cost,   t_price  = best_tempe(size, abbr, t_lk)
            bj_desc,  bj_price, bj_disc,  bj_promo,  _ = best_bj(size, abbr, bj_lk, t_desc)
            jax_desc, jax_price,jax_disc, jax_promo, _ = best_jax(size, abbr, jax_lk, t_desc)

            if not t_desc and not bj_desc and not jax_desc:
                continue

            bj_diff_p  = round(bj_price  - t_price, 2) if (bj_price  and t_price) else None
            bj_diff_d  = round(bj_disc   - t_price, 2) if (bj_disc   and t_price) else None
            jax_diff_p = round(jax_price - t_price, 2) if (jax_price and t_price) else None
            jax_diff_d = round(jax_disc  - t_price, 2) if (jax_disc  and t_price) else None

            dc(ws, row_num,  1, size,              bg=bg, align="center")
            dc(ws, row_num,  2, line,              bg=bg, align="center")
            dc(ws, row_num,  3, category,          bg=bg, align="center")
            dc(ws, row_num,  4, kumho_pat,         bg=bg, align="center")
            dc(ws, row_num,  5, f"{abbr} {BRANDS[abbr]}", bg=bg)
            dc(ws, row_num,  6, t_desc   or "—",  bg=bg)
            dc(ws, row_num,  7, t_cost,  bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num,  8, t_price, bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num,  9, bj_desc  or "—",  bg=bg)
            dc(ws, row_num, 10, bj_price,  bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 11, bj_disc,   bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 12, bj_promo,  bg=bg)
            dc(ws, row_num, 13, bj_diff_p, bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 14, bj_diff_d, bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 15, jax_desc or "—",  bg=bg)
            dc(ws, row_num, 16, jax_price,  bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 17, jax_disc,   bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 18, jax_promo,  bg=bg)
            dc(ws, row_num, 19, jax_diff_p, bg=bg, align="right", num_fmt="$#,##0.00")
            dc(ws, row_num, 20, jax_diff_d, bg=bg, align="right", num_fmt="$#,##0.00")
            row_num += 1

        for ci in range(1, 21):
            ws.cell(row=row_num, column=ci).fill = PatternFill("solid", fgColor="E8E8E8")
        row_num += 1

    autofit(ws)
    ws.column_dimensions["F"].width  = 38
    ws.column_dimensions["I"].width  = 38
    ws.column_dimensions["O"].width  = 38


def main():
    t_files   = sorted(glob.glob("Tempe_*.csv"),   reverse=True)
    bj_files  = sorted(glob.glob("BobJane_*.csv"), reverse=True)
    jax_files = sorted(glob.glob("JAX_*.csv"),     reverse=True)

    if not t_files:
        print("ERROR: No Tempe_*.csv found."); return
    if not bj_files:
        print("ERROR: No BobJane_*.csv found."); return
    if not jax_files:
        print("ERROR: No JAX_*.csv found."); return

    t_path   = t_files[0]
    bj_path  = bj_files[0]
    jax_path = jax_files[0]
    print(f"Tempe:   {t_path}")
    print(f"BobJane: {bj_path}")
    print(f"JAX:     {jax_path}")

    t_rows   = load_csv(t_path)
    bj_rows  = load_csv(bj_path)
    jax_rows = load_csv(jax_path)
    print(f"  Tempe rows:   {len(t_rows)}")
    print(f"  BobJane rows: {len(bj_rows)}")
    print(f"  JAX rows:     {len(jax_rows)}")

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, t_rows, bj_rows, jax_rows)
    sheet_bj_products(wb, bj_rows)
    sheet_jax_products(wb, jax_rows)
    sheet_match(wb, t_rows, bj_rows, jax_rows)

    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    out = f"JAX_Comparison_{ts}.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
