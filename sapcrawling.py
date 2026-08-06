import os
import time
import glob
import csv
import shutil
import win32com.client
from datetime import datetime, timedelta
from openpyxl import load_workbook

"""
ZSDM64300 "Daily stock from 3PL" auto-export.

Replaces the old MB52 crawl.  ZSDM64300 exposes DOT No. per pallet so
downstream code can compute stock aging.  This script:

  1. Loops over every plant in PLANTS (one call to ZSDM64300 each)
  2. Sets Interface Date to yesterday (SAP stamps the file overnight,
     so today's file usually isn't there yet in the morning)
  3. Selects "Display 3PL Interface Data" radio
  4. Executes → exports the ALV grid → SAP writes EXPORT_*.xlsx to C:\\temp
  5. Renames the per-plant XLSX to rawdata/unlock/dot_stock_<PLANT>.xlsx
  6. After the loop finishes, combines all four XLSXs into one CSV at
     rawdata/unlock/dot_stock.csv  (columns: Interface Date, Plant,
     Material, DOT No., Stock Qty — the yellow columns only, everything
     else in the SAP grid is ignored)

load_csv_to_mysql.py picks up rawdata/unlock/dot_stock.csv and loads
into the `stock` table.

Pre-req:
  - SAP GUI open AND logged in
  - pywin32, openpyxl

If the SAP field IDs on the ZSDM64300 selection screen differ from what
this script assumes, edit the ZSDM_FIELDS block below.  Get the correct
IDs by opening the selection screen and pressing F1 on each field →
Technical Info → "Screen Field".
"""

# ================= CONFIG =================
PLANTS = ["42R0", "42R1", "42R2", "42R4"]           # order run in
TRANSACTION_CODE = "ZSDM64300"

SAP_EXPORT_DIR  = r"C:\temp"                        # where EXPORT_*.xlsx appears
SAP_EXPORT_GLOB = "EXPORT_*.xlsx"

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
UNLOCK_DIR = os.path.join(BASE_DIR, "rawdata", "unlock")
COMBINED_CSV = os.path.join(UNLOCK_DIR, "dot_stock.csv")

DELETE_XLSX_AFTER_COMBINE = False                   # keep per-plant XLSX
MIN_CSV_SIZE = 200

# When True, dumps everything we can see on the result screen the first
# time export fails so it's clear which grid ID / menu path / okcd works
# on THIS SAP layout.  Log lands in the console; paste it here and we
# can pin the right IDs into the config permanently.
DEBUG_DUMP_ON_EXPORT_FAIL = True

# SAP GUI field IDs for the ZSDM64300 selection screen.  Adjust here if
# your SAP renders different names.  Every field is tried with .Text
# assignment — if the ID doesn't exist we skip it.
ZSDM_FIELDS = {
    "plant":              ["wnd[0]/usr/ctxtP_WERKS",
                            "wnd[0]/usr/ctxtS_WERKS-LOW",
                            "wnd[0]/usr/ctxtWERKS-LOW"],
    # Confirmed via Technical Information popup on the live SAP
    # layout (screen 1000 of ZSDM64300): Dynpro Field is
    # S_IFDAT-LOW.  Old guesses (S_DATE-LOW / P_DATE / S_INTFDT-LOW)
    # never matched — every prior crawl silently fell back to
    # SAP's default (today's data).
    "interface_date":     ["wnd[0]/usr/ctxtS_IFDAT-LOW",
                            "wnd[0]/usr/ctxtS_DATE-LOW",
                            "wnd[0]/usr/ctxtP_DATE",
                            "wnd[0]/usr/ctxtS_INTFDT-LOW"],
    # Radio "Display 3PL Interface Data" — usually a P_RADIO field.
    # Selection screen might already default to it; we still click.
    "display_interface":  ["wnd[0]/usr/radP_DISP",
                            "wnd[0]/usr/radP_INT"],
}

# Yellow columns from the ZSDM64300 ALV that we actually keep.  Header
# text is matched case-insensitively; if the SAP layout renames these
# in a future release, update here.
KEEP_COLUMNS = {
    "interface date": "interface_date",
    "plant":          "plant",
    "material":       "material",
    "dot no":         "dot_no",
    "dot no.":        "dot_no",
    "stock qty":      "stock_qty",
}
OUTPUT_HEADER = ["interface_date", "plant", "material", "dot_no", "stock_qty"]


# ================= HELPERS =================
def wait(sec=0.3):
    time.sleep(sec)

def ensure_dir(d):
    if d and not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)

def exists(session, wid):
    try:
        session.findById(wid)
        return True
    except:
        return False

def close_popups(session, max_steps=6):
    for _ in range(max_steps):
        if not exists(session, "wnd[1]"):
            return
        try:
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            wait(0.25); continue
        except: pass
        try:
            session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
            wait(0.25); continue
        except: pass
        try:
            session.findById("wnd[1]").sendVKey(0)
            wait(0.25); continue
        except: pass
        return


def _try_set_text(session, id_candidates, value):
    """Set .Text on the first field ID in the list that exists."""
    for wid in id_candidates:
        if exists(session, wid):
            try:
                session.findById(wid).Text = value
                return True
            except: pass
    return False


def _try_set_radio(session, id_candidates):
    """Turn on the first radio button that exists."""
    for wid in id_candidates:
        if exists(session, wid):
            try:
                session.findById(wid).select()
                return True
            except: pass
    return False


# ================= SAP driving =================
def start_transaction_fresh(session):
    session.findById("wnd[0]/tbar[0]/okcd").Text = "/n" + TRANSACTION_CODE
    session.findById("wnd[0]").sendVKey(0)
    wait(1.0)
    close_popups(session, 4)


def set_selection_screen(session, plant, interface_date_ddmmyyyy):
    _try_set_text(session, ZSDM_FIELDS["plant"], plant)
    _try_set_text(session, ZSDM_FIELDS["interface_date"], interface_date_ddmmyyyy)
    _try_set_radio(session, ZSDM_FIELDS["display_interface"])
    session.findById("wnd[0]").sendVKey(0)
    wait(0.4)
    close_popups(session, 3)


def newest_export_xlsx(after_ts: float) -> str:
    pattern = os.path.join(SAP_EXPORT_DIR, SAP_EXPORT_GLOB)
    candidates = []
    for p in glob.glob(pattern):
        try:
            mtime = os.path.getmtime(p)
            if mtime >= after_ts:
                candidates.append((mtime, p))
        except: pass
    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][1]


def try_export_via_alv_toolbar(session) -> bool:
    grid_ids = [
        # ZSDM64300 uses this specific container name — captured from
        # the on-screen dump.  Kept first so we hit it before probing
        # the generic MB52-style IDs below.
        "wnd[0]/usr/cntlG_CONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlGRID1/shellcont/shell",
        "wnd[0]/usr/cntlCONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlALV_CONTAINER/shellcont/shell",
        "wnd[0]/usr/cntlGRID1/shellcont/shellcont/shell",
    ]
    for gid in grid_ids:
        if not exists(session, gid):
            continue
        try:
            grid = session.findById(gid)
            opened = False
            for export_btn in ("&MB_EXPORT", "EXPORT", "&EXPORT"):
                try:
                    grid.pressToolbarContextButton(export_btn)
                    wait(0.4); opened = True; break
                except: continue
            if not opened:
                continue
            for item in ("&XXL", "XXL", "&SPREADSHEET", "SPREADSHEET", "&PC", "PC"):
                try:
                    grid.selectContextMenuItem(item)
                    wait(0.8); close_popups(session, 4)
                    return True
                except: continue
        except: continue
    return False


def try_export_via_menu(session) -> bool:
    """Fallback: List → Export → Spreadsheet (menu text lookup)."""
    try:
        mbar = session.findById("wnd[0]/mbar")
    except:
        return False
    target_path = ["List", "Export", "Spreadsheet"]
    level = [mbar.Children(i) for i in range(mbar.Children.Count)]
    for wanted in target_path:
        found = None
        for m in level:
            try:
                if m.Text.strip().lower() == wanted.lower():
                    found = m; break
            except: pass
        if not found: return False
        try:
            level = [found.Children(i) for i in range(found.Children.Count)]
        except:
            try:
                found.select(); wait(0.6); close_popups(session, 4)
                return True
            except:
                return False
    try:
        level[0].select(); wait(0.6); close_popups(session, 4)
        return True
    except:
        return False


def try_export_via_okcd(session) -> bool:
    """SAP has a hidden command '%pc' that opens the Local File save
    dialog on any ALV list.  Works even when the grid has no toolbar."""
    for cmd in ("%pc", "%%pc"):
        try:
            session.findById("wnd[0]/tbar[0]/okcd").Text = cmd
            session.findById("wnd[0]").sendVKey(0)
            wait(0.8)
            # Popup should ask for format; pick Spreadsheet
            if exists(session, "wnd[1]"):
                # Try radio "Spreadsheet"
                for rid in ("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]",
                            "wnd[1]/usr/radRB_SPRE",
                            "wnd[1]/usr/radR_XLS",
                            "wnd[1]/usr/radSPRE",
                            "wnd[1]/usr/radR_SPRE"):
                    if exists(session, rid):
                        try: session.findById(rid).select(); break
                        except: pass
                # Confirm
                try:
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    wait(0.6)
                except: pass
                # Save-file dialog will appear next (SAP will still
                # write the EXPORT_*.xlsx to C:\temp for our poller).
                close_popups(session, 4)
                return True
        except: pass
    return False


def try_export_via_toolbar_button(session) -> bool:
    """Some Z-transactions expose 'Local File' as a top-toolbar button
    with id btn[43], btn[45], or via a Shift+F9 key combo."""
    for vk in (45,):   # Ctrl+Shift+F9 → SAP standard "Local File" list menu
        try:
            session.findById("wnd[0]").sendVKey(vk)
            wait(0.6)
            if exists(session, "wnd[1]"):
                close_popups(session, 4)
                return True
        except: pass
    return False


def dump_screen(session):
    """Print everything we can see on the current SAP window.  Used
    once when export fails so the operator can share the output back
    to figure out the right IDs for this SAP layout."""
    print("  ==== SCREEN DUMP ====")
    def _walk(elem, depth=0):
        try:
            eid = elem.Id
        except: eid = "?"
        try:
            etype = elem.Type
        except: etype = "?"
        try:
            etext = (elem.Text or "").strip()[:40]
        except: etext = ""
        indent = "  " * depth
        print(f"    {indent}[{etype}] {eid}  {etext!r}")
        try:
            for i in range(elem.Children.Count):
                _walk(elem.Children(i), depth + 1)
        except: pass
    try:
        _walk(session.findById("wnd[0]"))
    except Exception as e:
        print(f"    dump failed: {e}")
    print("  ==== END DUMP ====")


def try_export_via_system_menu(session) -> bool:
    """SAP's universal 'System → List → Save → Local File' path.  On
    ZSDM64300 the dump proved this is at wnd[0]/mbar/menu[0]/menu[5]/
    menu[2]/menu[2] — hard-coded so we don't rely on Text-lookup that
    can fail if the label is translated or spaced differently."""
    for path in (
        "wnd[0]/mbar/menu[0]/menu[5]/menu[2]/menu[2]",  # ZSDM64300 layout
        "wnd[0]/mbar/menu[0]/menu[3]/menu[2]/menu[2]",  # slight variation
    ):
        if not exists(session, path):
            continue
        try:
            session.findById(path).select()
            wait(0.6)
        except:
            continue
        # Format popup — pick "Spreadsheet" (radio index 1).
        for rid in (
            "wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]",
            "wnd[1]/usr/radRB_SPRE",
            "wnd[1]/usr/radR_XLS",
        ):
            if exists(session, rid):
                try: session.findById(rid).select()
                except: pass
                break
        # Confirm the format dialog (green tick / Enter)
        try:
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            wait(0.6)
        except: pass
        # Now the Save-As dialog.  Fill the path + filename so the file
        # lands where the poller looks (C:\temp\EXPORT_ZSDM.xlsx).
        if exists(session, "wnd[1]"):
            try:
                if exists(session, "wnd[1]/usr/ctxtDY_PATH"):
                    session.findById("wnd[1]/usr/ctxtDY_PATH").Text = SAP_EXPORT_DIR + "\\"
                if exists(session, "wnd[1]/usr/ctxtDY_FILENAME"):
                    session.findById("wnd[1]/usr/ctxtDY_FILENAME").Text = "EXPORT_ZSDM.xlsx"
            except: pass
            # 'Replace' if a same-named file already exists — the poller
            # picks by mtime so the newest wins.
            for sid in ("wnd[1]/tbar[0]/btn[11]",   # Save
                        "wnd[1]/tbar[0]/btn[0]"):    # Continue / OK
                try:
                    session.findById(sid).press()
                    wait(0.8)
                    break
                except: pass
            close_popups(session, 4)
        return True
    return False


_DEBUG_DUMPED = False   # dump once per script run
def trigger_export(session) -> bool:
    global _DEBUG_DUMPED
    # Order: try the hard-coded ZSDM64300 System-menu path first (we
    # know it works from the dump), then fall back to the generic
    # strategies for other transactions.
    if try_export_via_system_menu(session):    return True
    if try_export_via_alv_toolbar(session):    return True
    if try_export_via_menu(session):           return True
    if try_export_via_okcd(session):           return True
    if try_export_via_toolbar_button(session): return True
    if DEBUG_DUMP_ON_EXPORT_FAIL and not _DEBUG_DUMPED:
        _DEBUG_DUMPED = True
        dump_screen(session)
    return False


def go_back_to_selection(session):
    """F3 twice tends to land back on the selection screen from anywhere."""
    for _ in range(3):
        try:
            session.findById("wnd[0]/tbar[0]/btn[3]").press()
            wait(0.3); close_popups(session, 3)
        except:
            break


# ================= XLSX → row extract =================
def read_xlsx_rows(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v).strip() for v in row])
    # drop trailing blank rows
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    return rows


def _norm_hdr(s):
    s = (s or "").strip().lower().replace("_", " ").replace(".", "")
    return " ".join(s.split())


def extract_keep_rows(xlsx_path: str):
    """From the raw ZSDM64300 grid, keep only the 5 output columns and
    return them in OUTPUT_HEADER order.  Skips summary rows (blank
    Material) and header repeats."""
    all_rows = read_xlsx_rows(xlsx_path)
    if not all_rows:
        return []
    header = all_rows[0]
    # Locate each keep-column index by header text
    col_idx = {}
    for i, h in enumerate(header):
        key = KEEP_COLUMNS.get(_norm_hdr(h))
        if key and key not in col_idx:
            col_idx[key] = i
    missing = [k for k in OUTPUT_HEADER if k not in col_idx]
    if missing:
        print(f"  [WARN] {os.path.basename(xlsx_path)}: missing columns {missing}. "
              f"Available headers: {[_norm_hdr(h) for h in header]}")
    out = []
    for r in all_rows[1:]:
        row = []
        keep = True
        for k in OUTPUT_HEADER:
            i = col_idx.get(k, -1)
            val = r[i].strip() if 0 <= i < len(r) else ""
            row.append(val)
        # Drop rows without a material or stock qty (totals / blanks)
        if not row[OUTPUT_HEADER.index("material")]:
            keep = False
        if keep:
            out.append(row)
    return out


def combine_to_csv(xlsx_paths, csv_path):
    ensure_dir(os.path.dirname(csv_path))
    total = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(OUTPUT_HEADER)
        for p in xlsx_paths:
            if not p or not os.path.exists(p):
                continue
            rows = extract_keep_rows(p)
            for r in rows:
                w.writerow(r)
            total += len(rows)
            print(f"  {os.path.basename(p)} → {len(rows):,} rows")
    print(f"  Combined CSV: {csv_path}  ({total:,} rows)")


# ================= MAIN =================
def yesterday_ddmmyyyy():
    return (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")


def main():
    ensure_dir(UNLOCK_DIR)
    interface_date = yesterday_ddmmyyyy()
    print(f"ZSDM64300 crawl · plants={PLANTS} · interface date={interface_date}")

    sap = win32com.client.GetObject("SAPGUI").GetScriptingEngine
    conn = sap.Children(0)
    session = conn.Children(0)
    try:
        session.findById("wnd[0]").maximize()
    except: pass

    xlsx_files = []
    for plant in PLANTS:
        print(f"\n── {plant} ──")
        start_transaction_fresh(session)
        set_selection_screen(session, plant, interface_date)

        export_start_ts = time.time()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()   # F8 execute
        wait(2.0); close_popups(session, 4)

        if not trigger_export(session):
            print(f"  [ERROR] {plant}: could not trigger export"); continue

        xlsx_path = ""
        for _ in range(50):
            xlsx_path = newest_export_xlsx(export_start_ts)
            if xlsx_path and os.path.exists(xlsx_path):
                s1 = os.path.getsize(xlsx_path); wait(0.8)
                s2 = os.path.getsize(xlsx_path)
                if s2 >= s1 and s2 > 500:
                    break
            wait(0.6)
        if not xlsx_path:
            print(f"  [ERROR] {plant}: SAP export XLSX not found in {SAP_EXPORT_DIR}")
            go_back_to_selection(session); continue

        dest = os.path.join(UNLOCK_DIR, f"dot_stock_{plant}.xlsx")
        shutil.copy2(xlsx_path, dest)
        print(f"  {plant} XLSX → {dest}")
        xlsx_files.append(dest)

        go_back_to_selection(session)

    print("\n── Combine to CSV ──")
    combine_to_csv(xlsx_files, COMBINED_CSV)

    if not os.path.exists(COMBINED_CSV) or os.path.getsize(COMBINED_CSV) < MIN_CSV_SIZE:
        raise RuntimeError("Combined CSV empty or too small.")
    print(f"\nDone.  Run load_csv_to_mysql.py next to push into MySQL.")

    if DELETE_XLSX_AFTER_COMBINE:
        for p in xlsx_files:
            try: os.remove(p)
            except: pass


if __name__ == "__main__":
    main()
