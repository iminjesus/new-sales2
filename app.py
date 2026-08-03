from flask import Flask, request, jsonify, send_file, send_from_directory, make_response, redirect
from werkzeug.middleware.proxy_fix import ProxyFix
import sqlite3
import mysql.connector
from mysql.connector import pooling
import threading
from time import time  # cache timestamps
import os
from flask_cors import CORS
from typing import Dict, List, Tuple, Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import traceback
USE_SQLITE = os.environ.get("USE_SQLITE") == "1"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# SA/TAS roll up to VIC, NT to WA, ACT to NSW.  Used both inside
# api_stock and by the claim portal's notification routing.
STATE_REMAP = {"SA": "VIC", "NT": "WA", "TAS": "VIC", "ACT": "NSW"}

# Pick up SMTP / DB creds from .env (gitignored).  Silent no-op if
# python-dotenv isn't installed — the platform env vars still apply.
try:
    from dotenv import load_dotenv as _load_dotenv  # type: ignore
    _load_dotenv(os.path.join(BASE_DIR, ".env"))
except ImportError:
    pass
SQLITE_PATH = os.path.join(BASE_DIR, "snapshot.db")

# MySQL connection pool (reduces connect() overhead per request)
_MYSQL_POOL = None
_MYSQL_POOL_LOCK = threading.Lock()


import sqlite3  # make sure this is at the top of app.py

class SQLiteCursorWrapper:
    def __init__(self, cursor):
        self._cursor = cursor
        self._empty_result = False  # flag for demo mode

    def execute(self, sql, params=None):
        # Replace MySQL-style "%s" with SQLite "?" placeholders
        if "%s" in sql:
            sql = sql.replace("%s", "?")

        self._empty_result = False
        try:
            if params is None:
                return self._cursor.execute(sql)
            return self._cursor.execute(sql, params)
        except sqlite3.OperationalError as e:
            # DEMO MODE: missing table or column ??pretend query returned nothing
            _e = str(e).lower()
            if "no such table" in _e or "no such column" in _e or "ambiguous column" in _e:
                print(f"[WARN] {e} -- returning empty result (demo mode)")
                self._empty_result = True
                return self
            raise

    def executemany(self, sql, seq_of_params):
        if "%s" in sql:
            sql = sql.replace("%s", "?")
        self._empty_result = False
        try:
            return self._cursor.executemany(sql, seq_of_params)
        except sqlite3.OperationalError as e:
            _e = str(e).lower()
            if "no such table" in _e or "no such column" in _e or "ambiguous column" in _e:
                print(f"[WARN] {e} -- ignoring executemany (demo mode)")
                self._empty_result = True
                return self
            raise

    def fetchall(self):
        if self._empty_result:
            return []  # no data instead of error
        rows = self._cursor.fetchall()
        return [dict(r) for r in rows]

    def fetchone(self):
        if self._empty_result:
            return None
        r = self._cursor.fetchone()
        return dict(r) if r is not None else None

    def __iter__(self):
        if self._empty_result:
            return iter([])
        for r in self._cursor:
            yield dict(r)

    def __getattr__(self, name):
        return getattr(self._cursor, name)


class SQLiteConnectionWrapper:
    def __init__(self, conn):
        self._conn = conn

    def cursor(self, *args, **kwargs):
        # ignore dictionary=True from mysql style
        kwargs.pop("dictionary", None)
        cur = self._conn.cursor(*args, **kwargs)
        return SQLiteCursorWrapper(cur)

    def __getattr__(self, name):
        return getattr(self._conn, name)


def _business_effective_ym():
    """Return (year, month) representing the "effective current month"
    for graph aggregations.

    Sales_thismonth (and, by extension, the nightly import job that
    pushes those rows into sales_2526) is populated *overnight* from
    the previous calendar day.  On the first business day of a new
    month the rows carry MONTH()=<current month> even though the
    underlying billing dates are all from the *prior* month — the
    Monthly Stacked chart then shows a phantom bar at the new month
    that shouldn't be there yet.

    Mirror the frontend's effectiveMonth() logic: if today is on or
    before the first weekday of the current month, treat the prior
    month as the effective current month."""
    from datetime import date, timedelta
    today = date.today()
    first = today.replace(day=1)
    while first.weekday() >= 5:   # 5=Saturday, 6=Sunday
        first += timedelta(days=1)
    if today <= first:
        # First business day (or before) — the current calendar month
        # is still populating overnight, so treat previous as current.
        return (today.year, 12) if today.month == 1 else (today.year, today.month - 1)
    return (today.year, today.month)


def parse_filters(req):
    """Uniform filter extraction."""
    return {
        "category":      (req.args.get("category") or "ALL").upper().strip(),
        "metric":        (req.args.get("metric") or "qty").lower().strip(),
        "year":          (req.args.get("year") or "").strip(),
        "month":         (req.args.get("month") or "").strip(),
        "region":        (req.args.get("region") or "ALL").strip(),
        "salesman":      (req.args.get("salesman") or "ALL").strip(),
        "channel":       (req.args.get("channel") or "ALL").strip(),
        "sold_to_group": (req.args.get("sold_to_group") or "ALL").strip(),
        "sold_to":       (req.args.get("sold_to") or "ALL").strip(),
        "ship_to":       (req.args.get("ship_to") or "ALL").strip(),
        "product_group": (req.args.get("product_group") or "ALL").strip(),
        "pattern":       (req.args.get("pattern") or "ALL").strip(),
        "material":       (req.args.get("material") or "ALL").strip(),
        "code":           (req.args.get("code")     or "ALL").strip(),
        # HK / LF (brand) — Product cascade layer between line and pg.
        "brand":         (req.args.get("brand") or "ALL").upper().strip(),
    }

# States that belong to each sales region.
# SA/TAS are part of VIC territory; NT is part of WA; ACT is part of NSW.
REGION_STATES = {
    "NSW": ["NSW", "ACT"],
    "QLD": ["QLD"],
    "VIC": ["VIC", "SA", "TAS"],
    "WA":  ["WA",  "NT"],
}

def build_customer_filters(alias_fact: str, f, *, use_sold_to_name: bool=False):
    """
    Returns (joins, wheres, params) to apply Region/Salesman/Group/Sold_to on a fact table.
    Customer JOIN is added only when needed (name-based filters or customer-dimension filters).
    Region/Group filters use EXISTS on customer (ship_to only) to avoid JOIN inflation.
    Salesman filter uses EXISTS on target_26 (authoritative bde?뭩hip_to mapping),
    because customer.salesman_name is often incomplete for WA/NT accounts.
    If use_sold_to_name=True, 'sold_to' will match customer.Sold_to_Name instead of id.
    """
    joins = []
    wh, p = [], []
    needs_cus = False   # only for name-based sold_to/ship_to lookups

    # ?? region: EXISTS on customer (ship_to only) ??no JOIN inflation ??
    if f["region"] != "ALL":
        states = REGION_STATES.get(f["region"].upper(), [f["region"]])
        ph = ",".join(["%s"] * len(states))
        wh.append(
            f"EXISTS (SELECT 1 FROM customer _cr"
            f" WHERE _cr.ship_to = {alias_fact}.ship_to"
            f" AND _cr.bde_state IN ({ph}))"
        )
        p.extend(states)

    # ?? salesman: EXISTS on target_26 (bde is authoritative; customer.salesman_name
    #    may be missing or inconsistent for some regions like WA) ??
    if f["salesman"] != "ALL":
        wh.append(
            f"EXISTS (SELECT 1 FROM target_26 _t"
            f" WHERE _t.ship_to = {alias_fact}.ship_to"
            f" AND UPPER(TRIM(_t.bde)) = UPPER(TRIM(%s)))"
        )
        p.append(f["salesman"])

    # ?? sold_to_group: EXISTS on customer (ship_to only) ??
    if f["sold_to_group"] != "ALL":
        wh.append(
            f"EXISTS (SELECT 1 FROM customer _cr"
            f" WHERE _cr.ship_to = {alias_fact}.ship_to"
            f" AND _cr.sold_to_group = %s)"
        )
        p.append(f["sold_to_group"])

    # Channel filter: EXISTS on customer (ship_to only), same shape as
    # sold_to_group above so a Channel pick narrows the fact table to
    # ship_tos whose customer-master row carries that channels value.
    if f.get("channel", "ALL") != "ALL":
        wh.append(
            f"EXISTS (SELECT 1 FROM customer _ch"
            f" WHERE _ch.ship_to = {alias_fact}.ship_to"
            f" AND TRIM(_ch.channels) = %s)"
        )
        p.append(f["channel"])

    # ?? sold_to: id ??direct filter on fact table; name ??subquery ??
    if f["sold_to"] != "ALL":
        sv = f["sold_to"]
        if not use_sold_to_name and (sv.isdigit() or sv.upper().startswith("A")):
            wh.append(f"{alias_fact}.sold_to = %s"); p.append(sv)
        else:
            wh.append(
                f"{alias_fact}.sold_to IN ("
                f"SELECT DISTINCT sold_to FROM customer WHERE sold_to_name = %s)"
            )
            p.append(sv)

    # ?? ship_to: code ??direct; name ??subquery (names now come as codes from frontend) ??
    if f["ship_to"] != "ALL":
        st = f["ship_to"].strip()
        if st.isdigit() or st.upper().startswith("A"):
            wh.append(f"{alias_fact}.ship_to = %s"); p.append(st)
        else:
            wh.append(
                f"{alias_fact}.ship_to IN ("
                f"SELECT DISTINCT ship_to FROM customer"
                f" WHERE UPPER(TRIM(ship_to_name)) = UPPER(TRIM(%s)))"
            )
            p.append(st)

    if needs_cus:
        joins.append(_customer_join(alias_fact))

    return joins, wh, p


def build_target_filters(alias: str, f):
    """
    Returns (joins, wheres, params) for target_26 table.
    target_26 has state, bde, ship_to, sold_to directly.
    When sold_to/ship_to is a name (not ID), joins customer table to resolve.
    """
    joins = []
    wh, p = [], []
    needs_cus = False

    if f["region"] != "ALL":
        states = REGION_STATES.get(f["region"].upper(), [f["region"]])
        if len(states) == 1:
            wh.append(f"{alias}.state = %s"); p.append(states[0])
        else:
            wh.append(f"{alias}.state IN ({','.join(['%s']*len(states))})"); p.extend(states)
    if f["salesman"] != "ALL":
        wh.append(f"UPPER(TRIM({alias}.bde)) = UPPER(TRIM(%s))"); p.append(f["salesman"])
    if f["sold_to_group"] != "ALL":
        needs_cus = True
        wh.append("cus.sold_to_group = %s"); p.append(f["sold_to_group"])
    if f["sold_to"] != "ALL":
        sv = f["sold_to"]
        if sv.isdigit() or sv.upper().startswith("A"):
            wh.append(f"{alias}.sold_to = %s"); p.append(sv)
        else:
            needs_cus = True
            wh.append("cus.sold_to_name = %s"); p.append(sv)
    if f["ship_to"] != "ALL":
        st = f["ship_to"].strip()
        if st.isdigit() or st.upper().startswith("A"):
            wh.append(f"{alias}.ship_to = %s"); p.append(st)
        else:
            needs_cus = True
            wh.append("UPPER(TRIM(cus.ship_to_name)) = UPPER(TRIM(%s))"); p.append(st)

    if needs_cus:
        joins.append(f"LEFT JOIN customer cus ON cus.ship_to = {alias}.ship_to")

    return joins, wh, p


def category_filters(alias: str, category: str):
    """
    Return (joins, wheres) for monthly-schema facts (sales2025, profit).
    - alias: table alias for the fact (e.g., "s" for sales2025, "p" for profit)
    - All predicates are index-friendly (equality / LIKE prefix).
    - Add optional JOINs only when that category needs them.
    """
    joins, wh = [], []
    cat = (category or "ALL").upper()

    if cat == "ALL":
        return joins, wh

    elif cat == "PCLT":
        # material codes starting with 1 or 2
        wh.append(f"{alias}.line = 'PCLT'")

    elif cat == "TBR":
        # example logic: material codes starting with 3 (adjust to your real rule)
        wh.append(f"{alias}.line = 'TBR'")

     # NEW: 18+ Inch means PCLT & inch > 18
    elif cat == "18PLUS":
        wh.append(f"{alias}.line = 'PCLT'")
        # inch is often stored as text; cast to numeric for safety
        wh.append(f"CAST({alias}.inch AS DECIMAL(10,2)) >= 18.0")
        
    elif cat == "ISEG":
        # ISEG mapping by Material
        # Ensure an index on iseg(Material)
        joins.append(f"JOIN iseg i ON cast(trim(i.Material) as unsigned) = {alias}.material")

    elif cat == "SUV":
        # SUV by Pattern
        # Ensure an index on suv(Pattern)
        joins.append(f"JOIN suv suv ON suv.Pattern = {alias}.pattern")

    elif cat == "LOWPROFILE":
        # Low profile / strategic by Material
        joins.append(f"JOIN lowprofile lp ON cast(trim(lp.Material) as unsigned) = {alias}.material")

    elif cat == "HM":
        # HM by Sold-To (use your customer join for Ship_To ??Sold_To; keep simple)
        # If your HM rule is customer-list based, prefer EXISTS against a keyed table.
        joins.append(f"JOIN hm hm ON hm.sold_to = {alias}.sold_to")

    elif cat == "443":
        wh.append(f"""
            EXISTS (
                SELECT 1
                FROM `443_25` p443
                WHERE p443.month         = {alias}.month
                AND p443.product_group = {alias}.product_group

            )
            """)

    return joins, wh


# ?? Helpers for normalised sales tables (line/product_group/pattern/inch live in carrying_26) ??

def _carrying_join(alias: str) -> str:
    """Returns the LEFT JOIN clause for carrying_26 using alias 'mat'."""
    return f"LEFT JOIN carrying_26 mat ON mat.m_code = {alias}.material"


def _ensure_carrying_join(alias: str, joins: list) -> None:
    """Adds the carrying_26 join to 'joins' if it is not already present."""
    j = _carrying_join(alias)
    if j not in joins:
        joins.append(j)


def _customer_join(alias: str) -> str:
    """LEFT JOIN against the pre-materialised customer_rollup table
    rather than re-aggregating customer per query.  customer is small
    (~5k rows) but every chart query was running its own GROUP BY
    ship_to over it, which adds up when a Group-By change fires
    ~10 endpoints in parallel.  customer_rollup is rebuilt at startup
    by _ensure_customer_rollup() and again whenever the operator
    refreshes via /api/admin/refresh_customer_rollup."""
    return f"LEFT JOIN customer_rollup cus ON cus.ship_to = {alias}.ship_to"


def _ensure_customer_join(alias: str, joins: list) -> None:
    """Adds the customer join to 'joins' if it is not already present."""
    j = _customer_join(alias)
    if j not in joins:
        joins.append(j)


def category_filters_sales(alias: str, category: str, has_brand: bool = False):
    """
    Like category_filters() but for the normalised sales fact tables
    (sales_2601 / sales_2526 / sales_21_25) where line / inch / pattern
    have been removed and now live in carrying_26 (alias: mat).

    Returns (joins, wheres) — same contract as category_filters().

    has_brand=True signals the fact table has its own `brand` column
    (currently only sales_thismonth).  For HK / LF that path filters
    directly on `{alias}.brand` so we don't drop rows where the
    material isn't in carrying_26 (a real case — Rebate brand-tagged
    sales for materials still missing from the master).
    """
    joins, wh = [], []
    cat = (category or "ALL").upper()

    if cat == "ALL":
        return joins, wh

    elif cat == "PCLT":
        joins.append(_carrying_join(alias))
        wh.append("mat.line = 'PCLT'")

    elif cat == "TBR":
        joins.append(_carrying_join(alias))
        wh.append("mat.line = 'TBR'")

    elif cat == "18PLUS":
        joins.append(_carrying_join(alias))
        wh.append("mat.line = 'PCLT'")
        # Extract rim inch from size string e.g. "225/45R18" ??18
        wh.append("CAST(SUBSTRING_INDEX(mat.size, 'R', -1) AS DECIMAL(5,2)) >= 18.0")

    elif cat == "ISEG":
        joins.append(f"JOIN iseg i ON cast(trim(i.Material) as unsigned) = {alias}.material")

    elif cat == "SUV":
        joins.append(_carrying_join(alias))
        joins.append("JOIN suv suv ON suv.Pattern = mat.pattern")

    elif cat == "LOWPROFILE":
        joins.append(f"JOIN lowprofile lp ON cast(trim(lp.Material) as unsigned) = {alias}.material")

    elif cat == "HM":
        joins.append(f"JOIN hm hm ON hm.sold_to = {alias}.sold_to")

    elif cat in ("HK", "LF"):
        # Brand filter via carrying_26.  sales_thismonth has its own
        # brand column but the other sales facts (monthly / yearly) don't,
        # so go through carrying for a single uniform path.
        if has_brand:
            wh.append(f"{alias}.brand = '{cat}'")
        else:
            joins.append(_carrying_join(alias))
            wh.append(f"mat.brand = '{cat}'")

    elif cat == "443":
        joins.append(_carrying_join(alias))
        wh.append(f"""
            EXISTS (
                SELECT 1
                FROM `443_25` p443
                WHERE p443.month         = {alias}.month
                AND p443.product_group = mat.product_group
            )
            """)

    return joins, wh


def category_target_filters(alias: str, category: str):
    """
    Returns (joins, wheres) for target_26 table.
    target_26 has a material column; line/inch/pattern attributes live in
    carrying_26 so we JOIN that table (alias: mat) for category filtering,
    exactly like category_filters_sales() does for sales fact tables.
    """
    joins, wh = [], []
    cat = (category or "ALL").upper()

    if cat == "ALL":
        return joins, wh

    carrying_join = f"LEFT JOIN carrying_26 mat ON mat.m_code = {alias}.material"

    if cat == "PCLT":
        joins.append(carrying_join)
        wh.append("mat.line = 'PCLT'")

    elif cat == "TBR":
        joins.append(carrying_join)
        wh.append("mat.line = 'TBR'")

    elif cat == "18PLUS":
        joins.append(carrying_join)
        wh.append("mat.line = 'PCLT'")
        wh.append("CAST(SUBSTRING_INDEX(mat.size, 'R', -1) AS DECIMAL(5,2)) >= 18.0")

    elif cat == "ISEG":
        joins.append(f"JOIN iseg i ON CAST(TRIM(i.Material) AS UNSIGNED) = {alias}.material")

    elif cat == "SUV":
        joins.append(carrying_join)
        joins.append("JOIN suv suv ON suv.Pattern = mat.pattern")

    elif cat == "LOWPROFILE":
        joins.append(f"JOIN lowprofile lp ON CAST(TRIM(lp.Material) AS UNSIGNED) = {alias}.material")

    elif cat == "HM":
        joins.append(f"JOIN hm hm ON hm.sold_to = {alias}.sold_to")

    elif cat in ("HK", "LF"):
        joins.append(carrying_join)
        wh.append(f"mat.brand = '{cat}'")

    elif cat == "443":
        # product_group lives in carrying_26 (alias: mat) for target_26
        joins.append(f"LEFT JOIN carrying_26 mat ON mat.m_code = {alias}.material")
        wh.append(f"""EXISTS (
            SELECT 1 FROM `443_25` p443
            WHERE p443.month = {alias}.month
            AND p443.product_group = mat.product_group
        )""")

    return joins, wh

app = Flask(__name__, static_folder="static")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
CORS(app, resources={r"/api/*": {"origins": "*"}})

@app.errorhandler(500)
def handle_500(e):
    return jsonify({"error": "internal_server_error", "detail": str(e)}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback as _tb
    _tb.print_exc()
    return jsonify({"error": type(e).__name__, "detail": str(e)}), 500

@app.after_request
def add_cache_headers(response):
    """Prevent Cloudflare / CDN from caching API responses."""
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
    else:
        # Static assets: prevent Cloudflare/CDN from caching
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
    return response

# ─── Per-request usage logging ───────────────────────────────────────
# Lightweight middleware that writes one row per HTTP request into
# request_log so we can answer questions Cloudflare Access logs can't:
# "which BDE opened /sales/rebate yesterday", "how many times did Asim
# open the meeting view this week", "which paths get the most traffic".
# Skips static + favicon to keep the table from filling with noise.
import time as _time

def _ensure_request_log_table():
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS request_log (
                id          BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                user_email  VARCHAR(255) NOT NULL DEFAULT '',
                path        VARCHAR(255) NOT NULL,
                method      VARCHAR(10)  NOT NULL DEFAULT 'GET',
                status      INT          NOT NULL DEFAULT 0,
                duration_ms INT          NOT NULL DEFAULT 0,
                ip          VARCHAR(64)  NOT NULL DEFAULT '',
                ua          VARCHAR(255) NOT NULL DEFAULT '',
                created_at  DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_rl_user_date (user_email, created_at),
                INDEX idx_rl_path_date (path, created_at),
                INDEX idx_rl_date      (created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Add country column if missing (Cloudflare CF-IPCountry, 2-letter
        # ISO code).  Older deployments created the table before this
        # column existed — keep adding it idempotently here so we don't
        # need a separate migration step.
        cur.execute("SHOW COLUMNS FROM request_log LIKE 'country'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE request_log "
                        "ADD COLUMN country VARCHAR(8) NOT NULL DEFAULT ''")
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"[request_log] schema init failed: {e}")

# Initial table creation is deferred to after get_connection() is defined
# (further down in the file).  The schema-init call happens alongside
# _ensure_meeting_plan_table() down there.

# Paths we never log — static assets, favicons, and the health-check-ish
# polls that would otherwise dominate the table and crowd out the
# user-meaningful navigation we actually care about.
_LOG_SKIP_PREFIXES = ("/static/", "/favicon")
_LOG_SKIP_EXACT    = {"/api/ai_status", "/api/whoami"}

@app.before_request
def _request_log_start():
    request._rl_t0 = _time.monotonic()

@app.after_request
def _request_log_save(response):
    try:
        path = request.path or ""
        if request.method == "OPTIONS":
            return response
        if any(path.startswith(p) for p in _LOG_SKIP_PREFIXES):
            return response
        if path in _LOG_SKIP_EXACT:
            return response
        email = (_bde_from_request() or "").strip().lower()
        dur = int((_time.monotonic() - getattr(request, "_rl_t0", _time.monotonic())) * 1000)
        ip = request.headers.get("CF-Connecting-IP") or request.remote_addr or ""
        ua = (request.user_agent.string or "")[:255]
        # Cloudflare injects CF-IPCountry on every request (2-letter ISO);
        # behind Tailscale / local dev it's absent so we just store "".
        country = (request.headers.get("CF-IPCountry") or "")[:8]
        conn = get_connection(); cur = conn.cursor()
        cur.execute(
            "INSERT INTO request_log "
            "(user_email, path, method, status, duration_ms, ip, ua, country) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (email[:255], path[:255], request.method[:10],
             int(response.status_code or 0), dur, ip[:64], ua, country)
        )
        conn.commit(); cur.close(); conn.close()
    except Exception:
        # Logging failure must never break the response — swallow quietly.
        pass
    return response

def category_filters_stock(alias: str, category: str):
    joins, wh = [], []
    cat = (category or "ALL").upper()

    if cat == "ALL":
        return joins, wh

    elif cat == "PCLT":
        wh.append(f"{alias}.line = 'PCLT'")  # stock ?뚯씠釉붿뿉 line ?덉쑝硫?OK

    elif cat == "TBR":
        wh.append(f"{alias}.line = 'TBR'")

    elif cat == "ISEG":
        joins.append("JOIN iseg i ON CAST(TRIM(i.Material) AS UNSIGNED) = s.material")

    elif cat == "SUV":
        # stock?먮뒗 pattern???놁쓣 ???덉쑝??carrying_26濡쒕???pattern 媛?몄?????        joins.append("JOIN carrying_26 c ON c.m_code = s.material")
        joins.append("JOIN suv suv ON suv.Pattern = c.pattern")

    elif cat == "LOWPROFILE":
        joins.append("JOIN lowprofile lp ON CAST(TRIM(lp.Material) AS UNSIGNED) = s.material")

    elif cat in ("HK", "LF"):
        joins.append("JOIN carrying_26 c ON c.m_code = s.material")
        wh.append(f"c.brand = '{cat}'")

    return joins, wh
def category_filters_orders(category: str):
    """
    Orders??移댄뀒怨좊━ ?꾪꽣.
    returns: (joins, wh, needs_carrying)
    - orders alias: o
    - carrying alias: c
    """
    joins, wh = [], []
    cat = (category or "ALL").strip().upper()
    needs_carrying = False

    if cat == "ALL":
        return joins, wh, needs_carrying

    if cat == "PCLT":
        wh.append("(CAST(o.material AS CHAR) LIKE '1%' OR CAST(o.material AS CHAR) LIKE '2%')")

    elif cat == "TBR":
        wh.append("CAST(o.material AS CHAR) LIKE '3%'")

    elif cat == "ISEG":
        joins.append("JOIN iseg i ON CAST(TRIM(i.Material) AS UNSIGNED) = o.material")

    elif cat == "LOWPROFILE":
        joins.append("JOIN lowprofile lp ON CAST(TRIM(lp.Material) AS UNSIGNED) = o.material")

    elif cat == "SUV":
        needs_carrying = True
        joins.append("JOIN suv suv ON suv.Pattern = c.pattern")

    elif cat in ("HK", "LF"):
        needs_carrying = True
        wh.append(f"c.brand = '{cat}'")

    return joins, wh, needs_carrying
# ------------------------------- v2 helpers -------------------------------
def _region_order_key(x: str) -> int:
    order = ["NSW", "QLD", "VIC", "WA", "SA", "NT", "TAS", "ACT", "COMMON"]
    try:
        return order.index((x or "").upper())
    except ValueError:
        return 999

def _to_cumulative(arr: List[float]) -> List[float]:
    out, run = [], 0.0
    for v in arr:
        run += float(v or 0)
        out.append(run)
    return out

def _stacks_from_rows(rows: List[Dict[str, Any]], idx_key: str, idx_count: int, *, group_sort: str="alpha"):
    """
    rows: [{idx_key: int, group_label: str, value: num}, ...]
    returns: (groups, value_by_group)
      value_by_group[group] = [len idx_count]
    """
    groups = sorted({(r.get("group_label") or "").strip() or "COMMON" for r in (rows or [])})
    if group_sort == "region":
        groups = sorted(groups, key=lambda g: (_region_order_key(g), g))

    by_group: Dict[str, List[float]] = {g: [0.0] * idx_count for g in groups}
    for r in (rows or []):
        g = (r.get("group_label") or "").strip() or "COMMON"
        i = int(r.get(idx_key) or 0) - 1
        if 0 <= i < idx_count:
            by_group[g][i] += float(r.get("value") or 0)

    return groups, by_group

def _pct_by_bucket(by_group: Dict[str, List[float]]) -> Dict[str, List[float]]:
    if not by_group:
        return {}
    keys = list(by_group.keys())
    n = len(by_group[keys[0]])
    totals = [0.0] * n
    for g in keys:
        for i, v in enumerate(by_group[g]):
            totals[i] += float(v or 0)
    out: Dict[str, List[float]] = {}
    for g in keys:
        out[g] = [ (0.0 if totals[i] <= 0 else round((float(by_group[g][i]) / totals[i]) * 100, 4)) for i in range(n) ]
    return out

# NOTE: simple process-local cache. Good enough for your current single-process setup.
# If you run multiple workers, each worker has its own cache (still fine for speed).
_V2_CACHE_DASH: Dict[str, Tuple[float, Any]] = {}
_V2_CACHE_DIMS: Dict[str, Tuple[float, Any]] = {}
_TOP_SOLD_TO_CACHE: Dict[str, Tuple[float, Any]] = {}
# Graph-view aggregation cache — short-TTL bucket shared across the
# heavy chart endpoints.  A single page load fires the same query
# many times (per region / per year / per group_by), and identical
# query strings will hit cache on every repeat after the first.
_GRAPH_CACHE: Dict[str, Tuple[float, Any]] = {}

# ---- Fixed Top list computed once at startup (Top 10/20/30) ----
def _cache_get(cache: Dict[str, Tuple[float, Any]], key: str):
    now = time()
    hit = cache.get(key)
    if not hit:
        return None
    exp, val = hit
    if exp < now:
        try:
            del cache[key]
        except Exception:
            pass
        return None
    return val

def _cache_set(cache: Dict[str, Tuple[float, Any]], key: str, val: Any, ttl_sec: int):
    cache[key] = (time() + ttl_sec, val)

def _norm(v: str) -> str:
    return (v or "").strip()

def _make_v2_key(prefix: str, req) -> str:
    # stable key: path + sorted query args
    items = sorted((k, _norm(v)) for k, v in req.args.items())
    return prefix + "|" + "&".join([f"{k}={v}" for k, v in items])


def cached_endpoint(ttl_sec: int = 30):
    """Wrap a Flask route so its JSON response is cached in _GRAPH_CACHE
    keyed by sorted query args.  Caches the decoded JSON body and
    re-jsonifies on cache hits, so each request gets a fresh Response
    object (Flask mutates response state during dispatch)."""
    from functools import wraps
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            key = _make_v2_key(fn.__name__, request)
            cached = _cache_get(_GRAPH_CACHE, key)
            if cached is not None:
                print(f"[cache HIT]  {fn.__name__}")
                return jsonify(cached)
            print(f"[cache MISS] {fn.__name__}")
            result = fn(*args, **kwargs)
            try:
                body = result.get_json(silent=True) if hasattr(result, "get_json") else None
            except Exception:
                body = None
            if body is not None:
                _cache_set(_GRAPH_CACHE, key, body, ttl_sec=ttl_sec)
            return result
        return wrapper
    return decorator


def _make_top_key(f: Dict[str, str], top_limit: int, value: str) -> str:
    # Top-N baseline is expensive; cache it a bit longer than chart payloads.
    parts = [
        f"top_limit={top_limit}",
        f"value={value}",
        f"year={_norm(f.get('year'))}",
        f"month={_norm(f.get('month'))}",
        f"category={_norm(f.get('category'))}",
        f"region={_norm(f.get('region'))}",
        f"salesman={_norm(f.get('salesman'))}",
        f"sold_to_group={_norm(f.get('sold_to_group'))}",
        f"sold_to={_norm(f.get('sold_to'))}",
        f"ship_to={_norm(f.get('ship_to'))}",
        f"product_group={_norm(f.get('product_group'))}",
        f"pattern={_norm(f.get('pattern'))}",
        f"material={_norm(f.get('material'))}",
    ]
    return "top|" + "&".join(parts)

def get_connection():
    """Get a DB connection.
    - SQLite: file-based (Render demo)
    - MySQL: pooled connections to avoid expensive TCP/auth handshake per request
    """
    # If USE_SQLITE=1 (on Render), use the local snapshot.db file
    if USE_SQLITE:
        conn = sqlite3.connect(SQLITE_PATH)
        conn.row_factory = sqlite3.Row  # rows behave like dicts
        return SQLiteConnectionWrapper(conn)

    # Otherwise use MySQL (your current local/remote setup)
    global _MYSQL_POOL
    cfg = {
        "host": os.getenv("DB_HOST", "127.0.0.1"),
        "port": int(os.getenv("DB_PORT", "3306")),
        "user": os.getenv("DB_USER", "root"),
        "password": os.getenv("DB_PASS", ""),
        "database": os.getenv("DB_NAME", "my_new_database"),
        "autocommit": True,
        "use_pure": True,
    }

    # pool_size 32: mysql-connector-python caps the pool at 32.  With
    # the teardown_request safety net below making sure no connection
    # leaks past a single request, 32 is comfortable for a single
    # dashboard tab's 15-20 concurrent calls.
    pool_size = int(os.getenv("DB_POOL_SIZE", "32"))
    pool_name = os.getenv("DB_POOL_NAME", "hka_pool")

    if _MYSQL_POOL is None:
        with _MYSQL_POOL_LOCK:
            if _MYSQL_POOL is None:
                _MYSQL_POOL = pooling.MySQLConnectionPool(
                    pool_name=pool_name,
                    pool_size=pool_size,
                    pool_reset_session=True,
                    **cfg,
                )

    # Retry up to 5 times: pool may be momentarily exhausted under burst load
    last_err = None
    for attempt in range(5):
        try:
            conn = _MYSQL_POOL.get_connection()
            try:
                conn.ping(reconnect=True, attempts=2, delay=1)
            except Exception:
                pass
            return conn
        except mysql.connector.errors.PoolError as e:
            last_err = e
            import time as _time
            _time.sleep(0.2 * (attempt + 1))   # 0.2s, 0.4s, 0.6s, 0.8s, 1.0s
        except mysql.connector.Error as e:
            last_err = e
            break

    print("DB connection failed:", last_err)
    raise RuntimeError(f"DB connection unavailable: {last_err}") from last_err

@app.get("/api/ping")
def ping():
    try:
        conn = get_connection()
        conn.close()
        return {"ok": True}
    except Exception:
        return {"ok": False}, 503

@app.get("/api/debug_salesman")
def debug_salesman():
    """Temporary: diagnose salesman/ship_to data for a given bde_state."""
    state = request.args.get("state", "WA")
    states = REGION_STATES.get(state.upper(), [state])
    ph = ",".join(["%s"] * len(states))
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        # 1. Distinct salesman names in customer for this state
        cur.execute(
            f"SELECT salesman_name, COUNT(*) AS cnt FROM customer"
            f" WHERE bde_state IN ({ph}) GROUP BY salesman_name ORDER BY cnt DESC",
            tuple(states))
        cus_salesmen = cur.fetchall()

        # 2. Distinct bde names in target_26 for this state
        cur.execute(
            f"SELECT bde, COUNT(DISTINCT ship_to) AS ships, COUNT(DISTINCT sold_to) AS soltos"
            f" FROM target_26 WHERE state IN ({ph})"
            f" GROUP BY bde ORDER BY ships DESC",
            tuple(states))
        tgt_bdes = cur.fetchall()

        # 3. Sample customer rows for state (shows ship_to, sold_to, salesman)
        cur.execute(
            f"SELECT ship_to, sold_to, salesman_name, bde_state FROM customer"
            f" WHERE bde_state IN ({ph}) LIMIT 20",
            tuple(states))
        cus_sample = cur.fetchall()

        # 4. Sample sales_2526 rows for a WA ship_to (check sold_to presence)
        cur.execute(
            f"SELECT s.ship_to, s.sold_to, YEAR(s.billing_date) AS year, SUM(s.amt) AS amt"
            f" FROM {_sales_2526_from('s', year=2025)}"
            f" WHERE s.ship_to IN (SELECT DISTINCT ship_to FROM customer WHERE bde_state IN ({ph}))"
            f" AND s.billing_date >= '2025-01-01' AND s.billing_date < '2026-01-01'"
            f" GROUP BY s.ship_to, s.sold_to, YEAR(s.billing_date)"
            f" LIMIT 20",
            tuple(states))
        sales_sample = cur.fetchall()

        # 5. Ship_to counts per salesman in target_26 for this state (shows split)
        cur.execute(
            f"SELECT t.bde, t.ship_to, t.sold_to, SUM(t.amt) AS tgt_amt"
            f" FROM target_26 t"
            f" WHERE t.state IN ({ph})"
            f" GROUP BY t.bde, t.ship_to, t.sold_to"
            f" LIMIT 30",
            tuple(states))
        tgt_sample = cur.fetchall()

        cur.close(); conn.close()
        return jsonify({
            "state_filter": states,
            "customer_salesmen": cus_salesmen,
            "target26_bdes": tgt_bdes,
            "customer_sample": cus_sample,
            "sales_sample": sales_sample,
            "target26_sample": tgt_sample,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500



def _hostname(url):
    from urllib.parse import urlparse
    try:
        return (urlparse(url or "").hostname or "").lower()
    except Exception:
        return ""

def _claim_portal_host():
    """Hostname of the customer-facing portal so the root route can
    decide whether to serve the Sales Dashboard or the claim form."""
    return _hostname(CLAIM_PORTAL_URL)

@app.route("/")
def index():
    portal_host = _claim_portal_host()
    dash_host   = _hostname(DASHBOARD_URL)
    # Only swap to the claim form when the portal subdomain is genuinely
    # distinct from the admin dashboard's hostname.  When both env vars
    # point at the same domain (e.g. customers reach the same
    # sales.hkaudashboard.com via a Cloudflare Access bypass on /claim/*),
    # the host check can't tell them apart and would otherwise shadow
    # the Sales Dashboard at /.
    if portal_host and portal_host != dash_host \
       and request.host.split(":")[0].lower() == portal_host:
        return send_from_directory("static", "claim.html")
    return app.send_static_file("index.html")

@app.route("/map")
def map_page():
    return app.send_static_file("map.html")

@app.route("/stock")
def stock_page():
    return app.send_static_file("stock.html")

@app.route("/order")
def orders_page():
    """Special Price Request Form — interactive layout.  Backing data
    comes from /api/orders/customer and /api/orders/material."""
    return app.send_static_file("orders.html")


# ══════════════════════════════════════════════════════════════════════
# Demo mode — anonymises customer names / BDE names / codes / PII on
# every JSON response, and paints a fixed orange banner across the top
# of every HTML page.  Turned on by visiting /demo (or /demo/<page>),
# which drops a session cookie; every request that comes back carrying
# that cookie (or an explicit ?demo=1) goes through the anonymiser in
# _demo_after_request below.  /demo_exit clears the cookie.
#
# Amounts / qtys / dates / product info stay real so the dashboards
# and charts still look meaningful — the demo just scrubs the "who".
# ══════════════════════════════════════════════════════════════════════
_DEMO_COOKIE = "SPRF_DEMO"

# Only these emails may enter demo mode (Cloudflare Access injects the
# authenticated email on every request in production).  Local-dev
# requests without any Cf-Access header are treated as allowed so a
# developer can test /demo on 127.0.0.1 without spoofing headers.
_DEMO_ALLOWED_EMAILS = {"jayden.bhang@hankooktyre.com.au"}


def _demo_user_allowed():
    """True when the caller may enter / stay in demo mode.
    Whitelisted email → allowed everywhere.
    No Cf-Access header at all → local dev, allowed too.
    Anyone else in prod → denied."""
    email = ""
    try:
        email = (request.headers.get("Cf-Access-Authenticated-User-Email")
                 or request.headers.get("cf-access-authenticated-user-email")
                 or "").strip().lower()
    except Exception:
        return False
    if email in _DEMO_ALLOWED_EMAILS:
        return True
    if not email:
        # Not behind Cloudflare Access — local / dev environment.
        return True
    return False

# Column-name buckets — matched case-insensitively.  Anything not in
# these sets passes through untouched.
_DEMO_NAME_FIELDS = {
    "sold_to_name", "ship_to_name", "customer_name", "shipname",
}
_DEMO_BDE_FIELDS = {
    "bde", "bde_name", "salesman", "salesman_name",
    "submitted_by_bde", "status_changed_by", "approved_by",
}
_DEMO_CODE_FIELDS = {
    "sold_to", "ship_to", "sold_to_code", "ship_to_code",
    "sold_to_group", "customer_group", "customer_grp",
    "bill_to_partner", "sap_customer", "material", "m_code",
}
_DEMO_HIDE_FIELDS = {
    "email", "contact_email", "submitted_by_email",
    "phone", "mobile", "telephone", "contact_phone", "mobile_phone",
    "phone_email", "address", "address_1", "ship_to_address",
    "city", "postcode", "post_code",
}
# Brand identifiers → generic labels (HK/LF are the real Hankook/
# Laufenn tyre lines that would leak in demo mode).  Consistent
# mapping so a customer's HK column stays "Brand-A" everywhere.
_DEMO_BRAND_FIELDS = {"brand", "brand_line", "tyre_brand"}
_DEMO_BRAND_MAP = {
    "HK": "Brand-A", "LF": "Brand-B", "KS": "Brand-C",
    "HANKOOK": "Brand-A", "LAUFENN": "Brand-B", "KUMHO": "Brand-C",
}


def _is_demo_mode():
    """True when the current request should be anonymised — either an
    explicit ?demo=1 query param (handy for API smoke-tests), the
    session cookie set by the /demo landing route, or a request-level
    flag stamped by the /demo route itself (needed for the very first
    request, whose cookie is only set BY the response — the request
    coming in doesn't carry it yet).

    Also gated by _demo_user_allowed() — an unauthorised viewer who
    somehow forges the cookie or query param won't be anonymised
    (their normal data flow just runs unchanged)."""
    try:
        if not _demo_user_allowed():
            return False
        if getattr(request, "_demo_mode_flag", False):
            return True
        if request.args.get("demo") == "1":
            return True
        if request.cookies.get(_DEMO_COOKIE) == "1":
            return True
    except Exception:
        pass
    return False


def _demo_short_hash(s, mod=10000):
    """Stable numeric id for a value — same input → same output — so a
    customer with three ship-tos consistently reads as the same
    Customer-042 across every table."""
    if s is None:
        return 0
    import hashlib as _h
    h = _h.md5(str(s).encode("utf-8", "ignore")).hexdigest()
    return int(h[:8], 16) % mod


def _demo_anon_value(key_lower, val):
    """Anonymise one field.  Names / BDE / codes / PII are rewritten
    with a stable-hash suffix; amounts / dates / booleans / etc. pass
    through so charts and rankings still look like real data."""
    if val is None or val == "":
        return val
    s = str(val)
    if key_lower in _DEMO_NAME_FIELDS:
        return f"Customer-{_demo_short_hash(s, 1000):03d}"
    if key_lower in _DEMO_BDE_FIELDS:
        # BDE identifiers can be a full name or an email — mask both
        # to the same "BDE-NN" label so the pool stays small.
        return f"BDE-{_demo_short_hash(s, 100):02d}"
    if key_lower in _DEMO_CODE_FIELDS:
        return f"D{_demo_short_hash(s, 100000):05d}"
    if key_lower in _DEMO_HIDE_FIELDS:
        return "—"
    if key_lower in _DEMO_BRAND_FIELDS:
        # Map real brand codes to generic labels; unknown values fall
        # back to a stable-hash label so nothing branded leaks.
        up = s.strip().upper()
        return _DEMO_BRAND_MAP.get(up, f"Brand-{_demo_short_hash(up, 100):02d}")
    return val


def _demo_anon_walk(obj):
    """Recursively rewrite dicts / lists in-place-safe style."""
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            k_lower = k.lower() if isinstance(k, str) else ""
            if isinstance(v, (dict, list)):
                out[k] = _demo_anon_walk(v)
            else:
                out[k] = _demo_anon_value(k_lower, v)
        return out
    if isinstance(obj, list):
        return [_demo_anon_walk(x) for x in obj]
    return obj


_DEMO_BANNER_HTML = """
<div id="SPRF_DEMO_BANNER" style="position:fixed;top:0;left:0;right:0;z-index:99999;background:#f97316;color:#fff;padding:6px 14px;text-align:center;font-weight:700;font-size:12px;letter-spacing:.3px;box-shadow:0 2px 4px rgba(0,0,0,.2)">
DEMO MODE &nbsp;·&nbsp; customer names, BDE names, codes and PII are anonymised
&nbsp;·&nbsp; <a href="/demo_exit" style="color:#fff;text-decoration:underline">exit demo</a>
</div>
<style>body{padding-top:30px !important}</style>
"""


@app.after_request
def _demo_after_request(response):
    """When demo mode is active, walk JSON responses and swap out
    identifying fields, and inject the orange banner into HTML pages.
    No-op on binary / attachment responses and on non-JSON / non-HTML
    content types."""
    if not _is_demo_mode():
        return response
    ctype = (response.mimetype or "").lower()
    if ctype == "application/json":
        try:
            import json as _json
            data = response.get_json(silent=True)
            if data is not None:
                data = _demo_anon_walk(data)
                response.set_data(_json.dumps(data, ensure_ascii=False, default=str))
                response.content_length = None
        except Exception:
            pass
    elif ctype in ("text/html", "text/html; charset=utf-8", "text/html;charset=utf-8"):
        try:
            # Skip file downloads, PDFs, etc. — only patch normal pages.
            if response.headers.get("Content-Disposition", "").startswith("attachment"):
                return response
            html = response.get_data(as_text=True)
            if "SPRF_DEMO_BANNER" not in html:
                if "</body>" in html:
                    html = html.replace("</body>", _DEMO_BANNER_HTML + "\n</body>", 1)
                else:
                    html += _DEMO_BANNER_HTML
                response.set_data(html)
                response.content_length = None
        except Exception:
            pass
    return response


# Mapping of /demo/<page> → static file.  Keys mirror the real routes
# so /demo/order lands on the same form as /order, etc.  Unlisted
# paths fall through to index.html.
_DEMO_PAGES = {
    "":            "index.html",
    "map":         "map.html",
    "stock":       "stock.html",
    "rebate":      "rebate.html",
    "price":       "price.html",
    "meeting":     "meeting.html",
    "claims":      "claims.html",
    "order":       "orders.html",
    "orders_list": "orders_list.html",
    "highlights":  "highlights.html",
    "fleet":       "fleet_chart.html",
}


_DEMO_FETCH_PATCH = """
<script>
/* Demo mode client-side patches:
    1. Wrap fetch / XHR so every same-origin API call carries ?demo=1
       — belt-and-braces with the SPRF_DEMO cookie so the server-side
       anonymiser fires even if the cookie is stripped.
    2. Override the REGION_SALESMEN constant in app.js with masked
       "BDE-NN" labels — those names are hardcoded on the client and
       leak straight into the KPI table's BDE column, so server-side
       JSON scrubbing alone can't hide them.
    3. Walk the DOM (existing + future via MutationObserver) and
       swap any occurrence of a known real name for its masked twin
       — catches names printed by app.js before the constant swap.
*/
(function(){
  const _addDemo = (url) => {
    if (typeof url !== "string") return url;
    if (url.startsWith("http") && !url.includes(location.host)) return url;
    if (/[?&]demo=1(&|$)/.test(url)) return url;
    return url + (url.includes("?") ? "&" : "?") + "demo=1";
  };
  const _f = window.fetch;
  window.fetch = function(u, o){ return _f.call(this, _addDemo(u), o); };
  const _o = XMLHttpRequest.prototype.open;
  XMLHttpRequest.prototype.open = function(m, u){
    arguments[1] = _addDemo(u);
    return _o.apply(this, arguments);
  };
  document.documentElement.dataset.demo = "1";

  // Stable hash — same input → same "BDE-42" label everywhere.
  const _hash = (s, mod) => {
    let h = 5381 >>> 0;
    for (let i = 0; i < s.length; i++) h = ((h << 5) + h + s.charCodeAt(i)) >>> 0;
    return h % mod;
  };
  const _maskBDE = (n) => "BDE-" + String(_hash(n, 100)).padStart(2, "0");

  // Wait for REGION_SALESMEN, then rewrite it and every leaf name in
  // the DOM.  50 ms poll is more than fast enough — app.js's first
  // render usually lands ~100-200 ms after DOMContentLoaded.
  const _nameMap = {};
  const _apply = () => {
    if (typeof REGION_SALESMEN === "undefined"
        || !REGION_SALESMEN || typeof REGION_SALESMEN !== "object") {
      return setTimeout(_apply, 50);
    }
    Object.entries(REGION_SALESMEN).forEach(([k, list]) => {
      if (!Array.isArray(list)) return;
      list.forEach(name => { if (name && !_nameMap[name]) _nameMap[name] = _maskBDE(name); });
      REGION_SALESMEN[k] = list.map(n => _nameMap[n] || n);
    });
    _scrubDom(document.body);
    new MutationObserver(muts => {
      muts.forEach(m => {
        m.addedNodes && m.addedNodes.forEach(_scrubDom);
        if (m.type === "characterData" && m.target) _scrubDom(m.target);
      });
    }).observe(document.body, { childList: true, subtree: true, characterData: true });
  };
  // Static substitutions for hardcoded labels the server-side JSON
  // scrubber can't reach (table headers, dropdown legends, DC-cell
  // captions).  Whole-word boundaries via regex so "HK" doesn't
  // accidentally rewrite "HKAU BDE Name".  Order matters: longer /
  // more specific patterns first.
  const _STATIC_RX = [
    // Longer / more specific patterns first so they don't get eaten
    // by the short-brand-code rules below.
    [/\bHK-PCLT\b/g,    "Brand-A-PCLT"],
    [/\bLF-PCLT\b/g,    "Brand-B-PCLT"],
    [/\bTBR\s*\(HK&LF\)/g, "TBR (Brand-A&B)"],
    [/\bHKAU BDE Name\b/g, "HKAU Salesmen Name"],
    [/\bHKAU BDE\b/g,      "HKAU Salesmen"],
    [/\bBDE Name\b/g,      "Salesmen Name"],
    [/\bBDEs\b/g,          "Salesmen"],
    [/\bBDE\b/g,           "Salesmen"],
    [/Hankook/gi,          "Brand-A"],
    [/Laufenn/gi,          "Brand-B"],
    [/Kumho/gi,            "Brand-C"],
    // Short brand codes — need explicit character-class boundaries
    // because \b treats "_" as a word char, so \bHK\b misses HK inside
    // compound identifiers like "HK_TBR_ATP_Q_SR" that the rebate
    // page renders in its structure-name column.  These patterns fire
    // whenever HK / LF / KS is surrounded by anything other than an
    // ASCII letter or digit (spaces, dashes, underscores, punctuation,
    // string boundaries all count as separators).
    [/(^|[^A-Za-z0-9])HK(?=[^A-Za-z0-9]|$)/g, "$1Brand-A"],
    [/(^|[^A-Za-z0-9])LF(?=[^A-Za-z0-9]|$)/g, "$1Brand-B"],
    [/(^|[^A-Za-z0-9])KS(?=[^A-Za-z0-9]|$)/g, "$1Brand-C"],
  ];
  const _applyStatic = (t) => {
    let out = t;
    for (const [rx, repl] of _STATIC_RX) out = out.replace(rx, repl);
    return out;
  };
  const _scrubDom = (root) => {
    if (!root) return;
    if (root.nodeType === 3) {
      let t = root.textContent;
      let changed = false;
      for (const real in _nameMap) {
        if (t.indexOf(real) !== -1) {
          t = t.split(real).join(_nameMap[real]);
          changed = true;
        }
      }
      const t2 = _applyStatic(t);
      if (t2 !== t) { t = t2; changed = true; }
      if (changed) root.textContent = t;
      return;
    }
    if (root.childNodes) root.childNodes.forEach(_scrubDom);
    // Also cover option labels in <select> (their textContent is on a
    // separate text node but the option itself carries .text too).
    if (root.tagName === "OPTION" && root.text) {
      let t = root.text;
      let changed = false;
      for (const real in _nameMap) {
        if (t.indexOf(real) !== -1) { t = t.split(real).join(_nameMap[real]); changed = true; }
      }
      const t2 = _applyStatic(t);
      if (t2 !== t) { t = t2; changed = true; }
      if (changed) { root.text = t; if (root.value && _nameMap[root.value]) root.value = _nameMap[root.value]; }
    }
  };
  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", _apply);
  } else {
    _apply();
  }
})();
</script>
"""


@app.route("/demo",           defaults={"page": ""})
@app.route("/demo/",          defaults={"page": ""})
@app.route("/demo/<path:page>")
def demo_page(page):
    """Enter demo mode — sets the SPRF_DEMO session cookie and serves
    the requested static page.  Reads the file directly (rather than
    going through send_static_file) so we always return a fresh 200
    with a body the after_request injector can modify, and stamps a
    request-level flag so the injector sees demo mode on THIS request
    too — the cookie only reaches the browser on the response, so a
    cookie check against the incoming request would miss it."""
    # Whitelist gate — only Jayden may enter demo mode in prod;
    # local dev (no Cf-Access header) still passes through.
    if not _demo_user_allowed():
        return ("<h1 style='font-family:sans-serif;padding:30px;color:#374151'>"
                "Demo mode is restricted.</h1>"
                "<p style='font-family:sans-serif;padding:0 30px;color:#6b7280'>"
                "Ask Jayden Bhang if you need access.</p>"), 403
    # Special-case: /demo/price has no static file.  The /price route
    # is dynamically registered from price_compare.price_dashboard at
    # startup and renders via render_template_string.  Set the demo
    # cookie + request flag, then 302 to /price so the demo cookie
    # + fetch-patch flow takes over from there.
    if page in ("price", "fleet"):
        resp = make_response(redirect(f"/{page}"))
        resp.set_cookie(_DEMO_COOKIE, "1", path="/", samesite="Lax")
        try: request._demo_mode_flag = True
        except Exception: pass
        return resp
    static_file = _DEMO_PAGES.get(page, "index.html")
    import os
    file_path = os.path.join(app.static_folder or "static", static_file)
    try:
        with open(file_path, "r", encoding="utf-8") as _fh:
            html = _fh.read()
    except Exception:
        return "Demo page not found", 404
    # Inject the fetch-patch script early so it applies before any
    # in-page JS fires its first API call.  The after_request injector
    # adds the orange banner on top of that.
    if "</head>" in html and "SPRF_DEMO_FETCH" not in html:
        html = html.replace("</head>",
                            _DEMO_FETCH_PATCH.replace("<script>", '<script id="SPRF_DEMO_FETCH">', 1) + "\n</head>",
                            1)
    resp = make_response(html)
    resp.headers["Content-Type"]  = "text/html; charset=utf-8"
    resp.headers["Cache-Control"] = "no-store"   # never let the browser stash a stale copy
    resp.set_cookie(_DEMO_COOKIE, "1", path="/", samesite="Lax")
    # Mark THIS request as demo so the after_request injector still
    # anonymises / banners the very first hit (cookie won't be in the
    # request yet — it's on this outgoing response).
    try: request._demo_mode_flag = True
    except Exception: pass
    return resp


@app.route("/demo_exit")
def demo_exit_page():
    """Turn off demo mode: clears the cookie and drops the user back
    on the main dashboard."""
    resp = make_response(redirect("/"))
    resp.set_cookie(_DEMO_COOKIE, "", path="/", expires=0)
    return resp


@app.route("/fleet")
def fleet_chart_page():
    """Fleet-by-rim bar chart.  Reads postcode_rim_demand.csv produced
    by postcode_rim_demand.py and shows total fleet units by rim
    family, with an optional state filter."""
    return app.send_static_file("fleet_chart.html")


# In-memory cache for postcode_rim_demand.csv so the /api endpoint
# doesn't re-read the file on every request.  Rebuild on mtime change.
_POSTCODE_RIM_CACHE: dict = {"mtime": 0.0, "rows": []}

def _load_postcode_rim_demand():
    """Load postcode_rim_demand.csv rows from disk (once, cached).
    Looks in out/rego/ next to app.py first, falls back to CWD."""
    import csv as _csv, os as _os, time as _time
    candidates = [
        _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                      "out", "rego", "postcode_rim_demand.csv"),
        _os.path.join(_os.getcwd(), "out", "rego", "postcode_rim_demand.csv"),
        _os.path.join(_os.getcwd(), "postcode_rim_demand.csv"),
    ]
    path = next((p for p in candidates if _os.path.exists(p)), None)
    if not path:
        return []
    try:
        mtime = _os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    if _POSTCODE_RIM_CACHE.get("path") == path \
       and _POSTCODE_RIM_CACHE.get("mtime") == mtime:
        return _POSTCODE_RIM_CACHE["rows"]
    rows = []
    try:
        with open(path, newline="", encoding="utf-8") as f:
            for r in _csv.DictReader(f):
                try:
                    units = int(r.get("fleet_units") or 0)
                except ValueError:
                    units = 0
                if units <= 0:
                    continue
                rows.append({
                    "state":       (r.get("state") or "").strip().upper(),
                    "postcode":    (r.get("postcode") or "").strip(),
                    "rim_family":  (r.get("rim_family") or "").strip(),
                    "fleet_units": units,
                })
    except Exception as _e:
        print(f"[fleet] failed to read {path}: {_e}")
        return []
    _POSTCODE_RIM_CACHE.update({"path": path, "mtime": mtime, "rows": rows})
    return rows


_HK_RIM_CACHE: dict = {"year": None, "ts": 0.0, "rows": []}
_HK_SIZE_CACHE: dict = {"year": None, "ts": 0.0, "rows": []}

# In-memory cache for postcode_size_demand.csv — same lazy-reload
# behaviour as _POSTCODE_RIM_CACHE above.  Read once per file mtime.
_POSTCODE_SIZE_CACHE: dict = {"mtime": 0.0, "rows": []}

def _load_postcode_size_demand():
    """Load postcode_size_demand.csv (state, postcode, size, gen,
    fleet_units, ...) once, cached until the file's mtime changes.
    Legacy / unknown rows (empty size) are kept in the raw list so
    callers can report coverage — the API endpoint filters them out
    before charting."""
    import csv as _csv, os as _os
    candidates = [
        _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                      "out", "rego", "postcode_size_demand.csv"),
        _os.path.join(_os.getcwd(), "out", "rego", "postcode_size_demand.csv"),
        _os.path.join(_os.getcwd(), "postcode_size_demand.csv"),
    ]
    path = next((p for p in candidates if _os.path.exists(p)), None)
    if not path:
        return []
    try:
        mtime = _os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    if _POSTCODE_SIZE_CACHE.get("path") == path \
       and _POSTCODE_SIZE_CACHE.get("mtime") == mtime:
        return _POSTCODE_SIZE_CACHE["rows"]
    rows = []
    try:
        with open(path, newline="", encoding="utf-8") as f:
            for r in _csv.DictReader(f):
                try:
                    units = int(r.get("fleet_units") or 0)
                except ValueError:
                    units = 0
                if units <= 0:
                    continue
                rows.append({
                    "state":       (r.get("state") or "").strip().upper(),
                    "postcode":    (r.get("postcode") or "").strip(),
                    "size":        (r.get("size") or "").strip(),
                    "gen":         (r.get("gen")  or "").strip(),
                    "fleet_units": units,
                })
    except Exception as _e:
        print(f"[fleet-size] failed to read {path}: {_e}")
        return []
    _POSTCODE_SIZE_CACHE.update({"path": path, "mtime": mtime, "rows": rows})
    return rows


# Per-postcode top-N (make, model) cache — the source is the same
# BITRE file postcode_size_demand.py reads, aggregated by
# (state, postcode, make, model) so /api/postcode_top_models can hand
# back the top-5 vehicles at a postcode without re-scanning 200 MB per
# request.  Loaded on first hit, invalidated on file mtime change.
_POSTCODE_MODELS_CACHE: dict = {"mtime": 0.0, "path": None,
                                "by_postcode": {}, "by_state": {}, "national": []}

def _load_postcode_top_models():
    """Return three keyed views over the BITRE fleet:
      by_postcode : {postcode → [(make, model, fleet), ...] DESC}
      by_state    : {state    → [(make, model, fleet), ...] DESC}
      national    : [(make, model, fleet), ...]              DESC
    Sorted DESC; caller slices to whatever top-N it needs.  Loaded
    once per BITRE file mtime — file is ~200MB so we cache aggressively.
    """
    import csv as _csv, os as _os
    candidates = [
        _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                      "out", "rego", "vehicle_postcode_make_model_year_estimate.csv"),
        _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                      "out", "rego", "vehicle_postcode_make_model.csv"),
    ]
    path = next((p for p in candidates if _os.path.exists(p)), None)
    if not path:
        return _POSTCODE_MODELS_CACHE
    try:
        mtime = _os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    if _POSTCODE_MODELS_CACHE.get("path") == path \
       and _POSTCODE_MODELS_CACHE.get("mtime") == mtime:
        return _POSTCODE_MODELS_CACHE

    # Aggregate {(state, pc, make, model) → fleet} first, then flatten
    # to the two secondary views.  ~20 M rows in the input; this scan
    # takes 10-20 s once but the result is <5 MB so it fits in RAM
    # easily and every subsequent request is O(1).
    print(f"[top_models] first-load scan of {path}", flush=True)
    from collections import defaultdict
    agg = defaultdict(int)
    natl = defaultdict(int)
    per_state = defaultdict(lambda: defaultdict(int))
    per_pc    = defaultdict(lambda: defaultdict(int))
    try:
        with open(path, newline="", encoding="utf-8") as f:
            for r in _csv.DictReader(f):
                mk = (r.get("make")  or "").strip().upper()
                md = (r.get("model") or "").strip().upper()
                st = (r.get("state") or "").strip().upper()
                pc = (r.get("postcode") or "").strip()
                if not mk or not md or not pc: continue
                if mk in {"-", "TOTAL"} or md in {"-", "TOTAL", "UNKNOWN", "OTHER"}:
                    continue
                try:    qty = int(float(r.get("qty") or 0))
                except: qty = 0
                if qty <= 0: continue
                if pc.isdigit(): pc = pc.zfill(4)
                key = (mk, md)
                natl[key]           += qty
                per_state[st][key]  += qty
                per_pc[pc][key]     += qty
    except Exception as _e:
        print(f"[top_models] read failed: {_e}")
        return _POSTCODE_MODELS_CACHE

    def _sorted(d): return sorted(
        [(mk, md, q) for (mk, md), q in d.items()],
        key=lambda t: -t[2],
    )
    _POSTCODE_MODELS_CACHE.update({
        "path":        path,
        "mtime":       mtime,
        "by_postcode": {pc: _sorted(d) for pc, d in per_pc.items()},
        "by_state":    {st: _sorted(d) for st, d in per_state.items()},
        "national":    _sorted(natl),
    })
    print(f"[top_models] cached: {len(per_pc)} postcodes, "
          f"{len(per_state)} states, {len(natl)} national (make, model) pairs",
          flush=True)
    return _POSTCODE_MODELS_CACHE


def _load_hk_size_sales(year: int = 2026, ttl: int = 300):
    """Same shape as _load_hk_rim_sales but keys HK sales by the FULL
    tyre size string (from carrying_26.size) instead of rim_family, so
    the size-level fleet chart can align HK bars to the same size
    buckets as the BITRE fleet."""
    import time as _t
    now = _t.monotonic()
    if _HK_SIZE_CACHE["year"] == year and (now - _HK_SIZE_CACHE["ts"]) < ttl:
        return _HK_SIZE_CACHE["rows"]
    try:
        conn = get_connection()
        cur  = conn.cursor(dictionary=True)
    except Exception as _e:
        print(f"[hk_size] DB connect failed: {_e}")
        return []
    try:
        cur.execute("SHOW COLUMNS FROM customer")
        cust_cols = {r["Field"] for r in cur.fetchall()}
        has_pc    = "postcode" in cust_cols
        state_col = "ship_to_state" if "ship_to_state" in cust_cols else "bde_state"
        pc_field  = "c.postcode AS postcode" if has_pc else "c.address_1 AS address_1"
        pc_group  = "c.postcode" if has_pc else "c.address_1"
        cur.execute(f"""
            SELECT c.{state_col} AS state, {pc_field},
                   cr.size AS size, SUM(s.qty) AS qty
            FROM sales_2526 s
            JOIN customer c    ON c.ship_to = s.ship_to
            JOIN carrying_26 cr ON cr.m_code = s.material
            WHERE s.billing_date >= '{year}-01-01'
              AND s.billing_date <  '{year + 1}-01-01'
              AND s.qty > 0
            GROUP BY c.{state_col}, {pc_group}, cr.size
        """)
        raw = cur.fetchall()
    except Exception as _e:
        print(f"[hk_size] SQL failed: {_e}")
        try: cur.close(); conn.close()
        except: pass
        return []
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    import re as _re_pc
    _PC_RX = _re_pc.compile(r"\b(\d{4})\b")
    rows = []
    for r in raw:
        state = (r.get("state") or "").strip().upper()
        if has_pc:
            pc = (r.get("postcode") or "").strip()
        else:
            addr = r.get("address_1") or ""
            m = _PC_RX.findall(addr)
            pc = m[-1] if m else ""
        if pc.isdigit(): pc = pc.zfill(4)
        # Normalise size to the same form the BITRE side emits — strip
        # load / speed ratings so "205/55R16 91V XL" collapses onto
        # "205/55R16".
        size_raw = (r.get("size") or "").strip()
        m = _re_pc.search(r"(\d{3}/\d{2}R\d{2}(?:\.\d)?)", size_raw)
        if not m:
            # Fall back to LT/C sizes like "205R16C"
            m = _re_pc.search(r"(\d{3}R\d{2}(?:\.\d)?C?)", size_raw)
        size = m.group(1) if m else size_raw
        rows.append({
            "state":    state,
            "postcode": pc,
            "size":     size,
            "qty":      int(r.get("qty") or 0),
        })
    _HK_SIZE_CACHE.update({"year": year, "ts": now, "rows": rows})
    return rows


# Rim-diameter parser (regex + bucket) — same shape as
# postcode_penetration.py so the two stay in sync.
_FLEET_RIM_RX = re.compile(r"R\s*(\d{2}(?:\.\d)?)", re.IGNORECASE) if 're' in globals() else None
if _FLEET_RIM_RX is None:
    import re as _re
    _FLEET_RIM_RX = _re.compile(r"R\s*(\d{2}(?:\.\d)?)", _re.IGNORECASE)

def _fleet_rim_from_size(s: str):
    """Rim diameter in inches — keeps the .5 for TBR sizes so the
    fleet chart's bucketing can separate truck from passenger."""
    if not s: return None
    m = _FLEET_RIM_RX.search(str(s))
    if not m: return None
    try:
        v = float(m.group(1))
        return v if v == int(v) + 0.5 else int(v)
    except:
        return None

def _fleet_rim_family(inches):
    if inches is None: return "UNKNOWN"
    # TBR half-inch sizes (Hino 300 R17.5, Isuzu FRR R19.5, prime
    # movers R22.5) get their own buckets so passenger R17/R19/R22
    # aren't polluted by truck volume.
    if inches == 17.5: return "R17.5 (TBR)"
    if inches == 19.5: return "R19.5 (TBR)"
    if inches == 22.5: return "R22.5 (TBR)"
    inches = int(inches)
    if inches <= 13: return "R13"
    if inches == 14: return "R14"
    if inches == 15: return "R15"
    if inches == 16: return "R16"
    if inches == 17: return "R17"
    if inches == 18: return "R18"
    if inches == 19: return "R19"
    if inches == 20: return "R20"
    if inches == 21: return "R21"
    if inches >= 22: return "R22+"
    return f"R{inches}"

def _load_hk_rim_sales(year: int = 2026, ttl: int = 300):
    """Aggregate Hankook sales (sales_2526) by (state, postcode,
    rim_family) so /api/fleet_by_rim can put HK volume next to BITRE
    fleet units.  Cached for `ttl` seconds; returns [] on any DB
    failure so the fleet chart still renders BITRE-only."""
    import time as _t
    now = _t.monotonic()
    if _HK_RIM_CACHE["year"] == year and (now - _HK_RIM_CACHE["ts"]) < ttl:
        return _HK_RIM_CACHE["rows"]
    try:
        conn = get_connection()
        cur  = conn.cursor(dictionary=True)
    except Exception as _e:
        print(f"[hk_rim] DB connect failed: {_e}")
        return []
    try:
        cur.execute("SHOW COLUMNS FROM customer")
        cust_cols = {r["Field"] for r in cur.fetchall()}
        has_pc     = "postcode" in cust_cols
        state_col  = "ship_to_state" if "ship_to_state" in cust_cols else "bde_state"
        pc_field   = "c.postcode AS postcode" if has_pc else "c.address_1 AS address_1"
        pc_group   = "c.postcode" if has_pc else "c.address_1"

        cur.execute(f"""
            SELECT c.{state_col} AS state, {pc_field},
                   cr.size AS size, SUM(s.qty) AS qty
            FROM sales_2526 s
            JOIN customer c    ON c.ship_to = s.ship_to
            JOIN carrying_26 cr ON cr.m_code = s.material
            WHERE s.billing_date >= '{year}-01-01'
              AND s.billing_date <  '{year + 1}-01-01'
              AND s.qty > 0
            GROUP BY c.{state_col}, {pc_group}, cr.size
        """)
        raw = cur.fetchall()
    except Exception as _e:
        print(f"[hk_rim] SQL failed: {_e}")
        try: cur.close(); conn.close()
        except: pass
        return []
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    import re as _re_pc
    _PC_RX = _re_pc.compile(r"\b(\d{4})\b")
    rows = []
    for r in raw:
        state = (r.get("state") or "").strip().upper()
        if has_pc:
            pc = (r.get("postcode") or "").strip()
        else:
            addr = r.get("address_1") or ""
            m = _PC_RX.findall(addr)
            pc = m[-1] if m else ""
        if pc.isdigit(): pc = pc.zfill(4)
        rim = _fleet_rim_from_size(r.get("size"))
        if rim is None: continue
        rows.append({
            "state":      state,
            "postcode":   pc,
            "rim_family": _fleet_rim_family(rim),
            "qty":        int(r.get("qty") or 0),
        })
    _HK_RIM_CACHE.update({"year": year, "ts": now, "rows": rows})
    return rows


@app.get("/api/fleet_by_rim")
def api_fleet_by_rim():
    """Return fleet_units aggregated by rim_family, optionally
    filtered / grouped by state or a single postcode.

      ?state=NSW          — one state; returns [{rim_family, fleet}, ...]
      ?state=ALL          — national totals (default)
      ?postcode=2000      — one postcode only; overrides state
      ?breakdown=state    — one series per state (used by the bar-chart
                            legend to show per-state stacks)
    """
    state     = (request.args.get("state") or "ALL").strip().upper()
    postcode  = (request.args.get("postcode") or "").strip()
    breakdown = (request.args.get("breakdown") or "").strip().lower()
    # region expands the 4 top-nav regions into their contiguous
    # states so the fleet chart follows the same grouping the rest of
    # the app uses.  ACT is treated as part of NSW; TAS/NT/SA fold
    # into VIC.  Explicit ?state= wins if provided.
    region   = (request.args.get("region") or "").strip().upper()
    _REGION_TO_STATES = {
        "NSW": {"NSW", "ACT"},
        "VIC": {"VIC", "TAS", "NT", "SA"},
        "QLD": {"QLD"},
        "WA":  {"WA"},
    }
    region_states = _REGION_TO_STATES.get(region)  # None → no expansion

    rows = _load_postcode_rim_demand()
    if not rows:
        return jsonify({
            "rim_order": [], "series": [],
            "warning": "postcode_rim_demand.csv not found — run "
                       "postcode_rim_demand.py first.",
        })

    # Zero-pad the incoming postcode to 4 digits so a boundary GeoJSON
    # ("0800", "2000") matches CSV rows that Excel may have re-saved as
    # bare ints ("800", "2000").  Applied to BOTH sides of the compare.
    if postcode:
        postcode = postcode.zfill(4) if postcode.isdigit() else postcode

    # Canonical rim ordering — passenger whole-inch first, then TBR
    # half-inch sizes grouped at the end so the chart visually reads
    # "cars … then trucks".
    rim_order = ["R13", "R14", "R15", "R16", "R17", "R18",
                 "R19", "R20", "R21", "R22+",
                 "R17.5 (TBR)", "R19.5 (TBR)", "R22.5 (TBR)",
                 "MOTORCYCLE", "UNKNOWN"]

    if breakdown == "state":
        # One series per state so a stacked / grouped bar chart can
        # render distribution across states side-by-side.
        agg: dict = {}
        for r in rows:
            agg.setdefault(r["state"], {})
            agg[r["state"]][r["rim_family"]] = agg[r["state"]].get(r["rim_family"], 0) + r["fleet_units"]
        series = []
        for st in sorted(agg.keys()):
            values = [agg[st].get(f, 0) for f in rim_order]
            series.append({"state": st, "values": values})
        return jsonify({"rim_order": rim_order, "series": series})

    # Single series — national, one-state / region, or one-postcode
    def _row_in_region(row_state: str) -> bool:
        if region_states is not None:
            return row_state in region_states
        return state == "ALL" or row_state == state

    totals: dict = {}
    matched = 0
    for r in rows:
        if postcode:
            # Compare on zero-padded 4-digit form so "0800" from a
            # boundary GeoJSON matches an Excel-mangled "800" in the CSV.
            row_pc = r["postcode"]
            if row_pc.isdigit():
                row_pc = row_pc.zfill(4)
            if row_pc != postcode:
                continue
        elif not _row_in_region(r["state"]):
            continue
        totals[r["rim_family"]] = totals.get(r["rim_family"], 0) + r["fleet_units"]
        matched += 1
    values = [totals.get(f, 0) for f in rim_order]

    # Overlay HK sales on the same rim buckets so the chart can put a
    # second bar (HK sold) and a line (penetration %) next to each
    # fleet bar.  Silent zero if the HK dataset isn't available on
    # this deployment — chart just shows BITRE bars.
    hk_totals: dict = {}
    for h in _load_hk_rim_sales():
        if postcode:
            hpc = h["postcode"]
            if hpc.isdigit(): hpc = hpc.zfill(4)
            if hpc != postcode: continue
        elif not _row_in_region(h["state"]):
            continue
        hk_totals[h["rim_family"]] = hk_totals.get(h["rim_family"], 0) + h["qty"]
    hk_sold = [hk_totals.get(f, 0) for f in rim_order]
    # Penetration % per rim family — HK sold ÷ fleet × 100.  Guard
    # against zero-fleet buckets (division-by-zero) with None so the
    # chart can render "no line point" instead of a spike to Infinity.
    penetration = [
        round(hk_sold[i] / values[i] * 100, 2) if values[i] > 0 else None
        for i in range(len(rim_order))
    ]

    if postcode:
        label = f"Postcode {postcode}"
    elif region_states is not None:
        label = region  # e.g. "NSW" — visible states are the group
    else:
        label = state
    return jsonify({
        "rim_order":    rim_order,
        "series":       [{
            "state":       label,
            "values":      values,
            "hk_sold":     hk_sold,
            "penetration": penetration,
        }],
        "matched_rows": matched,
        "total_units":  sum(values),
        "hk_total":     sum(hk_sold),
        "sample_postcodes": sorted(
            {r["postcode"] for r in rows[:200] if r["postcode"]})[:10],
    })


@app.get("/api/fleet_by_size")
def api_fleet_by_size():
    """Fleet-by-size version of /api/fleet_by_rim.  Same shape of
    request/response, but the bucket dimension is the full tyre size
    string (e.g. '215/60R16') instead of the coarse rim family (R16).

    Sizes with zero fleet in the current filter are dropped.  Result
    is sorted by fleet DESC so the chart's x-axis reads high → low
    naturally; the frontend can slice to a top-N view when the tail
    of ~20+ sizes doesn't fit."""
    state     = (request.args.get("state") or "ALL").strip().upper()
    postcode  = (request.args.get("postcode") or "").strip()
    region    = (request.args.get("region") or "").strip().upper()
    limit     = request.args.get("limit", type=int) or 0   # 0 = all
    _REGION_TO_STATES = {
        "NSW": {"NSW", "ACT"},
        "VIC": {"VIC", "TAS", "NT", "SA"},
        "QLD": {"QLD"},
        "WA":  {"WA"},
    }
    region_states = _REGION_TO_STATES.get(region)

    rows = _load_postcode_size_demand()
    if not rows:
        return jsonify({
            "size_order": [], "series": [],
            "warning": "postcode_size_demand.csv not found — run "
                       "postcode_size_demand.py first.",
        })

    if postcode:
        postcode = postcode.zfill(4) if postcode.isdigit() else postcode

    def _row_in_scope(row) -> bool:
        if postcode:
            row_pc = row["postcode"]
            if row_pc.isdigit():
                row_pc = row_pc.zfill(4)
            return row_pc == postcode
        if region_states is not None:
            return row["state"] in region_states
        return state == "ALL" or row["state"] == state

    # Aggregate fleet by size — skip empty-size rows (legacy /
    # unknown) so the chart only shows sizes we actually predicted.
    fleet_totals: dict[str, int] = {}
    unpredicted = 0
    matched = 0
    for r in rows:
        if not _row_in_scope(r):
            continue
        if not r["size"]:
            unpredicted += r["fleet_units"]
            continue
        fleet_totals[r["size"]] = fleet_totals.get(r["size"], 0) + r["fleet_units"]
        matched += 1

    # HK sales side — same scope filter, aggregate by full size.
    hk_totals: dict[str, int] = {}
    for h in _load_hk_size_sales():
        if not _row_in_scope(h):
            continue
        if not h["size"]:
            continue
        hk_totals[h["size"]] = hk_totals.get(h["size"], 0) + h["qty"]

    # Union of sizes seen on either side, sorted by fleet DESC (so the
    # x-axis lines up with tyre-demand priority).  Sizes only present
    # in HK sales but not fleet still surface — they read as
    # "penetration only", useful for cross-fleet opportunities.
    all_sizes = set(fleet_totals) | set(hk_totals)
    size_order_full = sorted(
        all_sizes,
        key=lambda s: (-fleet_totals.get(s, 0), -hk_totals.get(s, 0), s),
    )
    # Aggregate totals span the FULL scope (matched + unpredicted) so
    # the fleet card doesn't understate reality when the caller asked
    # for a top-N chart via ?limit.  chart_size_order is the sliced
    # view; totals stay whole.
    matched_total = sum(fleet_totals.values())
    hk_total_full = sum(hk_totals.values())
    size_order = size_order_full[:limit] if limit > 0 else size_order_full

    values      = [fleet_totals.get(s, 0) for s in size_order]
    hk_sold     = [hk_totals.get(s, 0)    for s in size_order]
    penetration = [
        round(hk_sold[i] / values[i] * 100, 2) if values[i] > 0 else None
        for i in range(len(size_order))
    ]

    if postcode:
        label = f"Postcode {postcode}"
    elif region_states is not None:
        label = region
    else:
        label = state
    # total_units = every registered vehicle in scope (matched sizes +
    # legacy + unknown) — matches the raw BITRE fleet number the reader
    # expects.  matched_units = subset with a predicted size (what the
    # chart shows on the x-axis when everything is expanded).
    total_all = matched_total + unpredicted
    return jsonify({
        "size_order": size_order,
        "series": [{
            "state":       label,
            "values":      values,
            "hk_sold":     hk_sold,
            "penetration": penetration,
        }],
        "matched_rows":       matched,
        "matched_units":      matched_total,  # matched to a real size
        "unpredicted_units":  unpredicted,    # LEGACY + UNKNOWN fleet
        "total_units":        total_all,      # everything registered
        "hk_total":           hk_total_full,  # full HK volume in scope
    })


@app.get("/api/postcode_top_models")
def api_postcode_top_models():
    """Return the top-N (make, model) pairs by fleet count.

      ?postcode=2000   → top-N for that postcode
      ?state=NSW       → top-N for that state
      (neither)        → top-N nationally

    ?limit=N (default 5).  Response: {rows: [{make, model, fleet}, ...]}.
    """
    postcode = (request.args.get("postcode") or "").strip()
    state    = (request.args.get("state") or "").strip().upper()
    limit    = request.args.get("limit", type=int) or 5
    if postcode.isdigit():
        postcode = postcode.zfill(4)

    cache = _load_postcode_top_models()
    if postcode:
        seq = cache.get("by_postcode", {}).get(postcode, [])
    elif state and state != "ALL":
        seq = cache.get("by_state", {}).get(state, [])
    else:
        seq = cache.get("national", [])
    rows = [{"make": mk, "model": md, "fleet": q}
            for (mk, md, q) in seq[:limit]]
    return jsonify({"rows": rows})


@app.get("/api/fleet_by_rim/hk_detail")
def api_fleet_hk_detail():
    """Drill-down behind an HK Sold bar on the popup chart.

      ?postcode=NNNN     one postcode  (usual case — a boundary was clicked)
      ?region=NSW|VIC|…  when the popup is scoped to a region instead
      ?state=XX          fallback single-state filter
      ?rim=R17           bucket to break down (matches rim_family label)
      ?year=2026         defaults to 2026 (matches the popup chart)

    Returns rows for that (scope, rim) combo, one per
    (size × pattern × sold_to), sorted by 2026 qty descending — the
    table the user sees expanded under the chart."""
    postcode = (request.args.get("postcode") or "").strip()
    region   = (request.args.get("region") or "").strip().upper()
    state    = (request.args.get("state") or "").strip().upper()
    rim_want = (request.args.get("rim") or "").strip()
    year     = int(request.args.get("year", 2026) or 2026)
    if postcode and postcode.isdigit():
        postcode = postcode.zfill(4)

    _REGION_TO_STATES = {
        "NSW": {"NSW", "ACT"},
        "VIC": {"VIC", "TAS", "NT", "SA"},
        "QLD": {"QLD"},
        "WA":  {"WA"},
    }
    region_states = _REGION_TO_STATES.get(region)

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cust_cols  = _list_columns(cur, "customer")
        has_pc     = "postcode" in cust_cols
        state_col  = "ship_to_state" if "ship_to_state" in cust_cols else "bde_state"
        pc_field   = "c.postcode AS postcode" if has_pc else "c.address_1 AS address_1"

        wh, params = [], []
        wh.append(f"s.billing_date >= '{year}-01-01'")
        wh.append(f"s.billing_date <  '{year + 1}-01-01'")
        wh.append("s.qty > 0")
        if not postcode:
            # Region / state filter only used when no postcode gate
            if region_states is not None:
                ph = ",".join(["%s"] * len(region_states))
                wh.append(f"c.{state_col} IN ({ph})")
                params.extend(sorted(region_states))
            elif state and state != "ALL":
                wh.append(f"c.{state_col} = %s")
                params.append(state)

        # If we have a real postcode column, gate in SQL — otherwise
        # pull address_1 and filter in Python.
        if postcode and has_pc:
            wh.append("c.postcode = %s")
            params.append(postcode)

        cur.execute(f"""
            SELECT
                cr.size    AS size,
                cr.pattern AS pattern,
                COALESCE(cus.sold_to_name, s.sold_to) AS sold_to_name,
                s.sold_to  AS sold_to,
                {pc_field},
                SUM(s.qty) AS qty
            FROM sales_2526 s
            JOIN customer c    ON c.ship_to = s.ship_to
            JOIN carrying_26 cr ON cr.m_code = s.material
            LEFT JOIN (
                SELECT sold_to, MIN(sold_to_name) AS sold_to_name
                FROM customer
                WHERE sold_to_name IS NOT NULL AND TRIM(sold_to_name) <> ''
                GROUP BY sold_to
            ) cus ON cus.sold_to = s.sold_to
            WHERE {' AND '.join(wh)}
            GROUP BY cr.size, cr.pattern, s.sold_to, sold_to_name
                     {', c.postcode' if has_pc else ', c.address_1'}
        """, tuple(params))
        raw = cur.fetchall()
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    import re as _re_pc
    _PC_RX = _re_pc.compile(r"\b(\d{4})\b")

    # Filter by postcode (address_1 path) and by rim family in Python.
    out = []
    for r in raw:
        if postcode and not has_pc:
            addr = r.get("address_1") or ""
            m = _PC_RX.findall(addr)
            row_pc = m[-1] if m else ""
            if row_pc.isdigit(): row_pc = row_pc.zfill(4)
            if row_pc != postcode:
                continue
        rim = _fleet_rim_from_size(r.get("size"))
        if rim_want:
            fam = _fleet_rim_family(rim)
            if fam != rim_want:
                continue
        out.append({
            "size":         (r.get("size") or "").strip(),
            "pattern":      (r.get("pattern") or "").strip(),
            "sold_to":      (r.get("sold_to") or "").strip(),
            "sold_to_name": (r.get("sold_to_name") or "").strip(),
            "qty":          int(r.get("qty") or 0),
        })
    out.sort(key=lambda r: r["qty"], reverse=True)
    return jsonify({
        "rim":     rim_want,
        "rows":    out,
        "total":   sum(r["qty"] for r in out),
    })


def _list_columns(cur, table):
    """Return the lower-cased set of columns on `table`, or empty set
    if the table can't be introspected (permissions, wrong schema, etc.).
    Used by the orders lookup endpoints so they gracefully skip columns
    that aren't in this deployment's customer / carrying_26 schema."""
    try:
        cur.execute(f"SHOW COLUMNS FROM {table}")
        cols = set()
        for r in cur.fetchall():
            name = r[0] if not isinstance(r, dict) else (r.get("Field") or r.get("field"))
            if name:
                cols.add(name.lower())
        return cols
    except Exception:
        return set()


@app.get("/api/orders/customer")
def api_orders_customer():
    """Look up a customer row by any of:
      ?sold_to=…      exact sold_to code
      ?ship_to=…      exact ship_to code
      ?sold_to_name=… exact sold_to_name
      ?ship_to_name=… exact ship_to_name

    Returns whatever address / phone / email columns exist plus a
    canonical `bde_name` alias for the salesman.  Missing columns
    come back as empty strings so the Orders form can consume the
    response without per-field null checks."""
    sold_to      = (request.args.get("sold_to")      or "").strip()
    ship_to      = (request.args.get("ship_to")      or "").strip()
    sold_to_name = (request.args.get("sold_to_name") or "").strip()
    ship_to_name = (request.args.get("ship_to_name") or "").strip()
    if not (sold_to or ship_to or sold_to_name or ship_to_name):
        return jsonify({"error": "one of sold_to / ship_to / sold_to_name / ship_to_name required"}), 400

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cols = _list_columns(cur, "customer")
        wanted = {
            "sold_to":         "sold_to",
            "sold_to_name":    "sold_to_name",
            "ship_to":         "ship_to",
            "ship_to_name":    "ship_to_name",
            "sold_to_group":   "sold_to_group",
            "bde_state":       "bde_state",
            "ship_to_state":   "ship_to_state",
            "channels":        "channels",
            "salesman_name":   "salesman_name",
            # optional / deployment-specific columns
            "address":         "address",
            "address_1":       "address_1",
            "ship_to_address": "ship_to_address",
            "city":            "city",
            "state":           "state",
            "phone":           "phone",
            "telephone":       "telephone",
            "contact_phone":   "contact_phone",
            "mobile":          "mobile",
            "mobile_phone":    "mobile_phone",
            "email":           "email",
            "contact_email":   "contact_email",
        }
        select_cols = [c for c in wanted if c in cols]
        if not select_cols:
            return jsonify({"error": "customer table has no known columns"}), 500

        # Choose the WHERE clause based on which query param was given.
        # Codes: exact.  Names: exact → normalised-LIKE so freely typed
        # "bobstyres" still resolves to "Bob's Tyres".
        row = None
        base_sql = f"SELECT {', '.join(select_cols)} FROM customer"
        if ship_to:
            cur.execute(f"{base_sql} WHERE ship_to = %s LIMIT 1", (ship_to,))
            row = cur.fetchone()
        elif sold_to:
            cur.execute(f"{base_sql} WHERE sold_to = %s LIMIT 1", (sold_to,))
            row = cur.fetchone()
        elif ship_to_name and "ship_to_name" in cols:
            cur.execute(f"{base_sql} WHERE TRIM(ship_to_name) = %s LIMIT 1", (ship_to_name,))
            row = cur.fetchone()
            if not row:
                n = _strip_noise_py(ship_to_name)
                if n:
                    cur.execute(
                        f"{base_sql} WHERE {_strip_noise_sql('ship_to_name')} LIKE %s LIMIT 1",
                        (f"%{n}%",),
                    )
                    row = cur.fetchone()
        elif sold_to_name and "sold_to_name" in cols:
            cur.execute(f"{base_sql} WHERE TRIM(sold_to_name) = %s LIMIT 1", (sold_to_name,))
            row = cur.fetchone()
            if not row:
                n = _strip_noise_py(sold_to_name)
                if n:
                    cur.execute(
                        f"{base_sql} WHERE {_strip_noise_sql('sold_to_name')} LIKE %s LIMIT 1",
                        (f"%{n}%",),
                    )
                    row = cur.fetchone()
        else:
            return jsonify({"error": "no usable lookup key"}), 400

        if not row:
            return jsonify({"error": "not found"}), 404

        out = {k: (row.get(k) or "") for k in select_cols}
        # Canonicalise cross-schema synonyms so the frontend only reads
        # one field name per concept, regardless of which column the
        # deployment actually stores the value in.
        addr_parts = [out.get("address"), out.get("address_1"), out.get("ship_to_address"), out.get("city")]
        out["address"] = ", ".join([p for p in addr_parts if p]) or ""
        out["phone"]   = out.get("phone")     or out.get("telephone")     or out.get("contact_phone") or ""
        out["mobile"]  = out.get("mobile")    or out.get("mobile_phone")  or ""
        out["email"]   = out.get("email")     or out.get("contact_email") or ""
        # ship_to_state is the authoritative jurisdiction on the customer
        # row; bde_state is the sales-org region and can differ from it.
        out["state"]   = out.get("ship_to_state") or out.get("state") or out.get("bde_state") or ""
        out["customer_group"] = out.get("sold_to_group") or ""
        out["bde_name"] = out.get("salesman_name") or ""
        return jsonify(out)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


# Punctuation the suggest layer treats as invisible so "185 70 R14" /
# "185/70R14" / "18570R14" all match the same tyre size, and "Bob's
# Tyres" / "Bobs Tyres" / "bobstyres" all match the same shop.
_NOISE_CHARS = " /,.-'&+()"
def _sql_lit(ch):
    """Quote a single character as a MySQL string literal, doubling
    the apostrophe when it IS the apostrophe.  Without this the
    generated SQL for the apostrophe entry ended up as
    REPLACE(..., ''', '') — a syntax error — which killed the whole
    suggest endpoint on any query."""
    if ch == "'":
        return "''''"      # '' = doubled apostrophe inside a '…' literal
    return "'" + ch + "'"
def _strip_noise_sql(col):
    """MySQL expression that lowers and strips the noise chars from
    a column so LIKE comparisons are punctuation-agnostic."""
    expr = f"LOWER({col})"
    for ch in _NOISE_CHARS:
        expr = f"REPLACE({expr}, {_sql_lit(ch)}, '')"
    return expr
def _strip_noise_py(s):
    s = (s or "").lower()
    for ch in _NOISE_CHARS:
        s = s.replace(ch, "")
    return s


@app.get("/api/orders/customer_suggest")
def api_orders_customer_suggest():
    """Autocomplete backend for the Sold-to / Ship-to inputs.
      ?q=…           1+ char, matched against code OR name (punctuation-agnostic)
      ?kind=sold|ship  which pair to search (default: sold)
      ?limit=15      max rows returned (hard cap 50)
    Returns [{code, name, sold_to, sold_to_name, ship_to, ship_to_name,
             bde_name}, …] so the frontend can populate a datalist with
    "code — name" labels and resolve the full row on select.

    Matching is normalized on BOTH sides — the noise chars in
    `_NOISE_CHARS` (spaces, /, ,, ., -, ', &, +, parens) are stripped
    from the column and from the query before LIKE runs, so typing
    "bobstyres" finds "Bob's Tyres" and typing "18570r14" finds a
    material stored as "185/70R14".  Multi-word queries are AND-ed
    per token (case-insensitive) so word order doesn't matter."""
    q    = (request.args.get("q") or "").strip()
    kind = (request.args.get("kind") or "sold").strip().lower()
    try:
        limit = max(1, min(int(request.args.get("limit", 15) or 15), 50))
    except ValueError:
        limit = 15
    if not q:
        return jsonify([])

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cols = _list_columns(cur, "customer")
        code_col = "ship_to" if kind == "ship" else "sold_to"
        name_col = "ship_to_name" if kind == "ship" else "sold_to_name"
        if code_col not in cols:
            return jsonify({"error": f"{code_col} column missing"}), 500

        # Discover which optional label columns exist so the response
        # can carry them for the frontend to use directly on pick.
        pick_cols = [c for c in ("sold_to", "sold_to_name", "ship_to", "ship_to_name",
                                 "salesman_name") if c in cols]
        if code_col not in pick_cols: pick_cols.append(code_col)

        # DISTINCT collapses the multi-ship_to-per-sold_to duplication
        # to a single suggestion per (code, name) pair.
        select_sql = "DISTINCT " + ", ".join(pick_cols)

        # Build per-token AND — each whitespace-separated token in q
        # must appear in the stripped code OR the stripped name.
        tokens = [t for t in q.split() if t.strip()]
        if not tokens:
            tokens = [q]
        code_expr = _strip_noise_sql(code_col)
        name_expr = _strip_noise_sql(name_col) if name_col in cols else None
        wh_and = []
        params = []
        for tok in tokens:
            tok_norm = _strip_noise_py(tok)
            if not tok_norm:
                continue
            per_tok = [f"{code_expr} LIKE %s"]
            params.append(f"%{tok_norm}%")
            if name_expr:
                per_tok.append(f"{name_expr} LIKE %s")
                params.append(f"%{tok_norm}%")
            wh_and.append("(" + " OR ".join(per_tok) + ")")
        if not wh_and:
            return jsonify([])

        sql = (
            f"SELECT {select_sql} FROM customer "
            f"WHERE {' AND '.join(wh_and)} "
            f"ORDER BY {code_col} LIMIT {limit}"
        )
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()

        # Shape a compact, uniform response.
        out = []
        for r in rows:
            code = (r.get(code_col) or "").strip()
            name = (r.get(name_col) or "").strip() if name_col in r else ""
            out.append({
                "code":         code,
                "name":         name,
                "sold_to":      (r.get("sold_to") or "").strip(),
                "sold_to_name": (r.get("sold_to_name") or "").strip(),
                "ship_to":      (r.get("ship_to") or "").strip(),
                "ship_to_name": (r.get("ship_to_name") or "").strip(),
                "bde_name":     (r.get("salesman_name") or "").strip(),
            })
        return jsonify(out)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.get("/api/orders/material_suggest")
def api_orders_material_suggest():
    """Autocomplete backend for the M-Code / Description inputs on the
    Orders form.  Same shape as /api/orders/customer_suggest — the
    frontend populates a datalist with "code — description" labels and
    calls /api/orders/material to fill the row on selection."""
    q = (request.args.get("q") or "").strip()
    try:
        limit = max(1, min(int(request.args.get("limit", 15) or 15), 50))
    except ValueError:
        limit = 15
    if not q:
        return jsonify([])

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cols = _list_columns(cur, "carrying_26")
        c_desc  = "description" if "description" in cols else ("size" if "size" in cols else None)
        c_brand = "brand" if "brand" in cols else None
        # carrying_26 doesn't ship a dedicated product_name column;
        # product_group is what the BDE reads as the product family
        # (e.g. "VENTUS S1 EVO3"), so alias it into the response as
        # product_name and keep the real product_name column as a
        # deployment-specific fallback for schemas that have it.
        c_prod  = next((c for c in ("product_name", "product_group") if c in cols), None)
        c_pat   = "pattern" if "pattern" in cols else None
        if "m_code" not in cols:
            return jsonify({"error": "m_code column missing"}), 500

        select_parts = ["m_code"]
        if c_desc:  select_parts.append(f"{c_desc} AS description")
        if c_brand: select_parts.append("brand")
        if c_prod:  select_parts.append(f"{c_prod} AS product_name")
        if c_pat:   select_parts.append("pattern")

        # Punctuation-agnostic multi-token match — see
        # customer_suggest for the rationale.  "185 70r14" and
        # "18570R14" both find "185/70R14…".  Pattern is added to the
        # OR set so typing "215 70r14 h724" narrows to the exact tread.
        code_expr = _strip_noise_sql("m_code")
        desc_expr = _strip_noise_sql(c_desc) if c_desc else None
        pat_expr  = _strip_noise_sql(c_pat)  if c_pat  else None
        tokens = [t for t in q.split() if t.strip()] or [q]
        wh_and = []
        params = []
        for tok in tokens:
            tok_norm = _strip_noise_py(tok)
            if not tok_norm: continue
            per_tok = [f"{code_expr} LIKE %s"]
            params.append(f"%{tok_norm}%")
            if desc_expr:
                per_tok.append(f"{desc_expr} LIKE %s")
                params.append(f"%{tok_norm}%")
            if pat_expr:
                per_tok.append(f"{pat_expr} LIKE %s")
                params.append(f"%{tok_norm}%")
            wh_and.append("(" + " OR ".join(per_tok) + ")")
        if not wh_and:
            return jsonify([])

        sql = (
            f"SELECT DISTINCT {', '.join(select_parts)} FROM carrying_26 "
            f"WHERE {' AND '.join(wh_and)} "
            f"ORDER BY m_code LIMIT {limit}"
        )
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
        out = [{
            "m_code":       (r.get("m_code") or "").strip(),
            "description":  (r.get("description") or "").strip(),
            "brand":        (r.get("brand") or "").strip() if c_brand else "",
            "product_name": (r.get("product_name") or "").strip() if c_prod else "",
            "pattern":      (r.get("pattern") or "").strip() if c_pat else "",
        } for r in rows]
        return jsonify(out)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.get("/api/orders/material")
def api_orders_material():
    """Look up a carrying_26 row by ?m_code=... or ?description=... .
    On description match, does an exact-then-LIKE fallback so half-
    typed sizes still resolve.  Returns the fields the Orders form
    needs (brand, product_name, pattern, load/speed, list_price).
    Falls back to empty strings for columns absent on this schema."""
    m_code = (request.args.get("m_code") or "").strip()
    desc   = (request.args.get("description") or "").strip()
    if not m_code and not desc:
        return jsonify({"error": "m_code or description required"}), 400

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cols = _list_columns(cur, "carrying_26")
        # Column-name aliases across deployments.
        pick = lambda *cands: next((c for c in cands if c in cols), None)
        c_m_code       = pick("m_code")
        c_description  = pick("description", "size", "material_desc")
        c_brand        = pick("brand")
        # product_group is what carrying_26 actually carries as the
        # readable product-family label ("VENTUS S1 EVO3", "KINERGY
        # ECO2"…), so use it as the product_name source when a real
        # product_name column isn't present.
        c_product_name = pick("product_name", "product", "product_group", "pattern_name")
        # product_group is a separate, structured product family label
        # (Dynapro, Kinergy, Ventus…) — the additional-DC lookup joins
        # against promo_plan on it, so surface it independently even
        # when it's also the product_name source.
        c_product_group = pick("product_group")
        c_pattern      = pick("pattern")
        c_line         = pick("line")
        c_load         = pick("load_speed", "load", "load_index")
        c_speed        = pick("speed", "speed_rating")
        c_list_price   = pick("list_price", "price")
        if not c_m_code:
            return jsonify({"error": "carrying_26 has no m_code column"}), 500

        parts = [f"{c_m_code} AS m_code"]
        if c_description:   parts.append(f"{c_description} AS description")
        if c_brand:         parts.append(f"{c_brand} AS brand")
        if c_product_name:  parts.append(f"{c_product_name} AS product_name")
        if c_product_group: parts.append(f"{c_product_group} AS product_group")
        if c_pattern:       parts.append(f"{c_pattern} AS pattern")
        if c_line:          parts.append(f"{c_line} AS line")
        # Combine load + speed into a single string like "88H" when
        # split — the source form always displays it joined.
        if c_load and c_speed and c_load != c_speed:
            parts.append(f"CONCAT_WS('', {c_load}, {c_speed}) AS load_speed")
        elif c_load:
            parts.append(f"{c_load} AS load_speed")
        if c_list_price:   parts.append(f"{c_list_price} AS list_price")

        select_sql = ", ".join(parts)

        if m_code:
            # Exact m_code first, then a normalised (case+noise-stripped)
            # fallback so "1017693 " or "1,017,693" still resolve.
            cur.execute(f"SELECT {select_sql} FROM carrying_26 WHERE {c_m_code} = %s LIMIT 1", (m_code,))
            row = cur.fetchone()
            if not row:
                mnorm = _strip_noise_py(m_code)
                if mnorm:
                    cur.execute(
                        f"SELECT {select_sql} FROM carrying_26 "
                        f"WHERE {_strip_noise_sql(c_m_code)} = %s LIMIT 1",
                        (mnorm,),
                    )
                    row = cur.fetchone()
        else:
            if not c_description:
                return jsonify({"error": "description column not available"}), 500
            # The Size dropdown now offers "size · pattern" as the
            # picked value so a single size with multiple patterns is
            # disambiguated at the point of pick.  Split on " · " and,
            # if a pattern component is present, gate matches on both
            # columns — otherwise fall through to size-only matching.
            desc_part, pat_part = desc, ""
            if " · " in desc:
                left, right = desc.split(" · ", 1)
                desc_part, pat_part = left.strip(), right.strip()

            row = None
            if pat_part and c_pattern:
                # size + pattern → the (size, pattern) pair is unique in
                # carrying_26, so this exact-match wins first.
                cur.execute(
                    f"SELECT {select_sql} FROM carrying_26 "
                    f"WHERE {c_description} = %s AND {c_pattern} = %s LIMIT 1",
                    (desc_part, pat_part),
                )
                row = cur.fetchone()

            if not row:
                # Exact size, then LIKE, then normalised-LIKE fallback —
                # the last step rescues freely-typed "18570R14" against
                # a raw "185/70R14…" via plain LIKE.
                cur.execute(
                    f"SELECT {select_sql} FROM carrying_26 WHERE {c_description} = %s LIMIT 1",
                    (desc_part,),
                )
                row = cur.fetchone()
            if not row:
                cur.execute(
                    f"SELECT {select_sql} FROM carrying_26 WHERE {c_description} LIKE %s LIMIT 1",
                    (f"%{desc_part}%",),
                )
                row = cur.fetchone()
            if not row:
                # Noise-stripped LIKE — the last-resort rescue for free
                # typing.  When a pattern was carried along (picked from
                # the "stripped-size · pattern" dropdown entry) gate on
                # the pattern too so a size shared by several treads
                # doesn't resolve to the wrong one.
                dnorm = _strip_noise_py(desc_part)
                if dnorm:
                    if pat_part and c_pattern:
                        cur.execute(
                            f"SELECT {select_sql} FROM carrying_26 "
                            f"WHERE {_strip_noise_sql(c_description)} LIKE %s "
                            f"  AND {c_pattern} = %s LIMIT 1",
                            (f"%{dnorm}%", pat_part),
                        )
                        row = cur.fetchone()
                    if not row:
                        cur.execute(
                            f"SELECT {select_sql} FROM carrying_26 "
                            f"WHERE {_strip_noise_sql(c_description)} LIKE %s LIMIT 1",
                            (f"%{dnorm}%",),
                        )
                        row = cur.fetchone()

        if not row:
            return jsonify({"error": "not found"}), 404

        # Normalise numeric list_price (comes as Decimal from MySQL)
        if row and row.get("list_price") is not None:
            try:
                row["list_price"] = float(row["list_price"])
            except Exception:
                row["list_price"] = None

        # Load/speed fallback — carrying_26 on this deployment doesn't
        # have a dedicated load_speed column, so hunt for the token in
        # any text column we already know about (description, size,
        # product_name) AND take a wide net across every other text-
        # like column in the row we already fetched.  Formats we've
        # seen: "205/55R16 91V", "265/70R16 112T", "205R16C 110/108T",
        # "11R22.5 148/145L", "205/55R16, 91V", "205/55R16-91V".
        if not row.get("load_speed"):
            import re as _re_ls
            # Fetch every text column for this m_code so a load-speed
            # stashed in, say, a `description_full` or `spec` we didn't
            # explicitly probe still gets found.
            wide_src = " ".join(str(v) for v in row.values() if v is not None)
            try:
                if c_m_code:
                    cur.execute(
                        f"SELECT * FROM carrying_26 WHERE {c_m_code} = %s LIMIT 1",
                        (row.get("m_code"),),
                    )
                    full = cur.fetchone() or {}
                    wide_src += " " + " ".join(str(v) for v in full.values() if v is not None)
            except Exception:
                pass
            # Loosen separator: any non-alphanumeric (space, comma,
            # dash, tab, punctuation) between the R-size and the
            # load/speed token.  Also accept load/speed WITHOUT a
            # trailing space (some datasets glue them together).
            m = _re_ls.search(
                r"R\s*\d{1,2}(?:\.\d)?C?[^A-Za-z0-9/]*"
                r"(\d{2,3}(?:/\d{2,3})?[A-Z]{1,2})\b",
                wide_src,
            )
            if m:
                row["load_speed"] = m.group(1)
        return jsonify(row)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

# plant 醫뚰몴 留ㅽ븨 (?덇? 以?媛믪쑝濡??낅뜲?댄듃)
# Plant → State mapping for the Orders form's per-state stock columns.
# Kept next to the endpoint that uses it so the schema-agnostic lookup
# doesn't have to reach into PLANT_GEO for something that's really
# about jurisdictions, not geometry.
_ORDER_PLANT_STATE = {"42R1": "NSW", "42R0": "QLD", "42R2": "VIC", "42R4": "WA"}


@app.get("/api/orders/aged_stock_at_state")
def api_orders_aged_stock_at_state():
    """Aged-stock list for one state — used by the /order form's right
    side panel to surface stock the BDE could push during a customer
    visit.  Only returns 19-24M / 25-36M / 37M+ tiers (fresh stock
    isn't a talking point).  Ordered by tier (oldest first within
    each tier)."""
    state = (request.args.get("state") or "").strip().upper()
    _STATE_TO_PLANT = {"NSW": "42R1", "QLD": "42R0",
                       "VIC": "42R2", "WA":  "42R4",
                       # SA/TAS fold into VIC territory; NT into WA;
                       # ACT into NSW — matches the rest of the app.
                       "SA":  "42R2", "TAS": "42R2",
                       "NT":  "42R4", "ACT": "42R1"}
    plant = _STATE_TO_PLANT.get(state)
    if not plant:
        return jsonify({"error": "unknown state", "rows": []}), 400

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT s.material, s.dot_no, "
            "       SUM(s.stock_qty) AS qty, "
            "       MAX(c.size)    AS size, "
            "       MAX(c.pattern) AS pattern "
            "FROM stock s "
            "LEFT JOIN carrying_26 c ON c.m_code = s.material "
            "WHERE s.plant = %s AND s.stock_qty > 0 "
            "GROUP BY s.material, s.dot_no",
            (plant,),
        )
        raw = cur.fetchall() or []
    finally:
        try: cur.close(); conn.close()
        except: pass

    # Only the three aged tiers surface — fresh (≤12M / 13-18M) is
    # ignored per BDE workflow: BDEs push OLD stock, not new.
    AGED_TIERS = {"19-24M", "25-36M", "37M+"}
    TIER_RANK  = {"19-24M": 0, "25-36M": 1, "37M+": 2}
    # Roll up per (material, dot_no) → (material, tier) so a material
    # with several old DOTs collapses to one row per tier.
    by_key: dict = {}
    for r in raw:
        age  = _dot_age_months(r["dot_no"])
        tier = _age_bucket(age)
        if tier not in AGED_TIERS: continue
        key = (str(r["material"] or "").strip(), tier)
        d = by_key.setdefault(key, {
            "m_code":  str(r["material"] or "").strip(),
            "size":    (r.get("size")    or "").strip(),
            "pattern": (r.get("pattern") or "").strip(),
            "tier":    tier,
            "qty":     0,
        })
        d["qty"] += int(r["qty"] or 0)
    rows = list(by_key.values())
    # Sort: 19-24M first, then 25-36M, then 37M+; within each tier
    # oldest / biggest wins.
    rows.sort(key=lambda r: (TIER_RANK[r["tier"]], -r["qty"]))
    return jsonify({
        "state":  state,
        "plant":  plant,
        "rows":   rows,
        "total":  sum(r["qty"] for r in rows),
    })


@app.get("/api/orders/stock_aging_by_material")
def api_orders_stock_aging_by_material():
    """Aging bucket breakdown for one (material, plant).  Called when
    a state stock cell on the Special Price Request form is clicked so
    the BDE can see how fresh / aged the on-hand stock is before
    quoting a discount.

    Groups every (material, plant, dot_no) row by the same 5 aging
    buckets the /stock aging popup uses (≤12M … 37M+, plus 'unknown'
    for malformed DOTs).  If plant isn't given, aggregates across all
    plants (NSW/QLD/VIC/WA) so the caller can also drive a nation-
    wide view from the same endpoint."""
    m_code = (request.args.get("m_code") or "").strip()
    plant  = (request.args.get("plant")  or "").strip()
    if not m_code:
        return jsonify({"error": "m_code required"}), 400

    where = ["s.material = %s", "s.stock_qty > 0"]
    ps    = [m_code]
    if plant:
        where.append("s.plant = %s")
        ps.append(plant)
    else:
        # Constrain to the four real plants so leaked test data
        # (fake 4-char plants) doesn't skew the totals.
        where.append(
            f"s.plant IN ({','.join(['%s'] * len(_ORDER_PLANT_STATE))})"
        )
        ps.extend(_ORDER_PLANT_STATE.keys())

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"SELECT s.dot_no, SUM(s.stock_qty) AS qty "
            f"FROM stock s "
            f"WHERE {' AND '.join(where)} "
            f"GROUP BY s.dot_no",
            tuple(ps),
        )
        rows = cur.fetchall() or []
    finally:
        try: cur.close(); conn.close()
        except: pass

    buckets = {b: 0 for b in _AGING_BUCKETS}
    buckets["unknown"] = 0
    for r in rows:
        age = _dot_age_months(r["dot_no"])
        b   = _age_bucket(age)
        buckets[b] += int(r["qty"] or 0)
    return jsonify({
        "m_code":       m_code,
        "plant":        plant or None,
        "state":        _ORDER_PLANT_STATE.get(plant, "") if plant else "",
        "buckets":      buckets,
        "bucket_order": _AGING_BUCKETS + ["unknown"],
        "total":        sum(buckets.values()),
    })


@app.get("/api/orders/stock_by_material")
def api_orders_stock_by_material():
    """Return {NSW, QLD, VIC, WA, TOTAL} stock for one or more materials.
    Batches multiple m_codes when ?m_code=... appears more than once so
    the form only pays for one round trip when a whole page of rows
    resolves at once.  Reads the same `stock_qty` column the Stock page
    shows (ZSDM64300-derived), aggregated per (material, plant) and
    folded to state.  Missing (material, plant) pairs come back as 0."""
    codes = [c.strip() for c in request.args.getlist("m_code") if c.strip()]
    if not codes:
        return jsonify({})

    ph = ",".join(["%s"] * len(codes))
    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"SELECT material, plant, SUM(stock_qty) AS qty "
            f"FROM stock "
            f"WHERE material IN ({ph}) "
            f"  AND plant IN ({','.join(['%s']*len(_ORDER_PLANT_STATE))}) "
            f"GROUP BY material, plant",
            tuple(codes) + tuple(_ORDER_PLANT_STATE.keys()),
        )
        rows = cur.fetchall() or []
        out = {c: {s: 0 for s in _ORDER_PLANT_STATE.values()} for c in codes}
        for r in rows:
            m = str(r.get("material") or "").strip()
            st = _ORDER_PLANT_STATE.get(r.get("plant"))
            if not m or not st or m not in out:
                continue
            out[m][st] = int(float(r.get("qty") or 0))
        for m, buckets in out.items():
            buckets["TOTAL"] = sum(buckets.values())
        return jsonify(out)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


# TBR (HK & LF) discount is a fixed 56% for every customer at the
# moment — the dc_basic_customer feed doesn't split PCLT vs TBR, so
# the two brand rows there feed the PCLT columns and TBR is stamped
# from this constant.  Move to a separate table once the feed is
# split by line.
_ORDERS_TBR_HKLF_PCT = 56.00


@app.get("/api/orders/base_dc")
def api_orders_base_dc():
    """Look up Base DC % for a sold-to from dc_basic_customer.

      ?sold_to=…        the bill_to_partner code on that table

    Returns
      { "HK_PCLT": 52.00, "LF_PCLT": 50.00, "TBR_HKLF": 56.00 }

    For each output cell (HK-PCLT, LF-PCLT, TBR) we score every
    candidate row and pick the highest.  Score is:
        customer_lvl * 100 + brand_match * 10 + line_match
    where each component is 2 for a row that matches the target
    exactly on that dimension and 1 for a blank-value fallback.
    A row whose brand / line is populated but doesn't equal the
    target is skipped (wrong bucket).

    This covers every combination in one pass:
      customer+brand+line  (2·100 + 2·10 + 2 = 222)   most specific
      customer+brand+blank (221) ▸ customer+blank+line (212)
      customer+blank+blank (211) ▸ group+brand+line (122)
      group+brand+blank    (121) ▸ group+blank+line  (112)
      group+blank+blank    (111)                     Group Basic DC

    HK-PCLT / LF-PCLT target (HK, PCLT) / (LF, PCLT); TBR targets
    ("", TBR) — for TBR only blank-brand rows score at all, so a
    branded row can't leak into the TBR cell.

    If no candidate scores, TBR falls back to _ORDERS_TBR_HKLF_PCT
    and the PCLT cells stay blank on the form.  When the table
    isn't loaded, the endpoint returns those same defaults silently."""
    sold_to = (request.args.get("sold_to") or "").strip()
    debug   = request.args.get("debug") in ("1", "true", "yes")
    out = {"HK_PCLT": None, "LF_PCLT": None, "TBR_HKLF": _ORDERS_TBR_HKLF_PCT}
    if debug:
        out["_debug"] = {"sold_to": sold_to}
    if not sold_to:
        return jsonify(out)

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # Silently no-op when the table isn't loaded on this deployment.
        try:
            cur.execute("SHOW TABLES LIKE 'dc_basic_customer'")
            if not cur.fetchone():
                return jsonify(out)
        except Exception:
            return jsonify(out)

        # Column probes — deployments differ on the group column name;
        # the line column was added in the most recent update.
        cols = _list_columns(cur, "dc_basic_customer")
        if debug:
            out["_debug"]["cols"] = sorted(cols)
        group_col = next((c for c in [
            "customer_grp", "customer_group", "sold_to_group",
            "bill_to_group", "customer", "customer_code",
            "group_code", "grouping", "cust_group", "cust_grp",
        ] if c in cols), None)
        has_line_col = "line" in cols
        if debug:
            out["_debug"]["group_col"]    = group_col
            out["_debug"]["has_line_col"] = has_line_col

        # Resolve customer group so tier 3/4/5 can match.  A single
        # sold_to can have several customer rows (one per ship_to);
        # some ship-only rows may carry a blank sold_to_group.  Take
        # the MAX of the non-blank values so a stray blank row doesn't
        # sink the whole lookup with LIMIT 1.
        group_code = ""
        try:
            cur.execute(
                "SELECT MAX(NULLIF(TRIM(sold_to_group), '')) AS grp "
                "FROM customer WHERE sold_to = %s",
                (sold_to,))
            gr = cur.fetchone()
            if gr:
                group_code = (gr.get("grp") or "").strip()
        except Exception as _e:
            if debug:
                out["_debug"]["group_code_error"] = str(_e)
        if debug:
            out["_debug"]["group_code"] = group_code

        # Pull every candidate row for this customer + their group in a
        # single sweep, then rank in Python.  Cheap even without an
        # index because dc_basic_customer is small and this query is
        # scoped to one sold_to and one group_code.
        select_cols = "bill_to_partner, brand, amount, valid_from, valid_to"
        if group_col:    select_cols += f", {group_col} AS customer_grp"
        else:            select_cols += ", NULL AS customer_grp"
        if has_line_col: select_cols += ", line"
        else:            select_cols += ", '' AS line"
        params = [sold_to]
        where_group = ""
        if group_col and group_code:
            where_group = (
                f" OR ((bill_to_partner IS NULL OR TRIM(bill_to_partner) = '')"
                f"     AND UPPER(TRIM({group_col})) = UPPER(%s))"
            )
            params.append(group_code)
        cur.execute(
            f"SELECT {select_cols} FROM dc_basic_customer "
            f"WHERE (TRIM(bill_to_partner) = %s {where_group}) "
            f"  AND (valid_from IS NULL OR valid_from <= CURDATE()) "
            f"  AND (valid_to   IS NULL OR valid_to   >= CURDATE())",
            tuple(params))
        raw_rows = cur.fetchall() or []

        # Normalise once, then apply the 5-tier priority for each cell.
        s_sold  = sold_to
        s_group = (group_code or "").upper()
        norm = []
        for r in raw_rows:
            r_bill  = (r.get("bill_to_partner") or "").strip()
            r_brand = (r.get("brand") or "").strip().upper()
            r_line  = (r.get("line")  or "").strip().upper()
            r_grp   = (r.get("customer_grp") or "").strip().upper()
            try:    r_amt = float(r.get("amount") or 0)
            except Exception: r_amt = 0.0
            norm.append({
                "bill": r_bill, "brand": r_brand, "line": r_line,
                "grp": r_grp, "amount": r_amt,
                "valid_from": r.get("valid_from"),
            })

        def _pick(target_brand, target_line):
            """Return (score, valid_from, amount, row) for the best row,
            or None if nothing matches.  Score is customer_lvl * 100 +
            brand_match * 10 + line_match, where each component is 2
            for an exact match to the target and 1 for a blank-value
            row that acts as fallback.  Higher score = more specific
            = wins.  A row's brand / line either matches the target
            exactly, is blank (fallback), or the row is skipped as
            not applicable.  This covers every dc_basic_customer
            combination (customer or group × brand-specific or blank
            × line-specific or blank) without hard-coding tiers."""
            tb = (target_brand or "").upper()
            tl = (target_line  or "").upper()
            best = None
            for r in norm:
                is_customer = (r["bill"] == s_sold)
                is_group    = (r["bill"] == "" and r["grp"] == s_group and s_group)
                if not (is_customer or is_group):
                    continue
                # Brand gate.
                if tb:  # HK or LF target — accept HK/LF exact OR blank
                    if r["brand"] == tb:  brand_score = 2
                    elif r["brand"] == "": brand_score = 1
                    else: continue        # wrong brand
                else:   # TBR target — only blank-brand rows apply
                    if r["brand"] == "":  brand_score = 2
                    else: continue        # branded row doesn't leak into TBR
                # Line gate.
                if tl:
                    if r["line"] == tl:   line_score = 2
                    elif r["line"] == "": line_score = 1
                    else: continue        # wrong line
                else:
                    if r["line"] == "":   line_score = 2
                    else: continue
                cust_score = 2 if is_customer else 1
                score = cust_score * 100 + brand_score * 10 + line_score
                key = r["valid_from"] or datetime(1900, 1, 1).date()
                cand = (score, key, abs(r["amount"]), r)
                if best is None:
                    best = cand
                elif cand[0] > best[0]:
                    best = cand
                elif cand[0] == best[0] and cand[1] > best[1]:
                    best = cand
            return best

        hk_pick  = _pick("HK", "PCLT")
        lf_pick  = _pick("LF", "PCLT")
        tbr_pick = _pick("",   "TBR")

        if hk_pick:  out["HK_PCLT"]  = hk_pick[2]
        if lf_pick:  out["LF_PCLT"]  = lf_pick[2]
        if tbr_pick: out["TBR_HKLF"] = tbr_pick[2]

        if debug:
            def _s(v):
                try: return v.isoformat()
                except Exception: return v
            def _dump(pk):
                if not pk: return None
                return {"score": pk[0], "valid_from": _s(pk[1]),
                        "amount": pk[2], "row": {k: _s(v) for k,v in pk[3].items()}}
            out["_debug"]["picks"] = {
                "HK_PCLT": _dump(hk_pick),
                "LF_PCLT": _dump(lf_pick),
                "TBR":     _dump(tbr_pick),
            }
            out["_debug"]["rows"] = [
                {k: _s(v) for k, v in r.items()} for r in raw_rows
            ]
        return jsonify(out)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.get("/api/orders/additional_dc")
def api_orders_additional_dc():
    """Look up an Additional Special DC % for one order line.

      ?sold_to=…        bill-to code on the customer master
      ?brand=…          HK | LF (row's brand from carrying_26)
      ?product_group=…  Dynapro | Kinergy | … (row's product_group)
      ?qty=…            row qty (min_qty threshold gate)

    Returns
      { "additional_dc": 30.00, "promo": "443", "min_qty": 4,
        "source": "sold_to" | "customer_grp", "matched": true }
      or { "additional_dc": null, "matched": false } when nothing
      qualifies (form leaves the add_dc cell alone).

    Lookup joins dc_additional_customer × promo_plan:
      1. Row keyed to this bill-to exactly beats a group row.
      2. Brand must match.
      3. min_qty must be <= this line's qty.
      4. The row's promo (443, iON, …) must have at least one
         promo_plan entry where the product_group matches (or is
         blank = all groups) AND CURDATE() falls in the plan's
         start_date / end_date window.
    If several rows still qualify, the highest additional_dc wins
    (best discount for the customer).  Missing table / missing
    plan / no match all return matched:false; the endpoint never
    500s under normal misses."""
    sold_to       = (request.args.get("sold_to") or "").strip()
    brand         = (request.args.get("brand") or "").strip()
    product_group = (request.args.get("product_group") or "").strip()
    try:
        qty = float((request.args.get("qty") or "0").replace(",", "").strip())
    except Exception:
        qty = 0.0
    debug = request.args.get("debug") in ("1", "true", "yes")

    out = {"additional_dc": None, "matched": False}
    if debug:
        out["_debug"] = {"sold_to": sold_to, "brand": brand,
                         "product_group": product_group, "qty": qty}
    if not sold_to or not brand:
        return jsonify(out)

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # Silently no-op when the table isn't loaded — the form still
        # functions with the cell blank.
        try:
            cur.execute("SHOW TABLES LIKE 'dc_additional_customer'")
            if not cur.fetchone():
                return jsonify(out)
        except Exception:
            return jsonify(out)

        # Resolve customer group for the group-fallback tier.
        group_code = ""
        try:
            cur.execute(
                "SELECT sold_to_group FROM customer WHERE sold_to = %s LIMIT 1",
                (sold_to,),
            )
            gr = cur.fetchone()
            if gr:
                group_code = (gr.get("sold_to_group") or "").strip()
        except Exception:
            pass
        if debug:
            out["_debug"]["group_code"] = group_code

        # Column names are pinned to what the CSV shipped with
        # (customer_grp / sold_to / promo / brand / min_qty /
        # additional_dc).  If a deployment renames one, this endpoint
        # will silently miss — small enough surface that a rename is
        # a code change here.
        ac_cols = _list_columns(cur, "dc_additional_customer")
        needed = {"customer_grp", "sold_to", "promo", "brand",
                  "min_qty", "additional_dc"}
        if not needed.issubset(ac_cols):
            if debug:
                out["_debug"]["missing_cols"] = sorted(needed - ac_cols)
                out["_debug"]["cols"] = sorted(ac_cols)
            return jsonify(out)

        # Also check promo_plan exists — the endpoint only makes sense
        # with the join.  If missing, ignore the promo gate (treat as
        # always in-plan) so an early rollout still surfaces something.
        pp_exists = False
        try:
            cur.execute("SHOW TABLES LIKE 'promo_plan'")
            pp_exists = bool(cur.fetchone())
        except Exception:
            pp_exists = False
        if debug:
            out["_debug"]["promo_plan"] = pp_exists

        # Build the promo gate — pp.product_group blank = "applies to
        # every product group" (matches how the existing promo filter
        # in _promo_filter_clauses treats blank).  Dates NULL/blank
        # treated as open-ended.
        promo_gate = ""
        if pp_exists:
            promo_gate = """
              AND EXISTS (
                SELECT 1 FROM promo_plan pp
                WHERE pp.promo = ac.promo
                  AND (pp.product_group IS NULL
                       OR TRIM(pp.product_group) = ''
                       OR UPPER(TRIM(pp.product_group)) = UPPER(%s))
                  AND (pp.start_date IS NULL OR pp.start_date <= CURDATE())
                  AND (pp.end_date   IS NULL OR pp.end_date   >= CURDATE())
              )"""
        # Group fallback fires only when we resolved a group code.
        group_join = ""
        params = [sold_to]
        if group_code:
            group_join = (
                " OR ((ac.sold_to IS NULL OR TRIM(ac.sold_to) = '')"
                "     AND UPPER(TRIM(ac.customer_grp)) = UPPER(%s))"
            )
            params.append(group_code)
        params += [brand, qty]
        if pp_exists:
            params.append(product_group)

        # Tier 1 (sold_to match) sorts ahead of tier 2 (group match).
        # Within the same tier, the biggest discount wins — the user
        # gets the best applicable rate.  min_qty DESC on ties so a
        # tighter-threshold row is preferred over a loose one at the
        # same %.
        sql = f"""
            SELECT ac.additional_dc, ac.min_qty, ac.promo,
                   ac.customer_grp, ac.sold_to, ac.brand,
                   CASE WHEN TRIM(ac.sold_to) = %s THEN 'sold_to'
                        ELSE 'customer_grp' END AS source,
                   CASE WHEN TRIM(ac.sold_to) = %s THEN 1 ELSE 2 END AS tier
            FROM dc_additional_customer ac
            WHERE (
                    TRIM(ac.sold_to) = %s
                    {group_join}
                  )
              AND UPPER(TRIM(ac.brand)) = UPPER(%s)
              AND (ac.min_qty IS NULL OR ac.min_qty <= %s)
              {promo_gate}
            ORDER BY tier ASC,
                     ABS(ac.additional_dc) DESC,
                     ac.min_qty DESC
            LIMIT 1
        """
        # Params order: source CASE, tier CASE, WHERE sold_to,
        # [group_code], brand, qty, [product_group].
        exec_params = [sold_to, sold_to] + params
        cur.execute(sql, tuple(exec_params))
        row = cur.fetchone()
        if debug:
            out["_debug"]["sql_params"] = exec_params
            out["_debug"]["row"] = row
        if not row:
            return jsonify(out)
        try:
            add_dc = float(row["additional_dc"])
        except Exception:
            add_dc = None
        if add_dc is None:
            return jsonify(out)
        # additional_dc is stored positive on the feed (30.00 = 30%
        # discount); pass it through as-is so the form adds it to the
        # DC stack directly.
        out["additional_dc"] = abs(add_dc)
        out["matched"]       = True
        out["promo"]         = row.get("promo")
        out["min_qty"]       = row.get("min_qty")
        out["source"]        = row.get("source")
        return jsonify(out)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e), "additional_dc": None, "matched": False}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


# ══════════════════════════════════════════════════════════════════════
# Submitted-orders workflow — BDE fills the Special Price Request Form,
# hits Submit, and the whole thing (header + lines + totals) lands in
# submitted_orders.  Harry (CS) then flips the SAP-entered flag to Y
# once he's keyed it in on the SAP side.  All state lives here — the
# order form and the list page both read/write via /api/orders/* below.
# ══════════════════════════════════════════════════════════════════════
HARRY_CS_EMAIL       = "harry.jallis@hankooktyre.com.au"
# Parallel approvers — either can approve independently; the front-end
# shows both statuses side-by-side.
MGMT_APPROVER_EMAILS = [
    "hayden.begbie@hankooktyre.com.au",
    "junjong.cho@hankooktyre.com.au",
]

def _ensure_submitted_orders_table():
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS submitted_orders (
                id                 BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                submitted_at       DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                submitted_by_bde   VARCHAR(120) NOT NULL DEFAULT '',
                submitted_by_email VARCHAR(255) NOT NULL DEFAULT '',
                sold_to            VARCHAR(40)  NOT NULL DEFAULT '',
                sold_to_name       VARCHAR(255) NOT NULL DEFAULT '',
                ship_to            VARCHAR(40)  NOT NULL DEFAULT '',
                ship_to_name       VARCHAR(255) NOT NULL DEFAULT '',
                state              VARCHAR(20)  NOT NULL DEFAULT '',
                po_number          VARCHAR(100) NOT NULL DEFAULT '',
                order_date         VARCHAR(20)  NOT NULL DEFAULT '',
                subtotal           DECIMAL(14,2) NOT NULL DEFAULT 0,
                total_inc_gst      DECIMAL(14,2) NOT NULL DEFAULT 0,
                freight_amount     DECIMAL(14,2) NOT NULL DEFAULT 0,
                grand_total        DECIMAL(14,2) NOT NULL DEFAULT 0,
                total_qty          INT          NOT NULL DEFAULT 0,
                sovd_qty           INT          NOT NULL DEFAULT 0,
                avg_dc_pct         DECIMAL(6,2) NOT NULL DEFAULT 0,
                status_sap         CHAR(1)      NOT NULL DEFAULT 'N',
                status_changed_at  DATETIME     NULL,
                status_changed_by  VARCHAR(255) NOT NULL DEFAULT '',
                payload_json       LONGTEXT     NOT NULL,
                INDEX idx_so_submitted (submitted_at),
                INDEX idx_so_status    (status_sap, submitted_at),
                INDEX idx_so_bde       (submitted_by_bde),
                INDEX idx_so_sold_to   (sold_to)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Parallel management-approval columns — added idempotently so
        # older deployments upgrade without a separate migration.
        # needs_mgmt_approval = 'Y' when the BDE ticked the box on the
        # form.  approved_a / approved_b track the two named approvers
        # independently (Hayden = approver A, JunJong = approver B).
        # mgmt_reason is a copy of the yellow "reason behind pricing"
        # textarea, surfaced as a column so it's queryable / visible in
        # the list without unpacking payload_json.
        _add_cols = [
            ("needs_mgmt_approval", "CHAR(1) NOT NULL DEFAULT 'N'"),
            ("mgmt_reason",         "TEXT NULL"),
            ("approved_a",          "CHAR(1) NOT NULL DEFAULT 'N'"),
            ("approved_a_at",       "DATETIME NULL"),
            ("approved_b",          "CHAR(1) NOT NULL DEFAULT 'N'"),
            ("approved_b_at",       "DATETIME NULL"),
        ]
        for col, ddl in _add_cols:
            cur.execute(f"SHOW COLUMNS FROM submitted_orders LIKE '{col}'")
            if not cur.fetchone():
                cur.execute(f"ALTER TABLE submitted_orders ADD COLUMN {col} {ddl}")
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"[submitted_orders] schema init failed: {e}")


def _submitted_order_email_html(oid, order, base_url):
    """HTML body of the notification email — mirrors the compact
    header + lines summary in the form.  Link lands on the Orders
    list page (harry can click through to the detail from there)."""
    header = order.get("header") or {}
    totals = order.get("totals") or {}
    lines  = order.get("lines")  or []
    def _row(lbl, val):
        return (f'<tr><td style="padding:4px 10px;color:#6b7280;font-size:11px;'
                f'text-transform:uppercase;letter-spacing:.4px;width:140px;">{_esc_html(lbl)}</td>'
                f'<td style="padding:4px 10px;color:#111;font-size:13px;">{_esc_html(val)}</td></tr>')
    line_rows = []
    for ln in lines:
        line_rows.append(
            f'<tr>'
            f'<td style="padding:6px 8px;border:1px solid #e5e7eb;font-family:monospace">{_esc_html(ln.get("m_code",""))}</td>'
            f'<td style="padding:6px 8px;border:1px solid #e5e7eb;text-align:right">{_esc_html(ln.get("qty",""))}</td>'
            f'<td style="padding:6px 8px;border:1px solid #e5e7eb">{_esc_html(ln.get("description","") or ln.get("product_name",""))}</td>'
            f'<td style="padding:6px 8px;border:1px solid #e5e7eb;text-align:right">{_esc_html(ln.get("list_price",""))}</td>'
            f'<td style="padding:6px 8px;border:1px solid #e5e7eb;text-align:right">{_esc_html(ln.get("proposed_dc",""))}</td>'
            f'<td style="padding:6px 8px;border:1px solid #e5e7eb;text-align:right">{_esc_html(ln.get("total_amount",""))}</td>'
            f'</tr>'
        )
    list_url   = f"{base_url}/orders_list"
    detail_url = f"{base_url}/order?id={oid}"
    # Approval-needed banner + reason block — only rendered when the
    # BDE ticked "Management Approval needed" so the two approvers
    # (Hayden + JunJong) immediately see WHAT they're being asked to
    # approve and WHY.
    approval_block = ""
    if (order.get("needs_mgmt_approval") == "Y"):
        reason_html = _esc_html(order.get("mgmt_reason") or "(no reason provided)")
        approval_block = f"""
        <div style="background:#fef3c7;border:1px solid #f59e0b;padding:10px 14px;margin-bottom:14px;border-radius:4px">
          <div style="font-weight:800;color:#92400e;margin-bottom:6px">⚑ Management Approval Requested</div>
          <div style="color:#78350f;font-size:12.5px;margin-bottom:6px">
            The BDE has flagged this order for parallel approval by Hayden Begbie and JunJong Cho.
          </div>
          <div style="background:#fff;border:1px solid #fbbf24;padding:8px 10px;border-radius:3px;font-size:12.5px;color:#111">
            <b>Reason:</b><br>{reason_html}
          </div>
        </div>
        """
    return f"""
    <div style="font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Arial,sans-serif;color:#111;max-width:820px">
      <div style="background:#f5c518;padding:12px 16px;font-weight:800;color:#000">
        Special Price Request Form &nbsp; #{oid}
      </div>
      <div style="padding:14px 16px;background:#fff;border:1px solid #e5e7eb">
        {approval_block}
        <table style="border-collapse:collapse;width:100%;margin-bottom:14px">
          {_row("Submitted",      order.get("submitted_at") or "")}
          {_row("BDE",            header.get("bde_name") or "")}
          {_row("Sold-to",        f'{header.get("sold_to","")} — {header.get("sold_to_name","")}')}
          {_row("Ship-to",        f'{header.get("ship_to","")} — {header.get("ship_to_name","")}')}
          {_row("State",          header.get("state") or "")}
          {_row("PO #",           header.get("po_number") or "—")}
          {_row("Order date",     header.get("order_date") or "")}
          {_row("Total qty",      totals.get("total_qty") or 0)}
          {_row("SOVD qty",       totals.get("sovd_qty") or 0)}
          {_row("Subtotal",       totals.get("subtotal") or "0.00")}
          {_row("Total inc GST",  totals.get("total_inc_gst") or "0.00")}
          {_row("Freight",        totals.get("freight_amount") or "0.00")}
          {_row("Grand total",    totals.get("grand_total") or "0.00")}
          {_row("Avg proposed DC", (str(totals.get("avg_dc_pct") or 0) + "%"))}
        </table>
        <table style="border-collapse:collapse;width:100%;font-size:12px">
          <thead>
            <tr style="background:#374151;color:#fff">
              <th style="padding:6px 8px;border:1px solid #374151;text-align:left">M-Code</th>
              <th style="padding:6px 8px;border:1px solid #374151;text-align:right">Qty</th>
              <th style="padding:6px 8px;border:1px solid #374151;text-align:left">Description</th>
              <th style="padding:6px 8px;border:1px solid #374151;text-align:right">List</th>
              <th style="padding:6px 8px;border:1px solid #374151;text-align:right">Proposed DC</th>
              <th style="padding:6px 8px;border:1px solid #374151;text-align:right">Total</th>
            </tr>
          </thead>
          <tbody>{"".join(line_rows) or '<tr><td colspan="6" style="padding:10px;color:#6b7280">No lines.</td></tr>'}</tbody>
        </table>
        <p style="margin:16px 0 4px;font-size:12px;color:#374151">
          Once this order is keyed into SAP, please flip the status flag
          to <b>Y</b> on the Orders page.
        </p>
        <p style="margin:2px 0;font-size:12px">
          <a href="{detail_url}" style="background:#2563eb;color:#fff;padding:8px 14px;border-radius:4px;text-decoration:none;font-weight:700">Open this order</a>
          &nbsp;
          <a href="{list_url}" style="color:#2563eb">View all orders</a>
        </p>
      </div>
    </div>
    """


@app.post("/api/orders/submit")
def api_orders_submit():
    """Persist a filled Special Price Request Form as a
    submitted_orders row and fire the notification email to Harry
    (CS).  Body is JSON:
      {
        header:  {bde_name, sold_to, sold_to_name, ship_to, ship_to_name,
                  state, po_number, order_date, ...},
        lines:   [{m_code, qty, description, product_name, brand, ...}],
        totals:  {total_qty, sovd_qty, subtotal, total_inc_gst,
                  freight_amount, grand_total, avg_dc_pct}
      }
    Returns { ok: true, id: 123 } on success."""
    import json as _json
    payload = request.get_json(silent=True) or {}
    header  = payload.get("header")  or {}
    totals  = payload.get("totals")  or {}
    lines   = payload.get("lines")   or []
    if not header.get("sold_to"):
        return jsonify({"error": "sold_to required"}), 400
    if not lines:
        return jsonify({"error": "at least one order line required"}), 400

    # Best-effort BDE identification — trust the header value the form
    # sends (the BDE picks their own name / it's auto-filled from the
    # customer master), fall back to the Cloudflare-injected email.
    submitted_by_bde   = (header.get("bde_name") or "").strip()
    submitted_by_email = (_bde_from_request() or "").strip().lower()

    def _num(v):
        try:
            s = str(v or "0").replace(",", "").replace("$", "").replace("%", "").strip()
            return float(s or 0)
        except Exception:
            return 0.0

    # Management-approval flag comes from the header block; the reason
    # is a copy of the yellow textarea (mgmt_reason).  Both are stored
    # as top-level columns so the list page and email templates can
    # read them without unpacking payload_json.
    needs_approval = "Y" if (header.get("needs_mgmt_approval") in ("Y", "y", True, "true", "1")) else "N"
    mgmt_reason    = (header.get("mgmt_reason") or "").strip()

    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO submitted_orders "
            "(submitted_by_bde, submitted_by_email, sold_to, sold_to_name, "
            " ship_to, ship_to_name, state, po_number, order_date, "
            " subtotal, total_inc_gst, freight_amount, grand_total, "
            " total_qty, sovd_qty, avg_dc_pct, status_sap, "
            " needs_mgmt_approval, mgmt_reason, payload_json) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'N',%s,%s,%s)",
            (
                submitted_by_bde[:120],
                submitted_by_email[:255],
                (header.get("sold_to")      or "")[:40],
                (header.get("sold_to_name") or "")[:255],
                (header.get("ship_to")      or "")[:40],
                (header.get("ship_to_name") or "")[:255],
                (header.get("state")        or "")[:20],
                (header.get("po_number")    or "")[:100],
                (header.get("order_date")   or "")[:20],
                _num(totals.get("subtotal")),
                _num(totals.get("total_inc_gst")),
                _num(totals.get("freight_amount")),
                _num(totals.get("grand_total")),
                int(_num(totals.get("total_qty"))),
                int(_num(totals.get("sovd_qty"))),
                _num(totals.get("avg_dc_pct")),
                needs_approval,
                mgmt_reason,
                _json.dumps(payload, ensure_ascii=False, default=str),
            ),
        )
        oid = cur.lastrowid
        conn.commit()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    # Send Harry the email (async — API returns immediately regardless
    # of SMTP / Graph outcome).  Base URL comes from DASHBOARD_URL for
    # production deploys and falls back to the request host in dev.
    try:
        base_url = DASHBOARD_URL.rstrip("/") or request.host_url.rstrip("/")
    except Exception:
        base_url = ""
    subject_prefix = "[SPRF APPROVAL NEEDED" if needs_approval == "Y" else "[SPRF"
    subject = (f"{subject_prefix} #{oid}] {submitted_by_bde or 'BDE'} → "
               f"{header.get('sold_to_name','')} ({header.get('sold_to','')})")
    payload_for_mail = dict(payload)
    payload_for_mail["submitted_at"]        = datetime.now().strftime("%Y-%m-%d %H:%M")
    payload_for_mail["needs_mgmt_approval"] = needs_approval
    payload_for_mail["mgmt_reason"]         = mgmt_reason
    # When approval is needed, put both approvers on To alongside
    # Harry so everyone who has to act sees it in their inbox
    # directly.  Otherwise it's Harry-only (BDE on Cc).
    to_list = [HARRY_CS_EMAIL]
    if needs_approval == "Y":
        to_list += MGMT_APPROVER_EMAILS
    cc = [submitted_by_email] if submitted_by_email else []
    try:
        _send_mail_async(to_list, cc, subject,
                         _submitted_order_email_html(oid, payload_for_mail, base_url))
    except Exception as e:
        print(f"[submit] mail queue failed: {e}")

    return jsonify({"ok": True, "id": oid})


@app.get("/api/orders/list")
def api_orders_list():
    """List submitted orders, most recent first.
      ?status=Y|N       filter by SAP-entered flag (default: all)
      ?limit=…          default 200
    Returns list of small rows suited for the /orders page table."""
    status = (request.args.get("status") or "").strip().upper()
    try:
        limit = min(int(request.args.get("limit") or 200), 1000)
    except Exception:
        limit = 200
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        wh, p = [], []
        if status in ("Y", "N"):
            wh.append("status_sap = %s"); p.append(status)
        where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""
        cur.execute(
            f"SELECT id, submitted_at, submitted_by_bde, submitted_by_email, "
            f"       sold_to, sold_to_name, ship_to, ship_to_name, state, "
            f"       total_qty, grand_total, status_sap, status_changed_at, "
            f"       status_changed_by, needs_mgmt_approval, approved_a, approved_b "
            f"FROM submitted_orders {where_sql} "
            f"ORDER BY submitted_at DESC LIMIT %s",
            tuple(p + [limit])
        )
        rows = cur.fetchall() or []
        # Normalise datetime -> string so JSON serialises cleanly.
        for r in rows:
            for k in ("submitted_at", "status_changed_at"):
                v = r.get(k)
                if v is not None:
                    try: r[k] = v.strftime("%Y-%m-%d %H:%M")
                    except Exception: r[k] = str(v)
            try: r["grand_total"] = float(r.get("grand_total") or 0)
            except Exception: r["grand_total"] = 0
        return jsonify({"rows": rows, "count": len(rows)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e), "rows": []}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.get("/api/orders/detail/<int:oid>")
def api_orders_detail(oid):
    """Return one submitted order's full payload_json plus the
    status metadata."""
    import json as _json
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT id, submitted_at, submitted_by_bde, submitted_by_email, "
            "       sold_to, sold_to_name, ship_to, ship_to_name, state, "
            "       total_qty, sovd_qty, subtotal, total_inc_gst, "
            "       freight_amount, grand_total, avg_dc_pct, "
            "       status_sap, status_changed_at, status_changed_by, "
            "       needs_mgmt_approval, mgmt_reason, "
            "       approved_a, approved_a_at, approved_b, approved_b_at, "
            "       payload_json "
            "FROM submitted_orders WHERE id = %s LIMIT 1",
            (oid,))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        for k in ("submitted_at", "status_changed_at",
                  "approved_a_at", "approved_b_at"):
            v = row.get(k)
            if v is not None:
                try: row[k] = v.strftime("%Y-%m-%d %H:%M")
                except Exception: row[k] = str(v)
        # Decimals to plain floats for JSON.
        for k in ("subtotal", "total_inc_gst", "freight_amount",
                  "grand_total", "avg_dc_pct"):
            try: row[k] = float(row.get(k) or 0)
            except Exception: row[k] = 0
        # Explode payload_json so the front-end can re-hydrate the form
        # from a single object instead of parsing an inline JSON string.
        try:
            row["payload"] = _json.loads(row.pop("payload_json") or "{}")
        except Exception:
            row["payload"] = {}
        return jsonify(row)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.post("/api/orders/detail/<int:oid>/status")
def api_orders_status(oid):
    """Toggle status_sap.  Only Harry (or an admin ALL-role user) can
    write.  Body: {status: 'Y' | 'N'}."""
    payload  = request.get_json(silent=True) or {}
    new_stat = (payload.get("status") or "").strip().upper()
    if new_stat not in ("Y", "N"):
        return jsonify({"error": "status must be Y or N"}), 400
    who = (_bde_from_request() or "").strip().lower()
    # Only Harry (CS) can flip the SAP-entered flag.  Leadership /
    # ALL-role users can view the panel but not write.
    if who != HARRY_CS_EMAIL:
        return jsonify({"error": "only Harry (CS) can change SAP status"}), 403
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE submitted_orders SET status_sap = %s, "
            "  status_changed_at = NOW(), status_changed_by = %s "
            "WHERE id = %s",
            (new_stat, who[:255], oid))
        if cur.rowcount == 0:
            return jsonify({"error": "not found"}), 404
        conn.commit()
        return jsonify({"ok": True, "status": new_stat, "by": who})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.post("/api/orders/detail/<int:oid>/update")
def api_orders_update(oid):
    """Update a previously-submitted order.  Only the BDE who
    originally submitted it can update (compared on
    submitted_by_email).  The content is refreshed, status_sap is
    reset to 'N' so Harry knows the SAP-side needs a re-check, and
    Harry gets a fresh notification email flagged as an update."""
    import json as _json
    payload = request.get_json(silent=True) or {}
    header  = payload.get("header")  or {}
    totals  = payload.get("totals")  or {}
    lines   = payload.get("lines")   or []
    if not header.get("sold_to"):
        return jsonify({"error": "sold_to required"}), 400
    if not lines:
        return jsonify({"error": "at least one order line required"}), 400

    who = (_bde_from_request() or "").strip().lower()
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT submitted_by_email, submitted_by_bde, status_sap "
            "FROM submitted_orders WHERE id = %s LIMIT 1",
            (oid,))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        # Ownership check — an ALL-role admin can edit on behalf of a
        # BDE who's stuck (e.g., left the company); everyone else must
        # be the original submitter.
        original = (row.get("submitted_by_email") or "").strip().lower()
        role     = _EMAIL_TO_DIR.get(who, (None, None, None))[2]
        if who != original and role != "ALL":
            return jsonify({"error": "only the original submitter can edit this order"}), 403

        def _num(v):
            try:
                s = str(v or "0").replace(",", "").replace("$", "").replace("%", "").strip()
                return float(s or 0)
            except Exception:
                return 0.0

        # Edited content invalidates any prior approvals — reset both
        # sides so the approvers see the new version cleanly.
        needs_approval = "Y" if (header.get("needs_mgmt_approval") in ("Y","y",True,"true","1")) else "N"
        mgmt_reason    = (header.get("mgmt_reason") or "").strip()
        cur2 = conn.cursor()
        try:
            cur2.execute(
                "UPDATE submitted_orders SET "
                "  sold_to=%s, sold_to_name=%s, ship_to=%s, ship_to_name=%s, "
                "  state=%s, po_number=%s, order_date=%s, "
                "  subtotal=%s, total_inc_gst=%s, freight_amount=%s, grand_total=%s, "
                "  total_qty=%s, sovd_qty=%s, avg_dc_pct=%s, "
                "  status_sap='N', status_changed_at=NOW(), status_changed_by=%s, "
                "  needs_mgmt_approval=%s, mgmt_reason=%s, "
                "  approved_a='N', approved_a_at=NULL, "
                "  approved_b='N', approved_b_at=NULL, "
                "  payload_json=%s "
                "WHERE id=%s",
                (
                    (header.get("sold_to")      or "")[:40],
                    (header.get("sold_to_name") or "")[:255],
                    (header.get("ship_to")      or "")[:40],
                    (header.get("ship_to_name") or "")[:255],
                    (header.get("state")        or "")[:20],
                    (header.get("po_number")    or "")[:100],
                    (header.get("order_date")   or "")[:20],
                    _num(totals.get("subtotal")),
                    _num(totals.get("total_inc_gst")),
                    _num(totals.get("freight_amount")),
                    _num(totals.get("grand_total")),
                    int(_num(totals.get("total_qty"))),
                    int(_num(totals.get("sovd_qty"))),
                    _num(totals.get("avg_dc_pct")),
                    who[:255],
                    needs_approval,
                    mgmt_reason,
                    _json.dumps(payload, ensure_ascii=False, default=str),
                    oid,
                ),
            )
            conn.commit()
        finally:
            cur2.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    # Notify Harry of the edit.  Subject prefixed with [UPDATE] so it's
    # easy to distinguish from the original submission thread.
    try:
        base_url = DASHBOARD_URL.rstrip("/") or request.host_url.rstrip("/")
    except Exception:
        base_url = ""
    submitted_by_bde = (row.get("submitted_by_bde") or "")
    subject_prefix = "[SPRF UPDATE + APPROVAL NEEDED" if needs_approval == "Y" else "[SPRF UPDATE"
    subject = (f"{subject_prefix} #{oid}] {submitted_by_bde or 'BDE'} → "
               f"{header.get('sold_to_name','')} ({header.get('sold_to','')})")
    payload_for_mail = dict(payload)
    payload_for_mail["submitted_at"]        = datetime.now().strftime("%Y-%m-%d %H:%M")
    payload_for_mail["needs_mgmt_approval"] = needs_approval
    payload_for_mail["mgmt_reason"]         = mgmt_reason
    to_list = [HARRY_CS_EMAIL]
    if needs_approval == "Y":
        to_list += MGMT_APPROVER_EMAILS
    cc = [who] if who else []
    try:
        _send_mail_async(to_list, cc, subject,
                         _submitted_order_email_html(oid, payload_for_mail, base_url))
    except Exception as e:
        print(f"[update] mail queue failed: {e}")

    return jsonify({"ok": True, "id": oid, "status": "N",
                    "note": "SAP flag + approvals reset to N after edit"})


@app.post("/api/orders/detail/<int:oid>/approve")
def api_orders_approve(oid):
    """Mark this order approved by the caller (Hayden or JunJong).
    Parallel approval — each slot moves independently.  Body: {}
    (approver is inferred from the request identity)."""
    who = (_bde_from_request() or "").strip().lower()
    approver_col = None
    if   who == MGMT_APPROVER_EMAILS[0]: approver_col = "approved_a"   # Hayden
    elif who == MGMT_APPROVER_EMAILS[1]: approver_col = "approved_b"   # JunJong
    if not approver_col:
        return jsonify({"error": "only the named approvers can approve"}), 403
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute(
            f"UPDATE submitted_orders SET {approver_col} = 'Y', "
            f"  {approver_col}_at = NOW() "
            f"WHERE id = %s AND needs_mgmt_approval = 'Y'",
            (oid,))
        if cur.rowcount == 0:
            return jsonify({"error": "not found or doesn't need approval"}), 404
        conn.commit()
        return jsonify({"ok": True, "slot": approver_col, "by": who})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


@app.get("/api/orders/whoami")
def api_orders_whoami():
    """Front-end reads this to decide whether to show the status
    toggle (Harry-only) and to auto-fill submitted_by.  Also
    surfaces is_approver_a / is_approver_b so the detail page can
    render the Approve button only for Hayden / JunJong."""
    who = (_bde_from_request() or "").strip().lower()
    name, state, role = _EMAIL_TO_DIR.get(who, (None, None, None))
    return jsonify({
        "email":     who,
        "name":      name or "",
        "state":     state or "",
        "role":      role or "",
        # Only Harry can flip Y/N — kept as a single field the front
        # end can key off (no ALL-role fallback here).
        "is_cs":         who == HARRY_CS_EMAIL,
        "is_harry":      who == HARRY_CS_EMAIL,
        "is_approver_a": who == MGMT_APPROVER_EMAILS[0],   # Hayden
        "is_approver_b": who == MGMT_APPROVER_EMAILS[1],   # JunJong
    })


@app.route("/orders_list")
def orders_list_page():
    """Submitted-orders list — small table with BDE / sold-to / ship-to
    / status; row click opens the detail view (reuses /order form)."""
    return app.send_static_file("orders_list.html")


PLANT_GEO = {
    "42R0": {"lat": -27.8688, "lon": 153.2093},
    "42R1": {"lat": -33.86,   "lon": 150.20},
    "42R2": {"lat": -37.85,   "lon": 144.21},
    "42R4": {"lat": -31.83,   "lon": 116.23},
}
# category -> mapping table
CATEGORY_TABLE = {
    "ISEG": "iseg",
    "SUV": "suv",
    "LOWPROFILE": "lowprofile",
    # ALL/PCLT/TBR???꾨옒?먯꽌 蹂꾨룄 泥섎━
}
@app.get("/api/stock")
def api_stock():
    # optional query params
    # ?metric=qty|unrestricted (湲곕낯 qty)
    # ?plants=42R0,42R1 (湲곕낯 42R0~42R4)
    # inputs
    category = (request.args.get("category") or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern = (request.args.get("pattern") or "").strip()
    material = (request.args.get("material") or "").strip()
    code    = (request.args.get("code")     or "").strip()
    # New stock table (ZSDM64300 daily 3PL feed) stores qty as
    # `stock_qty` — the old MB52 `unrestricted` column is gone.
    metric_col = "stock_qty"

    plants_param = (request.args.get("plants") or "").strip()
    if plants_param:
        plants = [p.strip().upper() for p in plants_param.split(",") if p.strip()]
    else:
        plants = ["42R0", "42R1", "42R2", "42R4"]

    # 醫뚰몴 ?녿뒗 plant???쒖쇅
    plants = [p for p in plants if p in PLANT_GEO]

    if not plants:
        return jsonify({"rows": [], "meta": {"metric": metric_col}})

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        joins = []
        wh = []
        params = []
        sql = f"""
            SELECT s.plant, SUM(s.{metric_col}) AS stock_value
            FROM stock s
        """

        # carrying join (for prod_group/pattern filters, and for PCLT/TBR if you implement via product_group)
        joins.append("JOIN carrying_26 c ON c.m_code = s.material")

        # plant filter
        wh.append(f"s.plant IN ({','.join(['%s']*len(plants))})")
        params += plants

        # product_group dropdown
        if prod_group and prod_group != "ALL":
            wh.append("c.product_group = %s")
            params.append(prod_group)

        # pattern search
        if pattern:
            wh.append("c.pattern LIKE %s")
            params.append(f"%{pattern}%")

        # material search
        if material:
            # allow partial
            wh.append("c.size LIKE %s")
            params.append(f"%{material}%")

        # code search (carrying_26.m_code) — exact match, same table
        if code and code != "ALL":
            wh.append("c.m_code = %s")
            params.append(code)

        # category chip handling ??carrying_26 c is already joined above
        if category == "PCLT":
            wh.append("c.line = 'PCLT'")
        elif category == "TBR":
            wh.append("c.line = 'TBR'")
        elif category == "18PLUS":
            wh.append("c.line = 'PCLT'")
            wh.append("CAST(SUBSTRING_INDEX(c.size, 'R', -1) AS DECIMAL(5,2)) >= 18.0")
        elif category == "ISEG":
            joins.append("JOIN iseg i ON CAST(TRIM(i.Material) AS UNSIGNED) = s.material")
        elif category == "SUV":
            joins.append("JOIN suv suv ON suv.Pattern = c.pattern")
        elif category == "LOWPROFILE":
            joins.append("JOIN lowprofile lp ON CAST(TRIM(lp.Material) AS UNSIGNED) = s.material")
        elif category == "HM":
            wh.append("""EXISTS (
                SELECT 1 FROM hm hm
                JOIN sales_2526 ss ON ss.sold_to = hm.sold_to
                WHERE ss.material = s.material
            )""")
        elif category == "443":
            wh.append("""EXISTS (
                SELECT 1 FROM `443_25` p443
                WHERE p443.product_group = c.product_group
            )""")
        # ALL: no extra filter

        sql += "\n" + "\n".join(joins)
        if wh:
            sql += "\nWHERE " + "\n  AND ".join(wh)
        sql += "\nGROUP BY s.plant"

        cur.execute(sql, params)
        rows = cur.fetchall() or []
        rows_out = []
        for r in rows:
            p = r.get("plant")
            g = PLANT_GEO.get(p)
            if not g:
                continue
            rows_out.append({
                "plant": p,
                "stock_value": float(r.get("stock_value") or 0),
                "lat": g["lat"],
                "lon": g["lon"],
            })
        return jsonify({
            "rows": rows_out,
            "meta": {
                "metric": "stock_qty",
                "category": category,
                "product_group": prod_group,
                "pattern": pattern,
                "material": material,
                "plants": plants
            }
        })
    finally:
        cur.close()
        conn.close()


# ─── Aging bucket helpers (shared with /api/stock_aging) ──────────────
_AGING_BUCKETS = ["≤12M", "13-18M", "19-24M", "25-36M", "37M+"]

def _dot_age_months(dot_str: str):
    """Convert a WWYY DOT string to age in months from today.  Returns
    None for unparsable / malformed DOTs so the caller can bucket them
    under 'unknown' instead of falsely-aging."""
    if not dot_str or len(dot_str) != 4 or not dot_str.isdigit():
        return None
    try:
        week = int(dot_str[:2])
        year = 2000 + int(dot_str[2:])
        if week < 1 or week > 53:
            return None
        from datetime import date, datetime
        d = datetime.fromisocalendar(year, week, 1).date()
        today = date.today()
        return (today.year - d.year) * 12 + (today.month - d.month)
    except Exception:
        return None

def _age_bucket(age_months):
    if age_months is None: return "unknown"
    if age_months <= 12:   return "≤12M"
    if age_months <= 18:   return "13-18M"
    if age_months <= 24:   return "19-24M"
    if age_months <= 36:   return "25-36M"
    return "37M+"


@app.get("/api/stock_aging")
def api_stock_aging():
    """Stock aging breakdown for one plant.  Called when a stock circle
    on /stock is clicked.  Groups every (material, dot_no) row by the
    5 aging buckets (≤12M, 13-18M, 19-24M, 25-36M, 37M+) plus an
    'unknown' bucket for malformed DOTs.

    Returns:
      plant, state, total, buckets{}, materials[]  — materials list is
      the top 30 by total qty with per-bucket breakdown so the popup
      can also show which sizes/patterns are contributing to which
      aging bucket.
    """
    plant = (request.args.get("plant") or "").strip()
    if not plant:
        return jsonify({"error": "plant required"}), 400

    # Same filter set as the cascade table — so clicking a Size-filtered
    # marker gives an aging popup that matches the map circle instead of
    # showing the whole plant's DOT population.
    category   = (request.args.get("category")      or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern    = (request.args.get("pattern")       or "").strip()
    material   = (request.args.get("material")      or "").strip()
    code       = (request.args.get("code")          or "").strip()

    car_wh = []
    car_p  = []
    if prod_group and prod_group != "ALL":
        car_wh.append("c.product_group = %s"); car_p.append(prod_group)
    if pattern:
        car_wh.append("c.pattern LIKE %s"); car_p.append(f"%{pattern}%")
    if material:
        car_wh.append("c.size = %s"); car_p.append(material)
    if code and code != "ALL":
        car_wh.append("c.m_code = %s"); car_p.append(code)
    if category == "PCLT":
        car_wh.append("c.line = 'PCLT'")
    elif category == "TBR":
        car_wh.append("c.line = 'TBR'")
    elif category == "18PLUS":
        car_wh.append("c.line = 'PCLT'")
        car_wh.append("CAST(SUBSTRING_INDEX(c.size,'R',-1) AS DECIMAL(5,2)) >= 18.0")

    has_filter = bool(car_wh)  # controls whether "top aged materials" is returned
    # When we have a filter, narrow via INNER JOIN to a dedup'd carrying
    # subquery so multi-row m_codes don't multiply stock quantities.
    if has_filter:
        inner_wh = " AND ".join(w.replace("c.", "") for w in car_wh)
        join_sql = (
            "JOIN (SELECT m_code,"
            "             MIN(size)          AS size,"
            "             MIN(pattern)       AS pattern,"
            "             MIN(brand)         AS brand"
            f"      FROM carrying_26 WHERE {inner_wh}"
            "      GROUP BY m_code) c ON c.m_code = s.material"
        )
    else:
        join_sql = "LEFT JOIN carrying_26 c ON c.m_code = s.material"

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        sql = (
            "SELECT s.material,"
            "       s.dot_no,"
            "       SUM(s.stock_qty) AS qty,"
            "       MAX(c.size)      AS size,"
            "       MAX(c.pattern)   AS pattern,"
            "       MAX(c.brand)     AS brand "
            f"FROM stock s {join_sql} "
            "WHERE s.plant = %s AND s.stock_qty > 0 "
            "GROUP BY s.material, s.dot_no"
        )
        cur.execute(sql, tuple(car_p) + (plant,))
        rows = cur.fetchall() or []
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    # Roll up per bucket (plant total) and per (material, bucket)
    plant_buckets = {b: 0 for b in _AGING_BUCKETS}
    plant_buckets["unknown"] = 0
    materials: dict = {}
    for r in rows:
        age = _dot_age_months(r["dot_no"])
        bucket = _age_bucket(age)
        qty = int(r["qty"] or 0)
        plant_buckets[bucket] += qty
        mat = str(r["material"] or "").strip()
        if mat not in materials:
            materials[mat] = {
                "material": mat,
                "size":     (r.get("size") or "").strip(),
                "pattern":  (r.get("pattern") or "").strip(),
                "brand":    (r.get("brand") or "").strip(),
                "buckets":  {b: 0 for b in _AGING_BUCKETS + ["unknown"]},
                "total":    0,
                "oldest_dot": r["dot_no"] if age is not None else "",
                "oldest_age_months": age if age is not None else -1,
            }
        m = materials[mat]
        m["buckets"][bucket] += qty
        m["total"] += qty
        if age is not None and age > m["oldest_age_months"]:
            m["oldest_age_months"] = age
            m["oldest_dot"] = r["dot_no"]

    # Sort materials by "aged first" — biggest 37M+ bucket wins, then
    # 25-36M, then total.  So the top of the list surfaces the most
    # concerning stock, not just the biggest-volume SKU.
    def _sort_key(m):
        return (-m["buckets"].get("37M+", 0),
                -m["buckets"].get("25-36M", 0),
                -m["total"])
    mat_list = sorted(materials.values(), key=_sort_key)[:30]

    # Top-aged materials list is only informative when we're looking at
    # the whole plant — once the user has picked a size / pattern the
    # per-material rows collapse into one and add nothing.  User asked
    # explicitly for it to disappear the moment any filter is applied.
    resp = {
        "plant":         plant,
        "state":         _ORDER_PLANT_STATE.get(plant, ""),
        "total":         sum(plant_buckets.values()),
        "buckets":       plant_buckets,
        "bucket_order":  _AGING_BUCKETS + ["unknown"],
        "has_filter":    has_filter,
    }
    if not has_filter:
        resp["materials"] = mat_list
    return jsonify(resp)


@app.route("/api/sales_stats")
def api_sales_stats():
    """Return 3M / 6M / 12M sales totals (qty) and their average (Base Sales)
    from sales_2526, filtered by the same category/product_group/pattern/material
    parameters used on the Stock page.
    Period boundaries are derived from the latest month present in the table.
    """
    category   = (request.args.get("category")      or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern    = (request.args.get("pattern")       or "").strip()
    material   = (request.args.get("material")      or "").strip()
    code       = (request.args.get("code")          or "").strip()

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # Determine the latest (year, month) present in sales_2526
        cur.execute("SELECT MAX(YEAR(billing_date)*100 + MONTH(billing_date)) AS ym FROM sales_2526")
        r = cur.fetchone()
        latest_ym = int((r or {}).get("ym") or 0)
        if not latest_ym:
            return jsonify({"qty_3m": 0, "qty_6m": 0, "qty_12m": 0, "base_sales": 0})

        latest_y = latest_ym // 100
        latest_m = latest_ym % 100

        def _months_back(n):
            """Return list of (year, month) tuples for the n months ending at latest."""
            result = []
            y, m = latest_y, latest_m
            for _ in range(n):
                result.append((y, m))
                m -= 1
                if m == 0:
                    m = 12
                    y -= 1
            return result

        periods_3  = _months_back(3)
        periods_6  = _months_back(6)
        periods_12 = _months_back(12)

        def _make_period_condition(periods, alias="s"):
            if not periods:
                return "1=0", []
            clauses = [f"({alias}.year=%s AND {alias}.month=%s)" for _ in periods]
            params  = [v for p in periods for v in p]
            return "(" + " OR ".join(clauses) + ")", params

        # Base joins / wheres shared across all three windows
        cat_joins, cat_wh = category_filters_sales("s", category)

        base_joins = list(cat_joins)
        base_wh    = list(cat_wh)
        base_params: list = []

        if prod_group and prod_group != "ALL":
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.product_group = %s")
            base_params.append(prod_group)
        if pattern:
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.pattern LIKE %s")
            base_params.append(f"%{pattern}%")
        if material:
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.size LIKE %s")
            base_params.append(f"%{material}%")
        if code and code != "ALL":
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.m_code = %s")
            base_params.append(code)

        results = {}
        for label, periods in (("3m", periods_3), ("6m", periods_6), ("12m", periods_12)):
            period_cond, period_params = _make_period_condition(periods)
            wh_all    = base_wh + [period_cond]
            params_all = base_params + period_params
            join_sql   = "\n".join(base_joins)
            where_sql  = ("WHERE " + " AND ".join(wh_all)) if wh_all else ""
            cur.execute(f"""
                SELECT SUM(s.qty) AS qty
                FROM {_sales_2526_from("s")}
                {join_sql}
                {where_sql}
            """, params_all)
            row = cur.fetchone()
            total = float((row or {}).get("qty") or 0)
            n = {"3m": 3, "6m": 6, "12m": 12}[label]
            results[label] = round(total / n)

        base_sales = round((results["3m"] + results["6m"] + results["12m"]) / 3)
        return jsonify({
            "qty_3m":      results["3m"],
            "qty_6m":      results["6m"],
            "qty_12m":     results["12m"],
            "base_sales":  base_sales,
            "latest_year":  latest_y,
            "latest_month": latest_m,
        })
    finally:
        cur.close()
        conn.close()


@app.route("/api/sales_stats_by_state")
def api_sales_stats_by_state():
    """Return 3M / 6M / 12M / Base Sales + stock/water/factory qty per state."""
    # COMMON excluded intentionally
    STATE_ORDER = ["NSW", "QLD", "VIC", "WA"]
    # SA is part of VIC territory ??merge into VIC
    STATE_REMAP = {"SA": "VIC", "NT": "WA", "TAS": "VIC", "ACT": "NSW"}
    # plant -> state mapping derived from plant geographic locations
    PLANT_STATE = {"42R1": "NSW", "42R0": "QLD", "42R2": "VIC", "42R4": "WA"}
    ALL_PLANTS  = list(PLANT_STATE.keys())

    category   = (request.args.get("category")      or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern    = (request.args.get("pattern")       or "").strip()
    material   = (request.args.get("material")      or "").strip()
    code       = (request.args.get("code")          or "").strip()

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT MAX(YEAR(billing_date)*100 + MONTH(billing_date)) AS ym FROM sales_2526")
        r = cur.fetchone()
        latest_ym = int((r or {}).get("ym") or 0)
        if not latest_ym:
            return jsonify({"rows": [], "latest_year": None, "latest_month": None})

        latest_y = latest_ym // 100
        latest_m = latest_ym % 100

        def _months_back(n):
            result = []
            y, m = latest_y, latest_m
            for _ in range(n):
                result.append((y, m))
                m -= 1
                if m == 0:
                    m = 12; y -= 1
            return result

        def _period_cond(periods, alias="s"):
            if not periods:
                return "1=0", []
            clauses = [f"({alias}.year=%s AND {alias}.month=%s)" for _ in periods]
            return "(" + " OR ".join(clauses) + ")", [v for p in periods for v in p]

        periods_3  = _months_back(3)
        periods_6  = _months_back(6)
        periods_12 = _months_back(12)

        # ?? category filter helpers ??????????????????????????????????
        def _cat_joins_wh_stock(tbl_alias):
            """Return (joins_list, wh_list, params_list) for stock/incoming/orders."""
            joins, wh, params = [], [], []
            has_code = bool(code and code != "ALL")
            needs_carrying = (
                (prod_group and prod_group != "ALL")
                or bool(pattern) or bool(material) or has_code
                or category in ("PCLT", "TBR", "18PLUS", "SUV", "443")
            )
            if needs_carrying:
                joins.append(f"JOIN carrying_26 c ON c.m_code = {tbl_alias}.material")
            if prod_group and prod_group != "ALL":
                wh.append("c.product_group = %s"); params.append(prod_group)
            if pattern:
                wh.append("c.pattern LIKE %s"); params.append(f"%{pattern}%")
            if material:
                wh.append("c.size LIKE %s"); params.append(f"%{material}%")
            if has_code:
                wh.append("c.m_code = %s"); params.append(code)
            if category == "PCLT":
                wh.append("c.line = 'PCLT'")
            elif category == "TBR":
                wh.append("c.line = 'TBR'")
            elif category == "18PLUS":
                wh.append("c.line = 'PCLT'")
                wh.append("CAST(SUBSTRING_INDEX(c.size,'R',-1) AS DECIMAL(5,2)) >= 18.0")
            elif category == "ISEG":
                joins.append(f"JOIN iseg i ON CAST(TRIM(i.Material) AS UNSIGNED) = {tbl_alias}.material")
            elif category == "SUV":
                joins.append("JOIN suv suv ON suv.Pattern = c.pattern")
            elif category == "LOWPROFILE":
                joins.append(f"JOIN lowprofile lp ON CAST(TRIM(lp.Material) AS UNSIGNED) = {tbl_alias}.material")
            elif category == "HM":
                wh.append(f"""EXISTS (
                    SELECT 1 FROM hm hm
                    JOIN sales_2526 ss ON ss.sold_to = hm.sold_to
                    WHERE ss.material = {tbl_alias}.material
                )""")
            elif category == "443":
                wh.append("EXISTS (SELECT 1 FROM `443_25` p443 WHERE p443.product_group = c.product_group)")
            return joins, wh, params

        def _plant_totals(table, val_col, extra_wh=None):
            """Return {plant: qty} for given table and value column, grouped by plant."""
            j, wh, p = _cat_joins_wh_stock("t")
            if extra_wh:
                wh = wh + list(extra_wh)
            join_sql  = "\n".join(j)
            where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""
            cur.execute(f"""
                SELECT t.plant, SUM(t.{val_col}) AS val
                FROM {table} t
                {join_sql}
                {where_sql}
                GROUP BY t.plant
            """, p)
            return {row["plant"]: float(row["val"] or 0) for row in (cur.fetchall() or [])}

        # Exclude orders whose po_no already appears in the incoming table
        # (those shipments are already counted as incoming, not open orders).
        # NOTE: this powers the "+ Factory Qty" column (Orders PO total)
        # and is intentionally NOT filtered to '42'-prefixed POs — that
        # filter only belongs on the Ready-to-Ship metric (confirm_qty),
        # applied in /api/orders below.
        _orders_extra = [
            "t.po_no NOT IN (SELECT DISTINCT po_no FROM incoming"
            " WHERE po_no IS NOT NULL AND TRIM(po_no) <> '')"
        ]

        # ?? stock / water (incoming) / factory (orders) per plant ???
        stock_by_plant   = _plant_totals("stock",    "stock_qty")
        water_by_plant   = _plant_totals("incoming", "po_qty")
        factory_by_plant = _plant_totals("orders",   "po_qty", extra_wh=_orders_extra)
        # CY = Ready-to-Ship factory confirmations.  Mirrors the PO-
        # prefix filter the /api/orders confirm metric uses — real
        # factory RTS POs start with '42'; sample / internal moves
        # otherwise inflate the number.
        cy_by_plant      = _plant_totals("orders",   "confirm_qty",
                                         extra_wh=["t.po_no LIKE '42%'"])

        # aggregate plant totals to state
        def _to_state(by_plant):
            d = {}
            for plant, val in by_plant.items():
                st = PLANT_STATE.get(plant)
                if st:
                    d[st] = d.get(st, 0) + val
            return d

        stock_by_state   = _to_state(stock_by_plant)
        water_by_state   = _to_state(water_by_plant)
        factory_by_state = _to_state(factory_by_plant)
        cy_by_state      = _to_state(cy_by_plant)

        # ?? sales per state ?????????????????????????????????????????
        # Use a deduplicated subquery for the customer join so that ship_tos
        # with multiple rows in customer (different sold_tos) don't inflate sums.
        cat_joins_s, cat_wh_s = category_filters_sales("s", category)
        base_joins = [
            "JOIN ("
            "  SELECT ship_to, MIN(bde_state) AS bde_state"
            "  FROM customer"
            "  WHERE bde_state IS NOT NULL AND bde_state != 'COMMON'"
            "  GROUP BY ship_to"
            ") cus ON cus.ship_to = s.ship_to"
        ] + list(cat_joins_s)
        base_wh    = list(cat_wh_s)
        base_params: list = []

        if prod_group and prod_group != "ALL":
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.product_group = %s"); base_params.append(prod_group)
        if pattern:
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.pattern LIKE %s"); base_params.append(f"%{pattern}%")
        if material:
            _ensure_carrying_join("s", base_joins)
            base_wh.append("mat.size LIKE %s"); base_params.append(f"%{material}%")

        state_data = {}
        for label, periods in (("3m", periods_3), ("6m", periods_6), ("12m", periods_12)):
            pcond, pparams = _period_cond(periods)
            wh_all     = base_wh + [pcond]
            params_all = base_params + pparams
            join_sql   = "\n".join(base_joins)
            where_sql  = ("WHERE " + " AND ".join(wh_all)) if wh_all else ""
            cur.execute(f"""
                SELECT cus.bde_state AS state, SUM(s.qty) AS qty
                FROM {_sales_2526_from("s")}
                {join_sql}
                {where_sql}
                GROUP BY state
            """, params_all)
            n = {"3m": 3, "6m": 6, "12m": 12}[label]
            for row in (cur.fetchall() or []):
                st  = row["state"]
                if not st or st == "COMMON":
                    continue
                st = STATE_REMAP.get(st, st)  # SA?뭋IC, NT?뭌A, etc.
                val = round(float(row["qty"] or 0) / n)
                sd = state_data.setdefault(st, {})
                sd[label] = sd.get(label, 0) + val

        rows_out = []
        for st in STATE_ORDER:
            d = state_data.get(st, {})
            q3  = d.get("3m",  0)
            q6  = d.get("6m",  0)
            q12 = d.get("12m", 0)
            stk = stock_by_state.get(st, 0)
            wtr = water_by_state.get(st, 0)
            fac = factory_by_state.get(st, 0)
            cy  = cy_by_state.get(st, 0)
            # Skip only when EVERY column would be zero — a state with
            # no sales for the current filter but plant inventory still
            # in transit (or sitting in stock) should stay visible so
            # the map dot matches a table row.
            if (q3 == 0 and q6 == 0 and q12 == 0
                and stk == 0 and wtr == 0 and fac == 0 and cy == 0):
                continue
            base = round((q3 + q6 + q12) / 3)
            rows_out.append({
                "state":       st,
                "qty_3m":      q3,
                "qty_6m":      q6,
                "qty_12m":     q12,
                "base_sales":  base,
                "stock_qty":   round(stk),
                "water_qty":   round(wtr),
                "cy_qty":      round(cy),
                "factory_qty": round(fac),
            })

        return jsonify({"rows": rows_out, "latest_year": latest_y, "latest_month": latest_m})
    finally:
        cur.close()
        conn.close()


@app.route("/api/cascade_ancestors")
def api_cascade_ancestors():
    """Given the most specific filter the user has narrowed to
    (material > pattern > product_group), return the cascade
    ancestor chain so the Stock page can auto-jump the cascade
    table to that level instead of forcing the user to drill
    Line → Product Group → Pattern → Size manually.

    Only resolves PCLT / TBR lines (the cascade table's universe).
    If the size/pattern/product_group belongs to another line
    (HM / HK / LF), returns level=line so the table falls back to
    the default top-level view.
    """
    material = (request.args.get("material")      or "").strip()
    pattern  = (request.args.get("pattern")       or "").strip()
    pg       = (request.args.get("product_group") or "").strip()
    code     = (request.args.get("code")          or "").strip()

    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        # Code is the most specific dimension — resolves to at most one
        # carrying_26 row.  Return the full ancestor chain so the /stock
        # UI can lock the Product Group / Pattern / Size dropdowns to
        # this code's parents in one round-trip.
        if code and code != "ALL":
            cur.execute(
                "SELECT line, product_group, pattern, size "
                "FROM carrying_26 "
                "WHERE line IN ('PCLT','TBR') AND m_code = %s "
                "LIMIT 1",
                (code,),
            )
            r = cur.fetchone()
            if r:
                return jsonify({
                    "level":         "size",
                    "line":          r["line"]          or "",
                    "product_group": r["product_group"] or "",
                    "pattern":       r["pattern"]       or "",
                    "size":          r["size"]          or "",
                    "code":          code,
                })
            # Code not in the PCLT/TBR universe — fall through to line
            # root and let the user re-narrow from scratch.
            return jsonify({"level": "line"})
        if material:
            # Narrow by every filter the user already set.  Picking
            # Pattern = RA18 AND Size = 185R14 should pin both —
            # picking ONLY Size = 185R14 should leave pattern / pg
            # empty so the cascade aggregates across every m_code
            # carrying that size (one size often appears under
            # multiple patterns).
            where  = ["line IN ('PCLT','TBR')", "size = %s"]
            params = [material]
            if pattern:
                where.append("pattern = %s"); params.append(pattern)
            if pg and pg != "ALL":
                where.append("product_group = %s"); params.append(pg)
            cur.execute(
                "SELECT DISTINCT line, product_group, pattern "
                "FROM carrying_26 "
                f"WHERE {' AND '.join(where)}",
                params,
            )
            rows = cur.fetchall()
            if rows:
                lines = {r["line"] for r in rows if r["line"]}
                pgs   = {r["product_group"] for r in rows if r["product_group"]}
                pats  = {r["pattern"] for r in rows if r["pattern"]}
                if len(lines) == 1:
                    return jsonify({
                        "level":         "size",
                        "line":          next(iter(lines)),
                        "product_group": next(iter(pgs))  if len(pgs)  == 1 else "",
                        "pattern":       next(iter(pats)) if len(pats) == 1 else "",
                        "size":          material,
                    })
                # Spans both PCLT and TBR — fall back to the line root
                # so the user can pick which side they meant.
                return jsonify({"level": "line"})
        if pattern:
            where  = ["line IN ('PCLT','TBR')", "pattern = %s"]
            params = [pattern]
            if pg and pg != "ALL":
                where.append("product_group = %s"); params.append(pg)
            cur.execute(
                "SELECT DISTINCT line, product_group "
                "FROM carrying_26 "
                f"WHERE {' AND '.join(where)}",
                params,
            )
            rows = cur.fetchall()
            if rows:
                lines = {r["line"] for r in rows if r["line"]}
                pgs   = {r["product_group"] for r in rows if r["product_group"]}
                if len(lines) == 1:
                    return jsonify({
                        "level":         "pattern",
                        "line":          next(iter(lines)),
                        "product_group": next(iter(pgs)) if len(pgs) == 1 else "",
                        "pattern":       pattern,
                    })
                return jsonify({"level": "line"})
        if pg and pg != "ALL":
            cur.execute(
                "SELECT line, product_group "
                "FROM carrying_26 "
                "WHERE line IN ('PCLT','TBR') AND product_group = %s "
                "LIMIT 1",
                (pg,),
            )
            r = cur.fetchone()
            if r:
                return jsonify({
                    "level":         "product_group",
                    "line":          r["line"]          or "",
                    "product_group": r["product_group"] or "",
                })
        return jsonify({"level": "line"})
    finally:
        cur.close(); conn.close()


@app.route("/api/sales_stats_by_product_level")
def api_sales_stats_by_product_level():
    """Same metric set as /api/sales_stats_by_state but grouped by a
    product dimension (line / product_group / pattern / size) so the
    user can drill PCLT → Kinergy → K125 → 205/55R16 on the stock page.

    Cascade params:
      level         = "line" (default) | "product_group" | "pattern" | "size"
      line          = ancestor filter (required when level != "line")
      product_group = ancestor filter (required for level in {pattern, size})
      pattern       = ancestor filter (required for level == "size")

    Plus the same UI filters as sales_stats_by_state (category /
    product_group / pattern / material) — the user's existing filter
    row still narrows the rows on every cascade level.
    """
    level    = (request.args.get("level") or "line").strip()
    bucket_map = {
        "line":          "c.line",
        "product_group": "c.product_group",
        "pattern":       "c.pattern",
        "size":          "c.size",
    }
    if level not in bucket_map:
        return jsonify({"error": "invalid level"}), 400
    bucket_col = bucket_map[level]

    line_anc = (request.args.get("line") or "").strip()
    pg_anc   = (request.args.get("product_group_anc") or "").strip()
    pat_anc  = (request.args.get("pattern_anc") or "").strip()

    # Compose ancestor filter (top-level cascade restricts to PCLT/TBR
    # to match the dashboard's Product cascade behaviour).
    anc_wh = []
    anc_p  = []
    if level == "line":
        anc_wh.append("c.line IN ('PCLT','TBR')")
    else:
        if not line_anc:
            return jsonify({"error": "line ancestor required"}), 400
        anc_wh.append("c.line = %s"); anc_p.append(line_anc)
        # pg_anc and pat_anc are OPTIONAL — when the user reaches a
        # deeper level via a search-box pick (e.g. Size = 185R14 with
        # no Product Group / Pattern picked), we want the cascade to
        # aggregate stock across every m_code carrying that size
        # rather than locking to one arbitrary pattern's m_code.
        if level in ("pattern", "size") and pg_anc:
            anc_wh.append("c.product_group = %s"); anc_p.append(pg_anc)
        if level == "size" and pat_anc:
            anc_wh.append("c.pattern = %s"); anc_p.append(pat_anc)
    # Drop empty bucket labels — usually carrying_26 rows missing the
    # dimension; we don't want a blank row.
    anc_wh.append(f"{bucket_col} IS NOT NULL AND TRIM({bucket_col}) <> ''")

    # Existing UI filters layered on top.
    category   = (request.args.get("category")      or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern    = (request.args.get("pattern")       or "").strip()
    material   = (request.args.get("material")      or "").strip()
    code       = (request.args.get("code")          or "").strip()

    extra_wh = []
    extra_p  = []
    if prod_group and prod_group != "ALL":
        extra_wh.append("c.product_group = %s"); extra_p.append(prod_group)
    if pattern:
        extra_wh.append("c.pattern LIKE %s"); extra_p.append(f"%{pattern}%")
    if material:
        # Exact match — the Size dropdown / cascade always emits a
        # DISTINCT carrying_26.size value, so LIKE %material% would
        # leak load-rating variants ("205/55R16 91V" alongside
        # "205/55R16") into the picked bucket.  User wants ONLY the
        # picked size to remain, so bind tightly.
        extra_wh.append("c.size = %s"); extra_p.append(material)
    if code and code != "ALL":
        extra_wh.append("c.m_code = %s"); extra_p.append(code)
    if category == "PCLT":
        extra_wh.append("c.line = 'PCLT'")
    elif category == "TBR":
        extra_wh.append("c.line = 'TBR'")
    elif category == "18PLUS":
        extra_wh.append("c.line = 'PCLT'")
        extra_wh.append("CAST(SUBSTRING_INDEX(c.size,'R',-1) AS DECIMAL(5,2)) >= 18.0")

    all_wh     = anc_wh + extra_wh
    all_params = anc_p  + extra_p
    # carrying_26 holds many rows per m_code (size / pattern / load-
    # rating variants).  We have to dedupe to one row per m_code so
    # the JOIN doesn't multiply stock / sales by N — but the filters
    # MUST be applied *inside* the dedup subquery, not after.
    #
    # Why: an outer WHERE on MIN(size) loses any m_code whose MIN
    # happens to be a different size than the one the user picked.
    # E.g. carrying has m_code 9999 with sizes {"14X4", "185R14"};
    # MIN(size) = "14X4" lexicographically, so an outer
    # "WHERE c.size LIKE '%185R14%'" would drop m_code 9999 even
    # though it really does carry 185R14 stock.  Moving the same
    # condition inside the SELECT-from-carrying-then-GROUP-BY ensures
    # MIN() is taken over rows that already match, so the dedup row
    # is guaranteed to satisfy the filter.
    inner_wh     = " AND ".join(w.replace("c.", "") for w in all_wh) or "1=1"
    inner_params = list(all_params)

    debug_mode = (request.args.get("debug") == "1")
    debug_dump = {} if debug_mode else None

    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        DEDUP_C = (
            "(SELECT m_code,"
            "        MIN(line)          AS line,"
            "        MIN(product_group) AS product_group,"
            "        MIN(pattern)       AS pattern,"
            "        MIN(size)          AS size"
            f" FROM carrying_26 WHERE {inner_wh}"
            " GROUP BY m_code) c"
        )
        if debug_mode:
            # Run the inner dedup standalone so we can see what
            # m_codes it actually returns for the user's filter.
            probe_sql = (
                "SELECT m_code, MIN(line) AS line, MIN(product_group) AS pg, "
                "MIN(pattern) AS pat, MIN(size) AS sz, COUNT(*) AS n_rows "
                f"FROM carrying_26 WHERE {inner_wh} GROUP BY m_code LIMIT 20"
            )
            cur.execute(probe_sql, inner_params)
            debug_dump["dedup_rows"] = cur.fetchall()
            debug_dump["inner_wh"]   = inner_wh
            debug_dump["inner_params"] = [str(p) for p in inner_params]
            debug_dump["bucket_col"]   = bucket_col
            debug_dump["level"]        = level

        # State table semantics carried over: plant code → physical
        # state, with SA / NT / TAS / ACT folded into the closest hub so
        # every bucket bottoms out at exactly NSW / QLD / VIC / WA.
        PLANT_STATE = {"42R1": "NSW", "42R0": "QLD",
                       "42R2": "VIC", "42R4": "WA"}
        STATE_ORDER = ["NSW", "QLD", "VIC", "WA"]
        STATE_REMAP = {"SA": "VIC", "NT": "WA", "TAS": "VIC", "ACT": "NSW"}

        def _agg(table_sql, val_col, extra_clause=""):
            """Returns {bucket: {state: qty}} grouped by both dims so
            the cascade can show 4 state sub-rows per product row.
            Filters live inside DEDUP_C; the outer query only needs
            extra_clause for table-specific predicates (incoming
            exclusion, '42%' PO prefix, etc.)."""
            sql = (
                f"SELECT {bucket_col} AS bucket, t.plant AS plant, "
                f"       SUM(t.{val_col}) AS val "
                f"FROM {table_sql} t "
                f"JOIN {DEDUP_C} ON c.m_code = t.material "
                f"WHERE 1=1{extra_clause} "
                f"GROUP BY {bucket_col}, t.plant"
            )
            cur.execute(sql, inner_params)
            out = {}
            for r in cur.fetchall():
                st = PLANT_STATE.get(r['plant'])
                if not st:
                    continue
                b = r['bucket'] or ''
                out.setdefault(b, {})
                out[b][st] = out[b].get(st, 0.0) + float(r['val'] or 0)
            return out

        stock_by   = _agg("stock",    "stock_qty")
        water_by   = _agg("incoming", "po_qty")

        # Aging breakdown — DOT-level pass over `stock`, then Python
        # bucketise on WWYY.  Keyed as {bucket: {state: {ab: qty}}}
        # so the frontend can render 5 aging sub-columns for the Stock
        # column when the user expands the header toggle.
        aging_by = {}
        cur.execute(
            f"SELECT {bucket_col} AS bucket, t.plant AS plant, "
            f"       t.dot_no AS dot_no, SUM(t.stock_qty) AS val "
            f"FROM stock t "
            f"JOIN {DEDUP_C} ON c.m_code = t.material "
            f"WHERE t.stock_qty > 0 "
            f"GROUP BY {bucket_col}, t.plant, t.dot_no",
            inner_params,
        )
        for r in cur.fetchall():
            st = PLANT_STATE.get(r['plant'])
            if not st:
                continue
            b   = r['bucket'] or ''
            age = _dot_age_months(r['dot_no'])
            ab  = _age_bucket(age)
            v   = int(float(r['val'] or 0))
            d   = aging_by.setdefault(b, {}).setdefault(
                st, {name: 0 for name in _AGING_BUCKETS + ["unknown"]}
            )
            d[ab] += v
        factory_by = _agg(
            "orders", "po_qty",
            extra_clause=(
                " AND t.po_no NOT IN (SELECT DISTINCT po_no FROM incoming"
                " WHERE po_no IS NOT NULL AND TRIM(po_no) <> '')"
            ),
        )
        cy_by      = _agg("orders", "confirm_qty",
                          extra_clause=" AND t.po_no LIKE '42%'")

        # Sales 3M / 6M / 12M — sales_2526 + billing_date date-range.
        cur.execute(
            "SELECT MAX(YEAR(billing_date)*100 + MONTH(billing_date)) AS ym "
            "FROM sales_2526"
        )
        r = cur.fetchone()
        latest_ym = int((r or {}).get("ym") or 0)
        if not latest_ym:
            return jsonify({"rows": [], "level": level})
        latest_y = latest_ym // 100
        latest_m = latest_ym % 100

        def _months_back(n):
            out = []
            y, m = latest_y, latest_m
            for _ in range(n):
                out.append((y, m))
                m -= 1
                if m == 0: m = 12; y -= 1
            return out

        def _period_clause(periods, alias="s"):
            if not periods:
                return "1=0"
            parts = []
            for y, m in periods:
                first = f"{y}-{m:02d}-01"
                ny, nm = (y+1, 1) if m == 12 else (y, m+1)
                nxt   = f"{ny}-{nm:02d}-01"
                parts.append(
                    f"({alias}.billing_date >= '{first}'"
                    f" AND {alias}.billing_date < '{nxt}')"
                )
            return "(" + " OR ".join(parts) + ")"

        # sales bucket col: same dimension but via mat alias on sales side.
        # Same filter-inside-the-dedup trick as DEDUP_C — we just
        # rename the alias to `mat` so the outer JOIN doesn't collide.
        assert DEDUP_C.endswith(") c"), "DEDUP_C alias contract broken"
        DEDUP_MAT = DEDUP_C[:-3] + ") mat"
        sales_bucket_col = bucket_col.replace("c.", "mat.")
        sales_joins = (
            f"JOIN {DEDUP_MAT} ON mat.m_code = s.material "
            f"JOIN ("
            f"  SELECT ship_to, MIN(bde_state) AS bde_state"
            f"  FROM customer"
            f"  WHERE bde_state IS NOT NULL AND bde_state != 'COMMON'"
            f"  GROUP BY ship_to"
            f") cus ON cus.ship_to = s.ship_to"
        )
        # sales_by_bucket: {bucket: {state: {label: monthly_avg_qty}}}
        sales_by_bucket = {}
        for label, periods in (("3m", _months_back(3)),
                                ("6m", _months_back(6)),
                                ("12m", _months_back(12))):
            pcond = _period_clause(periods, "s")
            cur.execute(
                f"SELECT {sales_bucket_col} AS bucket, "
                f"       cus.bde_state AS state, "
                f"       SUM(s.qty) AS qty "
                f"FROM sales_2526 s "
                f"{sales_joins} "
                f"WHERE {pcond} "
                f"GROUP BY {sales_bucket_col}, cus.bde_state",
                inner_params,
            )
            # Divide the period sum by N to match the State table's
            # monthly-average semantics.
            n_months = {"3m": 3, "6m": 6, "12m": 12}[label]
            for row in cur.fetchall():
                st = row['state']
                if not st or st == 'COMMON':
                    continue
                st = STATE_REMAP.get(st, st)
                if st not in STATE_ORDER:
                    continue
                b = row['bucket'] or ''
                d = sales_by_bucket.setdefault(b, {}).setdefault(st, {})
                d[label] = d.get(label, 0.0) + (
                    float(row['qty'] or 0) / n_months
                )

        all_buckets = set()
        all_buckets.update(stock_by.keys(), water_by.keys(),
                           factory_by.keys(), cy_by.keys(),
                           sales_by_bucket.keys())
        rows_out = []
        for b in sorted(b for b in all_buckets if b):
            by_state_sales   = sales_by_bucket.get(b, {})
            by_state_stock   = stock_by.get(b, {})
            by_state_water   = water_by.get(b, {})
            by_state_cy      = cy_by.get(b, {})
            by_state_factory = factory_by.get(b, {})
            by_state_aging   = aging_by.get(b, {})
            states_out = []
            tQ3 = tQ6 = tQ12 = 0.0
            tStk = tWtr = tCy = tFac = 0.0
            tot_aging = {name: 0 for name in _AGING_BUCKETS + ["unknown"]}
            for st in STATE_ORDER:
                sd  = by_state_sales.get(st, {})
                q3  = sd.get('3m',  0) or 0
                q6  = sd.get('6m',  0) or 0
                q12 = sd.get('12m', 0) or 0
                stk = by_state_stock.get(st, 0)
                wtr = by_state_water.get(st, 0)
                cy  = by_state_cy.get(st, 0)
                fac = by_state_factory.get(st, 0)
                ag  = by_state_aging.get(
                    st, {name: 0 for name in _AGING_BUCKETS + ["unknown"]}
                )
                for k, v in ag.items():
                    tot_aging[k] = tot_aging.get(k, 0) + int(v)
                tQ3  += q3;  tQ6  += q6;  tQ12 += q12
                tStk += stk; tWtr += wtr; tCy  += cy; tFac += fac
                states_out.append({
                    "state":       st,
                    "qty_3m":      int(round(q3)),
                    "qty_6m":      int(round(q6)),
                    "qty_12m":     int(round(q12)),
                    "base_sales":  int(round((q3 + q6 + q12) / 3)),
                    "stock_qty":   int(round(stk)),
                    "water_qty":   int(round(wtr)),
                    "cy_qty":      int(round(cy)),
                    "factory_qty": int(round(fac)),
                    "aging":       {k: int(v) for k, v in ag.items()},
                })
            if (tQ3 == 0 and tQ6 == 0 and tQ12 == 0
                and tStk == 0 and tWtr == 0
                and tCy == 0 and tFac == 0):
                continue
            rows_out.append({
                "bucket":      b,
                "qty_3m":      int(round(tQ3)),
                "qty_6m":      int(round(tQ6)),
                "qty_12m":     int(round(tQ12)),
                "base_sales":  int(round((tQ3 + tQ6 + tQ12) / 3)),
                "stock_qty":   int(round(tStk)),
                "water_qty":   int(round(tWtr)),
                "cy_qty":      int(round(tCy)),
                "factory_qty": int(round(tFac)),
                "by_state":    states_out,
                "aging":       tot_aging,
            })
        resp = {"rows": rows_out, "level": level}
        if debug_mode:
            debug_dump["stock_by"]   = stock_by
            debug_dump["water_by"]   = water_by
            debug_dump["factory_by"] = factory_by
            debug_dump["cy_by"]      = cy_by
            debug_dump["sales_by_bucket"] = sales_by_bucket
            debug_dump["all_buckets"] = sorted(all_buckets)
            resp["debug"] = debug_dump
        return jsonify(resp)
    finally:
        cur.close(); conn.close()


ORIGIN_GEO = {
    "CHN": {"name": "China", "lat": 33.8617, "lon": 104.1954},
    "KOR": {"name": "Korea", "lat": 33.5, "lon": 127.8},
    "HUN": {"name": "Hungary", "lat": 33.1625, "lon": 80.5033},
    "IDN": {"name": "Indonesia", "lat": -2.5489, "lon": 118.0149},
}        
@app.get("/api/orders")
def api_orders():
    # filters (same UI)
    category = (request.args.get("category") or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern = (request.args.get("pattern") or "").strip()
    material = (request.args.get("material") or "").strip()
    code     = (request.args.get("code")     or "").strip()

    # metric: po | confirm  (default po)
    metric = (request.args.get("metric") or "po").strip().lower()
    metric_col = "po_qty" if metric != "confirm" else "confirm_qty"

    # optional: plant filter if you want
    plants_param = (request.args.get("plants") or "").strip()
    plants = [p.strip().upper() for p in plants_param.split(",") if p.strip()] if plants_param else []

    # origin filter optional
    origins_param = (request.args.get("origins") or "").strip()
    origins = [x.strip().upper() for x in origins_param.split(",") if x.strip()] if origins_param else list(ORIGIN_GEO.keys())

    # keep only origins we can plot
    origins = [o for o in origins if o in ORIGIN_GEO]
    if not origins:
        return jsonify({"rows": [], "meta": {"metric": metric_col}})

    cat_joins, cat_wh, cat_needs_carrying = category_filters_orders(category)

    # carrying only when needed
    has_code = bool(code and code != "ALL")
    needs_carrying = cat_needs_carrying or (prod_group and prod_group != "ALL") or bool(pattern) or bool(material) or has_code

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        joins = []
        wh = []
        params = []

        sql = f"""
            SELECT o.origin, SUM(o.{metric_col}) AS order_value
            FROM orders o
        """

        if needs_carrying:
            joins.append("JOIN carrying_26 c ON c.m_code = o.material")

        joins += cat_joins

        # origin filter
        wh.append(f"o.origin IN ({','.join(['%s'] * len(origins))})")
        params += origins

        # optional plant filter
        if plants:
            wh.append(f"o.plant IN ({','.join(['%s'] * len(plants))})")
            params += plants

        # material filter
        # prod_group / pattern / material filters
        if needs_carrying:
            if material:
                wh.append("c.size LIKE %s")
                params.append(f"%{material}%")
            if prod_group and prod_group != "ALL":
                wh.append("c.product_group = %s")
                params.append(prod_group)
            if pattern:
                wh.append("c.pattern LIKE %s")
                params.append(f"%{pattern}%")
            if has_code:
                wh.append("c.m_code = %s")
                params.append(code)

        wh += cat_wh

        # Exclude orders whose po_no is already in the incoming table
        # (applies to both po_qty and confirmed_qty — the row is already
        # received).
        wh.append(
            "o.po_no NOT IN (SELECT DISTINCT po_no FROM incoming"
            " WHERE po_no IS NOT NULL AND TRIM(po_no) <> '')"
        )
        # Ready-to-Ship (confirm metric) only: real factory RTS POs all
        # start with '42'.  Other PO prefixes on the confirm side are
        # samples / internal moves that inflate the number.  Orders PO
        # (po metric) does NOT apply this filter — it still shows every
        # PO regardless of prefix.
        if metric == "confirm":
            wh.append("o.po_no LIKE '42%'")

        sql += "\n" + "\n".join(joins)
        if wh:
            sql += "\nWHERE " + "\n  AND ".join(wh)
        sql += "\nGROUP BY o.origin"

        cur.execute(sql, params)
        rows = cur.fetchall() or []

        by_origin = {r["origin"]: float(r.get("order_value") or 0) for r in rows}

        rows_out = []
        for og in origins:
            g = ORIGIN_GEO[og]
            rows_out.append({
                "origin": og,
                "origin_name": g.get("name", og),
                "order_value": by_origin.get(og, 0.0),
                "lat": g["lat"],
                "lon": g["lon"],
            })

        return jsonify({
            "rows": rows_out,
            "meta": {
                "metric": metric_col,
                "category": category,
                "product_group": prod_group,
                "pattern": pattern,
                "material": material,
                "origins": origins,
                "plants": plants,
                "needs_carrying": needs_carrying
            }
        })
    finally:
        cur.close()
        conn.close()
# origin name/code normalize
ORIGIN_CODE_MAP = {
    "CHN": "CHN", "CHINA": "CHN",
    "KOR": "KOR", "KOREA": "KOR",
    "HUN": "HUN", "HUNGARY": "HUN",
    "IDN": "IDN", "INDONESIA": "IDN",
}

def normalize_origin_code(x: str) -> str:
    s = (x or "").strip().upper()
    return ORIGIN_CODE_MAP.get(s, s)

def lerp(a: float, b: float, t: float) -> float:
    return a + (b - a) * t

from datetime import date, datetime

ETA_WINDOW_DAYS = 60  # 60?쇱쓣 ?대룞 援ш컙?쇰줈 媛??(?먰븯硫?30/90?쇰줈)

def parse_date_ymd(x):
    if x is None:
        return None
    if isinstance(x, (date, datetime)):
        return x.date() if isinstance(x, datetime) else x
    s = str(x).strip()
    if not s or s.startswith("0000-00-00"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except:
            pass
    return None

def clamp01(v):
    return 0.0 if v < 0 else (1.0 if v > 1 else v)

def lerp(a, b, t):
    return a + (b - a) * t

@app.get("/api/incoming")
def api_incoming():
    category = (request.args.get("category") or "ALL").strip().upper()
    prod_group = (request.args.get("product_group") or "ALL").strip()
    pattern = (request.args.get("pattern") or "").strip()
    material = (request.args.get("material") or "").strip()
    code    = (request.args.get("code")     or "").strip()

    metric = (request.args.get("metric") or "po").strip().lower()
    metric_col = "po_qty" if metric != "confirm" else "confirm_qty"

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        joins = []
        wh = []
        params = []

        cat_joins, cat_wh, cat_needs_carrying = category_filters_orders(category)
        has_code = bool(code and code != "ALL")
        needs_carrying = cat_needs_carrying or (prod_group and prod_group != "ALL") or bool(pattern) or bool(material) or has_code

        # ??ETA蹂??꾩튂瑜?諛붽씀?ㅻ㈃ eta_date瑜?洹몃９???ы븿?댁빞 ?댁꽌
        # origin/plant/eta_date ?⑥쐞濡?吏묎퀎(媛숈? ETA?쇰━ 臾띠엫)
        sql = f"""
            SELECT o.plant, o.origin, o.eta_date, SUM(o.{metric_col}) AS incoming_value
            FROM incoming o
        """

        if needs_carrying:
            joins.append("JOIN carrying_26 c ON c.m_code = o.material")

        joins += cat_joins

        if material:
            wh.append("c.size LIKE %s")
            params.append(f"%{material}%")

        if needs_carrying:
            if prod_group and prod_group != "ALL":
                wh.append("c.product_group = %s")
                params.append(prod_group)
            if pattern:
                wh.append("c.pattern LIKE %s")
                params.append(f"%{pattern}%")
            if has_code:
                wh.append("c.m_code = %s")
                params.append(code)

        wh += cat_wh

        sql += "\n" + "\n".join(joins)
        if wh:
            sql += "\nWHERE " + "\n  AND ".join(wh)
        sql += "\nGROUP BY o.plant, o.origin, o.eta_date"

        cur.execute(sql, params)
        rows = cur.fetchall() or []

        today = date.today()
        out = []

        for r in rows:
            plant = (r.get("plant") or "").strip().upper()
            origin_code = normalize_origin_code(r.get("origin"))

            gO = ORIGIN_GEO.get(origin_code)
            gP = PLANT_GEO.get(plant)
            if not gO or not gP:
                continue

            eta = parse_date_ymd(r.get("eta_date"))
            if eta:
                days_to_eta = (eta - today).days
                # progress 0(origin) -> 1(plant)
                progress = clamp01(1.0 - (days_to_eta / float(ETA_WINDOW_DAYS)))
            else:
                # ETA媛 ?놁쑝硫?以묎컙易?湲곗〈泥섎읆)
                progress = 0.75

            lat = lerp(gO["lat"], gP["lat"], progress)
            lon = lerp(gO["lon"], gP["lon"], progress)

            out.append({
                "plant": plant,
                "origin": origin_code,
                "origin_name": gO.get("name", origin_code),
                "eta_date": eta.isoformat() if eta else None,
                "progress": progress,
                "incoming_value": float(r.get("incoming_value") or 0),
                "lat": lat,
                "lon": lon,
                # polyline points
                "line": [
                    {"lat": gO["lat"], "lon": gO["lon"]},
                    {"lat": lat, "lon": lon},
                    {"lat": gP["lat"], "lon": gP["lon"]},
                ]
            })

        return jsonify({
            "rows": out,
            "meta": {
                "metric": metric_col,
                "category": category,
                "product_group": prod_group,
                "pattern": pattern,
                "material": material,
                "eta_window_days": ETA_WINDOW_DAYS
            }
        })
    finally:
        cur.close()
        conn.close()


def get_top_sold_to_from_baseline(cur, f, top_limit, value):
    """Top N sold_to based on 2026 sales (sales_2526 filtered to year=2026).
    Dynamic per filter: when region/salesman/category/etc are set, ranking
    runs within that slice — so clicking NSW + Top 10 gives the NSW top 10,
    not the country-wide top 10."""
    if not top_limit or top_limit <= 0:
        return None

    key = _make_top_key(f, int(top_limit), str(value))
    cached = _cache_get(_TOP_SOLD_TO_CACHE, key)
    if cached is not None:
        return cached

    joins, wh, params = build_customer_filters("sTop", f, use_sold_to_name=False)

    # category filter (same rule as sales) ??use normalised version
    if f.get("category") != "443":
        cat_joins, cat_where = category_filters_sales("sTop", f.get("category"))
        joins += cat_joins
        wh    += cat_where

    # product_group / pattern / size all live in carrying_26
    if (f.get("product_group") != "ALL" or f.get("pattern") != "ALL" or
        f.get("material") != "ALL"):
        _ensure_carrying_join("sTop", joins)
    if f.get("product_group") != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f.get("pattern") != "ALL":
        wh.append("mat.pattern = %s"); params.append(f["pattern"])
    if f.get("material") != "ALL":
        wh.append("mat.size = %s"); params.append(f["material"])

    # Always restrict to 2026 — the ranking is "this year's top sold_tos
    # within the selected slice", not last year's.
    wh.append("sTop.year = 2026")
    where_sql = "WHERE " + " AND ".join(wh)

    sql = f"""
      SELECT sTop.sold_to
        FROM {_sales_2526_from("sTop", year=2026)}
        {' '.join(joins)}
        {where_sql}
       GROUP BY sTop.sold_to
       ORDER BY SUM(sTop.{value}) DESC
       LIMIT %s
    """
    cur.execute(sql, tuple(params) + (int(top_limit),))
    rows = cur.fetchall()
    out = [r["sold_to"] for r in rows]

    # cache 60s (baseline doesn't change intra-minute for dashboards)
    _cache_set(_TOP_SOLD_TO_CACHE, key, out, ttl_sec=int(os.getenv("TOP_CACHE_TTL", "60")))
    return out


# ---------------------------------- v2: consolidated APIs ----------------------------------

@app.get("/api/v2/dimensions")
def v2_dimensions():
    """
    Consolidated lookups for the UI.
    Returns: sold_to_groups, sold_to_names, ship_to_names, product_groups, patterns, materials.
    Uses the same filtering logic as existing v1 endpoints.
    """
    key = _make_v2_key("dims", request)
    cached = _cache_get(_V2_CACHE_DIMS, key)
    if cached is not None:
        return jsonify(cached)

    f = parse_filters(request)
    value = "qty" if f.get("metric") == "qty" else "amt"
    top_limit = int(request.args.get("top_limit", 0) or 0)
    parent_group = (request.args.get("sold_to_group") or f.get("sold_to_group") or "ALL").strip()
    sold_to_name = (request.args.get("sold_to") or f.get("sold_to") or "ALL").strip()
    pg = (request.args.get("product_group") or f.get("product_group") or "ALL").strip()
    pat = (request.args.get("pattern") or f.get("pattern") or "ALL").strip()

    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        # Sold-to groups
        cur.execute("""
            SELECT DISTINCT TRIM(sold_to_group) AS v
            FROM customer
            WHERE sold_to_group IS NOT NULL AND TRIM(sold_to_group) <> ''
            ORDER BY TRIM(sold_to_group)
        """)
        sold_to_groups = [r["v"] for r in cur.fetchall()]

        # Sold-to names (optionally restricted by top_limit baseline)
        if top_limit <= 0:
            if parent_group != "ALL":
                cur.execute("""
                    SELECT DISTINCT TRIM(sold_to_name) AS v
                    FROM customer
                    WHERE sold_to_group = %s
                      AND sold_to_name IS NOT NULL AND TRIM(sold_to_name) <> ''
                    ORDER BY TRIM(sold_to_name)
                """, (parent_group,))
            else:
                cur.execute("""
                    SELECT DISTINCT TRIM(sold_to_name) AS v
                    FROM customer
                    WHERE sold_to_name IS NOT NULL AND TRIM(sold_to_name) <> ''
                    ORDER BY TRIM(sold_to_name)
                """)
            sold_to_names = [r["v"] for r in cur.fetchall()]
        else:
            top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value) or []
            if not top_sold_to:
                sold_to_names = []
            else:
                placeholders = ",".join(["%s"] * len(top_sold_to))
                params = list(top_sold_to)
                extra = ""
                if parent_group != "ALL":
                    extra = " AND c.sold_to_group = %s "
                    params.append(parent_group)
                cur.execute(f"""
                    SELECT DISTINCT TRIM(c.sold_to_name) AS v
                    FROM customer c
                    WHERE c.sold_to IN ({placeholders})
                      {extra}
                      AND c.sold_to_name IS NOT NULL AND TRIM(c.sold_to_name) <> ''
                    ORDER BY TRIM(c.sold_to_name)
                """, tuple(params))
                sold_to_names = [r["v"] for r in cur.fetchall()]

        # Ship-to names under selected sold-to or group
        where = ["ship_to_name IS NOT NULL", "TRIM(ship_to_name) <> ''"]
        params = []
        if sold_to_name.upper() != "ALL":
            where.append("TRIM(sold_to_name) = %s")
            params.append(sold_to_name)
        elif parent_group.upper() != "ALL":
            where.append("TRIM(sold_to_group) = %s")
            params.append(parent_group)
        where_sql = "WHERE " + " AND ".join(where)
        cur.execute(f"""
            SELECT DISTINCT TRIM(ship_to_name) AS v
            FROM customer
            {where_sql}
            ORDER BY TRIM(ship_to_name)
        """, tuple(params))
        ship_to_names = [r["v"] for r in cur.fetchall()]

        # Product groups ??from material master (carrying_26)
        cur.execute("""
            SELECT DISTINCT TRIM(product_group) AS v
            FROM carrying_26
            WHERE product_group IS NOT NULL AND TRIM(product_group) <> ''
            ORDER BY TRIM(product_group)
        """)
        product_groups = [r["v"] for r in cur.fetchall() if r.get("v") is not None]

        # Patterns (filtered by product_group) ??from material master
        if pg and pg != "ALL":
            cur.execute("""
                SELECT DISTINCT TRIM(pattern) AS v
                FROM carrying_26
                WHERE product_group = %s
                  AND pattern IS NOT NULL AND TRIM(pattern) <> ''
                ORDER BY TRIM(pattern)
            """, (pg,))
        else:
            cur.execute("""
                SELECT DISTINCT TRIM(pattern) AS v
                FROM carrying_26
                WHERE pattern IS NOT NULL AND TRIM(pattern) <> ''
                ORDER BY TRIM(pattern)
            """)
        patterns = [r["v"] for r in cur.fetchall() if r.get("v") is not None]

        # Materials (filtered by product_group/pattern)
        w2, p2 = [], []
        if pg and pg != "ALL":
            w2.append("product_group = %s"); p2.append(pg)
        if pat and pat != "ALL":
            w2.append("pattern = %s"); p2.append(pat)
        w2_sql = ("WHERE " + " AND ".join(w2)) if w2 else ""
        cur.execute(f"""
            SELECT DISTINCT size AS v
            FROM carrying_26
            {w2_sql}
            ORDER BY size
        """, tuple(p2))
        materials = [r["v"] for r in cur.fetchall() if r.get("v") is not None]

        payload = {
            "sold_to_groups": sold_to_groups,
            "sold_to_names": sold_to_names,
            "ship_to_names": ship_to_names,
            "product_groups": product_groups,
            "patterns": patterns,
            "materials": materials,
        }

        _cache_set(_V2_CACHE_DIMS, key, payload, ttl_sec=120)
        return jsonify(payload)

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cur.close(); conn.close()
        except Exception:
            pass


@app.get("/api/v2/dashboard")
def v2_dashboard():
    """
    Consolidated data endpoint to replace multiple v1 chart APIs.
    It returns daily/monthly/yearly totals + stacked breakdowns + cumulative + percent,
    using the same filter semantics as the existing v1 endpoints.
    """
    key = _make_v2_key("dash", request)
    cached = _cache_get(_V2_CACHE_DASH, key)
    if cached is not None:
        return jsonify(cached)

    f = parse_filters(request)
    value = "qty" if f.get("metric") == "qty" else "amt"
    top_limit = int(request.args.get("top_limit", 0) or 0)
    group_by = (request.args.get("group_by") or "region").strip()
    month = int(request.args.get("month", 1) or 1)  # daily page: month of 2026


    # Optional: limit work to only what the frontend needs.
    # sections=all (default) or comma list:
    #   daily,daily_stacked,monthly,monthly_stacked,monthly_target_stacked,yearly,yearly_stacked
    sections_raw = (request.args.get("sections") or "all").strip().lower()
    sections = {s.strip() for s in sections_raw.split(",") if s.strip()}
    if (not sections) or ("all" in sections):
        sections = {
            "daily",
            "daily_stacked",
            "monthly",
            "monthly_stacked",
            "monthly_target_stacked",
            "yearly",
            "yearly_stacked",
        }

    # group columns — for sold_to we GROUP BY the code (consistent across
    # sales and target tables) and SELECT the resolved name as label.
    group_cols_sales = {
        "product_group": "mat.product_group",
        "region":        "cus.bde_state",
        "salesman":      "cus.salesman_name",
        "sold_to_group": "cus.sold_to_group",
        "sold_to":       "s.sold_to",
        "pattern":       "mat.pattern",
        "brand":         "mat.brand",
    }
    group_cols_target = {
        "product_group": "mat.product_group",
        "region":        "t.state",
        "salesman":      "t.bde",
        "sold_to":       "t.sold_to",
        "pattern":       "mat.pattern",
        "brand":         "mat.brand",
    }
    if group_by not in group_cols_sales:
        return jsonify({"error": "invalid group_by"}), 400
    # scus / tcus = per-sold_to name resolvers — guarantee one label
    # per sold_to regardless of ship_to row joined.  See SOLD_TO_NAME_JOIN.
    label_col_sales  = ("MIN(COALESCE(scus.sold_to_name, s.sold_to))"
                       if group_by == "sold_to"
                       else f"COALESCE(NULLIF(TRIM({group_cols_sales[group_by]}),''), 'COMMON')")
    label_col_target = ("MIN(COALESCE(tcus.sold_to_name, t.sold_to))"
                       if group_by == "sold_to"
                       else f"COALESCE(NULLIF(TRIM({group_cols_target[group_by]}),''), 'COMMON')")

    import calendar
    days_in_month = calendar.monthrange(2026, month)[1]
    daily_labels = list(range(1, days_in_month + 1))
    monthly_labels = list(range(1, 13))
    yearly_labels = list(range(2021, 2026))

    conn = get_connection(); cur = conn.cursor(dictionary=True)
    t0 = time()
    try:
        # ---------------- top-limit (baseline) ----------------
        top_sold_to = None
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value) or []
            if not top_sold_to:
                # Nothing matched filters. Return empty but well-shaped payload.
                empty_daily = [0.0] * days_in_month
                empty_month = [0.0] * 12
                empty_year = [0.0] * len(yearly_labels)
                payload = {
                    "meta": {"filters": f, "group_by": group_by, "month": month, "top_limit": top_limit, "timing_ms": {"total": int((time()-t0)*1000)}},
                    "daily": {"labels": daily_labels, "total": empty_daily, "target": empty_daily, "cum_total": empty_daily, "cum_target": empty_daily, "achievement": [None]*days_in_month, "cum_achievement": [None]*days_in_month},
                    "daily_stacked": {"labels": daily_labels, "groups": [], "value": {}, "cum": {}, "pct": {}, "pct_cum": {}},
                    "monthly": {"labels": monthly_labels, "sales_2025": empty_month, "sales_2026": empty_month, "target_2026": empty_month, "cum_sales_2025": empty_month, "cum_sales_2026": empty_month, "cum_target_2026": empty_month, "ach_2026": [None]*12, "ach_cum_2026": [None]*12},
                    "monthly_stacked": {"labels": monthly_labels, "groups": [], "value_2025": {}, "value_2026": {}, "cum_2025": {}, "cum_2026": {}, "pct_2025": {}, "pct_2026": {}, "pct_cum_2025": {}, "pct_cum_2026": {}},
                    "monthly_target_stacked": {"labels": monthly_labels, "groups": [], "value": {}, "cum": {}, "pct": {}, "pct_cum": {}},
                    "yearly": {"labels": yearly_labels, "total": empty_year},
                    "yearly_stacked": {"labels": yearly_labels, "groups": [], "value": {}, "pct": {}},
                }
                _cache_set(_V2_CACHE_DASH, key, payload, ttl_sec=int(os.getenv("DASH_CACHE_TTL", "30")))
                return jsonify(payload)

        # ---------------- daily (sales_thismonth + target_26 month) ----------------
        joins_d, wh_d, params_d = build_customer_filters("s", f, use_sold_to_name=False)
        if f["category"] != "443":
            cj, cw = category_filters_sales("s", f["category"], has_brand=True)
            joins_d += cj; wh_d += cw
        # Ensure carrying join when product_group / pattern / material (size)
        # filter or group_by needs it.  Size is stored on carrying_26.size
        # (the dropdown value is the size string, not a material code), so
        # filtering by size requires the carrying join.
        if (group_by in ("line", "brand", "product_group", "pattern") or
            f["product_group"] != "ALL" or f["pattern"] != "ALL" or
            f["material"] != "ALL" or f["code"] != "ALL"):
            _ensure_carrying_join("s", joins_d)
        if group_by in ("region", "salesman", "sold_to_group", "sold_to"):
            _ensure_customer_join("s", joins_d)
        if f["product_group"] != "ALL":
            wh_d.append("mat.product_group = %s"); params_d.append(f["product_group"])
        if f["brand"] != "ALL":
            wh_d.append("mat.brand = %s"); params_d.append(f["brand"])
        if f["pattern"] != "ALL":
            wh_d.append("mat.pattern = %s"); params_d.append(f["pattern"])
        if f["material"] != "ALL":
            wh_d.append("mat.size = %s"); params_d.append(f["material"])
        if f["code"] != "ALL":
            wh_d.append("mat.m_code = %s"); params_d.append(f["code"])
        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh_d.append(f"s.sold_to IN ({placeholders})")
            params_d.extend(top_sold_to)
        # Add per-sold_to name resolver only for breakdown stack (used below).
        SCUS_JOIN = (
            "LEFT JOIN ("
            "  SELECT sold_to, MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name "
            "  FROM customer GROUP BY sold_to"
            ") scus ON scus.sold_to = s.sold_to"
        )
        joins_d_break = list(joins_d) + ([SCUS_JOIN] if group_by == "sold_to" else [])
        where_d = ("WHERE " + " AND ".join(wh_d)) if wh_d else ""

        cur.execute(f"""
            SELECT s.day AS k, SUM(s.{value}) AS v
            FROM sales_thismonth s
            {' '.join(joins_d)}
            {where_d}
            GROUP BY s.day
            ORDER BY s.day
        """, tuple(params_d))
        daily_rows = cur.fetchall()
        daily_map = {int(r["k"]): float(r["v"] or 0) for r in daily_rows}
        daily_total = [daily_map.get(d, 0.0) for d in daily_labels]
        daily_cum = _to_cumulative(daily_total)

        # daily breakdown stacks
        group_col_d = group_cols_sales[group_by]
        cur.execute(f"""
            SELECT s.day AS day, {label_col_sales} AS group_label, SUM(s.{value}) AS value
            FROM sales_thismonth s
            {' '.join(joins_d_break)}
            {where_d}
            GROUP BY s.day, {group_col_d}
            ORDER BY s.day
        """, tuple(params_d))
        d_break = cur.fetchall()
        d_groups, d_by = _stacks_from_rows(d_break, "day", days_in_month, group_sort=("region" if group_by=="region" else "alpha"))
        d_cum_by = {g: _to_cumulative(d_by[g]) for g in d_groups}
        d_pct_by = _pct_by_bucket(d_by)
        d_pct_cum_by = _pct_by_bucket(d_cum_by)

        # daily target for selected month (target_26)
        joins_t, wh_t, params_t = build_target_filters("t", f)
        tj, tw = category_target_filters("t", f["category"])
        joins_t += tj; wh_t += tw
        wh_t.append("t.month = %s"); params_t.append(month)
        carrying_join_t = "LEFT JOIN carrying_26 mat ON mat.m_code = t.material"
        needs_carrying_t = False
        if f["product_group"] != "ALL":
            needs_carrying_t = True
            wh_t.append("mat.product_group = %s"); params_t.append(f["product_group"])
        if f["brand"] != "ALL":
            needs_carrying_t = True
            wh_t.append("mat.brand = %s"); params_t.append(f["brand"])
        if f["pattern"] != "ALL":
            needs_carrying_t = True
            wh_t.append("mat.pattern = %s"); params_t.append(f["pattern"])
        if f["material"] != "ALL":
            needs_carrying_t = True
            wh_t.append("mat.size = %s"); params_t.append(f["material"])
        if f["code"] != "ALL":
            needs_carrying_t = True
            wh_t.append("mat.m_code = %s"); params_t.append(f["code"])
        if needs_carrying_t and carrying_join_t not in joins_t:
            joins_t.append(carrying_join_t)
        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh_t.append(f"t.sold_to IN ({placeholders})")
            params_t.extend(top_sold_to)
        where_t = ("WHERE " + " AND ".join(wh_t)) if wh_t else ""
        cur.execute(f"""
            SELECT SUM(t.{value}) AS monthly_total
            FROM target_26 t
            {' '.join(joins_t)}
            {where_t}
        """, tuple(params_t))
        row = cur.fetchone() or {}
        monthly_total_target = float(row.get("monthly_total") or 0)
        daily_target_val = (monthly_total_target / days_in_month) if days_in_month else 0.0
        daily_target = [daily_target_val for _ in daily_labels]
        daily_target_cum = _to_cumulative(daily_target)

        daily_ach = [None if daily_target[i] <= 0 else round((daily_total[i]/daily_target[i])*100.0, 2) for i in range(days_in_month)]
        daily_cum_ach = [None if daily_target_cum[i] <= 0 else round((daily_cum[i]/daily_target_cum[i])*100.0, 2) for i in range(days_in_month)]

        # ---------------- monthly (sales_2526 years 2025/2026 + target_26) ----------------
        joins_m, wh_m, params_m = build_customer_filters("s", f, use_sold_to_name=False)
        mj, mw = category_filters_sales("s", f["category"])
        joins_m += mj; wh_m += mw
        if (group_by in ("line", "brand", "product_group", "pattern") or
            f["product_group"] != "ALL" or f["pattern"] != "ALL" or
            f["material"] != "ALL" or f["code"] != "ALL"):
            _ensure_carrying_join("s", joins_m)
        if group_by in ("region", "salesman", "sold_to_group", "sold_to"):
            _ensure_customer_join("s", joins_m)
        if f["product_group"] != "ALL":
            wh_m.append("mat.product_group = %s"); params_m.append(f["product_group"])
        if f["brand"] != "ALL":
            wh_m.append("mat.brand = %s"); params_m.append(f["brand"])
        if f["pattern"] != "ALL":
            wh_m.append("mat.pattern = %s"); params_m.append(f["pattern"])
        if f["material"] != "ALL":
            wh_m.append("mat.size = %s"); params_m.append(f["material"])
        if f["code"] != "ALL":
            wh_m.append("mat.m_code = %s"); params_m.append(f["code"])
        # 2025/2026 window on billing_date (no s.year column on this
        # sales_2526 schema — everything is derived from billing_date).
        wh_m.append("s.billing_date >= '2025-01-01' AND s.billing_date < '2027-01-01'")
        # Suppress future-month rows.  The nightly import puts each
        # day's sales_thismonth batch into sales_2526 tagged with the
        # CURRENT calendar month — so on July 1 (before the first
        # business day) the June-30 batch shows up as a phantom July
        # bar.  Cap year*100+month to today's effective month.
        _eff_y, _eff_m = _business_effective_ym()
        wh_m.append(
            f"(YEAR(s.billing_date) * 100 + MONTH(s.billing_date)) <= "
            f"{_eff_y * 100 + _eff_m}"
        )
        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh_m.append(f"s.sold_to IN ({placeholders})")
            params_m.extend(top_sold_to)
        joins_m_break = list(joins_m) + ([SCUS_JOIN] if group_by == "sold_to" else [])
        where_m = ("WHERE " + " AND ".join(wh_m)) if wh_m else ""

        cur.execute(f"""
            SELECT YEAR(s.billing_date) AS year, MONTH(s.billing_date) AS month,
                   SUM(s.{value}) AS value
            FROM {_sales_2526_from("s")}
            {' '.join(joins_m)}
            {where_m}
            GROUP BY YEAR(s.billing_date), MONTH(s.billing_date)
            ORDER BY YEAR(s.billing_date), MONTH(s.billing_date)
        """, tuple(params_m))
        m_tot_rows = cur.fetchall()
        m25 = {int(r["month"]): float(r["value"] or 0) for r in m_tot_rows if int(r["year"]) == 2025}
        m26 = {int(r["month"]): float(r["value"] or 0) for r in m_tot_rows if int(r["year"]) == 2026}
        monthly_25 = [m25.get(m, 0.0) for m in monthly_labels]
        monthly_26 = [m26.get(m, 0.0) for m in monthly_labels]
        cum_25 = _to_cumulative(monthly_25)
        cum_26 = _to_cumulative(monthly_26)

        # monthly breakdown stacks (both years)
        group_col_m = group_cols_sales[group_by]
        cur.execute(f"""
            SELECT YEAR(s.billing_date) AS year, MONTH(s.billing_date) AS month,
                   {label_col_sales} AS group_label, SUM(s.{value}) AS value
            FROM {_sales_2526_from("s")}
            {' '.join(joins_m_break)}
            {where_m}
            GROUP BY YEAR(s.billing_date), MONTH(s.billing_date), {group_col_m}
            ORDER BY YEAR(s.billing_date), MONTH(s.billing_date)
        """, tuple(params_m))
        m_break = cur.fetchall()

        m_break_25 = [r for r in m_break if int(r.get("year") or 0) == 2025]
        m_break_26 = [r for r in m_break if int(r.get("year") or 0) == 2026]
        m_groups_25, m_by_25 = _stacks_from_rows(m_break_25, "month", 12, group_sort=("region" if group_by=="region" else "alpha"))
        m_groups_26, m_by_26 = _stacks_from_rows(m_break_26, "month", 12, group_sort=("region" if group_by=="region" else "alpha"))
        # unified group list (so chart legend is stable)
        groups_union = _stacks_from_rows(m_break, "month", 12, group_sort=("region" if group_by=="region" else "alpha"))[0]
        # ensure missing groups are present as zeros
        for g in groups_union:
            m_by_25.setdefault(g, [0.0]*12)
            m_by_26.setdefault(g, [0.0]*12)
        m_cum_25 = {g: _to_cumulative(m_by_25[g]) for g in groups_union}
        m_cum_26 = {g: _to_cumulative(m_by_26[g]) for g in groups_union}
        m_pct_25 = _pct_by_bucket(m_by_25)
        m_pct_26 = _pct_by_bucket(m_by_26)
        m_pct_cum_25 = _pct_by_bucket(m_cum_25)
        m_pct_cum_26 = _pct_by_bucket(m_cum_26)

        # monthly target totals & breakdown (target_26)
        # Build fresh for all months (don't reuse params_t because it includes a specific month)
        joins_mt, wh_mt, params_mt = build_target_filters("t", f)
        tj2, tw2 = category_target_filters("t", f["category"])
        joins_mt += tj2; wh_mt += tw2
        # carrying_26 needed for product_group/pattern (not stored in target_26 directly)
        carrying_join_mt = "LEFT JOIN carrying_26 mat ON mat.m_code = t.material"
        needs_carrying_mt = group_by in ("brand", "product_group", "pattern")
        if f["product_group"] != "ALL":
            needs_carrying_mt = True
            wh_mt.append("mat.product_group = %s"); params_mt.append(f["product_group"])
        if f["brand"] != "ALL":
            needs_carrying_mt = True
            wh_mt.append("mat.brand = %s"); params_mt.append(f["brand"])
        if f["pattern"] != "ALL":
            needs_carrying_mt = True
            wh_mt.append("mat.pattern = %s"); params_mt.append(f["pattern"])
        if f["material"] != "ALL":
            needs_carrying_mt = True
            wh_mt.append("mat.size = %s"); params_mt.append(f["material"])
        if f["code"] != "ALL":
            needs_carrying_mt = True
            wh_mt.append("mat.m_code = %s"); params_mt.append(f["code"])
        if needs_carrying_mt and carrying_join_mt not in joins_mt:
            joins_mt.append(carrying_join_mt)
        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh_mt.append(f"t.sold_to IN ({placeholders})")
            params_mt.extend(top_sold_to)
        if group_by == "sold_to":
            joins_mt.append(
                "LEFT JOIN ("
                "  SELECT sold_to, MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name "
                "  FROM customer GROUP BY sold_to"
                ") tcus ON tcus.sold_to = t.sold_to"
            )
        where_mt = ("WHERE " + " AND ".join(wh_mt)) if wh_mt else ""
        cur.execute(f"""
            SELECT t.month AS month, SUM(t.{value}) AS value
            FROM target_26 t
            {' '.join(joins_mt)}
            {where_mt}
            GROUP BY t.month
            ORDER BY t.month
        """, tuple(params_mt))
        mt_rows = cur.fetchall()
        mt_map = {int(r["month"]): float(r["value"] or 0) for r in mt_rows}
        target_26 = [mt_map.get(m, 0.0) for m in monthly_labels]
        cum_target_26 = _to_cumulative(target_26)

        ach_26 = [None if target_26[i] <= 0 else round((monthly_26[i]/target_26[i])*100.0, 2) for i in range(12)]
        ach_cum_26 = [None if cum_target_26[i] <= 0 else round((cum_26[i]/cum_target_26[i])*100.0, 2) for i in range(12)]

        # target breakdown stacks
        group_col_mt = group_cols_target[group_by]
        cur.execute(f"""
            SELECT t.month AS month, {label_col_target} AS group_label, SUM(t.{value}) AS value
            FROM target_26 t
            {' '.join(joins_mt)}
            {where_mt}
            GROUP BY t.month, {group_col_mt}
            ORDER BY t.month
        """, tuple(params_mt))
        mt_break = cur.fetchall()
        mt_groups, mt_by = _stacks_from_rows(mt_break, "month", 12, group_sort=("region" if group_by=="region" else "alpha"))
        mt_cum_by = {g: _to_cumulative(mt_by[g]) for g in mt_groups}
        mt_pct_by = _pct_by_bucket(mt_by)
        mt_pct_cum_by = _pct_by_bucket(mt_cum_by)

        # ---------------- yearly (sales_21_25) ----------------
        joins_y, wh_y, params_y = build_customer_filters("s", f, use_sold_to_name=False)
        if f["category"] != "443":
            yj, yw = category_filters_sales("s", f["category"])
            joins_y += yj; wh_y += yw
        if (group_by in ("line", "brand", "product_group", "pattern") or
            f["product_group"] != "ALL" or f["pattern"] != "ALL" or
            f["material"] != "ALL" or f["code"] != "ALL"):
            _ensure_carrying_join("s", joins_y)
        if group_by in ("region", "salesman", "sold_to_group", "sold_to"):
            _ensure_customer_join("s", joins_y)
        if f["product_group"] != "ALL":
            wh_y.append("mat.product_group = %s"); params_y.append(f["product_group"])
        if f["brand"] != "ALL":
            wh_y.append("mat.brand = %s"); params_y.append(f["brand"])
        if f["pattern"] != "ALL":
            wh_y.append("mat.pattern = %s"); params_y.append(f["pattern"])
        if f["material"] != "ALL":
            wh_y.append("mat.size = %s"); params_y.append(f["material"])
        if f["code"] != "ALL":
            wh_y.append("mat.m_code = %s"); params_y.append(f["code"])
        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh_y.append(f"s.sold_to IN ({placeholders})")
            params_y.extend(top_sold_to)
        wh_y.append("s.year BETWEEN 2021 AND 2025")
        joins_y_break = list(joins_y) + ([SCUS_JOIN] if group_by == "sold_to" else [])
        where_y = ("WHERE " + " AND ".join(wh_y)) if wh_y else ""
        cur.execute(f"""
            SELECT s.year AS year, SUM(s.{value}) AS value
            FROM sales_21_25 s
            {' '.join(joins_y)}
            {where_y}
            GROUP BY s.year
            ORDER BY s.year
        """, tuple(params_y))
        y_rows = cur.fetchall()
        y_map = {int(r["year"]): float(r["value"] or 0) for r in y_rows}
        y_total = [y_map.get(y, 0.0) for y in yearly_labels]

        group_col_y = group_cols_sales[group_by]
        cur.execute(f"""
            SELECT s.year AS year, {label_col_sales} AS group_label, SUM(s.{value}) AS value
            FROM sales_21_25 s
            {' '.join(joins_y_break)}
            {where_y}
            GROUP BY s.year, {group_col_y}
            ORDER BY s.year
        """, tuple(params_y))
        y_break = cur.fetchall()
        # NOTE: years are not 1..N indexes; map manually
        y_groups = sorted({(r.get("group_label") or "").strip() or "COMMON" for r in y_break}, key=lambda g: (_region_order_key(g), g) if group_by=="region" else (g.upper(), g))
        y_by2: Dict[str, List[float]] = {g: [0.0]*len(yearly_labels) for g in y_groups}
        year_idx = {y:i for i,y in enumerate(yearly_labels)}
        for r in y_break:
            g = (r.get("group_label") or "").strip() or "COMMON"
            yy = int(r.get("year") or 0)
            if yy not in year_idx:
                continue
            y_by2[g][year_idx[yy]] += float(r.get("value") or 0)
        y_pct = _pct_by_bucket(y_by2)

        payload = {
            "meta": {
                "filters": f,
                "group_by": group_by,
                "month": month,
                "top_limit": top_limit,
                "timing_ms": {"total": int((time()-t0)*1000)},
            },
            "daily": {
                "labels": daily_labels,
                "total": daily_total,
                "target": daily_target,
                "cum_total": daily_cum,
                "cum_target": daily_target_cum,
                "achievement": daily_ach,
                "cum_achievement": daily_cum_ach,
            },
            "daily_stacked": {
                "labels": daily_labels,
                "groups": d_groups,
                "value": d_by,
                "cum": d_cum_by,
                "pct": d_pct_by,
                "pct_cum": d_pct_cum_by,
            },
            "monthly": {
                "labels": monthly_labels,
                "sales_2025": monthly_25,
                "sales_2026": monthly_26,
                "target_2026": target_26,
                "cum_sales_2025": cum_25,
                "cum_sales_2026": cum_26,
                "cum_target_2026": cum_target_26,
                "ach_2026": ach_26,
                "ach_cum_2026": ach_cum_26,
            },
            "monthly_stacked": {
                "labels": monthly_labels,
                "groups": groups_union,
                "value_2025": {g: m_by_25.get(g, [0.0]*12) for g in groups_union},
                "value_2026": {g: m_by_26.get(g, [0.0]*12) for g in groups_union},
                "cum_2025": {g: m_cum_25.get(g, [0.0]*12) for g in groups_union},
                "cum_2026": {g: m_cum_26.get(g, [0.0]*12) for g in groups_union},
                "pct_2025": m_pct_25,
                "pct_2026": m_pct_26,
                "pct_cum_2025": m_pct_cum_25,
                "pct_cum_2026": m_pct_cum_26,
            },
            "monthly_target_stacked": {
                "labels": monthly_labels,
                "groups": mt_groups,
                "value": mt_by,
                "cum": mt_cum_by,
                "pct": mt_pct_by,
                "pct_cum": mt_pct_cum_by,
            },
            "yearly": {
                "labels": yearly_labels,
                "total": y_total,
            },
            "yearly_stacked": {
                "labels": yearly_labels,
                "groups": y_groups,
                "value": y_by2,
                "pct": y_pct,
            },
        }

        _cache_set(_V2_CACHE_DASH, key, payload, ttl_sec=int(os.getenv("DASH_CACHE_TTL", "30")))
        return jsonify(payload)

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cur.close(); conn.close()
        except Exception:
            pass


# ----------------------------- Daily Sales ---------------------------------
@app.get("/api/daily_sales")
@cached_endpoint(60)
def daily_sales():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    # Promo filter — sales_thismonth is already 2026 (current month),
    # so no FROM switch needed.  We just enforce PCLT + the promo
    # EXISTS when promos are passed.
    promos = request.args.getlist("promo")

    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)

    # category ??use normalised version for sales tables
    if f["category"] != "443":
        cat_joins, cat_where = category_filters_sales("s", f["category"], has_brand=True)
        joins += cat_joins
        wh    += cat_where

    # product_group / pattern / size all live in carrying_26 (alias: mat)
    if f["product_group"] != "ALL" or f["pattern"] != "ALL" or f["material"] != "ALL" or f["code"] != "ALL":
        _ensure_carrying_join("s", joins)
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s")
        params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s")
        params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s")
        params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s")
        params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s")
        params.append(f["code"])

    if promos:
        _ensure_carrying_join("s", joins)
        _ensure_customer_join("s", joins)
        # Pre-aggregated qty per (ship_to, day, brand) for TrueBlue's
        # ship_to+day SUM rule.  Same JOIN shape daily_breakdown uses.
        DAY_QTY_JOIN = (
            "LEFT JOIN ("
            "  SELECT ship_to, day, brand, SUM(qty) AS day_qty"
            "  FROM sales_thismonth"
            "  GROUP BY ship_to, day, brand"
            ") dq ON dq.ship_to = s.ship_to AND dq.day = s.day AND dq.brand = s.brand"
        )
        if DAY_QTY_JOIN not in joins:
            joins.append(DAY_QTY_JOIN)
        # sales_thismonth has no year/month columns of its own — it IS
        # the current calendar month — so feed those as integer literals
        # so the promo_plan period match still works.
        from datetime import date as _d
        _today = _d.today()
        promo_wh, promo_p = _promo_filter_clauses(
            promos,
            year_expr=str(_today.year),
            month_expr=str(_today.month),
            day_qty_alias="dq.day_qty",
        )
        wh.extend(promo_wh)
        params.extend(promo_p)

    base_where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        top_sold_to = None

        # 1) Get Top N sold_to from baseline table (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )

            if not top_sold_to:
                # no matching customers ??all days = 0
                return jsonify([{"day": d, "value": 0} for d in range(1, 32)])

        # 2) Daily totals, optionally restricted to the baseline top sold_to set
        wh2 = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        daily_sql = f"""
        SELECT s.day AS day_num, SUM(s.{value}) AS daily_total
            FROM sales_thismonth s
            {' '.join(joins)}
            {where_sql2}
        GROUP BY s.day
        ORDER BY s.day
        """
        cur.execute(daily_sql, tuple(params2))
        rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    day_map = {int(r["day_num"]): float(r["daily_total"] or 0) for r in rows}
    return jsonify([{"day": d, "value": day_map.get(d, 0)} for d in range(1, 32)])
     

#
# -------------------- Daily breakdown (stacked by group) -------------------
@app.get("/api/daily_breakdown")
@cached_endpoint(60)
def daily_breakdown():

    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"
    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    # Promo filter — same as daily_sales, no FROM switch since
    # sales_thismonth is always current month (2026).
    promos = request.args.getlist("promo")

    # Which dimension to group by?
    group_by = (request.args.get("group_by") or "region").strip()
    group_cols = {
        "line":          "mat.line",
        "brand":         "mat.brand",
        "product_group": "mat.product_group",
        "region":        "cus.bde_state",
        "salesman":      "cus.salesman_name",
        "channel":       "cus.channels",
        "sold_to_group": "cus.sold_to_group",
        "sold_to":       "cus.sold_to_name",
        "pattern":       "mat.pattern",
    }
    # Promotion buckets are computed from a CASE EXISTS against
    # promo_customer + promo_plan rather than a plain column reference.
    # Built below once we know the table's year/month context (sales_-
    # thismonth has neither so we synthesise today's date as literals).
    is_promo_group = group_by in ("promotion", "promotion_detail")
    if not is_promo_group and group_by not in group_cols:
        return jsonify({"error": "invalid group_by"}), 400

    # ---- Build base JOINs / WHEREs (same as daily_sales) ----
    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)

    if f["category"] != "443":
        cat_joins, cat_where = category_filters_sales("s", f["category"], has_brand=True)
        joins += cat_joins
        wh    += cat_where

    # Carrying/customer join needed for group_by or filter
    if (group_by in ("line", "brand", "product_group", "pattern") or
        is_promo_group or
        f["product_group"] != "ALL" or f["pattern"] != "ALL" or
        f["material"] != "ALL" or f["code"] != "ALL"):
        _ensure_carrying_join("s", joins)
    if group_by in ("region", "salesman", "channel", "sold_to_group", "sold_to") or is_promo_group:
        _ensure_customer_join("s", joins)
    # Force the Product top-level cascade to only show PCLT + TBR (the
    # two carrying lines that make sense as a product split).  HM and
    # any future lines are intentionally hidden because HM reads as a
    # customer-side pivot, not a product one.
    if group_by in ("line", "brand"):
        # Restrict the Product cascade's top two levels to PCLT + TBR
        # (the two carrying lines that make sense as a product split).
        # HM / HK-only / LF-only rows for other lines are intentionally
        # hidden because HM reads as a customer-side pivot, not a
        # product one.  Brand level (HK vs LF) inherits the same gate.
        wh.append("mat.line IN ('PCLT','TBR')")

    # Pre-aggregated qty per (ship_to, day, brand) for TrueBlue's
    # "X tires at this shop on this day" rule.  Added lazily — only
    # when promo logic is involved (group_by=promotion* or any promo
    # filter selected) so non-promo views skip the extra GROUP BY.
    # The alias `dq.day_qty` is what _promo_qty_match_sql checks
    # against pc.min_qty for TrueBlue rules.
    DAY_QTY_JOIN = (
        "LEFT JOIN ("
        "  SELECT ship_to, day, brand, SUM(qty) AS day_qty"
        "  FROM sales_thismonth"
        "  GROUP BY ship_to, day, brand"
        ") dq ON dq.ship_to = s.ship_to AND dq.day = s.day AND dq.brand = s.brand"
    )
    if is_promo_group or promos:
        if DAY_QTY_JOIN not in joins:
            joins.append(DAY_QTY_JOIN)

    if is_promo_group:
        # sales_thismonth has no year/month columns — feed today's as
        # literals (same reasoning as daily_sales).
        from datetime import date as _d
        _today = _d.today()
        group_col = _promotion_group_col_sql(
            detail=(group_by == "promotion_detail"),
            year_expr=str(_today.year),
            month_expr=str(_today.month),
            day_qty_alias="dq.day_qty",
        )
    else:
        group_col = group_cols[group_by]
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s");       params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s")
        params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s")
        params.append(f["code"])

    if promos:
        _ensure_carrying_join("s", joins)
        _ensure_customer_join("s", joins)
        # sales_thismonth has no year/month columns — feed today's as
        # literals (same reasoning as daily_sales).
        from datetime import date as _d
        _today = _d.today()
        promo_wh, promo_p = _promo_filter_clauses(
            promos,
            year_expr=str(_today.year),
            month_expr=str(_today.month),
            day_qty_alias="dq.day_qty",
        )
        wh.extend(promo_wh)
        params.extend(promo_p)

    base_where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        top_sold_to = None

        # 1) Get Top N sold_to from baseline table (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )

            # no matching customers ??nothing to show
            if not top_sold_to:
                return jsonify([])

        # 2) Daily breakdown, restricted to those top customers,
        #    but stacked by group_col (region / salesman / etc.)
        wh2 = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        # GROUP BY uses the SELECT alias when the bucket expression is
        # the giant CASE+EXISTS of the promotion grouping — duplicating
        # that correlated subquery in GROUP BY trips some MySQL builds.
        # Other group_by values keep the existing expression so the
        # COALESCE 'COMMON' fallback still applies.
        if is_promo_group:
            group_by_sql = "GROUP BY s.day, group_label"
        else:
            group_by_sql = f"GROUP BY s.day, COALESCE(NULLIF(TRIM({group_col}),''), 'COMMON')"
        sql = f"""
        SELECT s.day AS day,
                COALESCE(NULLIF(TRIM({group_col}),''), 'COMMON') AS group_label,
                SUM(s.{value}) AS value
            FROM sales_thismonth s
            {' '.join(joins)}
            {where_sql2}
        {group_by_sql}
        ORDER BY s.day
        """
        try:
            cur.execute(sql, tuple(params2))
        except Exception as _e:
            # Surface the real DB error so promotion-grouping regressions
            # show up in the logs instead of a bare 500.
            print(f"[daily_breakdown] group_by={group_by} SQL failed: {_e}\nSQL:\n{sql}")
            raise
        rows = cur.fetchall()

    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

    return jsonify(rows)


# ----------------------------- Daily Target (Oct) ---------------------------------
import calendar
@app.get("/api/daily_target")
def daily_target():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # which month? default to October (10) if nothing is passed
    month = int(request.args.get("month", 2))

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    joins, wh, params = build_target_filters("t", f)
    cat_joins, cat_where = category_target_filters("t", f["category"])
    joins += cat_joins
    wh    += cat_where

    # Apply product_group / pattern / size filters via carrying_26 join.
    carrying_join_dt = "LEFT JOIN carrying_26 mat ON mat.m_code = t.material"
    if (f["product_group"] != "ALL" or f["pattern"] != "ALL" or f["material"] != "ALL" or f["code"] != "ALL"):
        if carrying_join_dt not in joins:
            joins.append(carrying_join_dt)
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s"); params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s"); params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s"); params.append(f["code"])

    # restrict to the chosen month only
    wh.append("t.month = %s")
    params.append(month)

    base_where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        top_sold_to = None

        # 1) Get Top N sold_to from baseline table (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )                                               # CHANGED

            # nothing matched -> all days = 0
            if not top_sold_to:
                days_in_month = calendar.monthrange(2026, month)[1]   # CHANGED (2026)
                return jsonify([
                    {"day": d, "value": 0}
                    for d in range(1, days_in_month + 1)
                ])

        # 2) Monthly target, optionally restricted to baseline Top customers
        wh2     = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"t.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        sql = f"""
        SELECT t.month AS month_num, SUM(t.{value}) AS monthly_total
            FROM target_26 t
            {' '.join(joins)}
            {where_sql2}
        GROUP BY t.month
        ORDER BY t.month
        """

        cur.execute(sql, tuple(params2))
        row = cur.fetchone()

    finally:
        cur.close()
        conn.close()

    monthly_total = float(row["monthly_total"] or 0) if row else 0

    # how many days in that month? (target_26 is for 2026)
    days_in_month = calendar.monthrange(2026, month)[1]

    # ?? Determine working days ????????????????????????????????????????????????
    # Past days: working if company-wide total sales >= 10 (same threshold as frontend)
    # Future days: working if weekday (Mon?밊ri)
    conn2 = get_connection()
    cur2  = conn2.cursor(dictionary=True)
    try:
        cur2.execute("""
            SELECT day, SUM(qty) AS total_qty
            FROM sales_thismonth
            GROUP BY day
            ORDER BY day
        """)
        sales_by_day = {int(r["day"]): float(r["total_qty"] or 0) for r in cur2.fetchall()}
    finally:
        cur2.close()
        conn2.close()

    max_known_day = max(sales_by_day.keys()) if sales_by_day else 0

    from datetime import date as _date
    working_days = set()
    for d in range(1, days_in_month + 1):
        if d <= max_known_day:
            # past day: working if total company sales >= 10
            if sales_by_day.get(d, 0) >= 10:
                working_days.add(d)
        else:
            # future day: working if Mon?밊ri
            if _date(2026, month, d).weekday() < 5:
                working_days.add(d)

    total_working_days = len(working_days)
    daily_value = monthly_total / total_working_days if total_working_days else 0

    # return one entry per day: working days get daily_value, non-working days get 0
    return jsonify([
        {"day": d, "value": daily_value if d in working_days else 0}
        for d in range(1, days_in_month + 1)
    ])

def autosize_columns(ws, max_width=60):
    """
    Auto adjust column width based on content length
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)

        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def build_excel(rows, sheet_name="Data", header_order=None, meta_lines=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # 1) 留??꾩뿉 ?좏깮 議곌굔 ?쒖떆 (?듭뀡)
    start_row = 1
    if meta_lines:
        for line in meta_lines:
            ws.append([line])
        ws.append([])  # blank line
        start_row = ws.max_row + 1

    if not rows:
        ws.append(["No data"])
        return wb

    headers = header_order or list(rows[0].keys())

    ws.append(headers)
    header_row = ws.max_row
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r.get(h) for h in headers])

    # ?ㅻ뜑 ?ㅼ쓬 以꾨줈 freeze
    ws.freeze_panes = f"A{header_row+1}"

    autosize_columns(ws)
    # meta_lines in col A/B can be very long ??cap to keep sheet tidy
    if meta_lines:
        ws.column_dimensions['A'].width = min(ws.column_dimensions['A'].width, 10)
        ws.column_dimensions['B'].width = min(ws.column_dimensions['B'].width, 18)
    return wb
def fetch_table_rows(top_limit: int):
    f = parse_filters(request)
    metric = (request.args.get("metric") or f.get("metric") or "qty").lower()
    value_col = "qty" if metric == "qty" else "amt"

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    # pivot columns: 1~31
    day_cols = ",\n".join([
        f"SUM(CASE WHEN s.`day` = {d} THEN s.{value_col} ELSE 0 END) AS `{d}`"
        for d in range(1, 32)
    ])

    # ---------- STEP 1: Top-N sold_to (dynamic per current filter slice) ----------
    top_pairs = None
    if top_limit in (10, 20, 30):
        sold_to_list = get_top_sold_to_from_baseline(cur, f, int(top_limit), metric)
        if sold_to_list:
            top_pairs = [{"sold_to": s} for s in sold_to_list]

    # ---------- STEP 2: build filters (same as daily_sales) ----------
    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)
    _ensure_customer_join("s", joins)  # always needed: SELECT uses cus.* columns
    # Always join carrying_26 — the daily export splits rows by line.
    _ensure_carrying_join("s", joins)

    # category (same rule: skip 443) ??use normalised version
    if f.get("category", "ALL") != "443":
        cat_joins, cat_where = category_filters_sales("s", f.get("category", "ALL"), has_brand=True)
        joins += cat_joins
        wh    += cat_where
    if f.get("product_group", "ALL") != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f.get("pattern", "ALL") != "ALL":
        wh.append("mat.pattern = %s"); params.append(f["pattern"])
    if f.get("material", "ALL") != "ALL":
        wh.append("mat.size = %s"); params.append(f["material"])

    # top filter (sold_to 湲곗?)
    if top_pairs:
        conds = []
        for p in top_pairs:
            conds.append("s.sold_to = %s")
            params.append(p["sold_to"])
        wh.append("(" + " OR ".join(conds) + ")")

    where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    sql = f"""
        SELECT
            COALESCE(NULLIF(TRIM(cus.bde_state),''), 'COMMON') AS region,
            COALESCE(NULLIF(TRIM(cus.salesman_name),''), '') AS bde,
            COALESCE(NULLIF(TRIM(cus.channels),''), '') AS channel,
            COALESCE(NULLIF(TRIM(cus.sold_to_group),''), '') AS sold_to_group,
            COALESCE(NULLIF(TRIM(cus.sold_to_name),''), s.sold_to) AS sold_to_name,
            COALESCE(NULLIF(TRIM(cus.ship_to_name),''), s.ship_to) AS ship_to_name,
            s.sold_to AS sold_to_code,
            s.ship_to AS ship_to_code,
            COALESCE(NULLIF(TRIM(mat.line),''), '') AS line,
            {day_cols}
        FROM sales_thismonth s
        {' '.join(joins)}
        {where_sql}
        GROUP BY region, bde, channel, sold_to_group, sold_to_name, ship_to_name, s.sold_to, s.ship_to, line
        ORDER BY region DESC, bde DESC
    """

    cur.execute(sql, params)
    rows = cur.fetchall()

    cur.close()
    conn.close()
    return rows
@app.get("/api/export_excel")
def export_excel():
    try:
        top_limit = int(request.args.get("top_limit", "0") or "0")
        f = parse_filters(request)
        metric = (request.args.get("metric") or f.get("metric") or "qty").lower()
        value_col = "qty" if metric == "qty" else "amt"

        rows = fetch_table_rows(top_limit=top_limit)

        # Sort: NSW?뭂LD?뭋IC?뭌A, then BDE, then sold_to_name
        _REGION_ORDER = {"NSW": 0, "QLD": 1, "VIC": 2, "WA": 3}
        rows.sort(key=lambda r: (
            _REGION_ORDER.get((r.get("region") or "").upper(), 99),
            (r.get("bde") or "").lower(),
            (r.get("sold_to_name") or "").lower(),
        ))

        day_labels = [str(d) for d in range(1, 32)]
        for r in rows:
            r["Total"] = sum(float(r.get(c) or 0) for c in day_labels)

        # ?? Fetch monthly target ??ship_to level, fall back to sold_to ??
        conn2 = get_connection(); cur2 = conn2.cursor(dictionary=True)
        try:
            cur2.execute("SELECT MAX(year*100+month) AS ym FROM sales_thismonth")
            ym = int((cur2.fetchone() or {}).get("ym") or 0)
            cur_month = ym % 100 if ym else datetime.now().month
            # Fetch both ship_to and sold_to targets in one query
            cur2.execute(
                f"SELECT sold_to, ship_to, SUM({value_col}) AS tgt "
                f"FROM target_26 WHERE month=%s GROUP BY sold_to, ship_to",
                (cur_month,)
            )
            target_by_ship   = {}   # ship_to  -> target
            target_by_sold_to = {}  # sold_to  -> target
            for r2 in cur2.fetchall():
                st  = str(r2["ship_to"]  or "")
                so  = str(r2["sold_to"]  or "")
                tgt = float(r2["tgt"] or 0)
                if st:
                    target_by_ship[st]    = target_by_ship.get(st, 0) + tgt
                if so:
                    target_by_sold_to[so] = target_by_sold_to.get(so, 0) + tgt
        except Exception:
            target_by_ship = {}; target_by_sold_to = {}
        finally:
            cur2.close(); conn2.close()

        # Count ship_tos per sold_to (for proportional fallback)
        ship_count = {}
        for r in rows:
            sc = str(r.get("sold_to_code") or "")
            ship_count[sc] = ship_count.get(sc, 0) + 1

        for r in rows:
            ship = str(r.get("ship_to_code")  or "")
            sold = str(r.get("sold_to_code")  or "")
            # Use ship_to target if available; otherwise split sold_to target evenly
            tgt = target_by_ship.get(ship, 0)
            if tgt == 0 and sold in target_by_sold_to:
                cnt = ship_count.get(sold, 1)
                tgt = target_by_sold_to[sold] / cnt if cnt else 0
            r["Target"] = round(tgt, 1)
            r["Ach%"]   = round(r["Total"] / tgt * 100, 1) if tgt > 0 else None

        # ?? Pre-calculate state totals (sum sold_to targets once per sold_to) ??
        counted_sold_tos = set()
        state_totals = {}
        for r in rows:
            st = (r.get("region") or "").upper()
            sc = str(r.get("sold_to_code") or "")
            g  = state_totals.setdefault(st, {"Total": 0.0, "Target": 0.0})
            g["Total"] += r["Total"]
            # Add sold_to target only once per sold_to to avoid double-counting
            if sc not in counted_sold_tos:
                g["Target"] += r["Target"]
                counted_sold_tos.add(sc)

        # ?? Achievement color (matches KPI table) ????????????????????????
        def _ach_color(val):
            if val is None: return None
            if val >= 100:  return "16a34a"   # green
            if val >= 90:   return "f97316"   # orange
            return "dc2626"                    # red

        # ?? Build workbook ???????????????????????????????????????????????
        wb  = Workbook()
        ws  = wb.active
        ws.title = "sales_thismonth_by_day"

        # Meta info
        for line in [
            f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"metric={metric}, top_limit={top_limit if top_limit else 'ALL'}",
            f"category={f.get('category','ALL')}, region={f.get('region','ALL')}, salesman={f.get('salesman','ALL')}, sold_to_group={f.get('sold_to_group','ALL')}",
            f"product_group={f.get('product_group','ALL')}, pattern={f.get('pattern','ALL')}, material={f.get('material','ALL')}",
            f"sold_to={f.get('sold_to','ALL')}, ship_to={f.get('ship_to','ALL')}",
        ]:
            ws.append([line])
        ws.append([])

        # Header row: Total/Target/Ach% sit between ship_to_code and day columns
        HDR = (["region","bde","channel","sold_to_group","sold_to_name","ship_to_name",
                "sold_to_code","ship_to_code","line","Total","Target","Ach%"]
               + day_labels)
        ws.append(HDR)
        hdr_row = ws.max_row
        for c, h in enumerate(HDR, 1):
            cell = ws.cell(row=hdr_row, column=c)
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = f"A{hdr_row+1}"

        ACH_COL = HDR.index("Ach%") + 1   # 1-based column index

        GREY_FILL = PatternFill("solid", fgColor="D9D9D9")

        def _append_state_summary(region):
            g    = state_totals.get(region, {})
            tot  = g.get("Total",  0.0)
            tgt  = g.get("Target", 0.0)
            ach  = round(tot / tgt * 100, 1) if tgt > 0 else None
            vals = ([region, "", "", "", f"?? {region} TOTAL ??", "", "",
                     "", "", round(tot,1), round(tgt,1),
                     (f"{ach:.1f}%" if ach is not None else "-")]
                    + [""] * 31)
            ws.append(vals)
            sr = ws.max_row
            for c in range(1, len(HDR)+1):
                ws.cell(row=sr, column=c).font = Font(bold=True)
                ws.cell(row=sr, column=c).fill = GREY_FILL
            if ach is not None:
                ws.cell(row=sr, column=ACH_COL).font = Font(bold=True, color=_ach_color(ach))

        current_region = None
        for r in rows:
            region = (r.get("region") or "").upper()
            if region != current_region:
                current_region = region
                _append_state_summary(region)

            # Build row values
            vals = []
            for h in HDR:
                if h == "Ach%":
                    v = r.get("Ach%")
                    vals.append(f"{v:.1f}%" if v is not None else "-")
                else:
                    vals.append(r.get(h))
            ws.append(vals)

            # Colour the Ach% cell
            ach_val = r.get("Ach%")
            if ach_val is not None:
                ws.cell(row=ws.max_row, column=ACH_COL).font = Font(color=_ach_color(ach_val))

        # Column widths
        autosize_columns(ws)
        ws.column_dimensions["A"].width = min(ws.column_dimensions["A"].width, 10)
        ws.column_dimensions["B"].width = min(ws.column_dimensions["B"].width, 18)

        bio = BytesIO()
        wb.save(bio); bio.seek(0)
        stamp    = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"sales_thismonth_top{top_limit if top_limit else 'ALL'}_{metric}_{stamp}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print("export_excel error:", repr(e))
        print(traceback.format_exc())
        return jsonify({"error": "Internal Server Error", "message": str(e)}), 500


def _build_export_common_filters(f, joins, wh, params, alias="s"):
    """Apply category, product_group, pattern, material filters shared by all export endpoints."""
    if f.get("category", "ALL") != "ALL":
        cat_joins, cat_where = category_filters_sales(alias, f.get("category", "ALL"))
        joins += cat_joins
        wh    += cat_where

    if f.get("product_group", "ALL") != "ALL" or f.get("pattern", "ALL") != "ALL":
        _ensure_carrying_join(alias, joins)
    if f.get("product_group", "ALL") != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f.get("pattern", "ALL") != "ALL":
        wh.append("mat.pattern = %s"); params.append(f["pattern"])
    if f.get("material", "ALL") != "ALL":
        wh.append(f"{alias}.material = %s"); params.append(f["material"])


@app.get("/api/export_excel/sales2526")
def export_excel_sales2526():
    """Export 25/26 monthly sales pivoted by YYMM (2501..2512, 2601..2612)
    with matching Target columns for the 2026 months (T2601..T2612).
    Sales come from sales_2526, targets from target_26; the two pivots
    are merged in Python on (sold_to, ship_to) so a ship_to that only
    has target (no sales yet) still shows up on its own row."""
    try:
        f = parse_filters(request)
        metric = f.get("metric", "qty")
        value_col = "qty" if metric == "qty" else "amt"

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        # ------ 1. Sales pivot (sales_2526) ------
        joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)
        _ensure_customer_join("s", joins)
        _build_export_common_filters(f, joins, wh, params)
        # Force the carrying_26 join in even when no product filter is
        # active — we need mat.line to split rows into PCLT / TBR.
        _ensure_carrying_join("s", joins)

        # pivot columns: YEAR*100+MONTH from billing_date → label YYMM (2501…)
        pivot_cols = ",\n".join([
            f"SUM(CASE WHEN YEAR(s.billing_date)={y} AND MONTH(s.billing_date)={m} "
            f"THEN s.{value_col} ELSE 0 END) AS `{y % 100:02d}{m:02d}`"
            for y in [2025, 2026]
            for m in range(1, 13)
        ])
        col_labels = [f"{y % 100:02d}{m:02d}" for y in [2025, 2026] for m in range(1, 13)]

        where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""
        sql = f"""
            SELECT
                COALESCE(NULLIF(TRIM(cus.bde_state),''), 'COMMON') AS region,
                COALESCE(NULLIF(TRIM(cus.salesman_name),''), '') AS bde,
                COALESCE(NULLIF(TRIM(cus.channels),''), '') AS channel,
                COALESCE(NULLIF(TRIM(cus.sold_to_group),''), '') AS sold_to_group,
                COALESCE(NULLIF(TRIM(cus.sold_to_name),''), s.sold_to) AS sold_to_name,
                COALESCE(NULLIF(TRIM(cus.ship_to_name),''), s.ship_to) AS ship_to_name,
                s.sold_to AS sold_to_code,
                s.ship_to AS ship_to_code,
                COALESCE(NULLIF(TRIM(mat.line),''), '') AS line,
                {pivot_cols}
            FROM {_sales_2526_from("s")}
            {' '.join(joins)}
            {where_sql}
            GROUP BY region, bde, channel, sold_to_group, sold_to_name, ship_to_name, s.sold_to, s.ship_to, line
            ORDER BY region DESC, bde DESC
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        # ------ 2. Target pivot (target_26) ------
        # Same filter set the daily/monthly target endpoints use so the
        # exported Target columns match what's on the dashboard.
        t_joins, t_wh, t_params = build_target_filters("t", f)
        t_cj, t_cw = category_target_filters("t", f["category"])
        t_joins += t_cj; t_wh += t_cw

        # Product filters via carrying_26 on target_26.material.  We
        # always need this join now because the export splits rows by
        # mat.line (PCLT / TBR).
        carrying_join_t = "LEFT JOIN carrying_26 mat ON mat.m_code = t.material"
        if f["product_group"] != "ALL":
            t_wh.append("mat.product_group = %s"); t_params.append(f["product_group"])
        if f["brand"] != "ALL":
            t_wh.append("mat.brand = %s"); t_params.append(f["brand"])
        if f["pattern"] != "ALL":
            t_wh.append("mat.pattern = %s"); t_params.append(f["pattern"])
        if f["material"] != "ALL":
            t_wh.append("mat.size = %s"); t_params.append(f["material"])
        if f["code"] != "ALL":
            t_wh.append("mat.m_code = %s"); t_params.append(f["code"])
        if carrying_join_t not in t_joins:
            t_joins.append(carrying_join_t)

        target_pivot = ",\n".join([
            f"SUM(CASE WHEN t.month={m} THEN t.{value_col} ELSE 0 END) AS `T26{m:02d}`"
            for m in range(1, 13)
        ])
        target_labels = [f"T26{m:02d}" for m in range(1, 13)]

        # customer LEFT JOIN so a target-only row (no matching sales yet
        # for this ship_to) can still surface a readable name.
        cust_join_t = "LEFT JOIN customer tcus ON tcus.ship_to = t.ship_to"
        if cust_join_t not in t_joins:
            t_joins.append(cust_join_t)

        t_where_sql = ("WHERE " + " AND ".join(t_wh)) if t_wh else ""
        # Wrap every non-grouped column in MIN() so only_full_group_by
        # accepts the SELECT.  Each (sold_to, ship_to) resolves to a
        # single customer row, so MIN() is equivalent to picking the
        # only value; it just satisfies the strict-mode checker that
        # can't see the alias-equivalence through the COALESCE stack.
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(TRIM(MIN(tcus.bde_state)),''),
                         NULLIF(TRIM(MIN(t.state)),''), 'COMMON') AS region,
                COALESCE(NULLIF(TRIM(MIN(tcus.salesman_name)),''),
                         NULLIF(TRIM(MIN(t.bde)),''), '') AS bde,
                COALESCE(NULLIF(TRIM(MIN(tcus.channels)),''), '') AS channel,
                COALESCE(NULLIF(TRIM(MIN(tcus.sold_to_group)),''), '') AS sold_to_group,
                COALESCE(NULLIF(TRIM(MIN(tcus.sold_to_name)),''), t.sold_to) AS sold_to_name,
                COALESCE(NULLIF(TRIM(MIN(tcus.ship_to_name)),''), t.ship_to) AS ship_to_name,
                t.sold_to AS sold_to_code,
                t.ship_to AS ship_to_code,
                COALESCE(NULLIF(TRIM(mat.line),''), '') AS line,
                {target_pivot}
            FROM target_26 t
            {' '.join(t_joins)}
            {t_where_sql}
            GROUP BY t.sold_to, t.ship_to, line
        """, tuple(t_params))
        target_rows = cur.fetchall()
        cur.close(); conn.close()

        # ------ 3. Merge sales + target on (sold_to, ship_to) ------
        # Target numbers are always whole units — targets aren't set to
        # a fractional tyre — so round each pivot cell and the total to
        # int before they land on the sheet.
        as_int = lambda v: int(round(float(v or 0)))
        # Merge key includes line so PCLT sales match PCLT targets and
        # TBR sales match TBR targets — targets aren't cross-line.
        target_map = {
            (r.get("sold_to_code") or "", r.get("ship_to_code") or "",
             (r.get("line") or "").upper()): r
            for r in target_rows
        }
        # Per-year sales totals rather than a single combined Total —
        # 25 Total sits right after the 2025 columns, 26 Total right
        # after 2026 columns, and Target_Total after the T26MM block.
        labels_25 = [c for c in col_labels if c.startswith("25")]
        labels_26 = [c for c in col_labels if c.startswith("26")]

        for r in rows:
            key = ((r.get("sold_to_code") or ""),
                   (r.get("ship_to_code") or ""),
                   (r.get("line") or "").upper())
            t = target_map.pop(key, None)
            for lbl in target_labels:
                r[lbl] = as_int((t or {}).get(lbl))
            r["25 Total"]     = sum(float(r.get(c) or 0) for c in labels_25)
            r["26 Total"]     = sum(float(r.get(c) or 0) for c in labels_26)
            r["Target_Total"] = sum(r.get(l) or 0 for l in target_labels)
        # Target-only ship_tos (no sales row yet) — append with sales
        # months as zero so they still show up on the sheet.
        for key, t in target_map.items():
            new_row = {
                "region":         t.get("region")        or "",
                "bde":            t.get("bde")           or "",
                "channel":        t.get("channel")       or "",
                "sold_to_group":  t.get("sold_to_group") or "",
                "sold_to_name":   t.get("sold_to_name")  or "",
                "ship_to_name":   t.get("ship_to_name")  or "",
                "sold_to_code":   key[0],
                "ship_to_code":   key[1],
                "line":           t.get("line") or "",
            }
            for lbl in col_labels:
                new_row[lbl] = 0
            for lbl in target_labels:
                new_row[lbl] = as_int(t.get(lbl))
            new_row["25 Total"]     = 0
            new_row["26 Total"]     = 0
            new_row["Target_Total"] = sum(new_row.get(l) or 0 for l in target_labels)
            rows.append(new_row)

        header_order = (["region", "bde", "channel", "sold_to_group", "sold_to_name", "ship_to_name",
                         "sold_to_code", "ship_to_code", "line"]
                        + labels_25 + ["25 Total"]
                        + labels_26 + ["26 Total"]
                        + target_labels + ["Target_Total"])

        meta_lines = [
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"metric={metric}, category={f.get('category','ALL')}, region={f.get('region','ALL')}",
            f"salesman={f.get('salesman','ALL')}, sold_to_group={f.get('sold_to_group','ALL')}",
            f"product_group={f.get('product_group','ALL')}, pattern={f.get('pattern','ALL')}, material={f.get('material','ALL')}",
            f"sold_to={f.get('sold_to','ALL')}, ship_to={f.get('ship_to','ALL')}",
        ]

        wb = build_excel(rows, sheet_name="25_26_Sales", header_order=header_order, meta_lines=meta_lines)
        bio = BytesIO(); wb.save(bio); bio.seek(0)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        return send_file(bio, as_attachment=True,
                         download_name=f"sales_25_26_{metric}_{stamp}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print("export_excel_sales2526 error:", repr(e)); print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.get("/api/export_excel/yearly")
def export_excel_yearly():
    """Export yearly sales (2021-2025) pivoted by year."""
    try:
        f = parse_filters(request)
        metric = f.get("metric", "qty")
        value_col = "qty" if metric == "qty" else "amt"

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)
        _ensure_customer_join("s", joins)
        _build_export_common_filters(f, joins, wh, params)
        # Force carrying_26 join to surface mat.line (PCLT / TBR).
        _ensure_carrying_join("s", joins)

        years = [2021, 2022, 2023, 2024, 2025]
        col_labels = [str(y % 100) for y in years]
        pivot_cols = ",\n".join([
            f"SUM(CASE WHEN s.year={y} THEN s.{value_col} ELSE 0 END) AS `{y % 100}`"
            for y in years
        ])

        where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""
        sql = f"""
            SELECT
                COALESCE(NULLIF(TRIM(cus.bde_state),''), 'COMMON') AS region,
                COALESCE(NULLIF(TRIM(cus.salesman_name),''), '') AS bde,
                COALESCE(NULLIF(TRIM(cus.channels),''), '') AS channel,
                COALESCE(NULLIF(TRIM(cus.sold_to_group),''), '') AS sold_to_group,
                COALESCE(NULLIF(TRIM(cus.sold_to_name),''), s.sold_to) AS sold_to_name,
                COALESCE(NULLIF(TRIM(cus.ship_to_name),''), s.ship_to) AS ship_to_name,
                s.sold_to AS sold_to_code,
                s.ship_to AS ship_to_code,
                COALESCE(NULLIF(TRIM(mat.line),''), '') AS line,
                {pivot_cols}
            FROM sales_21_25 s
            {' '.join(joins)}
            {where_sql}
            GROUP BY region, bde, channel, sold_to_group, sold_to_name, ship_to_name, s.sold_to, s.ship_to, line
            ORDER BY region DESC, bde DESC
        """
        cur.execute(sql, params)
        rows = cur.fetchall()
        cur.close(); conn.close()

        for r in rows:
            r["Total"] = sum(float(r.get(c) or 0) for c in col_labels)

        header_order = ["region", "bde", "channel", "sold_to_group", "sold_to_name", "ship_to_name",
                        "sold_to_code", "ship_to_code", "line"] + col_labels + ["Total"]

        meta_lines = [
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"metric={metric}, category={f.get('category','ALL')}, region={f.get('region','ALL')}",
            f"salesman={f.get('salesman','ALL')}, sold_to_group={f.get('sold_to_group','ALL')}",
            f"product_group={f.get('product_group','ALL')}, pattern={f.get('pattern','ALL')}, material={f.get('material','ALL')}",
            f"sold_to={f.get('sold_to','ALL')}, ship_to={f.get('ship_to','ALL')}",
        ]

        wb = build_excel(rows, sheet_name="Yearly_Sales", header_order=header_order, meta_lines=meta_lines)
        bio = BytesIO(); wb.save(bio); bio.seek(0)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        return send_file(bio, as_attachment=True,
                         download_name=f"sales_yearly_{metric}_{stamp}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print("export_excel_yearly error:", repr(e)); print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


        # ----------------------------- Monthly Sales ---------------------------------
@app.get("/api/monthly_sales")
@cached_endpoint(60)
def monthly_sales():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # Year filter (default: current year if not provided)
    year = int(request.args.get("year", 2025) or 2025)

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    # Selected sub-promos from the new 443 / iON / TrueBlue buttons.
    # Both 2025 and 2026 are now in scope — promo_plan covers both
    # years (TrueBlue activates from July 2025, the others from start
    # of FY).  The promo helpers gate each sale on the plan period
    # via promo_plan.start_date / end_date, so a 2025 query simply
    # picks up the rules whose period covers that month.
    promos = request.args.getlist("promo")

    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)

    # category ??normalised version
    cat_joins, cat_where = category_filters_sales("s", f["category"])
    joins += cat_joins
    wh    += cat_where

    if f["product_group"] != "ALL" or f["pattern"] != "ALL" or f["material"] != "ALL" or f["code"] != "ALL":
        _ensure_carrying_join("s", joins)
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s")
        params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s")
        params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s")
        params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s")
        params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s")
        params.append(f["code"])

    # Both 2025 and 2026 read directly from sales_2526 using
    # billing_date (the base column that's guaranteed to be there and
    # indexed).  Grouping happens on MONTH(billing_date) so June's
    # sales_2526 rows land in the June bar as soon as the table is
    # loaded — no per-month union / no dependency on generated year /
    # month columns.
    year_expr  = "YEAR(s.billing_date)"
    month_expr = "MONTH(s.billing_date)"

    if promos:
        # Promo filter needs carrying_26 (mat) for line/product_group +
        # customer (cus) for the sold_to_group fallback in the EXISTS.
        # NOTE: monthly_sales stays on per-row qty for TrueBlue (the
        # ship_to+day SUM variant lives on daily endpoints only —
        # adding a self-referencing UNION subquery inside EXISTS at
        # this aggregation layer is too expensive).
        _ensure_carrying_join("s", joins)
        _ensure_customer_join("s", joins)
        promo_wh, promo_p = _promo_filter_clauses(
            promos,
            year_expr=year_expr,
            month_expr=month_expr,
        )
        wh.extend(promo_wh)
        params.extend(promo_p)

    # Year filter as a billing_date range so the index can prune.
    wh.append("s.billing_date >= %s AND s.billing_date < %s")
    params.extend([f"{year}-01-01", f"{year+1}-01-01"])
    # Cap at the last day of the business-effective current month so
    # the nightly sales_thismonth batch tagged with the CURRENT
    # calendar month doesn't spawn a phantom bar at the next month.
    from calendar import monthrange as _mr
    _eff_y, _eff_m = _business_effective_ym()
    _eff_last = _mr(_eff_y, _eff_m)[1]
    wh.append("s.billing_date <= %s")
    params.append(f"{_eff_y}-{_eff_m:02d}-{_eff_last:02d}")

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        from_sql = "FROM sales_2526 s"

        top_sold_to = None

        # 1) Get Top N sold_to from baseline table (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )

            if not top_sold_to:
                return jsonify([{"month": m, "value": 0} for m in range(1, 13)])

        # 2) Monthly totals, optionally restricted to baseline Top customers
        wh2 = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        monthly_sql = f"""
        SELECT {month_expr} AS month_num, SUM(s.{value}) AS monthly_total
            {from_sql}
            {' '.join(joins)}
            {where_sql2}
        GROUP BY {month_expr}
        ORDER BY {month_expr}
        """
        cur.execute(monthly_sql, tuple(params2))
        rows = cur.fetchall()

    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    month_map = {int(r["month_num"]): float(r["monthly_total"] or 0) for r in rows}
    return jsonify([{"month": m, "value": month_map.get(m, 0)} for m in range(1, 13)])


# -------------------- Monthly breakdown (stacked by group) -------------------
@app.get("/api/monthly_breakdown")
@cached_endpoint(60)
def monthly_breakdown():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # Year filter (default: 2025)
    year = int(request.args.get("year", 2025) or 2025)

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    # Promo filter — promo_plan now covers both 2025 and 2026, so a
    # query for either year picks up the rules whose period covers it.
    # When promos are selected, _promo_filter_clauses adds the PCLT
    # + customer + dc_rate-range + plan-period match conditions.
    promos = request.args.getlist("promo")

    # Which dimension to group by?
    group_by = (request.args.get("group_by") or "region").strip()
    # For sold_to: GROUP BY the code (s.sold_to) so sales and target rows
    # match by the same key, but SELECT the resolved name as the label
    # so the legend is readable.
    group_cols = {
        "line":          "mat.line",
        "product_group": "mat.product_group",
        "region":        "cus.bde_state",
        "salesman":      "cus.salesman_name",
        "channel":       "cus.channels",
        "sold_to_group": "cus.sold_to_group",
        "sold_to":       "s.sold_to",
        "pattern":       "mat.pattern",
        "brand":         "mat.brand",
    }
    is_promo_group = group_by in ("promotion", "promotion_detail")
    if not is_promo_group and group_by not in group_cols:
        return jsonify({"error": "invalid group_by"}), 400

    # Both 2025 and 2026 read directly from sales_2526 using
    # billing_date (the base column that's guaranteed to be there and
    # indexed).  Grouping happens on MONTH(billing_date) so June's
    # sales_2526 rows land in the June bar as soon as the table is
    # loaded — no per-month union / no dependency on generated year /
    # month columns.
    year_expr  = "YEAR(s.billing_date)"
    month_expr = "MONTH(s.billing_date)"

    if is_promo_group:
        # Promotion buckets: feed the year/month expression so the
        # promo_plan period check gates each row correctly.
        # NOTE: monthly_breakdown stays on per-row qty for TrueBlue (the
        # ship_to+day SUM variant lives on daily_breakdown only — adding
        # a self-referencing UNION subquery inside EXISTS at this layer
        # is too expensive on the monthly aggregation roll-up).
        group_col = _promotion_group_col_sql(
            detail=(group_by == "promotion_detail"),
            year_expr=year_expr,
            month_expr=month_expr,
        )
        label_col = group_col
    elif group_by == "sold_to":
        # scus alias = customer aggregated per-sold_to. Lets us resolve a
        # consistent name regardless of which ship_to a row points at.
        group_col = group_cols[group_by]
        label_col = "MIN(COALESCE(scus.sold_to_name, s.sold_to))"
    else:
        group_col = group_cols[group_by]
        label_col = f"COALESCE(NULLIF(TRIM({group_col}),''), 'COMMON')"

    # ---- Build base JOINs / WHEREs (same pattern as monthly_sales) ----
    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)
    cat_joins, cat_where = category_filters_sales("s", f["category"])
    joins += cat_joins
    wh    += cat_where

    if (group_by in ("line", "brand", "product_group", "pattern") or
        is_promo_group or
        f["product_group"] != "ALL" or f["pattern"] != "ALL" or
        f["material"] != "ALL" or f["code"] != "ALL"):
        _ensure_carrying_join("s", joins)
    if group_by in ("region", "salesman", "channel", "sold_to_group", "sold_to") or is_promo_group:
        _ensure_customer_join("s", joins)
    # Force the Product top-level cascade to only show PCLT + TBR
    # (the two carrying lines that make sense as a product split).
    if group_by in ("line", "brand"):
        # Restrict the Product cascade's top two levels to PCLT + TBR
        # (the two carrying lines that make sense as a product split).
        # HM / HK-only / LF-only rows for other lines are intentionally
        # hidden because HM reads as a customer-side pivot, not a
        # product one.  Brand level (HK vs LF) inherits the same gate.
        wh.append("mat.line IN ('PCLT','TBR')")
    # scus = per-sold_to name resolver (one row per sold_to). Used as
    # the label source when group_by == 'sold_to' so the resulting
    # legend doesn't have a mix of names and raw codes for the same
    # sold_to depending on which ship_to row was joined.
    if group_by == "sold_to":
        joins.append(
            "LEFT JOIN ("
            "  SELECT sold_to, MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name "
            "  FROM customer GROUP BY sold_to"
            ") scus ON scus.sold_to = s.sold_to"
        )
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s");       params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s");          params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s");          params.append(f["code"])

    if promos:
        # Promo filter needs carrying_26 (mat) for line/product_group +
        # customer (cus) for the sold_to_group fallback in the EXISTS.
        _ensure_carrying_join("s", joins)
        _ensure_customer_join("s", joins)
        promo_wh, promo_p = _promo_filter_clauses(
            promos,
            year_expr=year_expr,
            month_expr=month_expr,
        )
        wh.extend(promo_wh)
        params.extend(promo_p)

    # Year filter as a billing_date range so the index can prune.
    wh.append("s.billing_date >= %s AND s.billing_date < %s")
    params.extend([f"{year}-01-01", f"{year+1}-01-01"])
    # Cap at the last day of the business-effective current month so
    # the nightly sales_thismonth batch tagged with the CURRENT
    # calendar month doesn't spawn a phantom bar at the next month.
    from calendar import monthrange as _mr
    _eff_y, _eff_m = _business_effective_ym()
    _eff_last = _mr(_eff_y, _eff_m)[1]
    wh.append("s.billing_date <= %s")
    params.append(f"{_eff_y}-{_eff_m:02d}-{_eff_last:02d}")

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        from_sql = "FROM sales_2526 s"

        top_sold_to = None

        # 1) Get Top N sold_to from baseline table (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )                                               # CHANGED

            if not top_sold_to:
                return jsonify([])

        # 2) Monthly breakdown
        wh2     = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        # Same alias trick as daily_breakdown: avoid duplicating the
        # promo CASE/EXISTS in GROUP BY.  month_expr is whichever
        # alias matches the chosen FROM source.
        if is_promo_group:
            group_by_sql = f"GROUP BY {month_expr}, group_label"
        else:
            group_by_sql = f"GROUP BY {month_expr}, {group_col}"
        sql = f"""
        SELECT {month_expr} AS month,
                {label_col} AS group_label,
                SUM(s.{value}) AS value
            {from_sql}
            {' '.join(joins)}
            {where_sql2}
        {group_by_sql}
        ORDER BY {month_expr}
        """
        try:
            cur.execute(sql, tuple(params2))
        except Exception as _e:
            print(f"[monthly_breakdown] group_by={group_by} year={year} SQL failed: {_e}\nSQL:\n{sql}")
            raise
        rows = cur.fetchall()

    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

    return jsonify(rows)


# ----------------------------- Monthly Target ---------------------------------
@app.get("/api/monthly_target")
@cached_endpoint(60)
def monthly_target():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    joins, wh, params = build_target_filters("t", f)
    cat_joins, cat_where = category_target_filters("t", f["category"])
    joins += cat_joins
    wh    += cat_where

    # Apply product_group / pattern / size filters via carrying_26 join.
    carrying_join_mt = "LEFT JOIN carrying_26 mat ON mat.m_code = t.material"
    if (f["product_group"] != "ALL" or f["pattern"] != "ALL" or f["material"] != "ALL" or f["code"] != "ALL"):
        if carrying_join_mt not in joins:
            joins.append(carrying_join_mt)
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s"); params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s"); params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s"); params.append(f["code"])

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        top_sold_to = None

        # 1) Get Top N sold_to from baseline table (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )                                               # CHANGED

            # no matches -> all months zero
            if not top_sold_to:
                return jsonify([{"month": m, "value": 0} for m in range(1, 13)])

        # 2) Monthly target, optionally restricted to baseline Top customers
        wh2     = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"t.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        monthly_sql = f"""
        SELECT t.month AS month_num, SUM(t.{value}) AS monthly_total
            FROM target_26 t
            {' '.join(joins)}
            {where_sql2}
        GROUP BY t.month
        ORDER BY t.month
        """
        cur.execute(monthly_sql, tuple(params2))
        rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    month_map = {int(r["month_num"]): float(r["monthly_total"] or 0) for r in rows}
    return jsonify([{"month": m, "value": month_map.get(m, 0)} for m in range(1, 13)])


# ------------------------- Monthly Target Breakdown --------------------------
from mysql.connector import Error as MySQLError

@app.get("/api/admin/target_26_diag")
def admin_target_26_diag():
    """Diagnostic: for a given state, return distinct bde values and
    per-bde qty totals so we can see whether target_26 has real
    per-salesman rows or just a state-level lump.

    /api/admin/target_26_diag?state=VIC
    """
    state = (request.args.get("state") or "VIC").strip().upper()
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT COALESCE(NULLIF(TRIM(bde), ''), '(empty)') AS bde, "
            "       COUNT(*) AS rows, "
            "       SUM(qty) AS qty, "
            "       COUNT(DISTINCT ship_to) AS n_ship_to "
            "FROM target_26 "
            "WHERE state = %s "
            "GROUP BY COALESCE(NULLIF(TRIM(bde), ''), '(empty)') "
            "ORDER BY qty DESC",
            (state,),
        )
        rows = cur.fetchall()
        return jsonify({"state": state, "bde_breakdown": rows})
    finally:
        cur.close(); conn.close()


@app.get("/api/monthly_target_breakdown")
@cached_endpoint(60)
def monthly_target_breakdown():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    top_limit = int(request.args.get("top_limit", 0) or 0)

    group_by = (request.args.get("group_by") or "region").strip()
    # For sold_to: GROUP BY the code (stable, matches sales) but SELECT
    # the name (readable in the legend).  Other dimensions group + label
    # on the same column.  Channel + sold_to_group live on customer
    # master rather than target_26, so we route them through the
    # per-sold_to tcus subquery joined below.
    group_cols = {
        "line":          "mat.line",
        "brand":         "mat.brand",
        "product_group": "mat.product_group",
        "region":        "t.state",
        "salesman":      "t.bde",
        "channel":       "tcus.channels",
        "sold_to_group": "tcus.sold_to_group",
        "sold_to":       "t.sold_to",
        "pattern":       "mat.pattern",
    }
    # Promotion grouping doesn't apply to target — target_26 doesn't
    # carry promo membership, so the breakdown can't bucket by it.  The
    # frontend already skips the target fetch when grouped by Promotion
    # (via _skipSide), so we never expect "promotion" here.
    if group_by not in group_cols:
        return jsonify({"error": "invalid group_by"}), 400
    group_col = group_cols[group_by]
    # label_col: what shows up in the legend; for sold_to we resolve to name.
    if group_by == "sold_to":
        label_col = "MIN(COALESCE(tcus.sold_to_name, t.sold_to))"
    else:
        label_col = group_col

    joins, wh, params = build_target_filters("t", f)
    cat_joins, cat_where = category_target_filters("t", f["category"])
    joins += cat_joins
    wh    += cat_where

    # tcus = per-sold_to customer-master roll-up.  Exposes sold_to_name
    # (for the sold_to-by-name legend), sold_to_group, and channels so
    # the three customer-master groupings stay consistent with what the
    # sales endpoints produce.  NULLIF/TRIM mirrors _customer_join so
    # blank-string rows don't outrank a real value via MIN.
    if group_by in ("sold_to", "channel", "sold_to_group"):
        joins.append(
            "LEFT JOIN ("
            "  SELECT sold_to,"
            "         MIN(NULLIF(TRIM(sold_to_name),'' ))   AS sold_to_name,"
            "         MIN(NULLIF(TRIM(sold_to_group),'' ))  AS sold_to_group,"
            "         MIN(NULLIF(TRIM(channels),''))        AS channels"
            "  FROM customer GROUP BY sold_to"
            ") tcus ON tcus.sold_to = t.sold_to"
        )
    # carrying_26 join needed for group_by or filter on product_group/pattern/line
    carrying_join = "LEFT JOIN carrying_26 mat ON mat.m_code = t.material"
    needs_carrying = group_by in ("line", "brand", "product_group", "pattern")

    if f["product_group"] != "ALL":
        needs_carrying = True
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        needs_carrying = True
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f["pattern"] != "ALL":
        needs_carrying = True
        wh.append("mat.pattern = %s");       params.append(f["pattern"])
    if f["material"] != "ALL":
        needs_carrying = True
        wh.append("mat.size = %s");          params.append(f["material"])
    if f["code"] != "ALL":
        needs_carrying = True
        wh.append("mat.m_code = %s");        params.append(f["code"])

    if needs_carrying and carrying_join not in joins:
        joins.append(carrying_join)
    # Force the Product top-level cascade to only show PCLT + TBR
    if group_by in ("line", "brand"):
        # Restrict the Product cascade's top two levels to PCLT + TBR
        # (the two carrying lines that make sense as a product split).
        # HM / HK-only / LF-only rows for other lines are intentionally
        # hidden because HM reads as a customer-side pivot, not a
        # product one.  Brand level (HK vs LF) inherits the same gate.
        wh.append("mat.line IN ('PCLT','TBR')")

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)

    try:
        top_sold_to = None
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value)
            if not top_sold_to:
                return jsonify([])

        wh2     = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"t.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        sql = f"""
            SELECT
                t.month AS month,
                {label_col} AS group_label,
                SUM(t.{value}) AS value
            FROM target_26 t
            {' '.join(joins)}
            {where_sql2}
            GROUP BY t.month, {group_col}
            ORDER BY t.month
        """

        cur.execute(sql, tuple(params2))
        return jsonify(cur.fetchall())

    except MySQLError as e:
        # ?꾨줎?몄뿉??諛붾줈 ?먯씤 蹂댁씠?꾨줉 ?대젮以?(?댁쁺?대㈃ msg留??쒓굅?섍퀬 濡쒓렇濡쒕쭔 ?④린湲?
        return jsonify({
            "error": "mysql_error",
            "msg": str(e),
            "group_by": group_by,
        }), 400

    except Exception as e:
        return jsonify({
            "error": "server_error",
            "msg": str(e),
            "group_by": group_by,
        }), 500

    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass
# ----------------------------- Yearly Sales ---------------------------------
@app.get("/api/yearly_sales")
@cached_endpoint(60)
def yearly_sales():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)

    if f["category"] != "443":
        cat_joins, cat_where = category_filters_sales("s", f["category"])
        joins += cat_joins
        wh    += cat_where

    if f["product_group"] != "ALL" or f["pattern"] != "ALL" or f["material"] != "ALL" or f["code"] != "ALL":
        _ensure_carrying_join("s", joins)
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s")
        params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s")
        params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s")
        params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s")
        params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s")
        params.append(f["code"])
    base_where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        top_sold_to = None

        # 1) If top_limit > 0, get top N sold_to from baseline (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )

            if not top_sold_to:
                # no data ??return zeros for all years in range
                return jsonify([{"year": y, "value": 0} for y in range(2021, 2026)])

        # 2) Yearly totals, optionally restricted to those sold_to
        wh2 = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        yearly_sql = f"""
        SELECT s.year AS year_num, SUM(s.{value}) AS yearly_total
            FROM sales_21_25 s
            {' '.join(joins)}
            {where_sql2}
        GROUP BY s.year
        ORDER BY s.year
        """
        cur.execute(yearly_sql, tuple(params2))
        rows = cur.fetchall()

    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass

    year_map = {int(r["year_num"]): float(r["yearly_total"] or 0) for r in rows}
    return jsonify([{"year": y, "value": year_map.get(y, 0)} for y in range(2021, 2026)])


# -------------------- Yearly breakdown (stacked by group) -------------------
@app.get("/api/yearly_breakdown")
@cached_endpoint(60)
def yearly_breakdown():

    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    # 0 or missing = no top filter
    top_limit = int(request.args.get("top_limit", 0) or 0)

    # Which dimension to group by?
    group_by = (request.args.get("group_by") or "region").strip()
    # For sold_to: group by code, label as name (consistent with the
    # other breakdown endpoints).
    group_cols = {
        "line":          "mat.line",
        "brand":         "mat.brand",
        "product_group": "mat.product_group",
        "region":        "cus.bde_state",
        "salesman":      "cus.salesman_name",
        "channel":       "cus.channels",
        "sold_to_group": "cus.sold_to_group",
        "sold_to":       "s.sold_to",
        "pattern":       "mat.pattern",
    }
    is_promo_group = group_by in ("promotion", "promotion_detail")
    if not is_promo_group and group_by not in group_cols:
        return jsonify({"error": "invalid group_by"}), 400
    if is_promo_group:
        # sales_21_25 carries year + month, so the promo helper works as-is.
        # promo_plan now covers 2025 + 2026; pre-2025 years naturally
        # bucket under 'Non-Promotion' because no rule's period matches.
        group_col = _promotion_group_col_sql(
            detail=(group_by == "promotion_detail"),
        )
        label_col = group_col
    elif group_by == "sold_to":
        group_col = group_cols[group_by]
        label_col = "MIN(COALESCE(scus.sold_to_name, s.sold_to))"
    else:
        group_col = group_cols[group_by]
        label_col = f"COALESCE(NULLIF(TRIM({group_col}),''), 'COMMON')"

    # ---- Build base JOINs / WHEREs (same pattern as yearly_sales) ----
    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)
    if f["category"] != "443":
        cat_joins, cat_where = category_filters_sales("s", f["category"])
        joins += cat_joins
        wh    += cat_where
    if group_by == "sold_to":
        joins.append(
            "LEFT JOIN ("
            "  SELECT sold_to, MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name "
            "  FROM customer GROUP BY sold_to"
            ") scus ON scus.sold_to = s.sold_to"
        )

    if (group_by in ("line", "brand", "product_group", "pattern") or
        is_promo_group or
        f["product_group"] != "ALL" or f["pattern"] != "ALL" or
        f["material"] != "ALL" or f["code"] != "ALL"):
        _ensure_carrying_join("s", joins)
    if group_by in ("region", "salesman", "channel", "sold_to_group", "sold_to") or is_promo_group:
        _ensure_customer_join("s", joins)
    # Force the Product top-level cascade to only show PCLT + TBR.
    if group_by in ("line", "brand"):
        # Restrict the Product cascade's top two levels to PCLT + TBR
        # (the two carrying lines that make sense as a product split).
        # HM / HK-only / LF-only rows for other lines are intentionally
        # hidden because HM reads as a customer-side pivot, not a
        # product one.  Brand level (HK vs LF) inherits the same gate.
        wh.append("mat.line IN ('PCLT','TBR')")
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s"); params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s"); params.append(f["brand"])
    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s");       params.append(f["pattern"])
    if f["material"] != "ALL":
        wh.append("mat.size = %s")
        params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s")
        params.append(f["code"])
    base_where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        top_sold_to = None

        # 1) If top_limit > 0, get top N sold_to from baseline (sales_2526)
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(
                cur, f, top_limit, value
            )

            # no data ??nothing to show
            if not top_sold_to:
                return jsonify([])

        # 2) Yearly breakdown, restricted to those top customers,
        #    but stacked by group_col
        wh2     = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        # Same alias trick as the other breakdown endpoints — avoid
        # restating the promo CASE/EXISTS in GROUP BY.
        if is_promo_group:
            group_by_sql = "GROUP BY s.year, group_label"
        else:
            group_by_sql = f"GROUP BY s.year, {group_col}"
        sql = f"""
        SELECT s.year AS year,
                {label_col} AS group_label,
                SUM(s.{value}) AS value
            FROM sales_21_25 s
            {' '.join(joins)}
            {where_sql2}
        {group_by_sql}
        ORDER BY s.year
        """
        try:
            cur.execute(sql, tuple(params2))
        except Exception as _e:
            print(f"[yearly_breakdown] group_by={group_by} SQL failed: {_e}\nSQL:\n{sql}")
            raise
        rows = cur.fetchall()

    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

    return jsonify(rows)


# ---------------------- lookups used by the UI (optional) --------------------
@app.get("/api/sold_to_groups")
def sold_to_groups():
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT TRIM(sold_to_group)
            FROM customer
            WHERE sold_to_group IS NOT NULL AND TRIM(sold_to_group) <> ''
            ORDER BY TRIM(sold_to_group)
        """)
        groups = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(groups)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.get("/api/channels")
@cached_endpoint(300)
def api_channels():
    """Distinct customer.channels values for the Channel filter dropdown."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT DISTINCT TRIM(channels) AS v
            FROM customer
            WHERE channels IS NOT NULL AND TRIM(channels) <> ''
            ORDER BY TRIM(channels)
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([r["v"] for r in rows])
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.get("/api/sold_to_names")
def sold_to_names():
    parent = (request.args.get("sold_to_group") or "ALL").strip()
    top_limit = int(request.args.get("top_limit", 0) or 0)

    # Use your existing filter parser (metric/category/region/salesman/... etc.)
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # ----------------- 1) No top_limit -----------------
        # Only show sold_tos that have actual sales (avoids showing every
        # inactive account in the customer master ??far too many entries).
        if top_limit <= 0:
            params_s: list = []
            extra_wh = (
                "AND c.sold_to_group = %s" if parent != "ALL" else ""
            )
            if parent != "ALL":
                params_s.append(parent)
            cur.execute(f"""
                SELECT DISTINCT TRIM(c.sold_to_name) AS name
                  FROM customer c
                 WHERE c.sold_to IN (
                       SELECT DISTINCT sold_to FROM sales_2526
                        WHERE sold_to IS NOT NULL
                 )
                   {extra_wh}
                   AND c.sold_to_name IS NOT NULL
                   AND TRIM(c.sold_to_name) <> ''
                 ORDER BY TRIM(c.sold_to_name)
            """, params_s)
            rows = cur.fetchall()
            return jsonify([r["name"] for r in rows])

        # ----------------- 2) top_limit -> baseline top sold_to -> names -----------------
        # Get top sold_to list from baseline table (sales_2526)
        top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value)

        if not top_sold_to:
            return jsonify([])

        placeholders = ",".join(["%s"] * len(top_sold_to))

        # Filter names by sold_to_group if provided
        params = list(top_sold_to)
        where_parent = ""
        if parent != "ALL":
            where_parent = " AND c.sold_to_group = %s "
            params.append(parent)

        cur.execute(f"""
            SELECT DISTINCT TRIM(c.sold_to_name) AS name
              FROM customer c
             WHERE c.sold_to IN ({placeholders})
               {where_parent}
               AND c.sold_to_name IS NOT NULL
               AND TRIM(c.sold_to_name) <> ''
             ORDER BY TRIM(c.sold_to_name)
        """, tuple(params))

        rows = cur.fetchall()
        return jsonify([r["name"] for r in rows])

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

@app.get("/api/ship_to_names")
def ship_to_names():
    # parent (big group)
    stg3    = (request.args.get("sold_to_group") or "ALL").strip()
    # child (sold-to name that user picked)
    sold_to = (request.args.get("sold_to") or "ALL").strip()
    # Top N filter (same baseline as the rest of the dashboard) — when set,
    # restricts the ship_to list to those belonging to the top sold_to set.
    top_limit = int(request.args.get("top_limit", 0) or 0)

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        # Resolve top sold_to set first (if a top filter is active)
        top_sold_to = None
        if top_limit > 0:
            f = parse_filters(request)
            value = "qty" if f["metric"] == "qty" else "amt"
            top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value)
            if not top_sold_to:
                cur.close(); conn.close()
                return jsonify([])

        where = ["ship_to_name IS NOT NULL", "TRIM(ship_to_name) <> ''"]
        params = []

        if sold_to.upper() != "ALL":
            where.append("TRIM(sold_to_name) = %s")
            params.append(sold_to)
        elif stg3.upper() != "ALL":
            where.append("TRIM(sold_to_group) = %s")
            params.append(stg3)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            where.append(f"sold_to IN ({placeholders})")
            params.extend(top_sold_to)

        where_sql = "WHERE " + " AND ".join(where)

        cur.execute(f"""
            SELECT DISTINCT ship_to, TRIM(ship_to_name) AS ship_to_name
            FROM customer
            {where_sql}
            ORDER BY TRIM(ship_to_name)
        """, tuple(params))

        rows_out = [{"code": str(r["ship_to"]), "name": r["ship_to_name"]} for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(rows_out)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/api/product_group")
def product_group():
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT TRIM(product_group)
            FROM carrying_26
            WHERE product_group IS NOT NULL AND TRIM(product_group) <> ''
            ORDER BY TRIM(product_group)
        """)
        groups = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(groups)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.get("/api/patterns")
def patterns():
    product_group = request.args.get("product_group", "ALL")
    try:
        conn = get_connection(); cur = conn.cursor()
        if product_group and product_group != "ALL":
            cur.execute("""
                SELECT DISTINCT TRIM(pattern)
                FROM carrying_26
                WHERE product_group = %s
                  AND pattern IS NOT NULL AND TRIM(pattern) <> ''
                ORDER BY TRIM(pattern)
            """, (product_group,))
        else:
            cur.execute("""
                SELECT DISTINCT TRIM(pattern)
                FROM carrying_26
                WHERE pattern IS NOT NULL AND TRIM(pattern) <> ''
                ORDER BY TRIM(pattern)
            """)
        names = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(names)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/api/carrying_matrix")
def carrying_matrix():
    """Distinct (product_group, pattern, size) tuples from carrying_26.
    Used by the claim form so three independent dropdowns can cascade
    in any direction client-side without a round-trip per change."""
    return _carrying_matrix_response()

@app.get("/api/claim/product_matrix")
def claim_product_matrix():
    """Public mirror of /api/carrying_matrix that lives under /api/claim/*
    so it falls inside the Cloudflare Access bypass for the customer
    claim form (which has no Cf-Access-Authenticated-User-Email header)."""
    return _carrying_matrix_response()

def _carrying_matrix_response():
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT
                   TRIM(product_group) AS pg,
                   TRIM(pattern)       AS pat,
                   TRIM(size)          AS sz
            FROM carrying_26
            WHERE product_group IS NOT NULL AND TRIM(product_group) <> ''
              AND pattern       IS NOT NULL AND TRIM(pattern)       <> ''
              AND size          IS NOT NULL AND TRIM(size)          <> ''
            ORDER BY pg, pat, sz
        """)
        rows = [{"pg": r[0], "pattern": r[1], "size": r[2]} for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/api/codes")
def api_codes():
    """Return distinct carrying_26.m_code values, optionally narrowed by
    product_group / pattern / material (same shape as /api/materials).
    The Code dropdown on the Graph and Map views uses this to populate
    its picker.
    """
    product_group = (request.args.get("product_group") or "ALL").strip()
    pattern       = (request.args.get("pattern")       or "").strip()
    material      = (request.args.get("material")      or "").strip()

    where  = []
    params = []
    if product_group and product_group != "ALL":
        where.append("product_group = %s"); params.append(product_group)
    if pattern:
        where.append("pattern LIKE %s"); params.append(f"%{pattern}%")
    if material:
        where.append("size = %s"); params.append(material)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute(f"""
            SELECT DISTINCT m_code
            FROM carrying_26
            {where_sql}
            ORDER BY m_code
        """, tuple(params))
        codes = [str(r[0]) for r in cur.fetchall() if r[0] is not None]
        cur.close(); conn.close()
        return jsonify(codes)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.get("/api/materials")
def materials():
    """
    Return distinct Material list.
    - ?꾪꽣: product_group, pattern (????'ALL' ?대㈃ ?꾩껜)
    """
    product_group = (request.args.get("product_group") or "ALL").strip()
    pattern       = (request.args.get("pattern")       or "ALL").strip()

    try:
        conn = get_connection(); cur = conn.cursor()

        where = []
        params = []

        if product_group and product_group != "ALL":
            where.append("product_group = %s")
            params.append(product_group)

        if pattern and pattern != "ALL":
            where.append("pattern = %s")
            params.append(pattern)

        where_sql = ""
        if where:
            where_sql = "WHERE " + " AND ".join(where)

        cur.execute(f"""
            SELECT DISTINCT size
            FROM carrying_26
            {where_sql}
            ORDER BY size
        """, tuple(params))

        names = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(names)

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/carrying_price")
def carrying_price():
    """Return avg list_price and purchase_price from carrying_26 for current filter."""
    product_group = (request.args.get("product_group") or "ALL").strip()
    pattern       = (request.args.get("pattern")       or "").strip()
    material      = (request.args.get("material")      or "").strip()

    try:
        conn = get_connection()
        cur  = conn.cursor(dictionary=True)

        where  = []
        params = []

        if product_group and product_group != "ALL":
            where.append("product_group = %s")
            params.append(product_group)
        if pattern:
            where.append("pattern LIKE %s")
            params.append(f"%{pattern}%")
        if material:
            where.append("size LIKE %s")
            params.append(f"%{material}%")

        where_sql = ("WHERE " + " AND ".join(where)) if where else ""

        cur.execute(f"""
            SELECT AVG(NULLIF(list_price, 0)) AS list_price,
                   AVG(NULLIF(purchase_price, 0)) AS purchase_price
            FROM carrying_26
            {where_sql}
        """, tuple(params))

        row = cur.fetchone()
        cur.close(); conn.close()

        return jsonify({
            "list_price":      float(row["list_price"])      if row and row["list_price"]      is not None else None,
            "purchase_price":  float(row["purchase_price"])  if row and row["purchase_price"]  is not None else None,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/api/profit_monthly")
def profit_monthly():
    import traceback

    try:
        f = parse_filters(request)
        value = "qty" if f.get("metric") == "qty" else "amt"
        top_limit = int(request.args.get("top_limit", 0) or 0)

        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            top_sold_to = None

            # 1) Top-N sold_to from baseline sales table
            if top_limit > 0:
                top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value)
                if not top_sold_to:
                    return jsonify([
                        dict(month=m, gross=0, sd=0, cogs=0, op_cost=0)
                        for m in range(1, 13)
                    ])

            # 2) Build filters for the new `profit` table.
            #    profit has: month, sold_to, material, gross, sales_deduction, cogs, operating_cost, profit
            #    NO ship_to column ??ship_to filter resolves to sold_to via customer subquery.
            joins_p  = []
            wh_p     = []
            params_p = []

            # ?? region filter (using EXISTS to avoid JOIN inflation) ??
            if f["region"] != "ALL":
                states = REGION_STATES.get(f["region"].upper(), [f["region"]])
                ph = ",".join(["%s"] * len(states))
                wh_p.append(
                    f"EXISTS (SELECT 1 FROM customer c2 WHERE c2.sold_to = p.sold_to"
                    f" AND c2.bde_state IN ({ph}))"
                )
                params_p.extend(states)

            # ?? salesman filter ??
            if f["salesman"] != "ALL":
                wh_p.append(
                    "EXISTS (SELECT 1 FROM customer c2 WHERE c2.sold_to = p.sold_to"
                    " AND UPPER(TRIM(c2.salesman_name)) = UPPER(TRIM(%s)))"
                )
                params_p.append(f["salesman"])

            # ?? sold_to_group filter ??
            if f["sold_to_group"] != "ALL":
                wh_p.append(
                    "EXISTS (SELECT 1 FROM customer c2 WHERE c2.sold_to = p.sold_to"
                    " AND c2.sold_to_group = %s)"
                )
                params_p.append(f["sold_to_group"])

            # ?? sold_to filter (directly on p.sold_to) ??
            if f["sold_to"] != "ALL":
                sv = f["sold_to"]
                if sv.isdigit() or sv.upper().startswith("A"):
                    wh_p.append("p.sold_to = %s"); params_p.append(sv)
                else:
                    wh_p.append(
                        "p.sold_to IN (SELECT DISTINCT sold_to FROM customer WHERE sold_to_name = %s)"
                    )
                    params_p.append(sv)

            # ?? ship_to filter: no ship_to in profit ??resolve to its sold_to ??
            if f["ship_to"] != "ALL":
                st = f["ship_to"].strip()
                if st.isdigit() or st.upper().startswith("A"):
                    wh_p.append(
                        "p.sold_to IN (SELECT DISTINCT sold_to FROM customer WHERE ship_to = %s)"
                    )
                    params_p.append(st)
                else:
                    wh_p.append(
                        "p.sold_to IN (SELECT DISTINCT sold_to FROM customer"
                        " WHERE UPPER(TRIM(ship_to_name)) = UPPER(TRIM(%s)))"
                    )
                    params_p.append(st)

            # ?? category filter (via carrying_26, same as sales tables) ??
            cat_joins_p, cat_where_p = category_filters_sales("p", f.get("category", "ALL"))
            joins_p += cat_joins_p
            wh_p    += cat_where_p

            # ?? product_group / pattern filter (via carrying_26) ??
            if f.get("product_group", "ALL") != "ALL" or f.get("pattern", "ALL") != "ALL" or f.get("material", "ALL") != "ALL":
                _ensure_carrying_join("p", joins_p)
            if f.get("product_group", "ALL") != "ALL":
                wh_p.append("mat.product_group = %s"); params_p.append(f["product_group"])
            if f["brand"] != "ALL":
                wh_p.append("mat.brand = %s"); params_p.append(f["brand"])
            if f.get("pattern", "ALL") != "ALL":
                wh_p.append("mat.pattern = %s"); params_p.append(f["pattern"])
            if f["material"] != "ALL":
                wh_p.append("mat.size = %s"); params_p.append(f["material"])
            if f["code"] != "ALL":
                wh_p.append("mat.m_code = %s"); params_p.append(f["code"])

            # ?? top sold_to restriction ??
            if top_sold_to:
                placeholders = ",".join(["%s"] * len(top_sold_to))
                wh_p.append(f"p.sold_to IN ({placeholders})")
                params_p.extend(top_sold_to)

            where_sql2 = ("WHERE " + " AND ".join(wh_p)) if wh_p else ""
            monthly_sql = f"""
                SELECT CAST(p.month AS UNSIGNED) AS month,
                       SUM(p.gross)           AS gross,
                       SUM(p.sales_deduction) AS sd,
                       SUM(p.cogs)            AS cogs,
                       SUM(p.operating_cost)  AS op_cost
                  FROM profit p
                  {' '.join(joins_p)}
                  {where_sql2}
                 GROUP BY CAST(p.month AS UNSIGNED)
                 ORDER BY CAST(p.month AS UNSIGNED)
            """
            cur.execute(monthly_sql, tuple(params_p))
            rows = cur.fetchall()

        finally:
            cur.close()
            conn.close()

        out = [dict(month=m, gross=0, sd=0, cogs=0, op_cost=0) for m in range(1, 13)]
        for r in rows:
            m = int(r["month"] or 0)
            if 1 <= m <= 12:
                out[m - 1].update(
                    gross=float(r["gross"] or 0),
                    sd=float(r["sd"] or 0),
                    cogs=float(r["cogs"] or 0),
                    op_cost=float(r["op_cost"] or 0),
                )

        return jsonify(out)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/sales_map")
def sales_map():
    f = parse_filters(request)
    value = "qty" if f["metric"] == "qty" else "amt"

    top_limit = int(request.args.get("top_limit", 0) or 0)

    # ---- same base filter pattern as daily_sales ----
    joins, wh, params = build_customer_filters("s", f, use_sold_to_name=False)

    # join customer for lat/lng
    joins.append("""
        JOIN customer c ON c.ship_to = s.ship_to
    """)

    # category ??normalised version
    if f["category"] != "443":
        cat_joins, cat_where = category_filters_sales("s", f["category"])
        joins += cat_joins
        wh    += cat_where

    if (f["product_group"] != "ALL" or f["pattern"] != "ALL" or
        f["material"] != "ALL" or f["code"] != "ALL"):
        _ensure_carrying_join("s", joins)
    if f["product_group"] != "ALL":
        wh.append("mat.product_group = %s")
        params.append(f["product_group"])
    if f["brand"] != "ALL":
        wh.append("mat.brand = %s")
        params.append(f["brand"])

    if f["pattern"] != "ALL":
        wh.append("mat.pattern = %s")
        params.append(f["pattern"])

    if f["material"] != "ALL":
        wh.append("mat.size = %s")
        params.append(f["material"])
    if f["code"] != "ALL":
        wh.append("mat.m_code = %s")
        params.append(f["code"])

    # only customers with coordinates
    wh.append("c.latitude IS NOT NULL")
    wh.append("c.longitude IS NOT NULL")

    # 2026 cumulative only — filter on billing_date (the base column
    # that ships with sales_2526; no s.year on this schema).  Range
    # form so the billing_date index prunes instead of scanning.
    wh.append("s.billing_date >= %s AND s.billing_date < %s")
    params.extend(["2026-01-01", "2027-01-01"])

    base_where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)

    try:
        top_sold_to = None

        # ---- Top N logic (same baseline as other APIs) ----
        if top_limit > 0:
            top_sold_to = get_top_sold_to_from_baseline(cur, f, top_limit, value)
            if not top_sold_to:
                return jsonify([])

        # ---- apply top filter ----
        wh2     = list(wh)
        params2 = list(params)

        if top_sold_to:
            placeholders = ",".join(["%s"] * len(top_sold_to))
            wh2.append(f"s.sold_to IN ({placeholders})")
            params2.extend(top_sold_to)

        where_sql2 = ("WHERE " + " AND ".join(wh2)) if wh2 else ""

        # ---- Map totals ----
        map_sql = f"""
            SELECT
                c.ship_to,
                c.ship_to_name,
                c.latitude,
                c.longitude,
                MAX(c.bde_state)      AS region,
                MAX(c.salesman_name)  AS bde,
                SUM(s.{value})        AS total_value,
                SUM(CASE WHEN YEAR(s.billing_date) = 2026 THEN s.{value} ELSE 0 END) AS total_2026
            FROM {_sales_2526_from("s")}
            {' '.join(joins)}
            {where_sql2}
            GROUP BY
                c.ship_to,
                c.ship_to_name,
                c.latitude,
                c.longitude
            ORDER BY total_2026 DESC
        """

        cur.execute(map_sql, tuple(params2))
        rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    return jsonify(rows)

# ── GPS Visit count API ────────────────────────────────────────────────────────

# Candidate column names — the real schema is whatever the gps table actually
# uses; we pick the first one that exists so the endpoint isn't brittle.
_GPS_DATE_CANDIDATES = (
    "local_date", "date", "visit_date", "gps_date",
    "recorded_at", "created_at", "timestamp", "ts", "datetime",
)
_GPS_LAT_CANDIDATES = ("latitude", "lat", "y")
_GPS_LNG_CANDIDATES = ("longitude", "lng", "lon", "long", "x")
_GPS_SALESMAN_CANDIDATES = ("salesmen", "salesman", "salesman_name", "rep", "bde", "registration")

# BDE names to hide from the visit_summary by_bde table (e.g. test
# entries, old vehicles, names that aren't actual sales reps).
# Match is case/whitespace-insensitive.
BDE_EXCLUDE = frozenset({
    "JO TEDDY",
    "CHO JUNJONG",
})

def _resolve_gps_salesman_col(cur):
    """Optional column — returns the salesman/registration column name or
    None if the gps schema doesn't include one."""
    if USE_SQLITE:
        cur.execute("PRAGMA table_info(gps)")
        cols = [r["name"] if isinstance(r, dict) else r[1] for r in cur.fetchall()]
    else:
        cur.execute("SHOW COLUMNS FROM gps")
        cols = [r["Field"] for r in cur.fetchall()]
    cols_lc = {c.lower(): c for c in cols}
    for c in _GPS_SALESMAN_CANDIDATES:
        if c in cols_lc:
            return cols_lc[c]
    return None

# Australian national public holidays — used to exclude non-business days
# from visit counts.  State-specific holidays (Labour Day, Show Day, etc.)
# are intentionally not included to keep the rule consistent across
# territories. Add years here as needed.
AU_HOLIDAYS = frozenset({
    # 2025
    "2025-01-01", "2025-01-27", "2025-04-18", "2025-04-21",
    "2025-04-25", "2025-06-09", "2025-12-25", "2025-12-26",
    # 2026
    "2026-01-01", "2026-01-26", "2026-04-03", "2026-04-06",
    "2026-04-25", "2026-06-08", "2026-12-25", "2026-12-28",
    # 2027
    "2027-01-01", "2027-01-26", "2027-03-26", "2027-03-29",
    "2027-04-26", "2027-06-14", "2027-12-27", "2027-12-28",
})

def _business_day_filter_sql(date_col):
    """SQL fragment + params to keep only Mon–Fri non-holiday rows.
    MySQL DAYOFWEEK: 1=Sun, 7=Sat."""
    placeholders = ",".join(["%s"] * len(AU_HOLIDAYS))
    sql = (f"DAYOFWEEK({date_col}) NOT IN (1, 7) "
           f"AND DATE({date_col}) NOT IN ({placeholders})")
    return sql, list(AU_HOLIDAYS)

def _is_business_day(d):
    """d: date or datetime. True if Mon–Fri and not in AU_HOLIDAYS."""
    if d is None:
        return False
    if d.weekday() >= 5:  # 5=Sat, 6=Sun
        return False
    return d.strftime("%Y-%m-%d") not in AU_HOLIDAYS

def _resolve_gps_columns(cur):
    """Return (date_col, lat_col, lng_col) by inspecting the gps table.
    Raises RuntimeError if the table or required columns are missing."""
    if USE_SQLITE:
        cur.execute("PRAGMA table_info(gps)")
        cols = [r["name"] if isinstance(r, dict) else r[1] for r in cur.fetchall()]
    else:
        cur.execute("SHOW COLUMNS FROM gps")
        cols = [r["Field"] for r in cur.fetchall()]
    if not cols:
        raise RuntimeError("gps table has no columns or does not exist")
    cols_lc = {c.lower(): c for c in cols}
    def pick(cands, label):
        for c in cands:
            if c in cols_lc:
                return cols_lc[c]
        raise RuntimeError(f"no {label} column found in gps; tried {cands}; have {cols}")
    return pick(_GPS_DATE_CANDIDATES, "date"), pick(_GPS_LAT_CANDIDATES, "lat"), pick(_GPS_LNG_CANDIDATES, "lng")


@app.get("/api/monthly_visits")
def monthly_visits():
    """
    Count GPS visits within 300m of a given lat/lng per month.
    A 'visit' = at least one GPS record on a calendar day within the radius.
    Params: lat, lng, year, [radius=300]
    Returns: [{"m": 1, "visits": 3}, ...]   (m = month number 1-12)
    """
    try:
        lat = float(request.args.get("lat", ""))
        lng = float(request.args.get("lng", ""))
    except (TypeError, ValueError):
        return jsonify([])

    try:
        year = int(request.args.get("year", 2026) or 2026)
    except (TypeError, ValueError):
        year = 2026

    try:
        radius_m = float(request.args.get("radius", 300) or 300)
    except (TypeError, ValueError):
        radius_m = 300.0
    business_only = (request.args.get("business_days_only", "1").strip().lower()
                     not in ("0", "false", "no"))

    # bbox pre-filter sized from the radius (~111 km per degree lat,
    # ~96 km per degree lng at 30°S)
    LAT_D = (radius_m / 111000.0) * 1.2
    LNG_D = (radius_m / 96000.0)  * 1.2

    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
    except Exception as e:
        print("monthly_visits: DB connect failed:", e)
        return jsonify([])

    try:
        date_col, lat_col, lng_col = _resolve_gps_columns(cur)
        if USE_SQLITE:
            sql = f"""
                SELECT CAST(strftime('%m', {date_col}) AS INTEGER) AS m,
                       COUNT(DISTINCT date({date_col}))            AS visits
                FROM   gps
                WHERE  strftime('%Y', {date_col}) = ?
                  AND  {lat_col} BETWEEN ? AND ?
                  AND  {lng_col} BETWEEN ? AND ?
                GROUP  BY m
                ORDER  BY m
            """
            cur.execute(sql, [str(year),
                              lat - LAT_D, lat + LAT_D,
                              lng - LNG_D, lng + LNG_D])
        else:
            extra_sql = ""
            extra_params = []
            if business_only:
                bd_sql, bd_params = _business_day_filter_sql(date_col)
                extra_sql = f"  AND {bd_sql}\n"
                extra_params = bd_params
            sql = f"""
                SELECT MONTH({date_col})                  AS m,
                       COUNT(DISTINCT DATE({date_col}))   AS visits
                FROM   gps
                WHERE  YEAR({date_col}) = %s
                  AND  {lat_col} BETWEEN %s AND %s
                  AND  {lng_col} BETWEEN %s AND %s
                  AND  (6371000 * ACOS(LEAST(1.0,
                           COS(RADIANS(%s)) * COS(RADIANS({lat_col}))
                           * COS(RADIANS({lng_col}) - RADIANS(%s))
                           + SIN(RADIANS(%s)) * SIN(RADIANS({lat_col}))
                       ))) <= %s
                {extra_sql}
                GROUP  BY MONTH({date_col})
                ORDER  BY m
            """
            cur.execute(sql, [year,
                              lat - LAT_D, lat + LAT_D,
                              lng - LNG_D, lng + LNG_D,
                              lat, lng, lat, radius_m,
                              *extra_params])
        rows = cur.fetchall() or []
    except Exception as e:
        print("monthly_visits: query failed:", e)
        rows = []
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass

    return jsonify([{"m": r["m"], "visits": r["visits"]} for r in rows])


@app.get("/api/visit_for_shop")
def visit_for_shop():
    """Convenience: look up a shop's lat/lng by ship_to and return its
    monthly visit counts.  Saves you from copying coordinates manually.
    Params: ship_to (required), [year=2026], [radius=300]
    """
    ship_to = (request.args.get("ship_to") or "").strip()
    if not ship_to:
        return jsonify({"error": "ship_to required"})
    try:
        year = int(request.args.get("year", 2026) or 2026)
    except (TypeError, ValueError):
        year = 2026
    try:
        radius_m = float(request.args.get("radius", 300) or 300)
    except (TypeError, ValueError):
        radius_m = 300.0
    business_only = (request.args.get("business_days_only", "1").strip().lower()
                     not in ("0", "false", "no"))

    out = {"ship_to": ship_to, "year": year, "radius_m": radius_m,
           "business_days_only": business_only}
    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT ship_to_name, latitude, longitude FROM customer "
            "WHERE ship_to = %s LIMIT 1",
            [ship_to],
        )
        row = cur.fetchone()
        if not row:
            return jsonify({**out, "error": "ship_to not found in customer"})
        if row["latitude"] is None or row["longitude"] is None:
            return jsonify({**out, "ship_to_name": row["ship_to_name"],
                            "error": "shop has no lat/lng in customer table"})

        lat = float(row["latitude"])
        lng = float(row["longitude"])
        out["ship_to_name"] = row["ship_to_name"]
        out["shop_lat"] = lat
        out["shop_lng"] = lng

        date_col, lat_col, lng_col = _resolve_gps_columns(cur)
        LAT_D = (radius_m / 111000.0) * 1.2
        LNG_D = (radius_m / 96000.0)  * 1.2

        if USE_SQLITE:
            cur.execute(
                f"SELECT CAST(strftime('%m', {date_col}) AS INTEGER) AS m, "
                f"COUNT(DISTINCT date({date_col})) AS visits FROM gps "
                f"WHERE strftime('%Y', {date_col}) = ? "
                f"AND {lat_col} BETWEEN ? AND ? AND {lng_col} BETWEEN ? AND ? "
                f"GROUP BY m ORDER BY m",
                [str(year), lat-LAT_D, lat+LAT_D, lng-LNG_D, lng+LNG_D],
            )
        else:
            extra_sql = ""
            extra_params = []
            if business_only:
                bd_sql, bd_params = _business_day_filter_sql(date_col)
                extra_sql = f"AND {bd_sql} "
                extra_params = bd_params
            cur.execute(
                f"SELECT MONTH({date_col}) AS m, "
                f"COUNT(DISTINCT DATE({date_col})) AS visits, "
                f"GROUP_CONCAT(DISTINCT registration ORDER BY registration) AS regos "
                f"FROM gps WHERE YEAR({date_col}) = %s "
                f"AND {lat_col} BETWEEN %s AND %s AND {lng_col} BETWEEN %s AND %s "
                f"AND (6371000 * ACOS(LEAST(1.0, "
                f"COS(RADIANS(%s))*COS(RADIANS({lat_col}))"
                f"*COS(RADIANS({lng_col}) - RADIANS(%s)) "
                f"+ SIN(RADIANS(%s))*SIN(RADIANS({lat_col}))))) <= %s "
                f"{extra_sql}"
                f"GROUP BY MONTH({date_col}) ORDER BY m",
                [year, lat-LAT_D, lat+LAT_D, lng-LNG_D, lng+LNG_D,
                 lat, lng, lat, radius_m, *extra_params],
            )
        out["monthly_visits"] = cur.fetchall()
        out["total_visit_days"] = sum(r["visits"] for r in out["monthly_visits"])
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass
    return jsonify(out)


@app.get("/api/visit_debug")
def visit_debug():
    """Diagnose why /api/monthly_visits returns nothing for a shop.
    Params: lat, lng, [year=2026]
    Returns table existence, resolved column names, sample rows, bbox/radius
    match counts at several radii, and per-month visit counts.
    """
    out = {"params": dict(request.args)}
    try:
        lat = float(request.args.get("lat", ""))
        lng = float(request.args.get("lng", ""))
    except (TypeError, ValueError):
        out["error"] = "lat/lng required as floats"
        return jsonify(out)
    try:
        year = int(request.args.get("year", 2026) or 2026)
    except (TypeError, ValueError):
        year = 2026

    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
    except Exception as e:
        return jsonify({**out, "error": f"db connect: {e}"})

    try:
        # Table existence + column list
        if USE_SQLITE:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND lower(name)='gps'")
        else:
            cur.execute("SHOW TABLES LIKE 'gps'")
        out["gps_table_exists"] = bool(cur.fetchall())
        if not out["gps_table_exists"]:
            return jsonify(out)

        if USE_SQLITE:
            cur.execute("PRAGMA table_info(gps)")
            out["columns"] = [r["name"] if isinstance(r, dict) else r[1] for r in cur.fetchall()]
        else:
            cur.execute("SHOW COLUMNS FROM gps")
            out["columns"] = [{"name": r["Field"], "type": r["Type"]} for r in cur.fetchall()]

        try:
            date_col, lat_col, lng_col = _resolve_gps_columns(cur)
            out["resolved"] = {"date": date_col, "lat": lat_col, "lng": lng_col}
        except Exception as e:
            out["error"] = str(e)
            return jsonify(out)

        cur.execute(f"SELECT COUNT(*) AS n FROM gps")
        out["total_rows"] = cur.fetchone()["n"]

        cur.execute(f"SELECT * FROM gps LIMIT 3")
        out["sample"] = cur.fetchall()

        if USE_SQLITE:
            cur.execute(f"SELECT DISTINCT strftime('%Y', {date_col}) AS y FROM gps ORDER BY y")
        else:
            cur.execute(f"SELECT DISTINCT YEAR({date_col}) AS y FROM gps ORDER BY y")
        out["years_in_table"] = [r["y"] for r in cur.fetchall()]

        # Match counts at increasing radii
        match_counts = {}
        for r_m in (100, 250, 500, 1000, 2500, 5000):
            LAT_D = (r_m / 111000.0) * 1.2
            LNG_D = (r_m / 96000.0)  * 1.2
            if USE_SQLITE:
                cur.execute(
                    f"SELECT COUNT(*) AS n FROM gps "
                    f"WHERE strftime('%Y', {date_col}) = ? "
                    f"AND {lat_col} BETWEEN ? AND ? AND {lng_col} BETWEEN ? AND ?",
                    [str(year), lat-LAT_D, lat+LAT_D, lng-LNG_D, lng+LNG_D],
                )
            else:
                cur.execute(
                    f"SELECT COUNT(*) AS n FROM gps "
                    f"WHERE YEAR({date_col}) = %s "
                    f"AND {lat_col} BETWEEN %s AND %s AND {lng_col} BETWEEN %s AND %s "
                    f"AND (6371000 * ACOS(LEAST(1.0, "
                    f"COS(RADIANS(%s))*COS(RADIANS({lat_col}))"
                    f"*COS(RADIANS({lng_col}) - RADIANS(%s)) "
                    f"+ SIN(RADIANS(%s))*SIN(RADIANS({lat_col}))))) <= %s",
                    [year, lat-LAT_D, lat+LAT_D, lng-LNG_D, lng+LNG_D, lat, lng, lat, r_m],
                )
            match_counts[f"{r_m}m"] = cur.fetchone()["n"]
        out["match_counts_by_radius"] = match_counts

        # Geographic spread of the gps table — answers "does the data even
        # cover this region of Australia?"
        if not USE_SQLITE:
            cur.execute(
                f"SELECT MIN({lat_col}) AS lat_min, MAX({lat_col}) AS lat_max, "
                f"MIN({lng_col}) AS lng_min, MAX({lng_col}) AS lng_max FROM gps"
            )
            out["bbox_in_table"] = cur.fetchone()

            # Nearest 5 GPS points to the requested location, any year
            cur.execute(
                f"SELECT {date_col} AS d, {lat_col} AS lat, {lng_col} AS lng, "
                f"  registration, "
                f"  ROUND(6371000 * ACOS(LEAST(1.0, "
                f"    COS(RADIANS(%s))*COS(RADIANS({lat_col}))"
                f"    *COS(RADIANS({lng_col}) - RADIANS(%s)) "
                f"    + SIN(RADIANS(%s))*SIN(RADIANS({lat_col})) "
                f"  )), 0) AS dist_m "
                f"FROM gps "
                f"ORDER BY dist_m ASC LIMIT 5",
                [lat, lng, lat],
            )
            out["nearest_points_any_year"] = cur.fetchall()
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass

    return jsonify(out)


@app.get("/api/visit_debug_topshops")
def visit_debug_topshops():
    """List the top customer ship-tos by GPS visit count for a year, so we can
    confirm the Visit line renders for at least one shop and identify which
    territories actually have GPS coverage.
    Params: year=2026, radius=300, limit=20
    """
    try:
        year = int(request.args.get("year", 2026) or 2026)
    except (TypeError, ValueError):
        year = 2026
    try:
        radius_m = float(request.args.get("radius", 300) or 300)
    except (TypeError, ValueError):
        radius_m = 300.0
    try:
        limit = int(request.args.get("limit", 20) or 20)
    except (TypeError, ValueError):
        limit = 20

    if USE_SQLITE:
        # Skip — need trig, only Haversine works on MySQL here
        return jsonify({"note": "available only on MySQL"})

    out = {"params": {"year": year, "radius_m": radius_m, "limit": limit}}
    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        date_col, lat_col, lng_col = _resolve_gps_columns(cur)

        # For each customer, count distinct visit-days where any GPS sample
        # lies within radius_m of the ship-to location.  Bbox prefilter on
        # both customer and gps cuts the join down dramatically.
        LAT_D = (radius_m / 111000.0) * 1.2
        LNG_D = (radius_m / 96000.0)  * 1.2
        sql = f"""
            SELECT c.ship_to,
                   c.ship_to_name,
                   c.latitude  AS shop_lat,
                   c.longitude AS shop_lng,
                   COUNT(DISTINCT DATE(g.{date_col})) AS visits
            FROM   customer c
            JOIN   gps g
              ON   g.{lat_col} BETWEEN c.latitude  - %s AND c.latitude  + %s
              AND  g.{lng_col} BETWEEN c.longitude - %s AND c.longitude + %s
              AND  YEAR(g.{date_col}) = %s
              AND  (6371000 * ACOS(LEAST(1.0,
                       COS(RADIANS(c.latitude))*COS(RADIANS(g.{lat_col}))
                       *COS(RADIANS(g.{lng_col}) - RADIANS(c.longitude))
                       + SIN(RADIANS(c.latitude))*SIN(RADIANS(g.{lat_col}))
                   ))) <= %s
            WHERE  c.latitude IS NOT NULL AND c.longitude IS NOT NULL
            GROUP  BY c.ship_to, c.ship_to_name, c.latitude, c.longitude
            ORDER  BY visits DESC
            LIMIT  %s
        """
        cur.execute(sql, [LAT_D, LAT_D, LNG_D, LNG_D, year, radius_m, limit])
        out["top_shops"] = cur.fetchall()
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass
    return jsonify(out)


@app.get("/api/visit_debug_closest")
def visit_debug_closest():
    """Find the absolute closest customer-to-GPS distance — answers
    'is there a single customer that any salesperson got near at all?'.
    Also returns row counts per year so we can rule out the year filter,
    and the gps lat/lng range to spot coordinate-system issues.

    Strategy: load 30k gps points + a sample of customers into Python and
    do the nearest-neighbour scan in-memory.  A single SQL JOIN over
    customer × gps with bbox prefilter would time out without indexes
    on gps.latitude/longitude.
    Params: [sample=300] customers to scan, [limit=20]
    """
    if USE_SQLITE:
        return jsonify({"note": "available only on MySQL"})
    try:
        sample_n = int(request.args.get("sample", 300) or 300)
    except (TypeError, ValueError):
        sample_n = 300
    try:
        limit = int(request.args.get("limit", 20) or 20)
    except (TypeError, ValueError):
        limit = 20

    out = {"params": {"sample": sample_n, "limit": limit}}
    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        date_col, lat_col, lng_col = _resolve_gps_columns(cur)

        # Year distribution
        cur.execute(
            f"SELECT YEAR({date_col}) AS y, COUNT(*) AS n "
            f"FROM gps GROUP BY YEAR({date_col}) ORDER BY n DESC LIMIT 20"
        )
        out["rows_per_year_top20"] = cur.fetchall()

        # Bounding boxes side by side
        cur.execute(
            f"SELECT MIN({lat_col}) AS lat_min, MAX({lat_col}) AS lat_max, "
            f"MIN({lng_col}) AS lng_min, MAX({lng_col}) AS lng_max FROM gps"
        )
        out["gps_bbox"] = cur.fetchone()
        cur.execute(
            "SELECT MIN(latitude) AS lat_min, MAX(latitude) AS lat_max, "
            "MIN(longitude) AS lng_min, MAX(longitude) AS lng_max "
            "FROM customer WHERE latitude IS NOT NULL AND longitude IS NOT NULL"
        )
        out["customer_bbox"] = cur.fetchone()

        # Pull all gps points (~30k floats) into Python — small enough.
        cur.execute(
            f"SELECT {lat_col} AS la, {lng_col} AS lo FROM gps "
            f"WHERE {lat_col} IS NOT NULL AND {lng_col} IS NOT NULL"
        )
        gps_pts = [(float(r["la"]), float(r["lo"])) for r in cur.fetchall()]
        out["gps_points_loaded"] = len(gps_pts)

        # Sample customers with coords
        cur.execute(
            "SELECT ship_to, ship_to_name, ship_to_state, latitude, longitude "
            "FROM customer "
            "WHERE latitude IS NOT NULL AND longitude IS NOT NULL "
            "ORDER BY RAND() LIMIT %s",
            [sample_n],
        )
        sample = cur.fetchall()
        out["customers_scanned"] = len(sample)

        # In-memory Haversine: O(sample × 30k) ≈ 9M ops, well under a second
        from math import radians, sin, cos, asin, sqrt
        R = 6371000.0
        results = []
        for c in sample:
            cla = radians(float(c["latitude"]))
            clo = radians(float(c["longitude"]))
            best = None
            for la, lo in gps_pts:
                dla = radians(la) - cla
                dlo = radians(lo) - clo
                a = sin(dla / 2) ** 2 + cos(cla) * cos(radians(la)) * sin(dlo / 2) ** 2
                d = 2 * R * asin(min(1.0, sqrt(a)))
                if best is None or d < best:
                    best = d
            results.append({
                "ship_to":       c["ship_to"],
                "ship_to_name":  c["ship_to_name"],
                "ship_to_state": c["ship_to_state"],
                "shop_lat":      float(c["latitude"]),
                "shop_lng":      float(c["longitude"]),
                "min_dist_m":    round(best, 0) if best is not None else None,
            })
        results.sort(key=lambda r: (r["min_dist_m"] is None, r["min_dist_m"] or 0))

        out["closest_customer_to_any_gps"] = results[:limit]
        # Distribution: how many customers are within X metres
        out["distance_buckets"] = {
            "<=500m":   sum(1 for r in results if r["min_dist_m"] is not None and r["min_dist_m"] <=    500),
            "<=1000m":  sum(1 for r in results if r["min_dist_m"] is not None and r["min_dist_m"] <=  1000),
            "<=2500m":  sum(1 for r in results if r["min_dist_m"] is not None and r["min_dist_m"] <=  2500),
            "<=5000m":  sum(1 for r in results if r["min_dist_m"] is not None and r["min_dist_m"] <=  5000),
            "<=20000m": sum(1 for r in results if r["min_dist_m"] is not None and r["min_dist_m"] <= 20000),
        }
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass
    return jsonify(out)


@app.get("/api/visit_summary")
def visit_summary():
    """Total visit counts across every customer/ship-to.
    Params: [year=2026], [radius=300], [with_sales=1]
    Returns: total shops with ≥1 visit, total visit-days, monthly +
             per-state breakdown, and the top 20 shops.

    with_sales=1 (default) restricts the customer set to those that have
    at least one row in sales_thismonth — i.e. only count shops that are
    currently transacting, ignoring DNU / inactive accounts. Pass
    with_sales=0 to count every customer with coords regardless of sales.

    Strategy: load all gps points + all customers with coords into Python,
    bin gps into a coarse spatial grid, and for each customer scan only
    the nearby cells. Avoids the unindexed customer×gps SQL JOIN.
    """
    if USE_SQLITE:
        return jsonify({"note": "available only on MySQL"})
    try:
        year = int(request.args.get("year", 2026) or 2026)
    except (TypeError, ValueError):
        year = 2026
    try:
        radius_m = float(request.args.get("radius", 300) or 300)
    except (TypeError, ValueError):
        radius_m = 300.0
    with_sales = (request.args.get("with_sales", "0").strip().lower()
                  in ("1", "true", "yes"))
    business_only = (request.args.get("business_days_only", "1").strip().lower()
                     not in ("0", "false", "no"))

    out = {"params": {"year": year, "radius_m": radius_m,
                      "with_sales": with_sales,
                      "business_days_only": business_only}}
    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        date_col, lat_col, lng_col = _resolve_gps_columns(cur)
        salesman_col = _resolve_gps_salesman_col(cur)
        out["gps_salesman_column"] = salesman_col  # null if schema lacks it

        sm_select = f", {salesman_col} AS sm" if salesman_col else ""
        cur.execute(
            f"SELECT {lat_col} AS la, {lng_col} AS lo, {date_col} AS d{sm_select} "
            f"FROM gps WHERE YEAR({date_col}) = %s "
            f"AND {lat_col} IS NOT NULL AND {lng_col} IS NOT NULL",
            [year],
        )
        gps_rows = cur.fetchall()
        if business_only:
            gps_rows = [r for r in gps_rows if _is_business_day(r["d"])]
        out["gps_rows_in_year"] = len(gps_rows)

        # Always pull every customer — BDE assignment count must include
        # shops without sales and even shops without lat/lng.  The visit
        # check + with_sales filter is applied per-row inside the loop.
        cur.execute(
            "SELECT ship_to, ship_to_name, ship_to_state, "
            "       bde_state, salesman_name, latitude, longitude "
            "FROM customer"
        )
        customers = cur.fetchall()
        out["customers_total"] = len(customers)

        # When with_sales=1, only shops that appear in sales_thismonth count
        # toward the visit checking — but they still count toward total Shops.
        shops_with_sales = None
        if with_sales:
            cur.execute("SELECT DISTINCT ship_to FROM sales_thismonth")
            shops_with_sales = {r["ship_to"] for r in cur.fetchall()}
            out["customers_with_sales"] = len(shops_with_sales)

        # Spatial grid (cell ≈ 1.1 km).  Search 3×3 cells covers ≤ 1.5 km
        # which comfortably contains any radius up to ~1 km.  Bumps to 5×5
        # if the user passes a larger radius.
        from collections import defaultdict
        from math import radians, sin, cos, asin, sqrt
        GRID = 0.01
        # extend search ring so radius/cell_size_m is fully covered
        cell_size_m = 1100  # ~0.01° at 30°S
        ring = max(1, int((radius_m / cell_size_m) + 0.999))

        # Grid carries the salesman so we can attribute each GPS hit to the
        # actual BDE that recorded it.  Names normalised UPPER() / strip()
        # so customer.salesman_name and gps.salesmen match despite casing.
        def _norm_name(s):
            return (s or "").strip().upper()

        grid = defaultdict(list)
        for r in gps_rows:
            la, lo, d = float(r["la"]), float(r["lo"]), r["d"]
            sm = _norm_name(r.get("sm")) if salesman_col else ""
            grid[(int(la / GRID), int(lo / GRID))].append((la, lo, d, sm))

        # Active business days per BDE — denominator for No-Visit Days.
        # 'active' = a day where the BDE has any GPS recorded at all.
        active_days_by_bde = defaultdict(set)
        active_days_by_bde_by_month = defaultdict(lambda: defaultdict(set))  # bde -> month -> {date}
        if salesman_col:
            for r in gps_rows:
                sm = _norm_name(r.get("sm"))
                if sm and r["d"] is not None:
                    active_days_by_bde[sm].add(r["d"])
                    active_days_by_bde_by_month[sm][r["d"].month].add(r["d"])

        R = 6371000.0
        per_shop = []
        per_month = defaultdict(lambda: {"shops": set(), "visit_days": 0})
        per_state = defaultdict(lambda: {"shops": set(), "visit_days": 0})
        # per_bde keyed by NORMALIZED salesman name so customer.salesman_name
        # (territory) and gps.salesmen (visits) line up despite case/whitespace.
        per_bde = defaultdict(lambda: {
            "all_shops":    set(),                 # territory from customer.salesman_name
            "shops":        set(),                 # all shops their GPS hit (own + other)
            "shops_own":    set(),                 # shops in their own territory that they visited
            "shops_other":  set(),                 # shops outside their territory that they visited
            "visit_days":   0,
            "visited_dates": set(),
            "state_counts": defaultdict(int),
            "display_name": "",
            # Monthly sparkline data: each is keyed by calendar month (1-12).
            "shops_by_month":         defaultdict(set),  # month -> {ship_to}
            "visit_days_by_month":    defaultdict(int),
            "visited_dates_by_month": defaultdict(set),  # month -> {date}
            # Meeting-log-only tally (kept separate so the BDE table can
            # show a dedicated "Logs" metric next to the GPS-derived ones).
            "logs_total":             0,
            "logs_by_month":          defaultdict(int),
        })
        seen_ship_to = set()  # one ship_to can appear in multiple customer rows

        # Pre-seed display names from gps so a BDE who has no customer entries
        # (e.g. ex-territory) still shows up with their gps-side name.
        if salesman_col:
            for r in gps_rows:
                sm_raw = (r.get("sm") or "").strip()
                if sm_raw:
                    norm = sm_raw.upper()
                    if not per_bde[norm]["display_name"]:
                        per_bde[norm]["display_name"] = sm_raw

        for c in customers:
            if c["ship_to"] in seen_ship_to:
                continue
            seen_ship_to.add(c["ship_to"])

            # Territory: every assigned ship_to counts toward the BDE in
            # customer.salesman_name regardless of sales / coords / visits.
            cust_name_raw = (c["salesman_name"] or "").strip()
            cust_name_norm = cust_name_raw.upper() if cust_name_raw else "(UNASSIGNED)"
            per_bde[cust_name_norm]["all_shops"].add(c["ship_to"])
            # Customer master casing is authoritative for the display name.
            if cust_name_raw:
                per_bde[cust_name_norm]["display_name"] = cust_name_raw
            elif not per_bde[cust_name_norm]["display_name"]:
                per_bde[cust_name_norm]["display_name"] = "(unassigned)"
            if c["bde_state"]:
                per_bde[cust_name_norm]["state_counts"][c["bde_state"].strip()] += 1

            # Visit check needs coords; skip if missing.
            if c["latitude"] is None or c["longitude"] is None:
                continue
            # with_sales filter is applied ONLY to BDE attribution below.
            # Per-shop visit count (used by map popup + visits_by_ship_to)
            # should reflect every shop regardless of sales — otherwise the
            # popup says "0 visits" for a shop the chart clearly shows
            # had GPS visits.
            shop_has_sales = (shops_with_sales is None
                              or c["ship_to"] in shops_with_sales)

            try:
                cla_deg = float(c["latitude"]); clo_deg = float(c["longitude"])
            except (TypeError, ValueError):
                continue
            ci, cj = int(cla_deg / GRID), int(clo_deg / GRID)
            cla = radians(cla_deg); clo = radians(clo_deg)
            cos_cla = cos(cla)
            # Per-shop (any salesman) → drives map markers + popup totals.
            visit_dates_any = defaultdict(set)
            # Per-shop, per-gps-salesman → attribution for the BDE table.
            visit_dates_per_bde = defaultdict(lambda: defaultdict(set))  # bde_norm → month → dates

            for di in range(-ring, ring + 1):
                for dj in range(-ring, ring + 1):
                    for la, lo, d, gps_sm in grid.get((ci + di, cj + dj), ()):
                        dla = radians(la) - cla
                        dlo = radians(lo) - clo
                        h = sin(dla / 2) ** 2 + cos_cla * cos(radians(la)) * sin(dlo / 2) ** 2
                        dist = 2 * R * asin(min(1.0, sqrt(h)))
                        if dist <= radius_m:
                            visit_dates_any[d.month].add(d)
                            if salesman_col and gps_sm:
                                visit_dates_per_bde[gps_sm][d.month].add(d)

            # Per-shop (any salesman) — used by map / popup. Always populated.
            if visit_dates_any:
                total_days_any = sum(len(s) for s in visit_dates_any.values())
                # visited_by_norm = list of normalised BDE names who hit this
                # shop within radius (needed by the map BDE overlay to colour
                # "other-territory" visits when a BDE is filter-selected).
                visited_by_norm = sorted(visit_dates_per_bde.keys()) if salesman_col else []
                per_shop.append({
                    "ship_to":       c["ship_to"],
                    "ship_to_name":  c["ship_to_name"],
                    "ship_to_state": c["ship_to_state"],
                    "bde_state":     c["bde_state"],
                    "salesman_name": c["salesman_name"],
                    "latitude":      cla_deg,
                    "longitude":     clo_deg,
                    "visit_days":    total_days_any,
                    "visited_by":    visited_by_norm,
                    "by_month":      {m: len(s) for m, s in sorted(visit_dates_any.items())},
                })
                for m, s in visit_dates_any.items():
                    per_month[m]["shops"].add(c["ship_to"])
                    per_month[m]["visit_days"] += len(s)
                st = c["ship_to_state"] or "?"
                per_state[st]["shops"].add(c["ship_to"])
                per_state[st]["visit_days"] += total_days_any

            # Per-BDE — credit the visit to whoever's GPS came within radius.
            # Restricted to shops that have sales when with_sales=1 (default),
            # so the BDE table reflects effort on active accounts only.
            if shop_has_sales:
                for gps_norm, dates_by_month in visit_dates_per_bde.items():
                    total_days_this_bde = sum(len(s) for s in dates_by_month.values())
                    bde = per_bde[gps_norm]
                    bde["shops"].add(c["ship_to"])
                    bde["visit_days"] += total_days_this_bde
                    for m, s in dates_by_month.items():
                        bde["visited_dates"].update(s)
                        bde["visited_dates_by_month"][m].update(s)
                        bde["shops_by_month"][m].add(c["ship_to"])
                        bde["visit_days_by_month"][m] += len(s)

        per_shop.sort(key=lambda r: r["visit_days"], reverse=True)

        # ── meeting_log counts (separate "Logs" metric) ──────────────
        # Logged independently of GPS-derived visits — the BDE table
        # surfaces it as its own column so the user can compare effort
        # logged vs effort tracked.  Each row in meeting_log is one
        # tally point; no dedupe (a BDE who writes two memos for the
        # same call genuinely logged twice).
        logged_ship_tos = set()
        # Per-purpose ship-to lists so the map can offer a Purpose
        # filter row (All / Promotion / Product introduction / Claim
        # support / Rebate follow-up / Stock / Other).
        from collections import defaultdict as _dd
        logged_by_purpose = _dd(set)
        try:
            cur.execute(
                "SELECT bde_name, ship_to, visit_purpose, visit_date AS d "
                "FROM meeting_log WHERE YEAR(visit_date) = %s",
                [year],
            )
            for r in cur.fetchall():
                norm = (r.get("bde_name") or "").strip().upper()
                ship = (r.get("ship_to")  or "").strip()
                purp = (r.get("visit_purpose") or "").strip() or "Other"
                d    = r.get("d")
                if not norm or d is None:
                    continue
                if business_only and not _is_business_day(d):
                    continue
                bde = per_bde[norm]
                if not bde["display_name"]:
                    bde["display_name"] = r["bde_name"].strip()
                bde["logs_total"] += 1
                bde["logs_by_month"][d.month] += 1
                if ship:
                    logged_ship_tos.add(ship)
                    logged_by_purpose[purp].add(ship)
        except Exception as e:
            # meeting_log table may not exist on a brand-new install yet
            print(f"[visit_summary] meeting_log tally skipped: {e}")

        out["logged_ship_tos"] = sorted(logged_ship_tos)
        out["logged_ship_tos_by_purpose"] = {
            p: sorted(s) for p, s in logged_by_purpose.items()
        }
        out["total_shops_visited"] = len(per_shop)
        out["total_visit_days"]    = sum(r["visit_days"] for r in per_shop)
        out["by_month"] = [
            {"m": m, "shops_visited": len(per_month[m]["shops"]),
             "visit_days": per_month[m]["visit_days"]}
            for m in sorted(per_month)
        ]
        out["by_state"] = sorted(
            [{"state": s, "shops_visited": len(v["shops"]),
              "visit_days": v["visit_days"]} for s, v in per_state.items()],
            key=lambda r: r["visit_days"], reverse=True,
        )
        # Team-wide active business days (any salesman) — kept as a
        # reference number so the header line still shows it.
        active_business_days_team = {r["d"] for r in gps_rows
                                     if r["d"] is not None and _is_business_day(r["d"])}
        out["active_business_days"] = len(active_business_days_team)

        STATE_ORDER = ["NSW", "QLD", "VIC", "SA", "WA"]
        def _primary_state(counts):
            if not counts:
                return ""
            # Pick the most-common bde_state for this BDE; on ties, prefer
            # earlier in STATE_ORDER.
            return max(counts.items(),
                       key=lambda kv: (kv[1],
                                       -STATE_ORDER.index(kv[0]) if kv[0] in STATE_ORDER else -99))[0]
        def _state_sort_key(s):
            return STATE_ORDER.index(s) if s in STATE_ORDER else len(STATE_ORDER)

        bde_rows = []
        for norm, v in per_bde.items():
            if norm in BDE_EXCLUDE:
                continue
            # No-Visit Days denominator = days where THIS BDE had any GPS
            # recorded.  Subtract the days they actually visited a shop.
            bde_active = active_days_by_bde.get(norm, set())
            no_visit_days = max(0, len(bde_active) - len(v["visited_dates"]))
            display = v["display_name"] or norm
            # Monthly sparkline arrays (length 12, index 0 = January).
            active_by_month = active_days_by_bde_by_month.get(norm, {})
            shops_m  = [len(v["shops_by_month"].get(m, ()))      for m in range(1, 13)]
            visits_m = [int(v["visit_days_by_month"].get(m, 0))  for m in range(1, 13)]
            no_visit_m = [
                max(0, len(active_by_month.get(m, ())) -
                       len(v["visited_dates_by_month"].get(m, ())))
                for m in range(1, 13)
            ]
            logs_m   = [int(v["logs_by_month"].get(m, 0))        for m in range(1, 13)]
            bde_rows.append({
                "bde":           display,
                "state":         _primary_state(v["state_counts"]),
                "total_shops":   len(v["all_shops"]),
                "shops_visited": len(v["shops"]),
                "visit_days":    v["visit_days"],
                "active_days":   len(bde_active),
                "no_visit_days": no_visit_days,
                "logs_total":    v["logs_total"],
                "shops_by_month":         shops_m,
                "visit_days_by_month":    visits_m,
                "no_visit_days_by_month": no_visit_m,
                "logs_by_month":          logs_m,
            })
        out["by_bde"] = sorted(
            bde_rows,
            key=lambda r: (_state_sort_key(r["state"]), -r["visit_days"]),
        )
        out["top_shops"] = per_shop[:20]
        # Full set of visited ship_tos + per-shop visit counts so the map can
        # fade non-visited markers and show counts in popups without an extra
        # round-trip per shop.
        out["visited_ship_tos"] = [r["ship_to"] for r in per_shop]
        out["visits_by_ship_to"] = {r["ship_to"]: r["visit_days"] for r in per_shop}
        # Slim list of every visited shop with location + visited_by (norm
        # names) so the map can render "other-territory visits" when a single
        # BDE is filter-selected.  Trimmed fields keep payload small.
        out["visited_shops"] = [
            {
                "ship_to":       r["ship_to"],
                "ship_to_name":  r["ship_to_name"],
                "salesman_name": r["salesman_name"],
                "latitude":      r["latitude"],
                "longitude":     r["longitude"],
                "visit_days":    r["visit_days"],
                "visited_by":    r["visited_by"],
            }
            for r in per_shop
        ]
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass
    return jsonify(out)


@app.get("/api/visit_debug_gps_locality")
def visit_debug_gps_locality():
    """For each GPS point, find the nearest customer and bucket by
    distance.  Answers 'is the GPS data even taken AT customer locations,
    or mostly on highways / depots / homes?'.

    Also breaks down GPS records by registration so we can see the
    sampling rate (rows per vehicle per day).
    Params: [year=2026], [with_sales=1]
    """
    if USE_SQLITE:
        return jsonify({"note": "available only on MySQL"})
    try:
        year = int(request.args.get("year", 2026) or 2026)
    except (TypeError, ValueError):
        year = 2026
    with_sales = (request.args.get("with_sales", "1").strip().lower()
                  not in ("0", "false", "no"))

    out = {"params": {"year": year, "with_sales": with_sales}}
    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        date_col, lat_col, lng_col = _resolve_gps_columns(cur)

        # GPS points for the year
        cur.execute(
            f"SELECT {lat_col} AS la, {lng_col} AS lo, {date_col} AS d, "
            f"       registration AS rego "
            f"FROM gps WHERE YEAR({date_col}) = %s "
            f"AND {lat_col} IS NOT NULL AND {lng_col} IS NOT NULL",
            [year],
        )
        gps_rows = cur.fetchall()
        out["gps_rows_in_year"] = len(gps_rows)

        # Customers (optionally filtered to those with sales)
        if with_sales:
            cur.execute(
                "SELECT c.ship_to, c.latitude, c.longitude FROM customer c "
                "WHERE c.latitude IS NOT NULL AND c.longitude IS NOT NULL "
                "AND EXISTS (SELECT 1 FROM sales_thismonth s "
                "            WHERE s.ship_to = c.ship_to)"
            )
        else:
            cur.execute(
                "SELECT ship_to, latitude, longitude FROM customer "
                "WHERE latitude IS NOT NULL AND longitude IS NOT NULL"
            )
        customers = cur.fetchall()
        out["customers_considered"] = len(customers)

        # Build customer spatial grid for fast lookup
        from collections import defaultdict, Counter
        from math import radians, sin, cos, asin, sqrt
        GRID = 0.05  # ~5.5 km cells — search 3×3 covers ~15 km
        cust_grid = defaultdict(list)
        for c in customers:
            la, lo = float(c["latitude"]), float(c["longitude"])
            cust_grid[(int(la / GRID), int(lo / GRID))].append((la, lo, c["ship_to"]))

        R = 6371000.0
        buckets = Counter()
        BUCKET_EDGES = (100, 250, 500, 1000, 2500, 5000, 10000, 25000, 100000)
        no_match = 0  # GPS where there's no customer within the search ring

        for r in gps_rows:
            gla = float(r["la"]); glo = float(r["lo"])
            gi, gj = int(gla / GRID), int(glo / GRID)
            best = None
            for di in (-1, 0, 1):
                for dj in (-1, 0, 1):
                    for cla, clo, _ in cust_grid.get((gi + di, gj + dj), ()):
                        dla = radians(gla) - radians(cla)
                        dlo = radians(glo) - radians(clo)
                        h = sin(dla / 2) ** 2 + cos(radians(cla)) * cos(radians(gla)) * sin(dlo / 2) ** 2
                        d = 2 * R * asin(min(1.0, sqrt(h)))
                        if best is None or d < best:
                            best = d
            if best is None:
                no_match += 1
                continue
            # bucket: place into the smallest edge it fits under
            placed = False
            for edge in BUCKET_EDGES:
                if best <= edge:
                    buckets[f"<={edge}m"] += 1
                    placed = True
                    break
            if not placed:
                buckets[">100km"] += 1

        out["gps_no_match_within_15km"] = no_match
        out["gps_distance_to_nearest_customer"] = dict(buckets)
        out["gps_within_500m_of_some_customer_pct"] = round(
            100 * buckets.get("<=100m", 0) / max(1, len(gps_rows)) +
            100 * buckets.get("<=250m", 0) / max(1, len(gps_rows)) +
            100 * buckets.get("<=500m", 0) / max(1, len(gps_rows)),
            1,
        )

        # Per-registration sampling rate
        cur.execute(
            f"SELECT registration AS rego, COUNT(*) AS n, "
            f"COUNT(DISTINCT DATE({date_col})) AS days "
            f"FROM gps WHERE YEAR({date_col}) = %s "
            f"GROUP BY registration ORDER BY n DESC LIMIT 30",
            [year],
        )
        regs = cur.fetchall()
        for r in regs:
            r["rows_per_day"] = round(r["n"] / max(1, r["days"]), 2)
        out["top_registrations"] = regs
        cur.execute(
            f"SELECT COUNT(DISTINCT registration) AS n FROM gps "
            f"WHERE YEAR({date_col}) = %s",
            [year],
        )
        out["total_distinct_registrations"] = cur.fetchone()["n"]
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass
    return jsonify(out)

# ==============================================================================
# REBATE CALCULATOR
# ==============================================================================

@app.get("/rebate")
def rebate_page():
    return send_from_directory("static", "rebate.html")

@app.get("/api/rebate_assignment_check")
def api_rebate_assignment_check():
    """Diagnostic for 'why isn't BDE X showing shop Y on the rebate page?'.

    Query params:
      bde   – salesman_name to look up (e.g. 'Borghese Alessio')
      sold  – optional, restrict to ship_tos under a sold_to_name
              containing this string (e.g. 'JAX' or 'JAXQUICKFIT')

    Returns:
      total_ship_tos              – ship_to count assigned to that BDE
      ship_to_breakdown_by_sold   – sold_to_name → count of ship_tos
      filtered_sold_ship_tos      – matching ship_tos for ?sold=…
      jax_assignment_summary      – every BDE that owns any JAX ship_to
                                    (case insensitive sold_to_name LIKE %JAX%)
    """
    bde   = (request.args.get("bde")  or "").strip()
    sold  = (request.args.get("sold") or "").strip()
    out   = {}
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        # BDE-specific stats
        if bde:
            cur.execute(
                "SELECT COUNT(*) AS n FROM customer "
                "WHERE UPPER(TRIM(salesman_name)) = UPPER(TRIM(%s))",
                (bde,))
            out["total_ship_tos"] = (cur.fetchone() or {}).get("n", 0)

            cur.execute(
                "SELECT sold_to_name, COUNT(*) AS n FROM customer "
                "WHERE UPPER(TRIM(salesman_name)) = UPPER(TRIM(%s)) "
                "GROUP BY sold_to_name "
                "ORDER BY n DESC, sold_to_name "
                "LIMIT 80",
                (bde,))
            out["ship_to_breakdown_by_sold"] = cur.fetchall()

            if sold:
                like = f"%{sold}%"
                cur.execute(
                    "SELECT sold_to, sold_to_name, ship_to, ship_to_name, bde_state "
                    "FROM customer "
                    "WHERE UPPER(TRIM(salesman_name)) = UPPER(TRIM(%s)) "
                    "  AND UPPER(sold_to_name) LIKE UPPER(%s) "
                    "ORDER BY ship_to_name "
                    "LIMIT 200",
                    (bde, like))
                out["filtered_sold_ship_tos"] = cur.fetchall()

        # Cross-cutting summary: who owns JAX ship_tos right now?
        cur.execute(
            "SELECT TRIM(salesman_name) AS bde, bde_state, COUNT(*) AS n "
            "FROM customer "
            "WHERE UPPER(sold_to_name) LIKE '%JAX%' "
            "   OR UPPER(ship_to_name) LIKE 'JAX %' "
            "GROUP BY TRIM(salesman_name), bde_state "
            "ORDER BY n DESC")
        out["jax_assignment_summary"] = cur.fetchall()

        # All distinct salesman_name spellings — to catch case/format drift
        # (e.g. 'Borghese Alessio' vs 'BORGHESE Alessio' vs 'Alessio Borghese').
        cur.execute(
            "SELECT salesman_name, COUNT(*) AS n FROM customer "
            "GROUP BY salesman_name ORDER BY n DESC")
        out["all_salesman_names"] = cur.fetchall()

        cur.close(); conn.close()
        return jsonify(out)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── Rebate calc-basis helpers ───────────────────────────────────────────────
# The trailing token of a structure name designates how it is calculated:
#   _SR                       -> per SHIP_TO  (JAX/AJT, BJ/ABJ, TP/ATP, ACD, APP,
#                                              ATW, ABZ, Mobis, AVW …)
#   _HQ / _VR / _IT / _WTY    -> per SOLD_TO, on the full sold_to total
#   (no suffix)               -> per SOLD_TO  (base structure)
# This replaces the old hardcoded substring list {AJT,ABJ,ATP,APP,ACD}, which
# mis-classified HQ/VR variants (e.g. 731942's HK_ALL_AJT_Q_HQ) as per-ship_to
# and missed the SR variants of ATW/ABZ/Mobis/AVW.
_REBATE_SUFFIXES = ("SR", "HQ", "VR", "IT", "WTY")

def _rebate_suffix(struct):
    last = (struct or "").rsplit("_", 1)[-1].upper()
    return last if last in _REBATE_SUFFIXES else ""

def _rebate_is_ship_to(struct):
    return _rebate_suffix(struct) == "SR"

def _rebate_scope(struct):
    """Display/grouping scope: SR | HQ | VR | IT | WTY | BASE."""
    suf = _rebate_suffix(struct)
    return "SR" if suf == "SR" else (suf or "BASE")


# Sales order types (sales_thismonth.so_type) that count toward rebates.
# Only these are included in every rebate query / diagnostic below.
REBATE_SO_TYPES = ("ZWH1", "ZCR1", "ZDR1", "ZDF1", "ZRE1", "ZREN")
_REBATE_SO_TYPES_IN = "(" + ",".join("'%s'" % t for t in REBATE_SO_TYPES) + ")"

# ─── 2026 monthly sales tables — auto-detected at request time ─────────
# Promo features pull only from 2026 calendar-year sales (sales_2526
# carries no promo metadata) and the set of monthly tables grows by one
# every month — sales_2607, sales_2608, …  Discover them via
# INFORMATION_SCHEMA so a fresh month doesn't require a code change.
_SALES_2026_TABLES_CACHE = {"ts": 0, "names": []}

def _ensure_promo_indexes():
    """Adds the indexes the promo EXISTS predicate needs to stop
    full-scanning promo_customer / promo_plan per outer sales row.
    Without these, every Group By / Category switch on a chart that
    uses the promo logic re-walks both tables N × M times.

    Idempotent — skips when each index is already there.  Failures
    on any one don't block the others (different MySQL versions
    have slightly different syntax for IF NOT EXISTS on indexes,
    so we just swallow duplicate errors)."""
    try:
        conn = get_connection(); cur = conn.cursor()
    except Exception as e:
        print(f"[promo idx] connect skipped: {e}")
        return
    plans = [
        ("promo_customer", "idx_pc_promo",            "(promo)"),
        ("promo_customer", "idx_pc_brand_dc",         "(brand, dc_rate_start, dc_rate_end)"),
        ("promo_customer", "idx_pc_sold_to",          "(sold_to)"),
        ("promo_customer", "idx_pc_customer_group",   "(customer_group)"),
        ("promo_plan",     "idx_pp_promo",            "(promo)"),
        ("promo_plan",     "idx_pp_product_group",    "(product_group)"),
        ("promo_plan",     "idx_pp_dates",            "(start_date, end_date)"),
        # Customer master is hit by the per-ship_to roll-up subquery
        # on every chart query.  These two indexes turn that aggregate
        # into an index scan rather than a full table sort.
        ("customer",       "idx_cust_ship_to",        "(ship_to)"),
        ("customer",       "idx_cust_sold_to",        "(sold_to)"),
    ]
    for tbl, idx, cols in plans:
        try:
            cur.execute(f"CREATE INDEX {idx} ON {tbl} {cols}")
        except Exception as e:
            msg = str(e)
            # 'Duplicate key name' / errno 1061 just means the index
            # already exists; anything else gets surfaced once so the
            # operator knows to investigate.
            if "1061" not in msg and "Duplicate" not in msg:
                print(f"[promo idx] {tbl}.{idx} create skipped: {e}")
    try:
        conn.commit(); cur.close(); conn.close()
    except Exception:
        pass

_ensure_promo_indexes()


def _ensure_sales_2526_indexes():
    """Add the indexes monthly_sales / monthly_breakdown / map view /
    rebate roll-up etc. actually need on the rebuilt sales_2526.  The
    table only ships with billing_date, but every query gates either on
    billing_date (for the year filter) or joins on ship_to / sold_to /
    material.  Without these indexes those gates degrade to a full
    table scan, which is what was making Group-By switches still slow
    even after the wrapping-subquery rewrite.

    Idempotent — duplicate-key errors on existing indexes are swallowed
    so we can keep adding new ones in a single migration."""
    try:
        conn = get_connection(); cur = conn.cursor()
    except Exception as e:
        print(f"[sales_2526 idx] connect skipped: {e}")
        return
    plans = [
        ("idx_s2526_billing_date",         "(billing_date)"),
        ("idx_s2526_ship_to",              "(ship_to)"),
        ("idx_s2526_sold_to",              "(sold_to)"),
        ("idx_s2526_material",             "(material)"),
        ("idx_s2526_brand",                "(brand)"),
        # composite for any same-day ship_to roll-up done over
        # historical months (mirrors the sales_thismonth dq join).
        ("idx_s2526_billing_ship_brand",   "(billing_date, ship_to, brand)"),
    ]
    for idx, cols in plans:
        try:
            cur.execute(f"CREATE INDEX {idx} ON sales_2526 {cols}")
        except Exception as e:
            msg = str(e)
            if "1061" not in msg and "Duplicate" not in msg:
                print(f"[sales_2526 idx] {idx} create skipped: {e}")
    try:
        conn.commit(); cur.close(); conn.close()
    except Exception:
        pass

_ensure_sales_2526_indexes()


def _refresh_customer_rollup():
    """(Re)build customer_rollup as a materialised per-ship_to roll-up
    of customer.  Every chart query LEFT JOINs against it for
    bde_state / salesman / sold_to_group / channels / etc., so
    pre-computing it once means each query just does an index lookup
    on cus.ship_to instead of re-running GROUP BY ship_to over the
    whole customer master.

    Idempotent — drops and rebuilds.  customer is small (~5k rows) so
    a full rebuild takes milliseconds.  Call again from an admin
    endpoint after a customer-master refresh."""
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("DROP TABLE IF EXISTS customer_rollup")
        cur.execute(
            "CREATE TABLE customer_rollup AS "
            "SELECT ship_to, "
            "       MIN(NULLIF(TRIM(bde_state),''))      AS bde_state, "
            "       MIN(NULLIF(TRIM(salesman_name),'' )) AS salesman_name, "
            "       MIN(NULLIF(TRIM(sold_to_group),'' )) AS sold_to_group, "
            "       MIN(NULLIF(TRIM(sold_to_name),'' ))  AS sold_to_name, "
            "       MIN(NULLIF(TRIM(ship_to_name),'' ))  AS ship_to_name, "
            "       MIN(NULLIF(TRIM(channels),''))       AS channels, "
            "       MIN(NULLIF(TRIM(sold_to),''))        AS sold_to "
            "FROM customer GROUP BY ship_to"
        )
        try:
            cur.execute("ALTER TABLE customer_rollup ADD PRIMARY KEY (ship_to)")
        except Exception:
            pass
        for col in ("sold_to", "sold_to_group", "channels", "salesman_name", "bde_state"):
            try:
                cur.execute(
                    f"CREATE INDEX idx_cr_{col} ON customer_rollup ({col})"
                )
            except Exception:
                pass
        conn.commit(); cur.close(); conn.close()
        print("[customer_rollup] rebuilt")
    except Exception as e:
        print(f"[customer_rollup] rebuild skipped: {e}")

_refresh_customer_rollup()


@app.get("/api/admin/refresh_customer_rollup")
def admin_refresh_customer_rollup():
    """Manual hook: rebuild customer_rollup after editing the customer
    master.  Returns simple {"ok": True} on success."""
    _refresh_customer_rollup()
    return jsonify({"ok": True})


def _sales_2526_from(alias: str = "s", year=None) -> str:
    """Direct reference to sales_2526 with no derived-table wrap.

    Earlier iterations wrapped the table in a SELECT-with-aliases
    subquery so existing `s.year` / `s.month` / `s.day` references kept
    working, but MySQL had to materialise the entire 17-month table
    before JOIN / GROUP BY ran — slow enough that Group-By switching
    on the Monthly chart took minutes.  _ensure_sales_2526_generated_
    cols() now adds those same year / month / day aliases as VIRTUAL
    generated columns with a composite index on the table itself, so
    `FROM sales_2526 s` is enough and queries that gate on `s.year =
    2025` / `GROUP BY s.month` use the index path directly.

    `year` is kept on the signature for backward-compatible callsites;
    it's ignored here because the outer query is expected to gate on
    `s.year` (which is now indexed) or `s.billing_date` directly.
    """
    return f"sales_2526 {alias}"

def _sales_2026_tables(cur):
    """Return the ordered list of sales tables that hold 2026 data:
    every sales_26?? plus sales_thismonth.  Cached for 60s so the schema
    introspection doesn't repeat per request."""
    import time as _t
    now = _t.monotonic()
    if (now - _SALES_2026_TABLES_CACHE["ts"]) < 60 and _SALES_2026_TABLES_CACHE["names"]:
        return list(_SALES_2026_TABLES_CACHE["names"])
    try:
        cur.execute("""
            SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = DATABASE()
              AND (TABLE_NAME REGEXP '^sales_26[0-9]{2}$'
                   OR TABLE_NAME = 'sales_thismonth')
            ORDER BY TABLE_NAME
        """)
        names = [r[0] if not isinstance(r, dict) else r["TABLE_NAME"]
                 for r in cur.fetchall()]
        # Sort by (year, month) so sales_thismonth lands at the end.
        def _key(n):
            if n == "sales_thismonth":
                return (9999, 99)
            try:
                yy = int(n[6:8]); mm = int(n[8:10])
                return (2000 + yy, mm)
            except Exception:
                return (9999, 99)
        names.sort(key=_key)
        # Prevent double-counting: if sales_thismonth's effective month
        # (which _sales_table_year_month resolves to the business-
        # effective one) already has a matching sales_26MM archive in
        # the list, drop sales_thismonth from the union so June's data
        # doesn't count twice on July 1 (once from sales_2606, once
        # from sales_thismonth still holding June's rows).
        if "sales_thismonth" in names:
            eff_y, eff_m = _business_effective_ym()
            archive = f"sales_{eff_y % 100:02d}{eff_m:02d}"
            if archive in names:
                names = [n for n in names if n != "sales_thismonth"]
        _SALES_2026_TABLES_CACHE["ts"]    = now
        _SALES_2026_TABLES_CACHE["names"] = names
        return list(names)
    except Exception as e:
        print(f"[2026 sales] table discovery failed: {e}")
        # Fall back to the known set so the feature still works in an
        # environment that lacks INFORMATION_SCHEMA permissions.
        return ["sales_2601", "sales_2602", "sales_2603",
                "sales_2604", "sales_2605", "sales_thismonth"]

def _sales_table_year_month(table_name):
    """Derive (year, month) from a 2026 sales table name.  The per-month
    tables are named `sales_YYMM` (e.g. sales_2603 = March 2026), and
    `sales_thismonth` is the current calendar month — so we don't need
    the underlying table to expose year/month columns of its own.

    Uses the business-effective month for sales_thismonth so that on
    the first business day of a new month (when the table still holds
    the *previous* month's overnight-loaded rows) we don't spawn a
    phantom bar at the new month.  See _business_effective_ym()."""
    if table_name == "sales_thismonth":
        return _business_effective_ym()
    if table_name.startswith("sales_") and len(table_name) == len("sales_") + 4:
        yymm = table_name[len("sales_"):]
        try:
            yy = int(yymm[:2]); mm = int(yymm[2:])
            if 1 <= mm <= 12:
                return 2000 + yy, mm
        except Exception:
            pass
    return None, None

def _sales_2026_union(cur, alias="s", cols=None):
    """Build a `(SELECT … UNION ALL SELECT … …) AS <alias>` SQL fragment
    that exposes every 2026 monthly sales table as a single virtual
    source.  Use in place of `FROM sales_2526` for any 2026-only query.

    The per-month tables (sales_2601 … sales_2605) don't carry their
    own year/month columns — each table IS a single month, so it would
    be redundant.  We synthesise them as constants in each SELECT so
    the outer query can `GROUP BY s.month` exactly as it does on
    sales_2526.  All other columns come from SHOW COLUMNS (lowercase-
    compared so 'Month' vs 'month' don't break the intersection)."""
    tables = _sales_2026_tables(cur)
    if not tables:
        return "(SELECT * FROM sales_thismonth WHERE 1=0) AS " + alias

    if cols is None:
        # Discover non-year/month columns across all tables; everyone
        # gets the intersection so a UNION ALL can succeed.
        seen_in_all = None
        canonical   = {}
        for t in tables:
            try:
                cur.execute(f"SHOW COLUMNS FROM {t}")
                got_lc = set()
                for r in cur.fetchall():
                    name = r[0] if not isinstance(r, dict) else (r.get("Field") or r.get("field"))
                    if not name:
                        continue
                    lc = name.lower()
                    if lc in ("year", "month"):
                        # Synthesised below, ignore whatever the table
                        # happens to provide for these.
                        continue
                    got_lc.add(lc)
                    canonical.setdefault(lc, name)
                seen_in_all = got_lc if seen_in_all is None else (seen_in_all & got_lc)
            except Exception:
                pass

        required = ["qty", "amt", "sold_to", "ship_to", "brand", "material"]
        for lc in required:
            canonical.setdefault(lc, lc)
        if seen_in_all:
            cols_lc = sorted(seen_in_all | set(required))
        else:
            cols_lc = sorted(set(required))
        non_ym_cols = ", ".join(canonical[lc] for lc in cols_lc)
    else:
        non_ym_cols = cols

    parts = []
    for t in tables:
        y, m = _sales_table_year_month(t)
        if y is None or m is None:
            # Unknown naming scheme — fall back to NULL year/month so
            # the SELECT still parses; the outer query will just see
            # zero rows for that month bucket.
            parts.append(f"SELECT NULL AS year, NULL AS month, {non_ym_cols} FROM {t}")
        else:
            parts.append(f"SELECT {y} AS year, {m} AS month, {non_ym_cols} FROM {t}")
    return "(\n  " + "\n  UNION ALL ".join(parts) + f"\n) AS {alias}"

def _ensure_promo_customer_category():
    """Earlier iterations added a `category` column to promo_customer;
    the actual setup has a separate `promo_category` table with the
    mapping, so drop the now-unused column if it lingers from a prior
    deploy.  Safe to leave the column too — keeping the migration
    idempotent in both directions."""
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SHOW TABLES LIKE 'promo_customer'")
        if not cur.fetchone():
            cur.close(); conn.close()
            return
        cur.execute("SHOW COLUMNS FROM promo_customer LIKE 'category'")
        if cur.fetchone():
            try:
                cur.execute("DROP INDEX idx_promo_customer_cat ON promo_customer")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE promo_customer DROP COLUMN category")
            except Exception as e:
                # Non-fatal — column stays in place, queries below ignore it.
                print(f"[promo_customer] could not drop legacy category col: {e}")
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"[promo_customer] category cleanup skipped: {e}")

_ensure_promo_customer_category()

def _promo_category_of(promo_name):
    """Auto-derive the top-level button name from a sub-promo string by
    splitting on the first underscore.  No mapping table required.
        '443'          → '443'
        '443_30%'      → '443'
        'iON_70%'      → 'iON'
        'TrueBlue_12%' → 'TrueBlue'
    A bare promo with no underscore acts as both category and its sole
    sub-button (e.g. '443' renders as a category with a single sub '443')."""
    s = (promo_name or "").strip()
    if not s:
        return ""
    return s.split("_", 1)[0]

def _promo_qty_match_sql(sales_alias="s", day_qty_alias=None):
    """SQL fragment that gates a sale against a promo_customer row's
    min_qty threshold.

    Default (day_qty_alias=None) is the per-row gate:
        s.qty >= pc.min_qty

    When the caller pre-aggregates qty per (ship_to, day, brand) and
    exposes it through a LEFT JOIN alias (e.g. "dq.day_qty"), TrueBlue
    rules compare min_qty against that pre-summed day total instead —
    the threshold the sales team actually quotes for TrueBlue ("X
    tires that day at that shop", not "X of a single SKU").  Other
    promos (443 / iON) keep the per-row gate.

    Why a JOIN alias and not an inline correlated subquery?  EXISTS
    already iterates promo_customer per outer row; a self-referencing
    SUM subquery inside that EXISTS would scan the sales table once
    per match attempt and make daily_breakdown take minutes.  Pre-
    aggregating via LEFT JOIN runs the SUM once.
    """
    s = sales_alias
    if not day_qty_alias:
        return f"{s}.qty >= pc.min_qty"
    return (
        f"((pc.promo = 'TrueBlue' AND COALESCE({day_qty_alias}, 0) >= pc.min_qty)"
        f" OR (pc.promo <> 'TrueBlue' AND {s}.qty >= pc.min_qty))"
    )

def _promo_filter_clauses(promos, sales_alias="s",
                          carrying_alias="mat", customer_alias="cus",
                          year_expr=None, month_expr=None,
                          day_qty_alias=None):
    """Build the WHERE clauses + params needed to constrain a sales query
    to rows that qualify for any of the selected sub-promos.

    Caller responsibility:
      • Ensure the `mat` (carrying_26) join is present — promos require
        product_group + line columns.
      • Ensure the `cus` (customer) join is present — sold_to_group is
        used for customer_group fallback when promo_customer.sold_to is
        empty.
      • Sales table aliased as `s` (or pass sales_alias).

    `year_expr` / `month_expr` default to `<sales_alias>.year` and
    `<sales_alias>.month` — fine for sales_2526 and for the 2026
    monthly-tables union (which synthesises those columns).  For a
    single-month table like sales_thismonth that has neither, pass in
    integer literals (e.g. year_expr="2026", month_expr="6") so the
    period match against promo_plan still works.

    Returns: (wh_list, params_list).  Empty when promos is empty.
    """
    if not promos:
        return [], []
    placeholders = ",".join(["%s"] * len(promos))
    s = sales_alias; c = carrying_alias; cu = customer_alias
    y_e = year_expr  if year_expr  is not None else f"{s}.year"
    m_e = month_expr if month_expr is not None else f"{s}.month"
    qty_check = _promo_qty_match_sql(sales_alias=s, day_qty_alias=day_qty_alias)
    wh = [
        # PCLT is always the umbrella for promos.
        f"{c}.line = 'PCLT'",
        # Sale qualifies for at least one selected sub-promo.
        f"""EXISTS (
            SELECT 1
            FROM promo_customer pc
            LEFT JOIN promo_plan pp ON pp.promo = pc.promo
            WHERE pc.promo IN ({placeholders})
              AND ({qty_check})
              AND {s}.dc_rate  BETWEEN pc.dc_rate_start AND pc.dc_rate_end
              AND {s}.brand    = pc.brand
              AND (
                   (pc.sold_to <> '' AND pc.sold_to = {s}.sold_to)
                OR (pc.sold_to = '' AND pc.customer_group = {cu}.sold_to_group)
              )
              AND (
                pp.promo IS NULL
                OR (
                  (pp.product_group = '' OR {c}.product_group = pp.product_group)
                  AND (pp.start_date IS NULL
                       OR ({y_e} * 100 + {m_e}) >=
                          (YEAR(pp.start_date) * 100 + MONTH(pp.start_date)))
                  AND (pp.end_date IS NULL
                       OR ({y_e} * 100 + {m_e}) <=
                          (YEAR(pp.end_date) * 100 + MONTH(pp.end_date)))
                  AND (pp.material = '' OR pp.material = {s}.material)
                )
              )
        )"""
    ]
    return wh, list(promos)

def _promotion_group_col_sql(detail=False, sales_alias="s",
                             carrying_alias="mat", customer_alias="cus",
                             year_expr=None, month_expr=None,
                             day_qty_alias=None):
    """SQL expression that buckets each sales row by promotion membership.

    detail=False → returns 'Promotion' / 'Non-Promotion' (binary stack).
    detail=True  → returns the matching sub-promo name (e.g. '443',
                   '443_30%', 'iON_70%') or 'Non-Promotion' when no
                   promo qualifies.

    Mirrors the same EXISTS predicate used by `_promo_filter_clauses`
    so the bucket aligns exactly with the promo-button filter:
      • carrying line must be PCLT (the only line promos live on),
      • qty / dc_rate / brand match the promo_customer row,
      • either sold_to matches or customer_group matches (group fallback),
      • promo_plan period covers the sale (year × month BETWEEN),
      • product_group + optional material align with the plan.

    Caller must ensure the `mat` (carrying_26) and `cus` (customer)
    joins are added — both are referenced inside the predicate.  For
    single-month tables like sales_thismonth that lack year / month
    columns, pass integer-literal `year_expr` / `month_expr` strings.
    """
    s = sales_alias; c = carrying_alias; cu = customer_alias
    y_e = year_expr  if year_expr  is not None else f"{s}.year"
    m_e = month_expr if month_expr is not None else f"{s}.month"
    qty_check = _promo_qty_match_sql(sales_alias=s, day_qty_alias=day_qty_alias)
    match_sql = f"""
        SELECT pc.promo
        FROM promo_customer pc
        LEFT JOIN promo_plan pp ON pp.promo = pc.promo
        WHERE ({qty_check})
          AND {s}.dc_rate  BETWEEN pc.dc_rate_start AND pc.dc_rate_end
          AND {s}.brand    = pc.brand
          AND (
               (pc.sold_to <> '' AND pc.sold_to = {s}.sold_to)
            OR (pc.sold_to = '' AND pc.customer_group = {cu}.sold_to_group)
          )
          AND (
            pp.promo IS NULL
            OR (
              (pp.product_group = '' OR {c}.product_group = pp.product_group)
              AND (pp.start_date IS NULL
                   OR ({y_e} * 100 + {m_e}) >=
                      (YEAR(pp.start_date) * 100 + MONTH(pp.start_date)))
              AND (pp.end_date IS NULL
                   OR ({y_e} * 100 + {m_e}) <=
                      (YEAR(pp.end_date) * 100 + MONTH(pp.end_date)))
              AND (pp.material = '' OR pp.material = {s}.material)
            )
          )"""
    if not detail:
        return (
            f"CASE WHEN {c}.line = 'PCLT' AND EXISTS ({match_sql}) "
            f"THEN 'Promotion' ELSE 'Non-Promotion' END"
        )
    # detail = the matched promo name as-is.  promo_customer.promo
    # carries the bucket label directly (443 / iON / TrueBlue /
    # 443_beforeTrueBlue / ...) so each surfaces as its own stack on
    # the Detail chart.  MIN() collapses the rare case where one sale
    # qualifies under two rows; the user has confirmed the categories
    # don't overlap so MIN is safe.
    return (
        f"CASE WHEN {c}.line = 'PCLT' THEN "
        f"COALESCE((SELECT MIN(pc.promo) FROM promo_customer pc "
        f"LEFT JOIN promo_plan pp ON pp.promo = pc.promo "
        f"WHERE ({qty_check}) AND {s}.dc_rate BETWEEN pc.dc_rate_start AND pc.dc_rate_end "
        f"AND {s}.brand = pc.brand AND ("
        f"(pc.sold_to <> '' AND pc.sold_to = {s}.sold_to) OR "
        f"(pc.sold_to = '' AND pc.customer_group = {cu}.sold_to_group)"
        f") AND (pp.promo IS NULL OR ("
        f"(pp.product_group = '' OR {c}.product_group = pp.product_group) AND "
        f"(pp.start_date IS NULL OR "
        f"  ({y_e} * 100 + {m_e}) >= (YEAR(pp.start_date) * 100 + MONTH(pp.start_date))) AND "
        f"(pp.end_date IS NULL OR "
        f"  ({y_e} * 100 + {m_e}) <= (YEAR(pp.end_date) * 100 + MONTH(pp.end_date))) AND "
        f"(pp.material = '' OR pp.material = {s}.material)"
        f"))), 'Non-Promotion') ELSE 'Non-Promotion' END"
    )

@app.get("/api/promo/buttons")
def api_promo_buttons():
    """Return the active promo categories derived live from
    promo_customer.promo.  After the schema simplification, promo IS
    the category (443 / iON / TrueBlue); the per-tier split lives in
    dc_rate_start..dc_rate_end on each row instead of in the promo
    name suffix.  Each category surfaces as a single sub-button so the
    response shape stays compatible with the existing button row.

    Shape:
      { "categories": [
          { "name": "443",      "subs": ["443"] },
          { "name": "iON",      "subs": ["iON"] },
          { "name": "TrueBlue", "subs": ["TrueBlue"] },
        ] }
    """
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT DISTINCT promo
            FROM promo_customer
            WHERE promo IS NOT NULL AND TRIM(promo) <> ''
            ORDER BY promo
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        out = {}
        for r in rows:
            pr  = r["promo"]
            cat = _promo_category_of(pr)
            out.setdefault(cat, [])
            if pr not in out[cat]:
                out[cat].append(pr)
        # Stable ordering — numeric categories first ('443'), then
        # alphabetic.  Keeps the button row predictable across deploys.
        cats = sorted(out.keys(),
                      key=lambda k: (0 if k and k[:1].isdigit() else 1, k))
        return jsonify({
            "categories": [{"name": k, "subs": out[k]} for k in cats]
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ─── Monthly Highlights ───────────────────────────────────────────────
# Per-month executive snapshot with five objective views: overall
# summary, by region, by product group, by promotion, and sold-to
# movement (gainers / losers / newly active / newly silent).
@app.get("/highlights")
def highlights_page():
    return send_from_directory("static", "highlights.html")

def _highlights_src_for(cur, year, month):
    """Resolve (year, month) → (table_name, where_clause, params_tuple).
    Picks the 2026 monthly table when available, sales_thismonth for
    the *business-effective* current month, sales_2526 for other years
    (with year+month filter applied).

    The business-effective month matches the frontend's effectiveMonth()
    logic: on the first business day of a new month, sales_thismonth
    still holds the previous month's overnight-loaded rows.  So on
    July 1 a request for June 2026 should read from sales_thismonth,
    NOT fall through to an empty sales_2606 archive."""
    if year == 2026:
        eff_y, eff_m = _business_effective_ym()
        if year == eff_y and month == eff_m:
            # Only fall back to sales_thismonth if the "closed" archive
            # for that month doesn't exist yet.  Prevents double-counting
            # once the batch job has moved June into sales_2606.
            candidate = f"sales_{year % 100:02d}{month:02d}"
            try:
                cur.execute("SHOW TABLES LIKE %s", (candidate,))
                if cur.fetchone():
                    return (candidate, "", ())
            except Exception:
                pass
            return ("sales_thismonth", "", ())
        candidate = f"sales_{year % 100:02d}{month:02d}"
        try:
            cur.execute("SHOW TABLES LIKE %s", (candidate,))
            if cur.fetchone():
                return (candidate, "", ())
        except Exception:
            pass
        # Archive isn't split out yet — fall through to sales_2526 with
        # a YEAR/MONTH gate.  Without this the page rendered as all
        # -100% because the resolver returned (None, ...) and the
        # caller silently skipped the query.
    return (
        "sales_2526",
        "WHERE YEAR(s.billing_date) = %s AND MONTH(s.billing_date) = %s",
        (year, month),
    )

def _highlights_pct(this_v, prev_v):
    """% change from prev → this, or None when prev is zero/missing
    (objective: no fake denominators)."""
    if not prev_v or prev_v == 0:
        return None
    return round((this_v - prev_v) / prev_v * 100, 1)

def _narrate_change(pct):
    """Turn a raw MoM percentage into a phrase the narrative generator
    can splice into a sentence.  None (missing baseline) reads as
    'from a zero base' so the reader knows the % is meaningless there."""
    if pct is None:
        return "up from a zero base"
    if pct >= 15:      return f"up {pct:+.1f}%"
    if pct >= 5:       return f"up a solid {pct:+.1f}%"
    if pct >= 1:       return f"up {pct:+.1f}%"
    if pct >= -1:      return "essentially flat"
    if pct >= -5:      return f"down {abs(pct):.1f}%"
    if pct >= -15:     return f"down a meaningful {abs(pct):.1f}%"
    return f"down sharply {abs(pct):.1f}%"

@app.get("/api/monthly_highlights")
def api_monthly_highlights():
    """Per-month KPI snapshot.  Body: ?month=YYYY-MM&metric=qty|amt.
    Returns a single JSON shape with five sections so the page renders
    in one round trip."""
    month_str = (request.args.get("month") or "").strip()
    metric    = (request.args.get("metric") or "qty").strip().lower()
    metric    = "qty" if metric == "qty" else "amt"
    # Optional NSW / QLD / VIC / WA filter.  Folds SA/TAS into VIC,
    # NT into WA, ACT into NSW the same way the Region section does
    # so the totals reconcile.
    state_filter = (request.args.get("state") or "").strip().upper()
    _STATE_FILTER_EXPAND = {
        "NSW": ("NSW", "ACT"),
        "QLD": ("QLD",),
        "VIC": ("VIC", "SA", "TAS"),
        "WA":  ("WA",  "NT"),
    }
    state_filter_members = _STATE_FILTER_EXPAND.get(state_filter) if state_filter else None

    # ── Period resolution ────────────────────────────────────────
    # Supports both single-month (YYYY-MM) and half-year (YYYY-H1 /
    # YYYY-H2) selectors.  Range mode aggregates sales across the
    # months and disables the "vs 3M avg" comparison (it has no
    # natural meaning when the period itself is multi-month).
    is_range = False
    period_label = ""
    try:
        if "-H" in month_str.upper():
            y_s, h_s = month_str.upper().split("-H")
            year = int(y_s); half = int(h_s)
            assert half in (1, 2)
            if half == 1:
                this_months = [(year,     m) for m in range(1, 7)]
                prev_months = [(year - 1, m) for m in range(7, 13)]
                py_months   = [(year - 1, m) for m in range(1, 7)]
                period_label = f"{year} 1st Half (Jan–Jun)"
            else:
                this_months = [(year,     m) for m in range(7, 13)]
                prev_months = [(year,     m) for m in range(1, 7)]
                py_months   = [(year - 1, m) for m in range(7, 13)]
                period_label = f"{year} 2nd Half (Jul–Dec)"
            is_range = True
            # Single-month back-compat (some helpers still want a
            # single (year, month) for things like the day-column
            # probe).  Point them at the last month of the period.
            year, month = this_months[-1]
            prev_year, prev_month = prev_months[-1]
            py_year,   py_month   = py_months[-1]
        else:
            y_s, m_s = month_str.split("-")
            year, month = int(y_s), int(m_s)
            assert 1 <= month <= 12
            if month == 1:
                prev_year, prev_month = year - 1, 12
            else:
                prev_year, prev_month = year, month - 1
            py_year, py_month = year - 1, month
            this_months = [(year, month)]
            prev_months = [(prev_year, prev_month)]
            py_months   = [(py_year, py_month)]
            period_label = f"{year}-{month:02d}"
    except Exception:
        return jsonify({"error": "month must be YYYY-MM or YYYY-H1 / YYYY-H2"}), 400

    # Trailing 3-month window (M-1, M-2, M-3) for the single-month
    # path.  Skipped in range mode — "vs 3M avg" doesn't make sense
    # for a 6-month period, so the response leaves those fields None
    # and the frontend hides the column.
    def _months_back(n, anchor_y, anchor_m):
        out, y, m = [], anchor_y, anchor_m
        for _ in range(n):
            m -= 1
            if m == 0:
                m = 12; y -= 1
            out.append((y, m))
        return out
    trailing3 = [] if is_range else _months_back(3, year, month)

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        # Customer-side state filter.  When a state button is on,
        # splice an extra JOIN between `FROM {tbl} s` and the existing
        # WHERE so totals / product groups / sold-tos / promo
        # participants all narrow together.  No alias collision: the
        # customer join uses cf_state; everything else still uses `s`.
        if state_filter_members:
            placeholders = ",".join(["%s"] * len(state_filter_members))
            state_join_sql = (
                f" JOIN customer cf_state ON cf_state.ship_to = s.ship_to "
                f"AND cf_state.bde_state IN ({placeholders}) "
            )
            state_join_ps  = list(state_filter_members)
        else:
            state_join_sql = ""
            state_join_ps  = []

        def _src_with_state(y, m):
            """Per-month source + WHERE; injects the optional state
            JOIN before the existing WHERE so every helper reuses
            one filter path."""
            tbl, where, ps = _highlights_src_for(cur, y, m)
            if tbl and state_join_sql:
                where = state_join_sql + (where or "")
                ps    = list(state_join_ps) + list(ps)
            return tbl, where, ps

        # ── helpers that re-use the cursor ──────────────────────────
        def totals(y, m):
            tbl, where, ps = _src_with_state(y, m)
            if not tbl:
                return {"qty": 0.0, "amt": 0.0, "days": 0}
            # Probe for a `day` column — per-month tables (sales_2601…)
            # carry one; sales_2526 doesn't but has `billing_date` we
            # can DATE() on for the same distinct-day count.  Without
            # this second branch the daily_avg for any month served
            # from sales_2526 was 0 → None → dash in the UI.
            day_expr = None
            try:
                cur.execute(f"SHOW COLUMNS FROM {tbl} LIKE 'day'")
                if cur.fetchone():
                    day_expr = "s.day"
                else:
                    cur.execute(f"SHOW COLUMNS FROM {tbl} LIKE 'billing_date'")
                    if cur.fetchone():
                        day_expr = "DATE(s.billing_date)"
            except Exception:
                pass
            day_col = (f", COUNT(DISTINCT {day_expr}) AS days"
                       if day_expr else ", 0 AS days")
            cur.execute(
                f"SELECT COALESCE(SUM(s.qty),0) AS qty, "
                f"       COALESCE(SUM(s.amt),0) AS amt {day_col} "
                f"FROM {tbl} s {where}",
                ps,
            )
            r = cur.fetchone() or {}
            return {"qty": float(r.get("qty") or 0),
                    "amt": float(r.get("amt") or 0),
                    "days": int(r.get("days") or 0)}

        def per_region(y, m):
            tbl, where, ps = _src_with_state(y, m)
            if not tbl:
                return {}
            cur.execute(
                f"SELECT cus.bde_state AS state, "
                f"       COALESCE(SUM(s.qty),0) AS qty, "
                f"       COALESCE(SUM(s.amt),0) AS amt "
                f"FROM {tbl} s "
                f"LEFT JOIN customer cus ON cus.ship_to = s.ship_to "
                f"{where} "
                f"GROUP BY cus.bde_state",
                ps,
            )
            # Roll SA/TAS into VIC, NT into WA, ACT into NSW so the
            # numbers reconcile with the 4-region buckets used
            # everywhere else on the dashboard.
            remap = {"SA": "VIC", "TAS": "VIC", "NT": "WA", "ACT": "NSW"}
            out = {}
            for r in cur.fetchall():
                st = (r["state"] or "").strip().upper()
                if not st or st == "COMMON":
                    st = "COMMON"
                else:
                    st = remap.get(st, st)
                d = out.setdefault(st, {"qty": 0.0, "amt": 0.0})
                d["qty"] += float(r["qty"] or 0)
                d["amt"] += float(r["amt"] or 0)
            return out

        def per_product_group(y, m):
            tbl, where, ps = _src_with_state(y, m)
            if not tbl:
                return {}
            cur.execute(
                f"SELECT mat.product_group AS pg, "
                f"       COALESCE(SUM(s.qty),0) AS qty, "
                f"       COALESCE(SUM(s.amt),0) AS amt "
                f"FROM {tbl} s "
                f"LEFT JOIN carrying_26 mat ON mat.m_code = s.material "
                f"{where} "
                f"GROUP BY mat.product_group",
                ps,
            )
            out = {}
            for r in cur.fetchall():
                pg = (r["pg"] or "—").strip() or "—"
                out[pg] = {"qty": float(r["qty"] or 0),
                           "amt": float(r["amt"] or 0)}
            return out

        def per_sold_to(y, m):
            tbl, where, ps = _src_with_state(y, m)
            if not tbl:
                return {}
            cur.execute(
                f"SELECT s.sold_to AS sold_to, "
                f"       COALESCE(SUM(s.qty),0) AS qty, "
                f"       COALESCE(SUM(s.amt),0) AS amt "
                f"FROM {tbl} s "
                f"{where} "
                f"GROUP BY s.sold_to",
                ps,
            )
            return {(r["sold_to"] or ""): {"qty": float(r["qty"] or 0),
                                            "amt": float(r["amt"] or 0)}
                    for r in cur.fetchall()}

        # ── Period summing wrappers ───────────────────────────────
        # Half-year mode = sum the per-month helpers over the 6
        # months in the period.  Single-month mode = just the one
        # month (so the loops run exactly once).
        def totals_period(months):
            agg = {"qty": 0.0, "amt": 0.0, "days": 0}
            for (y, m) in months:
                t = totals(y, m)
                agg["qty"]  += t["qty"]
                agg["amt"]  += t["amt"]
                agg["days"] += t["days"]
            return agg
        def per_region_period(months):
            out = {}
            for (y, m) in months:
                for k, qd in per_region(y, m).items():
                    o = out.setdefault(k, {"qty": 0.0, "amt": 0.0})
                    o["qty"] += qd.get("qty", 0)
                    o["amt"] += qd.get("amt", 0)
            return out
        def per_pg_period(months):
            out = {}
            for (y, m) in months:
                for k, qd in per_product_group(y, m).items():
                    o = out.setdefault(k, {"qty": 0.0, "amt": 0.0})
                    o["qty"] += qd.get("qty", 0)
                    o["amt"] += qd.get("amt", 0)
            return out
        def per_sold_to_period(months):
            out = {}
            for (y, m) in months:
                for k, qd in per_sold_to(y, m).items():
                    o = out.setdefault(k, {"qty": 0.0, "amt": 0.0})
                    o["qty"] += qd.get("qty", 0)
                    o["amt"] += qd.get("amt", 0)
            return out

        # ── ① Summary ──────────────────────────────────────────────
        this_t = totals_period(this_months)
        prev_t = totals_period(prev_months)
        py_t   = totals_period(py_months)
        # 3-month trailing average — single-month mode only.
        if is_range:
            trailing3_totals = []
            avg3_v = None
        else:
            trailing3_totals = [totals(y, m) for (y, m) in trailing3]
            avg3_v = sum(t[metric] for t in trailing3_totals) / 3.0
        this_v = this_t[metric]; prev_v = prev_t[metric]; py_v = py_t[metric]
        summary = {
            "this":         this_t,
            "prev":         prev_t,
            "py":           py_t,
            "avg3":         ({"qty": sum(t["qty"] for t in trailing3_totals) / 3.0,
                              "amt": sum(t["amt"] for t in trailing3_totals) / 3.0}
                             if trailing3_totals else None),
            "mom_pct":      _highlights_pct(this_v, prev_v),
            "yoy_pct":      _highlights_pct(this_v, py_v),
            "vs_avg3_pct":  _highlights_pct(this_v, avg3_v) if avg3_v is not None else None,
            "working_days": this_t["days"],
            "daily_avg":    (this_v / this_t["days"]) if this_t["days"] > 0 else None,
        }

        # ── ② Regions ──────────────────────────────────────────────
        this_reg = per_region_period(this_months)
        prev_reg = per_region_period(prev_months)
        py_reg   = per_region_period(py_months)
        trailing3_reg = [] if is_range else [per_region(y, m) for (y, m) in trailing3]
        def _reg_avg3(r):
            if not trailing3_reg:
                return None
            return sum(d.get(r, {}).get(metric, 0) for d in trailing3_reg) / 3.0
        all_regions = set(this_reg) | set(prev_reg) | set(py_reg)
        for d in trailing3_reg:
            all_regions |= set(d)
        # COMMON is a placeholder bucket (customers with no state
        # assignment); exclude it from the region breakdown so it
        # doesn't show up in the analysis as a real geography.
        all_regions.discard("COMMON")
        # Stable order: standard 4 first, then anything unexpected.
        order = ["NSW", "QLD", "VIC", "WA"]
        ordered = [r for r in order if r in all_regions] + [r for r in all_regions if r not in order]
        regions = []
        total_metric = sum(this_reg.get(r, {}).get(metric, 0) for r in ordered) or 1
        for r in ordered:
            tv = this_reg.get(r, {}).get(metric, 0)
            pv = prev_reg.get(r, {}).get(metric, 0)
            yv = py_reg.get(r, {}).get(metric, 0)
            av = _reg_avg3(r)
            regions.append({
                "region":  r,
                "this":    round(tv, 2),
                "prev":    round(pv, 2),
                "py":      round(yv, 2),
                "avg3":    round(av, 2) if av is not None else None,
                "mom_pct": _highlights_pct(tv, pv),
                "yoy_pct": _highlights_pct(tv, yv),
                "vs_avg3_pct": _highlights_pct(tv, av) if av is not None else None,
                "share":   round(tv / total_metric * 100, 1),
            })

        # ── ③ Product Groups (Top 10 by this-month metric) ─────────
        this_pg = per_pg_period(this_months)
        prev_pg = per_pg_period(prev_months)
        py_pg   = per_pg_period(py_months)
        trailing3_pg = [] if is_range else [per_product_group(y, m) for (y, m) in trailing3]
        def _pg_avg3(pg):
            if not trailing3_pg:
                return None
            return sum(d.get(pg, {}).get(metric, 0) for d in trailing3_pg) / 3.0
        pg_total = sum(d.get(metric, 0) for d in this_pg.values()) or 1
        pg_rows = []
        for pg, d in this_pg.items():
            tv = d.get(metric, 0)
            pv = prev_pg.get(pg, {}).get(metric, 0)
            yv = py_pg.get(pg, {}).get(metric, 0)
            av = _pg_avg3(pg)
            pg_rows.append({
                "product_group": pg,
                "this":   round(tv, 2),
                "prev":   round(pv, 2),
                "py":     round(yv, 2),
                "avg3":   round(av, 2) if av is not None else None,
                "mom_pct": _highlights_pct(tv, pv),
                "yoy_pct": _highlights_pct(tv, yv),
                "vs_avg3_pct": _highlights_pct(tv, av) if av is not None else None,
                "share":   round(tv / pg_total * 100, 1),
            })
        pg_rows.sort(key=lambda r: -(r["this"] or 0))
        product_groups = pg_rows[:10]

        # ── ④ Promotion ─ MoM per promo + Promo vs Non-Promo share ─
        # Per request, restricted to the three real promos the team
        # actively analyses (443 / iON / TrueBlue).  443_beforeTrueblue
        # is a backfill artefact for pre-TrueBlue 2025 months and is
        # excluded from the highlights view.
        ANALYZED_PROMOS = ["443", "iON", "TrueBlue"]
        promotions = []
        promo_aggregate = None   # {promo: X, non_promo: Y, share, ...}
        this_tbl, this_where, this_ps = _src_with_state(year, month)
        if this_tbl:
            ym_year_lit  = year
            ym_month_lit = month

            def _participants_for(promo_name, src_tbl, src_where, src_ps,
                                  yy, mm):
                """Sold_tos qualifying for `promo_name` in the given month
                source.  Same match logic as the original ④ query."""
                try:
                    cur.execute(
                        f"""SELECT DISTINCT s.sold_to AS sold_to
                            FROM {src_tbl} s
                            LEFT JOIN carrying_26 mat ON mat.m_code = s.material
                            LEFT JOIN customer    cus ON cus.ship_to = s.ship_to
                            {src_where + (' AND ' if src_where else 'WHERE ')}
                              mat.line = 'PCLT'
                              AND EXISTS (
                                SELECT 1 FROM promo_customer pc
                                LEFT JOIN promo_plan pp ON pp.promo = pc.promo
                                WHERE pc.promo = %s
                                  AND s.qty     >= pc.min_qty
                                  AND s.dc_rate  BETWEEN pc.dc_rate_start AND pc.dc_rate_end
                                  AND s.brand    = pc.brand
                                  AND (
                                    (pc.sold_to <> '' AND pc.sold_to = s.sold_to)
                                    OR (pc.sold_to = '' AND pc.customer_group = cus.sold_to_group)
                                  )
                                  AND (
                                    pp.promo IS NULL
                                    OR (
                                      (pp.product_group = '' OR mat.product_group = pp.product_group)
                                      AND (pp.start_date IS NULL
                                           OR ({yy} * 100 + {mm}) >=
                                              (YEAR(pp.start_date)*100 + MONTH(pp.start_date)))
                                      AND (pp.end_date IS NULL
                                           OR ({yy} * 100 + {mm}) <=
                                              (YEAR(pp.end_date)*100 + MONTH(pp.end_date)))
                                      AND (pp.material = '' OR pp.material = s.material)
                                    )
                                  )
                              )
                        """,
                        tuple(src_ps) + (promo_name,)
                    )
                    return {r["sold_to"] for r in cur.fetchall() if r["sold_to"]}
                except Exception:
                    return set()

            # Per-sold_to qty for the whole period (single month or H1).
            tm = per_sold_to_period(this_months)
            pm = per_sold_to_period(prev_months)
            all_this = sum((d.get(metric, 0) for d in tm.values()))
            all_prev = sum((d.get(metric, 0) for d in pm.values()))

            # Resolve participants across every month in this_months /
            # prev_months — same shop can be in different months; we
            # take the union so a shop qualifying any month counts.
            def _participants_over(months, promo_name):
                parts = set()
                for (y, m) in months:
                    src = _src_with_state(y, m)
                    if not src or not src[0]:
                        continue
                    parts |= _participants_for(
                        promo_name, src[0], src[1], src[2], y, m)
                return parts

            # Trailing 3 (single-month only).
            trailing3_pst = [] if is_range else [per_sold_to(y, m) for (y, m) in trailing3]
            trailing3_src = [] if is_range else [_src_with_state(y, m)
                                                 for (y, m) in trailing3]
            trailing3_all = [sum(d.get(metric, 0) for d in pst.values())
                             for pst in trailing3_pst]
            avg3_all = (sum(trailing3_all) / 3.0) if trailing3_all else None

            promo_participants_this = set()
            promo_participants_prev = set()
            trailing3_promo_part = [set() for _ in trailing3]

            for promo_name in ANALYZED_PROMOS:
                participants_this = _participants_over(this_months, promo_name)
                participants_prev = _participants_over(prev_months, promo_name)

                promo_participants_this |= participants_this
                promo_participants_prev |= participants_prev

                p_this = sum(tm.get(st, {}).get(metric, 0) for st in participants_this)
                p_prev = sum(pm.get(st, {}).get(metric, 0) for st in participants_prev)

                # Trailing 3M (single-month mode only).
                p_avg3 = None
                if not is_range:
                    p_trailing = []
                    for idx, (y, m) in enumerate(trailing3):
                        src = trailing3_src[idx]
                        if not src or not src[0]:
                            p_trailing.append(0.0); continue
                        parts = _participants_for(
                            promo_name, src[0], src[1], src[2], y, m)
                        trailing3_promo_part[idx] |= parts
                        pst = trailing3_pst[idx]
                        p_trailing.append(
                            sum(pst.get(st, {}).get(metric, 0) for st in parts))
                    p_avg3 = sum(p_trailing) / 3.0

                # 443 / TrueBlue are always-on; iON kicks in only some
                # months.  Hide the row when there's literally nothing
                # to analyse (no participants either side, no qty).
                if (promo_name == "iON"
                        and not participants_this and not participants_prev
                        and p_this == 0 and p_prev == 0
                        and (p_avg3 in (None, 0))):
                    continue
                promotions.append({
                    "promo":        promo_name,
                    "participants": len(participants_this),
                    "total_shops":  len({k for k in tm.keys() if k}),
                    "this":         round(p_this, 2),
                    "prev":         round(p_prev, 2),
                    "avg3":         round(p_avg3, 2) if p_avg3 is not None else None,
                    "mom_pct":      _highlights_pct(p_this, p_prev),
                    "vs_avg3_pct":  _highlights_pct(p_this, p_avg3) if p_avg3 is not None else None,
                })

            promo_this = sum(tm.get(st, {}).get(metric, 0) for st in promo_participants_this)
            promo_prev = sum(pm.get(st, {}).get(metric, 0) for st in promo_participants_prev)
            non_promo_this = max(all_this - promo_this, 0)
            non_promo_prev = max(all_prev - promo_prev, 0)
            if avg3_all is not None:
                trailing3_promo_qty = [
                    sum(trailing3_pst[i].get(st, {}).get(metric, 0)
                        for st in trailing3_promo_part[i])
                    for i in range(3)
                ]
                avg3_promo     = sum(trailing3_promo_qty) / 3.0
                avg3_non_promo = max(avg3_all - avg3_promo, 0)
            else:
                avg3_promo = None
                avg3_non_promo = None

            def _share(n, d):
                return round(n / d * 100, 1) if d else None

            promo_aggregate = {
                "total_this":     round(all_this, 2),
                "total_prev":     round(all_prev, 2),
                "total_avg3":     round(avg3_all, 2) if avg3_all is not None else None,
                "promo_this":     round(promo_this, 2),
                "promo_prev":     round(promo_prev, 2),
                "promo_avg3":     round(avg3_promo, 2) if avg3_promo is not None else None,
                "non_promo_this": round(non_promo_this, 2),
                "non_promo_prev": round(non_promo_prev, 2),
                "non_promo_avg3": round(avg3_non_promo, 2) if avg3_non_promo is not None else None,
                "promo_mom_pct":          _highlights_pct(promo_this, promo_prev),
                "promo_vs_avg3_pct":      _highlights_pct(promo_this, avg3_promo) if avg3_promo is not None else None,
                "non_promo_mom_pct":      _highlights_pct(non_promo_this, non_promo_prev),
                "non_promo_vs_avg3_pct":  _highlights_pct(non_promo_this, avg3_non_promo) if avg3_non_promo is not None else None,
                "promo_share_this":      _share(promo_this,     all_this),
                "promo_share_prev":      _share(promo_prev,     all_prev),
                "promo_share_avg3":      _share(avg3_promo,     avg3_all) if avg3_promo is not None else None,
                "non_promo_share_this":  _share(non_promo_this, all_this),
                "non_promo_share_prev":  _share(non_promo_prev, all_prev),
                "non_promo_share_avg3":  _share(avg3_non_promo, avg3_all) if avg3_non_promo is not None else None,
            }

            # Per-promo share within total promotion (this / prev / avg3).
            for p in promotions:
                p["share_of_promo_this"] = _share(p["this"], promo_this)
                p["share_of_promo_prev"] = _share(p["prev"], promo_prev)
                p["share_of_promo_avg3"] = (_share(p["avg3"], avg3_promo)
                                            if avg3_promo is not None and p.get("avg3") is not None
                                            else None)
                p["share_of_total_this"] = _share(p["this"], all_this)
                p["share_of_total_prev"] = _share(p["prev"], all_prev)
                p["share_of_total_avg3"] = (_share(p["avg3"], avg3_all)
                                            if avg3_all is not None and p.get("avg3") is not None
                                            else None)

        # ── ⑤ Sold-to movement ────────────────────────────────────
        # Use period-aware sums so half-year mode compares H1 vs prior
        # half across the full sold-to set.
        tm = per_sold_to_period(this_months)
        pm = per_sold_to_period(prev_months)

        # Map sold_to → name via customer master (one query, all sold_tos).
        name_map = {}
        candidates = set(tm.keys()) | set(pm.keys())
        if candidates:
            ph = ",".join(["%s"] * len(candidates))
            try:
                cur.execute(
                    f"SELECT sold_to, MIN(NULLIF(TRIM(sold_to_name),'')) AS name "
                    f"FROM customer WHERE sold_to IN ({ph}) "
                    f"GROUP BY sold_to",
                    tuple(candidates),
                )
                name_map = {r["sold_to"]: (r["name"] or r["sold_to"]) for r in cur.fetchall()}
            except Exception:
                pass

        # Compute deltas across the full set, then take the top 30 each side.
        rows = []
        for st in candidates:
            if not st:
                continue
            tv = tm.get(st, {}).get(metric, 0)
            pv = pm.get(st, {}).get(metric, 0)
            rows.append({
                "sold_to":  st,
                "name":     name_map.get(st, st),
                "this":     round(tv, 2),
                "prev":     round(pv, 2),
                "delta":    round(tv - pv, 2),
                "delta_pct": _highlights_pct(tv, pv),
            })
        gainers = sorted([r for r in rows if (r["delta"] or 0) > 0],
                         key=lambda r: -r["delta"])[:30]
        losers  = sorted([r for r in rows if (r["delta"] or 0) < 0],
                         key=lambda r: r["delta"])[:30]

        # Sold-to vs 3M-avg variant: per-sold_to 3-month average over
        # M-1 / M-2 / M-3.  Skipped in half-year mode — the prev half
        # block already provides the "prior period" comparison.
        trailing3_sold  = [] if is_range else [per_sold_to(y, m) for (y, m) in trailing3]
        gainers_vs_avg3 = []
        losers_vs_avg3  = []
        if trailing3_sold:
            rows_avg3 = []
            candidates_a = set()
            for d in trailing3_sold:
                candidates_a |= set(d.keys())
            candidates_a |= set(tm.keys())
            for st in candidates_a:
                if not st:
                    continue
                tv = tm.get(st, {}).get(metric, 0)
                avg3 = sum(d.get(st, {}).get(metric, 0) for d in trailing3_sold) / 3.0
                rows_avg3.append({
                    "sold_to":  st,
                    "name":     name_map.get(st, st),
                    "this":     round(tv, 2),
                    "avg3":     round(avg3, 2),
                    "delta":    round(tv - avg3, 2),
                    "delta_pct": _highlights_pct(tv, avg3),
                })
            gainers_vs_avg3 = sorted(
                [r for r in rows_avg3 if (r["delta"] or 0) > 0],
                key=lambda r: -r["delta"])[:30]
            losers_vs_avg3 = sorted(
                [r for r in rows_avg3 if (r["delta"] or 0) < 0],
                key=lambda r: r["delta"])[:30]

        # Newly Active = no sales in last 3 months, positive this month.
        # Newly Silent = regular buyer that just dropped to zero —
        # qty > 100 in EACH of the trailing 3 months AND zero this
        # month.  Threshold prevents one-off shops from polluting
        # the list; we want to flag steady customers that suddenly
        # stopped.
        NEWLY_SILENT_MIN_QTY = 100
        newly_active = []
        newly_silent = []
        try:
            # Per-sold_to qty in each of the trailing 3 months.
            # trailing3_sold was already collected above for the
            # vs-3M-avg gainers/losers; reuse it.
            three_back_sold = set()
            for d in trailing3_sold:
                for st, qd in d.items():
                    if st and (qd.get(metric, 0) or 0) > 0:
                        three_back_sold.add(st)
            this_sold = {st for st, d in tm.items()
                         if st and (d.get(metric, 0) or 0) > 0}
            for st in (this_sold - three_back_sold):
                newly_active.append({
                    "sold_to": st,
                    "name":    name_map.get(st, st),
                    "this":    round(tm.get(st, {}).get(metric, 0), 2),
                })

            # Newly silent: must have hit > MIN_QTY in EVERY trailing
            # month, then zero this month.
            this_qty = {st: (d.get(metric, 0) or 0) for st, d in tm.items() if st}
            steady_candidates = set(trailing3_sold[0].keys()) if trailing3_sold else set()
            for d in trailing3_sold[1:]:
                steady_candidates &= set(d.keys())
            for st in steady_candidates:
                if not st:
                    continue
                qs = [trailing3_sold[i].get(st, {}).get(metric, 0) or 0
                      for i in range(3)]
                if all(q > NEWLY_SILENT_MIN_QTY for q in qs) and (this_qty.get(st, 0) or 0) == 0:
                    newly_silent.append({
                        "sold_to":   st,
                        "name":      name_map.get(st, st),
                        "prev":      round(qs[0], 2),
                        "trailing3": [round(q, 2) for q in qs],
                        "avg3":      round(sum(qs) / 3.0, 2),
                    })
            newly_active.sort(key=lambda r: -(r["this"] or 0))
            newly_silent.sort(key=lambda r: -(r.get("avg3", 0) or 0))
        except Exception:
            pass

        # ── ⑥ Decline analysis ───────────────────────────────────────
        # Even when the total ROSE this month, the reader still wants
        # to see which slices fell — and when the total fell, which
        # slices drove the drop.  This section synthesises the region
        # / product-group / customer breakdowns we already computed,
        # ranks negative contributors by absolute change, and pushes
        # a cross-attribution (top-declining region × product groups
        # that fell inside it) so a "VIC dropped 8%" line can be
        # followed by "concentrated in Ventus (-800) and Kinergy
        # (-400)".  Uses the same metric (qty | amt) the rest of the
        # page runs on.
        def _decl(items, key):
            """From a list of dicts with `this` and `prev`, return the
            negative ones sorted by most-negative delta."""
            out = []
            for r in items:
                d = (r.get("this") or 0) - (r.get("prev") or 0)
                if d < 0:
                    out.append({key: r.get(key), "this": r.get("this"),
                                "prev": r.get("prev"), "change": round(d, 2),
                                "pct": r.get("mom_pct")})
            out.sort(key=lambda x: x["change"])   # most negative first
            return out
        # COMMON isn't a real BDE region — it's the catch-all for
        # customers without a state assignment (HQ / cross-border /
        # unresolved).  Excluding it from every analysis path so
        # reports never call out "COMMON dropped 60%" as if it were
        # actionable.
        region_declines = [r for r in _decl(regions, "region")
                           if r.get("region") != "COMMON"]
        pg_declines     = _decl(pg_rows, "product_group")   # full pg list, not the top-10 slice
        # Customer losses: reuse `losers` (already sorted by most-negative
        # delta) — trim to 5 for the narrative and expose the delta
        # under a common `change` key so the frontend can use one
        # renderer for every declines row.
        customer_losses = [
            {"sold_to": r["sold_to"], "name": r["name"],
             "this": r["this"], "prev": r["prev"],
             "change": r["delta"], "pct": r["delta_pct"]}
            for r in losers[:5]
        ]
        # Cross-attribution: for the top-3 declining regions, pull the
        # product groups that fell WITHIN that region so the user can
        # see WHY the region dropped, not just that it did.  Uses the
        # per-region-per-pg cross-table we build lazily here.
        cross = []
        try:
            def per_pg_by_region(y, m):
                tbl, where, ps = _src_with_state(y, m)
                if not tbl: return {}
                cur.execute(
                    f"SELECT COALESCE(cus.bde_state,'COMMON') AS state, "
                    f"       IFNULL(cr.product_group,'(none)') AS pg, "
                    f"       SUM(s.qty) AS qty, SUM(s.amt) AS amt "
                    f"FROM {tbl} s "
                    f"LEFT JOIN customer cus ON cus.ship_to = s.ship_to "
                    f"LEFT JOIN carrying_26 cr ON cr.m_code = s.material "
                    f"{where} "
                    f"GROUP BY state, pg",
                    ps,
                )
                out = {}
                for r in cur.fetchall():
                    st = (r["state"] or "COMMON").upper()
                    # Fold state → region same as the Region section
                    reg = ({"SA":"VIC","TAS":"VIC","NT":"WA","ACT":"NSW"}
                           .get(st, st))
                    out.setdefault(reg, {})[r["pg"] or "(none)"] = (
                        float(r["qty"] or 0) if metric == "qty"
                        else float(r["amt"] or 0)
                    )
                return out
            this_pg_by_reg = {}
            prev_pg_by_reg = {}
            for (y, m) in this_months:
                for reg, pgs in per_pg_by_region(y, m).items():
                    for pg, v in pgs.items():
                        this_pg_by_reg.setdefault(reg, {})
                        this_pg_by_reg[reg][pg] = this_pg_by_reg[reg].get(pg, 0) + v
            for (y, m) in prev_months:
                for reg, pgs in per_pg_by_region(y, m).items():
                    for pg, v in pgs.items():
                        prev_pg_by_reg.setdefault(reg, {})
                        prev_pg_by_reg[reg][pg] = prev_pg_by_reg[reg].get(pg, 0) + v
            for r in region_declines[:3]:
                reg = r["region"]
                pg_deltas = []
                seen = set(this_pg_by_reg.get(reg, {}).keys()) | set(prev_pg_by_reg.get(reg, {}).keys())
                for pg in seen:
                    tv = this_pg_by_reg.get(reg, {}).get(pg, 0)
                    pv = prev_pg_by_reg.get(reg, {}).get(pg, 0)
                    if tv - pv < 0:
                        pg_deltas.append({"pg": pg, "change": round(tv - pv, 2)})
                pg_deltas.sort(key=lambda x: x["change"])
                cross.append({"region": reg, "region_change": r["change"],
                              "product_declines": pg_deltas[:5]})
        except Exception:
            pass   # cross-attribution is a nice-to-have; skip on failure

        overall_change = (this_v or 0) - (prev_v or 0)
        declines = {
            "overall": {
                "direction": "down" if overall_change < 0
                             else ("up" if overall_change > 0 else "flat"),
                "change":    round(overall_change, 2),
                "pct":       summary["mom_pct"],
                "this":      round(this_v or 0, 2),
                "prev":      round(prev_v or 0, 2),
            },
            "regions":         region_declines,
            "product_groups":  pg_declines[:10],
            "customers":       customer_losses,
            "cross":           cross,
        }

        # ═══════════════════════════════════════════════════════════════
        #  Strategic insights layer (trajectory / mix shift / narrative /
        #  watch list).  Single-month mode only — the range view already
        #  aggregates half-years so a trailing 6-month regression on top
        #  of that would be meaningless.  Every block is wrapped so a
        #  failure never breaks the KPI page.
        # ═══════════════════════════════════════════════════════════════
        trajectory = None
        mix_shift  = None
        narrative  = []
        watch_list = None
        if not is_range:
            # ── ⑦ Trajectory: 6-month series + linear-fit slope ────────
            try:
                back6 = list(reversed(_months_back(6, year, month)))  # oldest → newest
                series_months = back6 + [(year, month)]
                series = []
                for (y, m) in series_months:
                    t = totals(y, m)
                    series.append({
                        "month": f"{y}-{m:02d}",
                        "value": round(t[metric], 2),
                    })
                vals = [s["value"] for s in series]
                # Simple linear regression: y = a·x + b where x is 0..n-1.
                # slope in "units per month" is `a`; intercept isn't
                # interesting.  Handles zero-variance rows gracefully.
                n = len(vals)
                if n >= 3 and sum(vals) > 0:
                    xs = list(range(n))
                    mean_x = sum(xs) / n
                    mean_y = sum(vals) / n
                    num = sum((xs[i] - mean_x) * (vals[i] - mean_y) for i in range(n))
                    den = sum((xs[i] - mean_x) ** 2 for i in range(n))
                    slope = num / den if den else 0
                    # Slope-to-latest-value ratio gives a normalised
                    # % / month growth trajectory the narrative can
                    # quote instead of raw units.
                    slope_pct = (slope / mean_y * 100) if mean_y else 0
                else:
                    slope = 0
                    slope_pct = 0
                # Run pattern — direction of each MoM step, then longest
                # consecutive run at the tail (positive or negative).
                run_pattern = []
                for i in range(1, n):
                    d = vals[i] - vals[i-1]
                    run_pattern.append("up" if d > 0 else "down" if d < 0 else "flat")
                # Consecutive-from-end tally
                cons_up = cons_down = 0
                for step in reversed(run_pattern):
                    if step == "up" and cons_down == 0:
                        cons_up += 1
                    elif step == "down" and cons_up == 0:
                        cons_down += 1
                    else:
                        break
                # Direction verdict — is the trajectory accelerating or
                # decelerating?  Compare the last-3-months mean to the
                # first-3-months mean; positive gap = accelerating.
                if n >= 6:
                    first_avg = sum(vals[:3]) / 3.0
                    last_avg  = sum(vals[-3:]) / 3.0
                    momentum  = last_avg - first_avg
                else:
                    momentum = 0
                if abs(slope_pct) < 1.0:
                    direction = "stable"
                elif slope > 0 and momentum > 0:
                    direction = "accelerating growth"
                elif slope > 0 and momentum <= 0:
                    direction = "growing but slowing"
                elif slope < 0 and momentum < 0:
                    direction = "accelerating decline"
                else:
                    direction = "declining but stabilising"
                trajectory = {
                    "series":            series,
                    "slope_units":       round(slope, 1),
                    "slope_pct":         round(slope_pct, 2),
                    "direction":         direction,
                    "run_pattern":       run_pattern,
                    "consecutive_up":    cons_up,
                    "consecutive_down":  cons_down,
                }
            except Exception:
                trajectory = None

            # ── ⑧ Mix shift: product-group + rim-family + brand share ──
            # Compare this month's shares against 6 months ago so
            # slow-moving mix drift becomes visible.
            try:
                anchor6 = _months_back(6, year, month)[-1]   # T-6 point
                anchor6_pg = per_pg_period([anchor6])
                # Shares now
                total_now  = sum(d.get(metric, 0) for d in this_pg.values())
                total_prev = sum(d.get(metric, 0) for d in anchor6_pg.values())
                pg_mix = []
                all_pgs = set(this_pg) | set(anchor6_pg)
                for pg in all_pgs:
                    now_v  = this_pg.get(pg, {}).get(metric, 0)
                    prev_v_pg = anchor6_pg.get(pg, {}).get(metric, 0)
                    if total_now == 0 and total_prev == 0:
                        continue
                    sh_now  = (now_v / total_now * 100)  if total_now  else 0
                    sh_prev = (prev_v_pg / total_prev * 100) if total_prev else 0
                    pg_mix.append({
                        "product_group": pg,
                        "share_now":     round(sh_now, 1),
                        "share_prev":    round(sh_prev, 1),
                        "share_change":  round(sh_now - sh_prev, 1),
                    })
                pg_mix.sort(key=lambda r: -abs(r["share_change"]))
                pg_mix = pg_mix[:6]
                # Rim family mix — quick-and-dirty regex on carrying size
                import re as _re_r
                def _rim_shares(months):
                    """Group units by rim family for the given month list.
                    Runs one query per month to keep the join simple."""
                    agg = {}
                    for (yy, mm) in months:
                        tbl, where, ps = _src_with_state(yy, mm)
                        if not tbl: continue
                        cur.execute(
                            f"SELECT c.size AS sz, "
                            f"       SUM(s.{metric}) AS q "
                            f"FROM {tbl} s "
                            f"JOIN carrying_26 c ON c.m_code = s.material "
                            f"{where} "
                            f"GROUP BY c.size", ps,
                        )
                        for r in cur.fetchall():
                            sz = r.get("sz") or ""
                            mm2 = _re_r.search(r"R\s*(\d{2}(?:\.\d)?)", sz, _re_r.I)
                            if not mm2: continue
                            inch = float(mm2.group(1))
                            if inch == 17.5: fam = "R17.5(TBR)"
                            elif inch == 19.5: fam = "R19.5(TBR)"
                            elif inch == 22.5: fam = "R22.5(TBR)"
                            elif inch >= 22:   fam = "R22+"
                            else:              fam = f"R{int(inch)}"
                            agg[fam] = agg.get(fam, 0) + float(r.get("q") or 0)
                    return agg
                rim_now  = _rim_shares(this_months)
                rim_prev = _rim_shares([anchor6])
                rim_total_now  = sum(rim_now.values())
                rim_total_prev = sum(rim_prev.values())
                rim_mix = []
                for rim in set(rim_now) | set(rim_prev):
                    sh_n = (rim_now.get(rim, 0)  / rim_total_now  * 100) if rim_total_now  else 0
                    sh_p = (rim_prev.get(rim, 0) / rim_total_prev * 100) if rim_total_prev else 0
                    rim_mix.append({
                        "rim":          rim,
                        "share_now":    round(sh_n, 1),
                        "share_prev":   round(sh_p, 1),
                        "share_change": round(sh_n - sh_p, 1),
                    })
                rim_mix.sort(key=lambda r: -abs(r["share_change"]))
                rim_mix = rim_mix[:6]
                mix_shift = {
                    "anchor_prev":    f"{anchor6[0]}-{anchor6[1]:02d}",
                    "anchor_now":     f"{year}-{month:02d}",
                    "product_groups": pg_mix,
                    "rim_family":     rim_mix,
                }
            except Exception:
                mix_shift = None

            # ── ⑨ Narrative: 3-sentence auto-generated summary ─────────
            try:
                lines = []
                # Sentence 1 — overall movement, framed against trajectory
                if trajectory and trajectory["direction"]:
                    lines.append(
                        f"This month totalled {int(this_v):,} units, "
                        f"{_narrate_change(summary['mom_pct'])} MoM — "
                        f"{trajectory['direction']} on a 6-month view "
                        f"(slope {trajectory['slope_pct']:+.1f}%/month)."
                    )
                else:
                    lines.append(
                        f"This month totalled {int(this_v):,} units, "
                        f"{_narrate_change(summary['mom_pct'])} MoM."
                    )
                # Sentence 2 — dominant driver (product + biggest region)
                bits = []
                if product_groups:
                    top_pg_up = next((p for p in product_groups if (p.get("vs_avg3_pct") or 0) > 0), None)
                    top_pg_dn = next((p for p in product_groups if (p.get("vs_avg3_pct") or 0) < 0), None)
                    if top_pg_up and abs(top_pg_up.get("vs_avg3_pct") or 0) >= 10:
                        bits.append(
                            f"{top_pg_up['product_group']} surged "
                            f"{top_pg_up['vs_avg3_pct']:+.1f}% vs 3M avg "
                            f"({top_pg_up['share']:.1f}% share)"
                        )
                    if top_pg_dn and abs(top_pg_dn.get("vs_avg3_pct") or 0) >= 10:
                        bits.append(
                            f"{top_pg_dn['product_group']} pulled back "
                            f"{top_pg_dn['vs_avg3_pct']:+.1f}%"
                        )
                if bits:
                    lines.append(f"Driver: {'; '.join(bits)}.")
                # Sentence 3 — mix-shift or regional counter-movement
                if mix_shift and mix_shift.get("product_groups"):
                    top_mix = mix_shift["product_groups"][0]
                    if abs(top_mix["share_change"]) >= 1.5:
                        arrow = "up" if top_mix["share_change"] > 0 else "down"
                        lines.append(
                            f"Mix is shifting: {top_mix['product_group']} share "
                            f"{arrow} {abs(top_mix['share_change']):.1f} points "
                            f"vs six months ago ({top_mix['share_prev']:.1f}% → "
                            f"{top_mix['share_now']:.1f}%)."
                        )
                narrative = lines
            except Exception:
                narrative = []

            # ── ⑩ Watch list: actionable to-dos ─────────────────────────
            try:
                call = []
                escalate = []
                # Newly silent big customers — already computed; pick
                # the top-5 by trailing-3-avg (most valuable to recover).
                for n_r in (newly_silent or [])[:5]:
                    call.append({
                        "reason": f"Newly silent — avg {int(n_r.get('avg3') or 0):,}/mo",
                        "sold_to": n_r.get("sold_to"),
                        "name":    n_r.get("name"),
                    })
                # Concentration escalation — top-10 share this month
                # vs 6 months ago.  Watch >3pt jumps.
                if trajectory:
                    try:
                        # Top-10 share this month
                        top10_this = sorted(
                            [(k, v.get(metric, 0)) for k, v in tm.items()],
                            key=lambda kv: -kv[1])[:10]
                        top10_share_this = (sum(v for _, v in top10_this) / (this_v or 1) * 100)
                        # 6 months ago
                        anchor_tm = per_sold_to(anchor6[0], anchor6[1])
                        anchor_total = sum(v.get(metric, 0) for v in anchor_tm.values())
                        top10_anchor = sorted(
                            [(k, v.get(metric, 0)) for k, v in anchor_tm.items()],
                            key=lambda kv: -kv[1])[:10]
                        top10_share_prev = (sum(v for _, v in top10_anchor)
                                            / (anchor_total or 1) * 100)
                        share_jump = top10_share_this - top10_share_prev
                        if share_jump >= 3.0:
                            escalate.append({
                                "reason": "Concentration risk",
                                "detail": f"Top-10 customers = {top10_share_this:.1f}% "
                                          f"(was {top10_share_prev:.1f}% six months ago, "
                                          f"+{share_jump:.1f}pt)",
                            })
                    except Exception:
                        pass
                # Region-level escalation — regions that dropped >5% MoM
                for r in region_declines[:2]:
                    if r.get("pct") is not None and r["pct"] <= -5:
                        escalate.append({
                            "reason": f"{r['region']} pulled back",
                            "detail": f"MoM {r['pct']:+.1f}% ({int(r['change']):,}) — "
                                      f"tour with the state manager",
                        })
                # Product-level escalation — top-line pg dropping vs 3M avg
                for p in product_groups[:5]:
                    if (p.get("vs_avg3_pct") or 0) <= -15:
                        escalate.append({
                            "reason": f"{p['product_group']} weakening",
                            "detail": f"{p['vs_avg3_pct']:+.1f}% vs 3M avg — "
                                      f"check inventory & promo alignment",
                        })
                watch_list = {"call": call, "escalate": escalate}
            except Exception:
                watch_list = None

        cur.close(); conn.close()
        return jsonify({
            "month":          period_label,
            "period_label":   period_label,
            "is_range":       is_range,
            "metric":         metric,
            "summary":        summary,
            "regions":        regions,
            "product_groups": product_groups,
            "promotions":     promotions,
            "promo_aggregate": promo_aggregate,
            "declines":       declines,
            "trajectory":     trajectory,
            "mix_shift":      mix_shift,
            "narrative":      narrative,
            "watch_list":     watch_list,
            "sold_to": {
                "gainers":          gainers,
                "losers":           losers,
                "gainers_vs_avg3":  gainers_vs_avg3,
                "losers_vs_avg3":   losers_vs_avg3,
                "newly_active":     newly_active,
                "newly_silent":     newly_silent,
                "notes": {
                    "gainers_limit": 30,
                    "losers_limit":  30,
                    "newly_active":  "Sold-to with zero sales in the previous 3 months AND positive sales this month.",
                    "newly_silent":  "Steady customers that just stopped buying — qty > 100 in EACH of the trailing 3 months AND zero sales this month.",
                }
            },
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# Sales source table per ?month=... button on the rebate page.  A whitelist —
# the resolved value is the only thing interpolated into the FROM clause, so an
# unknown month can never inject a table name.
REBATE_SALES_TABLES = {
    "thismonth": "sales_thismonth",
}

def _rebate_sales_table(month_arg):
    """Resolve a Month-button key to the physical sales table.
    - "thismonth"            → sales_thismonth
    - 4-digit YYMM (e.g.2606) → sales_2606 (if the table exists)
    - unknown / missing       → sales_thismonth (last-resort fallback)

    Existence check runs once per key and caches so the front-end
    stays snappy across button clicks; a table that appears later
    (end-of-month archive job) is picked up on the next process
    restart."""
    key = (month_arg or "thismonth").strip().lower()
    hit = REBATE_SALES_TABLES.get(key)
    if hit:
        return hit
    # YYMM pattern like "2606" → try sales_2606.  Verify the table
    # exists so we don't blow up on an as-yet-unloaded month.
    if len(key) == 4 and key.isdigit():
        candidate = f"sales_{key}"
        try:
            conn = get_connection(); cur = conn.cursor()
            cur.execute("SHOW TABLES LIKE %s", (candidate,))
            exists = cur.fetchone() is not None
            cur.close(); conn.close()
        except Exception:
            exists = False
        if exists:
            REBATE_SALES_TABLES[key] = candidate
            return candidate
        # Table not there yet — cache the miss so we don't re-probe
        # every request in the same process.
        REBATE_SALES_TABLES[key] = "__MISSING__"
        return "__MISSING__"
    if hit == "__MISSING__":
        return "__MISSING__"
    return "sales_thismonth"


_REBATE_REGION_MAP = {"SA": "VIC", "TAS": "VIC", "NT": "VIC", "ACT": "NSW"}

def _rebate_region(state):
    """Region used on the rebate view.  The 4 region cards are NSW/QLD/VIC/WA;
    the smaller states/territories roll into one of them (SA/TAS/NT -> VIC,
    ACT -> NSW) so their sales/rebate don't fall outside every card."""
    s = (state or "").strip()
    return _REBATE_REGION_MAP.get(s.upper(), s)

def _or_default(val, dflt):
    """val unless it's blank or the '-' placeholder, in which case dflt.  Used
    so ship_tos missing from the customer master (no region/BDE) fall back to
    the sold_to's region/BDE instead of becoming an orphan '-' that drops out
    of the 4 region cards."""
    v = (val or "").strip()
    return v if v and v != "-" else dflt


@app.get("/api/rebate_structure_check")
def api_rebate_structure_check():
    """Diagnostic: what rebate structure is a sold_to mapped to, and how
    does the rebate calc bucket its ship_tos by BDE?  Hit like:
      /api/rebate_structure_check?sold_to=731942
    Returns:
      structures           – every (brand, structure_name) row from
                              rebate_customer_map for this sold_to
      structure_path       – 'PER_SHIP_TO' (AJT/ABJ/ATP/APP/ACD) or
                              'BDE_GROUPED' (everything else)
      bde_breakdown        – per-BDE counts:
                              owned_in_customer  – customer.salesman_name
                                                    rows for this sold_to
                              with_sales_thismonth – of those, how many
                                                    have a sales_thismonth
                                                    row this period
    """
    sold = (request.args.get("sold_to") or "").strip()
    if not sold:
        return jsonify({"error": "sold_to is required"}), 400
    sales_tbl = _rebate_sales_table(request.args.get("month"))
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        cur.execute(
            "SELECT brand, structure_name FROM rebate_customer_map "
            "WHERE sold_to = %s", (sold,))
        structures = cur.fetchall()

        # Per-ship_to if ANY mapped structure carries the _SR suffix.
        path = "BDE_GROUPED"
        for s in structures:
            if _rebate_is_ship_to(s.get("structure_name") or ""):
                path = "PER_SHIP_TO"; break
        # Annotate each structure with its own calc basis (mixed SR + HQ/VR
        # accounts like 731942 carry both).
        for s in structures:
            s["calc_basis"] = ("SHIP_TO" if _rebate_is_ship_to(s.get("structure_name") or "")
                               else "SOLD_TO")
            s["scope"] = _rebate_scope(s.get("structure_name") or "")

        cur.execute(
            "SELECT TRIM(c.salesman_name) AS bde, "
            "       COUNT(DISTINCT c.ship_to) AS owned_in_customer, "
            "       COUNT(DISTINCT CASE WHEN EXISTS("
            "           SELECT 1 FROM " + sales_tbl + " s "
            "           WHERE s.ship_to = c.ship_to "
            "             AND s.brand IN ('HK','LF') "
            "             AND s.so_type IN " + _REBATE_SO_TYPES_IN + " "
            "       ) THEN c.ship_to END) AS with_sales_thismonth "
            "FROM customer c "
            "WHERE c.sold_to = %s "
            "GROUP BY TRIM(c.salesman_name) "
            "ORDER BY owned_in_customer DESC",
            (sold,))
        breakdown = cur.fetchall()

        # Cross-check from the sales-side: which ship_tos in
        # sales_thismonth carry this sold_to, and who do they map back to
        # via customer.salesman_name (this is exactly what the rebate
        # BDE-grouping loop sees).
        cur.execute(
            "SELECT TRIM(c.salesman_name) AS bde, "
            "       COUNT(DISTINCT s.ship_to) AS sales_ship_tos "
            "FROM " + sales_tbl + " s "
            "LEFT JOIN customer c ON c.ship_to = s.ship_to "
            "WHERE s.sold_to = %s "
            "  AND s.brand IN ('HK','LF') "
            "  AND s.so_type IN " + _REBATE_SO_TYPES_IN + " "
            "GROUP BY TRIM(c.salesman_name) "
            "ORDER BY sales_ship_tos DESC",
            (sold,))
        sales_side = cur.fetchall()

        cur.close(); conn.close()
        return jsonify({
            "sold_to":           sold,
            "structures":        structures,
            "structure_path":    path,
            "bde_breakdown":     breakdown,
            "sales_side_bde":    sales_side,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/rebate_data")
def api_rebate_data():
    """
    Return rebate status per SHIP_TO 횞 territory ??server-side paginated.
    Query params:
      territory     (TTL | HK | LF | ALL, default ALL)
      sold_to_group (default ALL)
      search        free-text filter on name / code
      show          ALL | NEXT | MAX | ZERO
      sort          actual | est_rebate | needed | ship_to_name (default: actual)
      dir           desc | asc (default desc)
      page          0-indexed page of sold_to groups (default 0)
      page_size     groups per page (default 40)
    unit=A ??measure in $ amount
    unit=Q ??measure in qty
    brand=TTL ??sum HK + LF sales
    """
    brand_filter  = request.args.get("territory",     "ALL").upper()  # UI still sends 'territory'
    stg_filter    = request.args.get("sold_to_group", "ALL")
    region_filter = request.args.get("region",        "ALL").upper()
    sales_tbl     = _rebate_sales_table(request.args.get("month"))  # source table per month button
    # Month table not present in the DB yet (e.g. Jun clicked before the
    # end-of-month archive job runs) → return an empty payload with a
    # friendly note so the UI just shows "no data for this month" rather
    # than showing the same numbers as This Month.
    if sales_tbl == "__MISSING__":
        return jsonify({
            "rows": [], "total": {}, "page": 0, "page_size": 0,
            "note": ("Sales table for the requested month has not been "
                     "archived yet.  It becomes available after the "
                     "end-of-month archive job runs."),
        })
    export_fmt    = (request.args.get("export") or "").strip().lower()  # 'xlsx' -> download
    # BDE filter — matches the role-scope lock used by graph/map views.
    # Comparison is case-insensitive on salesman_name so name casing in
    # the customer master doesn't make a BDE invisible to themselves.
    bde_filter    = (request.args.get("bde") or "").strip()
    line_filter   = (request.args.get("line") or "ALL").upper()        # PCLT | TBR | ALL
    sold_to_filter = (request.args.get("sold_to") or "").strip()       # exact name or code
    ship_to_filter = (request.args.get("ship_to") or "").strip()
    search      = request.args.get("search",  "").strip().lower()
    show        = request.args.get("show",    "ALL").upper()
    sort_col    = request.args.get("sort",    "actual")
    sort_dir    = request.args.get("dir",     "desc")
    try:
        page      = int(request.args.get("page",      0))
        page_size = int(request.args.get("page_size", 40))
    except ValueError:
        page, page_size = 0, 40

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # ?? 1. Rebate-mapped customers (sold_to level) ???????????????????????
        cur.execute("""
            SELECT m.sold_to, m.brand, m.structure_name,
                   MIN(c.sold_to_name)  AS sold_to_name,
                   MIN(c.sold_to_group) AS sold_to_group
            FROM rebate_customer_map m
            LEFT JOIN customer c ON c.sold_to = m.sold_to
            WHERE (%s = 'ALL' OR m.brand = %s)
              AND (%s = 'ALL' OR m.line = %s)
              AND (%s = 'ALL' OR c.sold_to_group = %s)
            GROUP BY m.sold_to, m.brand, m.structure_name
        """, (brand_filter, brand_filter, line_filter, line_filter, stg_filter, stg_filter))
        customers = cur.fetchall()
        if not customers:
            return jsonify([])

        # Ship_to -> Group map (for _SR group summing).  The map lists
        # individual ship_tos with a manual Group number for SR structures.
        # ship_tos sharing a Group are summed and the tier is applied to the
        # group total; every member then displays that same group figure
        # (even when members sit under different BDEs).
        ship_grp = {}   # (sold_to, structure_name, ship_to) -> group label
        cur.execute("""
            SELECT sold_to, structure_name, ship_to, grp
            FROM rebate_customer_map
            WHERE grp IS NOT NULL AND grp <> '' AND ship_to IS NOT NULL AND ship_to <> ''
        """)
        for r in cur.fetchall():
            ship_grp[(str(r["sold_to"]), r["structure_name"], str(r["ship_to"]))] = str(r["grp"])

        # Sold_to-level grouping: rows with grp filled and ship_to EMPTY bundle
        # multiple sold_tos together for tier qualification on non-_SR
        # structures (e.g. HK_ALL_DIA_A_VR with 3 sold_tos sharing group "1").
        # tier basis = sum of all group members; each sold_to/store still earns
        # on its own amount x the resulting rate, so totals are sum-correct.
        sold_to_grp  = {}   # (structure_name, sold_to) -> grp label
        grp_to_solds = {}   # (structure_name, grp)     -> set of sold_tos
        cur.execute("""
            SELECT sold_to, structure_name, grp
            FROM rebate_customer_map
            WHERE grp IS NOT NULL AND grp <> ''
              AND (ship_to IS NULL OR ship_to = '')
        """)
        for r in cur.fetchall():
            k = (r["structure_name"], str(r["sold_to"]))
            sold_to_grp[k] = str(r["grp"])
            grp_to_solds.setdefault((r["structure_name"], str(r["grp"])), set()).add(str(r["sold_to"]))

        # sold_tos that carry a Store rebate (_SR) structure.  For these, the
        # secondary HQ/VR/IT/WTY rebates are pulled out of the per-BDE/State
        # table and combined into the top summary box instead.  sold_tos with
        # NO _SR structure keep their HQ/VR in the table (BDE/State as before).
        sr_sold_tos = {str(c["sold_to"]) for c in customers
                       if _rebate_is_ship_to(c["structure_name"])}
        hq_box = {}   # sold_to -> {sold_to, sold_to_name, sold_to_group, rebate, parts:[...]}
        hq_by_region = {}   # region -> HQ/VR rebate (each store's amount x rate, by its region)

        # Pre-load the set of "Promo" ship_tos so the rebate calc can
        # skip them in every aggregation downstream — the customer
        # master flags promotional stores by including "Promo" in the
        # ship_to_name (case-insensitive substring), and those don't
        # count toward rebate qualification or earning.
        cur.execute(
            "SELECT ship_to FROM customer "
            "WHERE UPPER(ship_to_name) LIKE '%PROMO%'"
        )
        promo_ship_tos = {str(r["ship_to"]) for r in cur.fetchall()}

        # ?? 2. Sales from sales_thismonth by (sold_to, ship_to, brand, line) ??
        # Brand comes from sales_thismonth.brand directly (SAP loader fills it).
        # Line is derived from material prefix: 1xxx/2xxx → PCLT, 3xxx → TBR.
        # No carrying join needed.
        cur.execute("""
            SELECT s.sold_to, s.ship_to, s.brand AS brand,
                   CASE
                     WHEN LEFT(s.material,1) IN ('1','2') THEN 'PCLT'
                     WHEN LEFT(s.material,1) = '3'        THEN 'TBR'
                     ELSE ''
                   END AS line,
                   SUM(s.qty) AS qty, SUM(s.amt) AS amt
            FROM """ + sales_tbl + """ s
            WHERE s.brand IN ('HK','LF')
              AND s.so_type IN """ + _REBATE_SO_TYPES_IN + """
            GROUP BY s.sold_to, s.ship_to, s.brand,
                     CASE
                       WHEN LEFT(s.material,1) IN ('1','2') THEN 'PCLT'
                       WHEN LEFT(s.material,1) = '3'        THEN 'TBR'
                       ELSE ''
                     END
        """)
        ship_sales      = {}   # (sold_to, ship_to, brand) -> {qty, amt}  all lines
        ship_sales_line = {}   # (sold_to, ship_to, brand, line) -> {qty, amt}
        ship_idx        = {}   # (sold_to, brand) -> set{ship_to}
        ship_idx_line   = {}   # (sold_to, brand, line) -> set{ship_to}
        sold_brand_tot      = {}   # (sold_to, brand) -> {qty, amt}  per sold_to
        sold_brand_line_tot = {}   # (sold_to, brand, line) -> {qty, amt}
        for r in cur.fetchall():
            st, sh, br, ln = str(r["sold_to"]), str(r["ship_to"]), r["brand"], r["line"]
            if sh in promo_ship_tos:
                continue   # Promo ship_tos are excluded from rebate calc.
            qty, amt = float(r["qty"] or 0), float(r["amt"] or 0)
            # aggregate all lines ??brand-level totals
            agg = ship_sales.setdefault((st, sh, br), {"qty": 0.0, "amt": 0.0})
            agg["qty"] += qty; agg["amt"] += amt
            # store by line
            ship_sales_line[(st, sh, br, ln)] = {"qty": qty, "amt": amt}
            ship_idx.setdefault((st, br), set()).add(sh)
            ship_idx_line.setdefault((st, br, ln), set()).add(sh)
            # per-sold_to roll-ups (for sold_to-group tier basis)
            agg2 = sold_brand_tot.setdefault((st, br), {"qty": 0.0, "amt": 0.0})
            agg2["qty"] += qty; agg2["amt"] += amt
            agg3 = sold_brand_line_tot.setdefault((st, br, ln), {"qty": 0.0, "amt": 0.0})
            agg3["qty"] += qty; agg3["amt"] += amt

        # ?? 3. Customer lookup (ship_to ??name, bde_state, salesman) ??????????
        cur.execute("SELECT ship_to, ship_to_name, bde_state, salesman_name, sold_to FROM customer")
        ship_cust_map  = {}   # ship_to str -> {name, state, bde}
        sold_to_ships  = {}   # sold_to str -> set(ship_to) from customer master.
                              # Lets the rebate calc surface a BDE row even when
                              # all of that BDE's ship_tos for a sold_to have
                              # zero sales this month (otherwise they're hidden
                              # because ship_idx is sales-based).
        for r in cur.fetchall():
            sh = str(r["ship_to"])
            ship_cust_map[sh] = {
                "name":  r["ship_to_name"] or sh,
                "state": _rebate_region(r["bde_state"]) or "-",
                "bde":   (r["salesman_name"] or "").strip() or "-",
            }
            if sh in promo_ship_tos:
                continue   # Don't surface Promo ships as zero-sales BDE rows.
            stk = str(r["sold_to"] or "")
            if stk:
                sold_to_ships.setdefault(stk, set()).add(sh)
        name_map = {sh: v["name"] for sh, v in ship_cust_map.items()}

        # Build BDE ??region mapping: for each BDE, use the most common state among
        # their ship_tos in ship_cust_map (actual BDE region, not sold_to's bde_state)
        from collections import Counter
        _bde_state_counter = {}
        for v in ship_cust_map.values():
            bde = v.get("bde", "-")
            st  = v.get("state", "-")
            if bde and bde != "-" and st and st != "-":
                _bde_state_counter.setdefault(bde, Counter())[st] += 1
        bde_region_map = {bde: cnt.most_common(1)[0][0]
                          for bde, cnt in _bde_state_counter.items()}

        # ?? 4. Tier definitions (only meaningful tiers: tier_order <= top_order) ?
        cur.execute("""
            SELECT structure_name, unit, tier_order, top_order, threshold, rate
            FROM rebate_structure
            WHERE tier_order <= top_order
            ORDER BY structure_name, tier_order
        """)
        tiers_map = {}
        for r in cur.fetchall():
            sd = tiers_map.setdefault(r["structure_name"],
                                      {"unit": r["unit"], "top_order": int(r["top_order"]), "tiers": []})
            sd["tiers"].append({"tier":      int(r["tier_order"]),
                                 "threshold": float(r["threshold"]),
                                 "rate":      float(r["rate"])})

        # ?? 5. Build result ??one row per SHIP_TO ????????????????????????????
        def _calc_tier(actual, tiers, top_order):
            """
            Returns (curr_tier, next_tier).
            next_tier is None if actual has reached top_order tier.
            tiers list is already trimmed to top_order entries.
            """
            curr = {"tier": 0, "threshold": 0, "rate": 0}
            nxt  = None
            for t in tiers:
                if t["threshold"] <= actual and t["rate"] > 0:
                    curr = t
                elif t["threshold"] > actual and t["rate"] > 0 and nxt is None:
                    nxt = t
            # If curr tier has reached top_order, there is no next tier
            if curr["tier"] >= top_order:
                nxt = None
            return curr, nxt

        # Calc basis is driven by the structure-name suffix (see _rebate_*
        # helpers above): _SR -> per ship_to; _HQ/_VR/_IT/_WTY -> per sold_to
        # on the full sold_to total; no suffix -> per sold_to (BDE-grouped).

        def _rebate_brand_line(struct_name):
            """From the structure name tokens decide which sales to count:
              token0 brand : HK | LF | ALL(=HK+LF)
              token1 line  : PCLT | TBR | ALL(=all lines)
            A 'NOLFT' token anywhere in the rest of the name flags an
            exclusion ('all selected sales MINUS LF/TBR') for the rare
            case where the basis should ignore that one (brand, line)
            cell.  Other tokens stay free-form (used by scope suffixes).
            Returns (brands_list, line_filter_or_None, brand_key_for_display, excludes).
            """
            tk = struct_name.split("_")
            bt = (tk[0] if len(tk) > 0 else "ALL").upper()
            lt = (tk[1] if len(tk) > 1 else "ALL").upper()
            brands    = ["HK", "LF"] if bt == "ALL" else [bt]
            line_filt = lt if lt in ("PCLT", "TBR") else None
            brand_key = "TTL" if bt == "ALL" else bt
            excludes  = set()
            if "NOLFT" in [t.upper() for t in tk[2:]]:
                excludes.add(("LF", "TBR"))
            return brands, line_filt, brand_key, excludes

        def _sold_to_qa(st_id, brand_list, line, excludes=()):
            """(qty, amt) for one sold_to summed over the given brands/line,
            subtracting any (brand, line) cells flagged as excluded."""
            q = a = 0.0
            for br in brand_list:
                if line:
                    d = sold_brand_line_tot.get((st_id, br, line), {"qty": 0.0, "amt": 0.0})
                else:
                    d = sold_brand_tot.get((st_id, br), {"qty": 0.0, "amt": 0.0})
                q += d["qty"]; a += d["amt"]
                for (eb, el) in excludes:
                    if eb != br: continue
                    if line == el:
                        # Whole bucket already matches the excluded cell.
                        q -= d["qty"]; a -= d["amt"]
                    elif line is None:
                        de = sold_brand_line_tot.get((st_id, eb, el), {"qty": 0.0, "amt": 0.0})
                        q -= de["qty"]; a -= de["amt"]
            return q, a


        rows = []
        for c in customers:
            struct        = c["structure_name"]
            brand         = c["brand"]
            sold_to       = str(c["sold_to"])
            sold_to_group = c["sold_to_group"] or "-"
            sd            = tiers_map.get(struct)
            if not sd:
                continue
            unit      = sd["unit"]       # A=Amount, Q=Qty
            tiers     = sd["tiers"]
            top_order = sd["top_order"]
            scope     = _rebate_scope(struct)            # SR | HQ | VR | IT | WTY | BASE
            is_ship_to_struct = (scope == "SR")
            is_secondary      = scope in ("HQ", "VR", "IT", "WTY")

            for brands, line_filt, brand_key, excludes in [_rebate_brand_line(struct)]:

                # ship_tos with sales for this structure's brand(s) and line.
                # line_filt PCLT/TBR -> only that line (material prefix 1,2=PCLT
                # 3=TBR); None -> all lines (brand-level totals).
                ship_set = set()
                for _br in brands:
                    if line_filt:
                        ship_set |= ship_idx_line.get((sold_to, _br, line_filt), set())
                    else:
                        ship_set |= ship_idx.get((sold_to, _br), set())

                # Keep every ship_to that has sales for this sold_to, even if it
                # isn't in the customer master (e.g. A018377 — qty 0 but an
                # amount, or credits/returns) — those still count toward the
                # rebate.  Region/BDE fall back to the sold_to's when unknown.

                if not ship_set:
                    ship_set.add(sold_to)   # show zero row so sold_to is visible

                # Determine sold_to's canonical BDE and region.
                # Start from the self-referencing record (ship_to == sold_to code),
                # then fall back to individual ship_tos if BDE/region is still missing.
                st_info = ship_cust_map.get(sold_to, {})
                sold_to_bde    = (st_info.get("bde",   "-") or "-") if st_info else "-"
                sold_to_region = (st_info.get("state", "-") or "-") if st_info else "-"

                # If sold_to's own record is missing BDE or region, infer from
                # the actual ship_tos (customer table joined via ship_to).
                if sold_to_bde == "-" or sold_to_region == "-":
                    real_ships = [sh for sh in ship_set if sh != sold_to]
                    state_cnt = Counter(
                        ship_cust_map[sh].get("state", "") for sh in real_ships
                        if ship_cust_map.get(sh, {}).get("state", "")
                        and ship_cust_map[sh]["state"] != "-"
                    )
                    bde_cnt = Counter(
                        ship_cust_map[sh].get("bde", "") for sh in real_ships
                        if ship_cust_map.get(sh, {}).get("bde", "")
                        and ship_cust_map[sh]["bde"] != "-"
                    )
                    if sold_to_bde == "-":
                        sold_to_bde    = bde_cnt.most_common(1)[0][0]   if bde_cnt   else "-"
                    if sold_to_region == "-":
                        sold_to_region = state_cnt.most_common(1)[0][0] if state_cnt else "-"

                # Badge labels for UI: brand, then line (PCLT/TBR) if filtered,
                # then unit; secondary scopes (HQ/VR/IT/WTY) get a scope tag.
                badges = [brand_key] + ([line_filt] if line_filt else []) + [unit]
                if is_secondary:
                    badges = badges + [scope]

                def _get_sales(sh, _brands=brands, _line=line_filt, _excl=excludes):
                    """(qty, amt) summed over this structure's brand(s)/line,
                    minus any (brand, line) cells flagged as excluded."""
                    q = a = 0.0
                    for _br in _brands:
                        if _line:
                            d = ship_sales_line.get((sold_to, sh, _br, _line), {"qty": 0.0, "amt": 0.0})
                        else:
                            d = ship_sales.get((sold_to, sh, _br), {"qty": 0.0, "amt": 0.0})
                        q += d["qty"]; a += d["amt"]
                        for (eb, el) in _excl:
                            if eb != _br: continue
                            if _line == el:
                                q -= d["qty"]; a -= d["amt"]
                            elif _line is None:
                                de = ship_sales_line.get((sold_to, sh, eb, el), {"qty": 0.0, "amt": 0.0})
                                q -= de["qty"]; a -= de["amt"]
                    return q, a

                if is_ship_to_struct:
                    # _SR → per ship_to, but ship_tos sharing a map Group are
                    # summed and the tier applied to the group total.  Every
                    # member then displays that same group figure (qty/amt/
                    # rate/rebate) while staying under its own BDE.  A ship_to
                    # with no Group is its own singleton.
                    grp_totals = {}   # group_key -> {qty, amt, members:[...]}
                    for sh in sorted(ship_set):
                        q, a = _get_sales(sh)
                        sh_info_local = ship_cust_map.get(sh, {})
                        gk = ship_grp.get((sold_to, struct, sh))
                        group_key = ("G", gk) if gk else ("S", sh)
                        gt = grp_totals.setdefault(group_key, {"qty": 0.0, "amt": 0.0, "members": []})
                        gt["qty"] += q
                        gt["amt"] += a
                        gt["members"].append({
                            "sh":     sh,
                            "bde":    _or_default(sh_info_local.get("bde"),   sold_to_bde),
                            "region": _or_default(sh_info_local.get("state"), sold_to_region),
                        })
                    calc_items = []
                    for group_key, gt in grp_totals.items():
                        members = gt["members"]
                        # Count the group's rebate ONCE toward region/BDE totals,
                        # attributed to the BDE that owns the most members; the
                        # other members show the same figure but are display-only
                        # so a cross-BDE group isn't summed twice.
                        dom_bde = Counter(m["bde"] for m in members).most_common(1)[0][0]
                        primary_taken = False
                        for m in members:
                            is_primary = (not primary_taken) and (m["bde"] == dom_bde)
                            if is_primary:
                                primary_taken = True
                            calc_items.append({
                                "sh":            m["sh"],
                                "qty":           gt["qty"],   # group total → same number for all members
                                "amt":           gt["amt"],
                                "bde":           m["bde"],
                                "region":        m["region"],
                                "sold_to_basis": False,
                                "ship_details":  [],
                                "rollup":        is_primary,
                            })
                elif is_secondary and sold_to in sr_sold_tos:
                    # _HQ/_VR/_IT/_WTY for an account that ALSO has a Store
                    # rebate (_SR): compute one rebate on the FULL sold_to total
                    # and fold it into the top summary box (combined across HQ/
                    # VR/etc).  No per-BDE/State table rows for these.
                    full_ship_set = ship_set | sold_to_ships.get(sold_to, set())
                    tot_q = tot_a = 0.0
                    store_amts = []   # (region, amt) per store, to split the HQ/VR by region
                    for sh in full_ship_set:
                        q, a = _get_sales(sh)
                        tot_q += q; tot_a += a
                        reg = _or_default(ship_cust_map.get(sh, {}).get("state"), sold_to_region)
                        store_amts.append((reg, a))
                    # Tier qualifies on the sold_to GROUP total when this
                    # sold_to is bundled with others via the map's Group column;
                    # otherwise on this sold_to's own total.  Each sold_to still
                    # earns on its own amount x the resulting rate.
                    grp_label = sold_to_grp.get((struct, sold_to))
                    if grp_label:
                        bq = ba = 0.0
                        for _gs in grp_to_solds.get((struct, grp_label), {sold_to}):
                            _q, _a = _sold_to_qa(_gs, brands, line_filt, excludes)
                            bq += _q; ba += _a
                    else:
                        bq, ba = tot_q, tot_a
                    hq_actual = bq if unit == "Q" else ba
                    hq_curr, _ = _calc_tier(hq_actual, tiers, top_order)
                    hq_rebate = round(tot_a * hq_curr["rate"] / 100, 2)
                    for reg, a in store_amts:   # split this HQ/VR across regions by store sales
                        hq_by_region[reg] = hq_by_region.get(reg, 0.0) + a * hq_curr["rate"] / 100
                    box = hq_box.setdefault(sold_to, {
                        "sold_to":       sold_to,
                        "sold_to_name":  c["sold_to_name"] or st_info.get("name") or sold_to,
                        "sold_to_group": sold_to_group,
                        "rebate":        0.0,
                        "parts":         [],
                    })
                    box["rebate"] += hq_rebate
                    box["parts"].append({
                        "scope":          scope,
                        "brand":          brand_key,
                        "structure_name": struct,
                        "rate":           hq_curr["rate"],
                        "actual_qty":     round(tot_q, 2),
                        "actual_amt":     round(tot_a, 2),
                        "rebate":         hq_rebate,
                    })
                    calc_items = []   # nothing in the per-BDE/State table
                else:
                    # sold_to-level rebate (base, or HQ/VR of a non-Store-rebate
                    # account).  The tier QUALIFIES on the full sold_to total, but
                    # each Store earns the rebate on its OWN sales at that rate —
                    # so every store shows its own qty/amt/rebate (not one shared
                    # group number).  No rollup gating needed: each store is
                    # counted once with its own amount, summing to the correct
                    # sold_to total.
                    full_ship_set = ship_set | sold_to_ships.get(sold_to, set())
                    full_q = full_a = 0.0
                    per_store = []
                    for sh in sorted(full_ship_set):
                        q, a = _get_sales(sh)
                        full_q += q; full_a += a
                        info = ship_cust_map.get(sh, {})
                        per_store.append({
                            "sh":     sh,
                            "qty":    q,
                            "amt":    a,
                            "bde":    _or_default(info.get("bde"),   sold_to_bde),
                            "region": _or_default(info.get("state"), sold_to_region),
                        })
                    # tier basis = sold_to GROUP total if grouped, else own total
                    grp_label = sold_to_grp.get((struct, sold_to))
                    if grp_label:
                        bq = ba = 0.0
                        for _gs in grp_to_solds.get((struct, grp_label), {sold_to}):
                            _q, _a = _sold_to_qa(_gs, brands, line_filt, excludes)
                            bq += _q; ba += _a
                    else:
                        bq, ba = full_q, full_a
                    calc_items = [{
                        "sh":            st["sh"],
                        "qty":           st["qty"],
                        "amt":           st["amt"],
                        "bde":           st["bde"],
                        "region":        st["region"],
                        "sold_to_basis": False,
                        "ship_details":  [],
                        "tier_basis_q":  bq,   # tier qualifies on the (group) total
                        "tier_basis_a":  ba,
                        "rollup":        True,
                    } for st in per_store]
                    if not calc_items:
                        calc_items = [{
                            "sh":           sold_to,
                            "qty":          0.0,
                            "amt":          0.0,
                            "bde":          sold_to_bde,
                            "region":       sold_to_region,
                            "sold_to_basis": False,
                            "ship_details": [],
                            "tier_basis_q": 0.0,
                            "tier_basis_a": 0.0,
                            "rollup":       True,
                        }]

                for item in calc_items:
                    sh          = item["sh"]
                    actual_qty  = item["qty"]
                    actual_amt  = item["amt"]
                    row_bde     = item["bde"]
                    row_region  = item["region"]
                    row_stbasis = item["sold_to_basis"]
                    row_details = item["ship_details"]
                    row_rollup  = item.get("rollup", True)

                    actual = actual_qty if unit == "Q" else actual_amt

                    # The tier (rate / next / needed) qualifies on the basis —
                    # the group/sold_to total for sold_to-level rebates — while
                    # the rebate is earned on THIS row's own amount.  Rows that
                    # don't set a basis (SR group members, box) use their own
                    # value, so the rate matches what they display.
                    basis_qty = item.get("tier_basis_q", actual_qty)
                    basis_amt = item.get("tier_basis_a", actual_amt)
                    basis = basis_qty if unit == "Q" else basis_amt

                    curr_tier, next_tier = _calc_tier(basis, tiers, top_order)
                    curr_rebate = round(actual_amt * curr_tier["rate"] / 100, 2)
                    est_rebate  = round(next_tier["threshold"] * next_tier["rate"] / 100, 2) if next_tier else None
                    needed_qty = round(next_tier["threshold"] - basis_qty, 2) if next_tier and unit == "Q" else None
                    needed_amt = round(next_tier["threshold"] - basis_amt, 2) if next_tier and unit == "A" else None

                    sh_info = ship_cust_map.get(sh, {})
                    rows.append({
                        "sold_to":        sold_to,
                        "sold_to_name":   c["sold_to_name"] or st_info.get("name") or sold_to,
                        "sold_to_group":  sold_to_group,
                        "region":         row_region,
                        "bde":            row_bde,
                        "ship_to":        sh,
                        "ship_to_name":   sh_info.get("name") or (c["sold_to_name"] or sh),
                        "brand":          brand_key,
                        "badges":         badges,
                        "structure_name": struct,
                        "scope":          scope,
                        "rollup":         row_rollup,
                        "sold_to_basis":  row_stbasis,
                        "ship_details":   row_details,
                        "unit":           unit,
                        "actual_qty":     round(actual_qty, 2),
                        "actual_amt":     round(actual_amt, 2),
                        "actual":         round(actual, 2),
                        "tier_actual":    round(basis, 2),
                        "curr_rate":      curr_tier["rate"],
                        "curr_threshold": curr_tier["threshold"],
                        "next_threshold": next_tier["threshold"] if next_tier else None,
                        "next_rate":      next_tier["rate"]      if next_tier else None,
                        "needed_qty":     needed_qty,
                        "needed_amt":     needed_amt,
                        "curr_rebate":    curr_rebate,
                        "est_rebate":     est_rebate,
                        "tiers":          tiers,
                    })

        # ?? 6. Client-side-style filters applied server-side ?????????????????
        # Role-scope BDE filter applied first so the summary cards (region
        # totals etc.) only reflect what this BDE actually owns.  Region
        # buttons still let them switch between regions, but they only
        # ever see their own slice.
        if bde_filter:
            bf = bde_filter.strip().upper()
            rows = [r for r in rows if (r["bde"] or "").strip().upper() == bf]
        if sold_to_filter and sold_to_filter.upper() != "ALL":
            sf = sold_to_filter.upper()
            rows = [r for r in rows if (r["sold_to_name"] or "").strip().upper() == sf
                    or str(r["sold_to"]).upper() == sf]
        if ship_to_filter and ship_to_filter.upper() != "ALL":
            shf = ship_to_filter.upper()
            rows = [r for r in rows if (r["ship_to_name"] or "").strip().upper() == shf
                    or str(r["ship_to"]).upper() == shf]
        if search:
            rows = [r for r in rows if
                    search in r["sold_to_name"].lower() or
                    search in str(r["sold_to"]) or
                    search in r["ship_to_name"].lower() or
                    search in str(r["ship_to"])]
        if show == "NEXT":
            rows = [r for r in rows if r["next_rate"] is not None and r["actual"] > 0]
        elif show == "MAX":
            rows = [r for r in rows if r["next_rate"] is None and r["curr_rate"] > 0]
        elif show == "ZERO":
            rows = [r for r in rows if r["actual"] == 0]

        # Top-box HQ/VR rebate (Store-rebate accounts only).  Respects the
        # search filter (sold_to name/code) but NOT region/BDE — HQ ignores
        # State/BDE by design.
        hq_items = sorted(hq_box.values(), key=lambda x: x["rebate"], reverse=True)
        if search:
            hq_items = [h for h in hq_items
                        if search in (h["sold_to_name"] or "").lower()
                        or search in str(h["sold_to"])]
        hq_box_out = {
            "total": round(sum(h["rebate"] for h in hq_items), 2),
            "items": [{**h, "rebate": round(h["rebate"], 2)} for h in hq_items],
        }

        # ?? 7. Summary stats (over all filtered rows) ?????????????????????????
        REGION_KEYS = ["NSW", "QLD", "VIC", "WA"]
        region_totals = {rk: {"rebate": 0.0, "qty": 0.0, "amt": 0.0} for rk in REGION_KEYS}
        for r in rows:
            if not r.get("rollup", True):
                continue   # display-only _SR group duplicate
            rk = (r["region"] or "").strip().upper()
            if rk in region_totals:
                region_totals[rk]["rebate"] += r["curr_rebate"]
                region_totals[rk]["qty"]    += r["actual_qty"]
                region_totals[rk]["amt"]    += r["actual_amt"]
        for rk in region_totals:
            region_totals[rk] = {k: round(v, 2) for k, v in region_totals[rk].items()}

        # Store-rebate (table) total, HQ/VR total, and the grand total split by
        # region: region card = Store/base rebate in that region + the HQ/VR
        # allocated to it (each store's amount x rate).  The 4 region cards sum
        # to the grand total.
        store_total = round(sum(rt["rebate"] for rt in region_totals.values()), 2)
        hq_total    = hq_box_out["total"]
        grand_total = round(store_total + hq_total, 2)
        region_grand = {rk: round(region_totals[rk]["rebate"] + hq_by_region.get(rk, 0.0), 2)
                        for rk in REGION_KEYS}

        summary = {
            "total_ship_to": len(rows),
            "has_next":  sum(1 for r in rows if r["next_rate"] is not None and r["actual"] > 0),
            "max_tier":  sum(1 for r in rows if r["next_rate"] is None and r["curr_rate"] > 0),
            "zero_sales": sum(1 for r in rows if r["actual"] == 0),
            "est_total":  round(sum(r["curr_rebate"] for r in rows if r.get("rollup", True)), 2),
            "region_totals": region_totals,
            "region_grand":  region_grand,
            "hq_by_region":  {rk: round(hq_by_region.get(rk, 0.0), 2) for rk in REGION_KEYS},
            "store_total":   store_total,
            "hq_total":      hq_total,
            "grand_total":   grand_total,
            "hq_box": hq_box_out,
        }

        # Apply region filter to rows (after computing region_totals)
        if region_filter != "ALL":
            rows = [r for r in rows if (r["region"] or "").strip().upper() == region_filter]

        # Excel download of exactly what the current filters produced (reuses
        # the same computed rows + HQ/VR box, so it matches the screen).
        if export_fmt == "xlsx":
            from io import BytesIO
            from datetime import date
            from flask import Response
            wb = Workbook()
            ws = wb.active; ws.title = "Rebate"
            hdr = ["Region", "BDE", "Sold-To", "Sold-To Name", "Ship-To",
                   "Ship-To Name", "Brand", "Type", "Structure", "Qty",
                   "Amount", "Rate %", "Next %", "Need Qty", "Need Amt",
                   "Rebate"]
            ws.append(hdr)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E3A5F")
            for r in sorted(rows, key=lambda x: (x["region"] or "", x["bde"] or "",
                                                 x["sold_to_name"] or "", -x["actual"])):
                ws.append([
                    r["region"], r["bde"], r["sold_to"], r["sold_to_name"],
                    r["ship_to"], r["ship_to_name"], r["brand"],
                    "Store" if r["scope"] == "SR" else r["scope"],
                    r["structure_name"], r["actual_qty"], r["actual_amt"],
                    r["curr_rate"],
                    r["next_rate"]   if r["next_rate"]   is not None else "",
                    r["needed_qty"]  if r["needed_qty"]  is not None else "",
                    r["needed_amt"]  if r["needed_amt"]  is not None else "",
                    r["curr_rebate"],
                ])
            # Sum-safe TOTAL: matches the screen's Grand Total exactly by
            # applying the same two rules used there:
            #   1. rollup=False rows are display-only duplicates of an _SR
            #      group, so they're excluded entirely.
            #   2. Secondary HQ/VR/IT/WTY rows cover the same underlying
            #      sales as the primary (BASE/SR) rows on the same group,
            #      so their qty/amt is excluded when a primary row exists.
            # Rebate stays a straight sum — each rebate is earned independently.
            SECONDARY_SCOPES = {"HQ", "VR", "IT", "WTY"}
            scopes_by_group = {}
            for r in rows:
                if not r.get("rollup", True): continue
                k = (r.get("region") or "", r.get("bde") or "", r.get("sold_to") or "")
                scopes_by_group.setdefault(k, set()).add(r.get("scope") or "BASE")
            tot_qty = tot_amt = tot_reb = 0.0
            for r in rows:
                if not r.get("rollup", True): continue
                k = (r.get("region") or "", r.get("bde") or "", r.get("sold_to") or "")
                has_primary = any(s not in SECONDARY_SCOPES for s in scopes_by_group.get(k, ()))
                is_secondary = (r.get("scope") or "BASE") in SECONDARY_SCOPES
                if not (has_primary and is_secondary):
                    tot_qty += r["actual_qty"]
                    tot_amt += r["actual_amt"]
                tot_reb += r["curr_rebate"]
            tot_qty = round(tot_qty, 2)
            tot_amt = round(tot_amt, 2)
            store_reb = round(tot_reb, 2)
            hq_reb    = round(hq_box_out["total"], 2)
            grand_reb = round(store_reb + hq_reb, 2)
            ws.append([])
            # Grand Total row: matches the screen's top-right GRAND TOTAL box
            # exactly (Store rebate + HQ/VR rebate).  Qty and Amount sum the
            # underlying sales the rebates were earned on.
            ws.append(["TOTAL", "", "", "", "", "", "", "", "",
                       tot_qty, tot_amt, "", "", "", "", grand_reb])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
            for label, val in (("  · Store rebate (per-store program)", store_reb),
                               ("  · HQ / VR rebate (account-level program)", hq_reb)):
                ws.append([label, "", "", "", "", "", "", "", "",
                           "", "", "", "", "", "", val])
                for cell in ws[ws.max_row]:
                    cell.font = Font(italic=True, color="64748B")

            # HQ / VR rebate (the top box) on its own sheet
            ws2 = wb.create_sheet("HQ_VR Rebate")
            ws2.append(["Sold-To", "Sold-To Name", "Group", "Rebate", "Detail"])
            for c in ws2[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="BE123C")
            for h in hq_box_out["items"]:
                detail = "; ".join(f'{p["structure_name"]} {p["rate"]}% = {p["rebate"]}'
                                   for p in h.get("parts", []))
                ws2.append([h["sold_to"], h["sold_to_name"], h.get("sold_to_group", ""),
                            h["rebate"], detail])
            ws2.append([])
            ws2.append(["TOTAL", "", "", hq_box_out["total"], ""])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            mon = (request.args.get("month") or "thismonth")
            fname = f"rebate_{mon}_{date.today().isoformat()}.xlsx"
            return Response(
                bio.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment;filename={fname}"})

        # ?? 8. Group by (region, sold_to) with brand sub-groups ??????????????
        grp_map = {}
        brand_order = {"HK": 0, "LF": 1, "TTL": 2}
        for r in rows:
            # Include BDE in the group key.  For PER_SHIP_TO structures
            # (AJT etc.) each ship_to row already carries its own BDE; the
            # frontend keys its display tree by region+bde.  Without BDE
            # here, every NSW row for JAX 731942 collapses into a single
            # group whose label is whichever row landed first — so
            # Makris's 26 ship_tos and Alessio's 21 show up under one
            # BDE only.
            key = r["region"] + "|" + r["sold_to"] + "|" + (r["bde"] or "-")
            if key not in grp_map:
                grp_map[key] = {
                    "key": key,
                    "sold_to": r["sold_to"], "sold_to_name": r["sold_to_name"],
                    "sold_to_group": r["sold_to_group"], "region": r["region"],
                    "bde": r["bde"],
                    "grp_actual": 0.0, "grp_actual_qty": 0.0, "grp_actual_amt": 0.0,
                    "grp_curr_rebate": 0.0, "grp_est": 0.0,
                    "brands": {},
                }
            g = grp_map[key]
            # rollup=False rows are display-only duplicates: members of an _SR
            # Group that aren't the group's primary BDE row.  They still render
            # (so every ship_to shows the shared group figure) but must not be
            # summed into the totals, else a multi-member group would multiply.
            # (HQ/VR of _SR accounts never reach here — they go to the top box.)
            do_rollup = r.get("rollup", True)
            if do_rollup:
                g["grp_curr_rebate"] += r["curr_rebate"]
                g["grp_est"]         += r["est_rebate"] if r["est_rebate"] else 0.0
                g["grp_actual"]      += r["actual"]
                g["grp_actual_qty"]  += r["actual_qty"]
                g["grp_actual_amt"]  += r["actual_amt"]
            # One display block per structure, so different structures that
            # share a brand_key (e.g. two TTL/HK_TBR/… or an _SR vs _HQ of the
            # same brand) never merge into one block.
            bkey = r["structure_name"]
            if bkey not in g["brands"]:
                g["brands"][bkey] = {
                    "brand": r["brand"], "unit": r["unit"],
                    "scope": r["scope"],
                    "badges": r["badges"],
                    "structure_name": r["structure_name"],
                    "grp_actual": 0.0, "grp_actual_qty": 0.0, "grp_actual_amt": 0.0,
                    "grp_curr_rebate": 0.0, "grp_est": 0.0, "items": [],
                }
            b = g["brands"][bkey]
            if do_rollup:
                b["grp_actual"]      += r["actual"]
                b["grp_actual_qty"]  += r["actual_qty"]
                b["grp_actual_amt"]  += r["actual_amt"]
                b["grp_curr_rebate"] += r["curr_rebate"]
                b["grp_est"]         += r["est_rebate"] if r["est_rebate"] else 0.0
            b["items"].append(r)

        # Convert brands dict to sorted list (HK → LF → TTL; within a brand,
        # the SR/base block first, then the HQ/VR/IT/WTY programs).
        scope_order = {"SR": 0, "BASE": 0, "HQ": 1, "VR": 2, "IT": 3, "WTY": 4}
        for g in grp_map.values():
            g["brands"] = sorted(
                g["brands"].values(),
                key=lambda b: (brand_order.get(b["brand"], 99),
                               scope_order.get(b.get("scope", "BASE"), 9)))
            # When a sold_to has both a Store/BASE rebate AND a covering
            # HQ/VR/IT/WTY rebate, the secondary structure's qty/amt is
            # the same underlying sales already counted in the base.
            # Subtract the secondary contributions so the sold_to summary
            # row shows the actual sales once (not doubled).
            secondary_scopes = {"HQ", "VR", "IT", "WTY"}
            has_primary = any(b.get("scope", "BASE") not in secondary_scopes
                              for b in g["brands"])
            if has_primary:
                for b in g["brands"]:
                    if b.get("scope", "BASE") in secondary_scopes:
                        g["grp_actual_qty"] -= b["grp_actual_qty"]
                        g["grp_actual_amt"] -= b["grp_actual_amt"]
                        g["grp_actual"]     -= b["grp_actual"]

        groups = list(grp_map.values())
        summary["total_groups"] = len(groups)

        # ?? 9. Sort groups ????????????????????????????????????????????????????
        rev = (sort_dir != "asc")
        if sort_col == "est_rebate":
            groups.sort(key=lambda g: g["grp_est"], reverse=rev)
        elif sort_col == "actual":
            groups.sort(key=lambda g: g["grp_actual"], reverse=rev)
        elif sort_col == "sold_to_name":
            groups.sort(key=lambda g: g["sold_to_name"].lower(), reverse=rev)
        else:
            groups.sort(key=lambda g: (g["region"].lower(), g["bde"].lower(), g["sold_to_name"].lower()))

        # ?? 10. Sort items within each brand sub-group by actual desc ?????????
        for g in groups:
            for b in g["brands"]:
                b["items"].sort(key=lambda r: r["actual"], reverse=True)

        # ?? 11. Paginate ??????????????????????????????????????????????????????
        total_pages = max(1, -(-len(groups) // page_size))  # ceil div
        page = max(0, min(page, total_pages - 1))
        page_groups = groups[page * page_size : (page + 1) * page_size]

        stg_list = sorted({r["sold_to_group"] for r in rows if r["sold_to_group"] != "-"})
        return jsonify({
            "summary":     summary,
            "page":        page,
            "page_size":   page_size,
            "total_pages": total_pages,
            "groups":      page_groups,
            "stg_list":    stg_list,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.get("/api/rebate_export")
def api_rebate_export():
    """Stream a CSV of all rows matching current filters (no pagination)."""
    import csv, io
    brand_filter = request.args.get("territory",     "ALL").upper()
    stg_filter   = request.args.get("sold_to_group", "ALL")
    search       = request.args.get("search",  "").strip().lower()
    show         = request.args.get("show",    "ALL").upper()
    sort_dir     = request.args.get("dir",     "desc")
    sales_tbl    = _rebate_sales_table(request.args.get("month"))

    conn = get_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT m.sold_to, m.brand, m.structure_name,
                   c.sold_to_name, c.sold_to_group, c.bde_state AS region
            FROM rebate_customer_map m
            LEFT JOIN customer c ON c.sold_to = m.sold_to
            WHERE (%s='ALL' OR m.brand=%s)
              AND (%s='ALL' OR c.sold_to_group=%s)
        """, (brand_filter, brand_filter, stg_filter, stg_filter))
        customers = cur.fetchall()

        # Brand comes from sales_thismonth.brand (no carrying join needed
        # since this query doesn't read line/product_group/pattern).
        cur.execute("""
            SELECT s.sold_to, s.ship_to, s.brand AS brand,
                   SUM(s.qty) AS qty, SUM(s.amt) AS amt
            FROM """ + sales_tbl + """ s
            WHERE s.brand IN ('HK','LF')
              AND s.so_type IN """ + _REBATE_SO_TYPES_IN + """
            GROUP BY s.sold_to, s.ship_to, s.brand
        """)
        ship_sales = {}; ship_idx = {}
        for r in cur.fetchall():
            st,sh,br = str(r["sold_to"]),str(r["ship_to"]),r["brand"]
            ship_sales[(st,sh,br)] = {"qty":float(r["qty"] or 0),"amt":float(r["amt"] or 0)}
            ship_idx.setdefault((st,br),set()).add(sh)

        cur.execute("SELECT ship_to, ship_to_name, bde_state, salesman_name FROM customer")
        ship_cust_map_ex = {}
        for r in cur.fetchall():
            sh = str(r["ship_to"])
            ship_cust_map_ex[sh] = {
                "name":  r["ship_to_name"] or sh,
                "state": _rebate_region(r["bde_state"]) or "-",
                "bde":   (r["salesman_name"] or "").strip() or "-",
            }
        name_map = {sh: v["name"] for sh, v in ship_cust_map_ex.items()}

        cur.execute("""
            SELECT structure_name, unit, tier_order, top_order, threshold, rate
            FROM rebate_structure
            WHERE tier_order <= top_order
            ORDER BY structure_name, tier_order
        """)
        tiers_map = {}
        for r in cur.fetchall():
            sd = tiers_map.setdefault(r["structure_name"],{"unit":r["unit"],"top_order":int(r["top_order"]),"tiers":[]})
            sd["tiers"].append({"tier":int(r["tier_order"]),"threshold":float(r["threshold"]),"rate":float(r["rate"])})

        def _calc(actual, tiers, top_order):
            curr={"tier":0,"threshold":0,"rate":0}; nxt=None
            for t in tiers:
                if t["threshold"]<=actual and t["rate"]>0: curr=t
                elif t["threshold"]>actual and t["rate"]>0 and nxt is None: nxt=t
            if curr["tier"] >= top_order: nxt=None
            return curr, nxt

        rows=[]
        for c in customers:
            struct=c["structure_name"]; brand=c["brand"]; sold_to=str(c["sold_to"])
            sold_to_group = c["sold_to_group"] or "-"
            sd=tiers_map.get(struct)
            if not sd: continue
            unit=sd["unit"]; tiers=sd["tiers"]; top_order=sd["top_order"]
            if brand=="TTL":
                ship_set=(ship_idx.get((sold_to,"HK"),set())|ship_idx.get((sold_to,"LF"),set()))
            else:
                ship_set=ship_idx.get((sold_to,brand),set()).copy()
            ship_set = {sh for sh in ship_set if sh in ship_cust_map_ex}
            if not ship_set: ship_set.add(sold_to)

            # Resolve region and BDE from sold_to's own record, falling back to ship_tos
            st_info_ex = ship_cust_map_ex.get(sold_to, {})
            sold_to_bde_ex    = (st_info_ex.get("bde",   "-") or "-") if st_info_ex else "-"
            sold_to_region_ex = (st_info_ex.get("state", "-") or "-") if st_info_ex else "-"
            if sold_to_bde_ex == "-" or sold_to_region_ex == "-":
                real_ships = [sh for sh in ship_set if sh != sold_to]
                if real_ships:
                    from collections import Counter as _Counter
                    sc = _Counter(ship_cust_map_ex[sh].get("state","") for sh in real_ships
                                  if ship_cust_map_ex.get(sh,{}).get("state","") and ship_cust_map_ex[sh]["state"]!="-")
                    bc = _Counter(ship_cust_map_ex[sh].get("bde","") for sh in real_ships
                                  if ship_cust_map_ex.get(sh,{}).get("bde","") and ship_cust_map_ex[sh]["bde"]!="-")
                    if sold_to_bde_ex == "-":
                        sold_to_bde_ex    = bc.most_common(1)[0][0] if bc else "-"
                    if sold_to_region_ex == "-":
                        sold_to_region_ex = sc.most_common(1)[0][0] if sc else "-"

            if _rebate_is_ship_to(struct):
                # _SR → one row per ship_to
                calc_items=[]
                for sh in sorted(ship_set):
                    if brand=="TTL":
                        hk=ship_sales.get((sold_to,sh,"HK"),{"qty":0,"amt":0})
                        lf=ship_sales.get((sold_to,sh,"LF"),{"qty":0,"amt":0})
                        calc_items.append((sh,(hk["qty"]+lf["qty"]) if unit=="Q" else (hk["amt"]+lf["amt"])))
                    else:
                        d=ship_sales.get((sold_to,sh,brand),{"qty":0,"amt":0})
                        calc_items.append((sh,d["qty"] if unit=="Q" else d["amt"]))
            else:
                # Aggregate all ship_tos ??one row per sold_to
                total=0.0
                for sh in ship_set:
                    if brand=="TTL":
                        hk=ship_sales.get((sold_to,sh,"HK"),{"qty":0,"amt":0})
                        lf=ship_sales.get((sold_to,sh,"LF"),{"qty":0,"amt":0})
                        total+=(hk["qty"]+lf["qty"]) if unit=="Q" else (hk["amt"]+lf["amt"])
                    else:
                        d=ship_sales.get((sold_to,sh,brand),{"qty":0,"amt":0})
                        total+=d["qty"] if unit=="Q" else d["amt"]
                calc_items=[(sold_to,total)]

            for sh,actual in calc_items:
                sh_info = ship_cust_map_ex.get(sh, {})
                curr_tier,next_tier=_calc(actual,tiers,top_order)
                rows.append({
                    "sold_to":sold_to,"sold_to_name":c["sold_to_name"] or sold_to,
                    "sold_to_group":sold_to_group,
                    "region": sold_to_region_ex,
                    "bde":    sold_to_bde_ex,
                    "ship_to":sh,"ship_to_name":sh_info.get("name") or (c["sold_to_name"] or sh),
                    "brand":brand,"structure_name":struct,"unit":unit,
                    "actual":round(actual,2),"curr_rate":curr_tier["rate"],
                    "next_rate":next_tier["rate"] if next_tier else None,
                    "needed":round(next_tier["threshold"]-actual,2) if next_tier else None,
                    "est_rebate":round(actual*curr_tier["rate"]/100,2),
                })

        # filters
        if search:
            rows=[r for r in rows if search in r["sold_to_name"].lower() or search in str(r["sold_to"]) or search in r["ship_to_name"].lower() or search in str(r["ship_to"])]
        if show=="NEXT": rows=[r for r in rows if r["next_rate"] is not None and r["actual"]>0]
        elif show=="MAX": rows=[r for r in rows if r["next_rate"] is None and r["curr_rate"]>0]
        elif show=="ZERO": rows=[r for r in rows if r["actual"]==0]

        rows.sort(key=lambda r:(r["sold_to_group"],r["sold_to_name"],r["brand"],r["ship_to"]))

        # build CSV
        out=io.StringIO()
        w=csv.writer(out)
        w.writerow(["Sold-To","Sold-To Name","Group","Region","BDE","Ship-To","Ship-To Name","Brand","Type","Actual","Curr Rate%","Next Rate%","Need to Reach","Est Rebate","Structure"])
        for r in rows:
            w.writerow([r["sold_to"],r["sold_to_name"],r["sold_to_group"],r["region"],r["bde"],r["ship_to"],r["ship_to_name"],r["brand"],"Annual $" if r["unit"]=="A" else "QTR Qty",r["actual"],r["curr_rate"],r["next_rate"] if r["next_rate"] is not None else "",r["needed"] if r["needed"] is not None else "",r["est_rebate"],r["structure_name"]])

        from flask import Response
        from datetime import date
        fname=f"rebate_{date.today().isoformat()}.csv"
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition":f"attachment;filename={fname}"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error":str(e)}),500
    finally:
        cur.close(); conn.close()

# ------------------------------------------------------------------------------
# BDE Visit / Meeting Log — page at /meeting, backed by meeting_log table.
# ------------------------------------------------------------------------------

MEETING_PHOTO_DIR = os.path.join(BASE_DIR, "static", "meeting_photos")

# ── Mail config ────────────────────────────────────────────────────
# All settings overridable via .env so SMTP / Graph credentials never
# live in the repo.  Two delivery paths are supported, in order of
# preference: Microsoft Graph API (modern auth, recommended) →
# Exchange SMTP (legacy, blocked by Security Defaults).
MAIL_FROM      = os.getenv("MAIL_FROM",      "dashboard@hankooktyre.com.au")
SMTP_HOST      = os.getenv("SMTP_HOST",      "smtp.office365.com")
SMTP_PORT      = int(os.getenv("SMTP_PORT",  "587"))
SMTP_USER      = os.getenv("SMTP_USER",      MAIL_FROM)
SMTP_PASSWORD  = os.getenv("SMTP_PASSWORD",  "")
SMTP_USE_TLS   = (os.getenv("SMTP_USE_TLS",  "1") == "1")
GRAPH_TENANT_ID     = os.getenv("GRAPH_TENANT_ID",     "")
GRAPH_CLIENT_ID     = os.getenv("GRAPH_CLIENT_ID",     "")
GRAPH_CLIENT_SECRET = os.getenv("GRAPH_CLIENT_SECRET", "")
DASHBOARD_URL  = os.getenv("DASHBOARD_URL",  "https://sales.hkaudashboard.com")
# Customer-facing portal (claim form + QR codes). Falls back to
# DASHBOARD_URL so existing setups keep working until a dedicated
# subdomain (e.g. https://claim.hkaudashboard.com.au) is provisioned.
CLAIM_PORTAL_URL = os.getenv("CLAIM_PORTAL_URL", DASHBOARD_URL)
MAIL_DEBUG     = (os.getenv("MAIL_DEBUG",    "0") == "1")

# ── BDE / state-manager directory ──────────────────────────────────
# (name, email, state, role).
#   role = "BDE"  → locked to their own salesman name (own territory
#                   + amber overlay markers on the map for shops they
#                   visited outside that territory).
#         "SM"   → locked to their state's region filter.
#         "ALL"  → no scope restriction (leadership / dashboard ops).
_BDE_DIRECTORY = [
    # Name MUST match the customer.salesman_name format used by the
    # database / REGION_SALESMEN constant in app.js — generally
    # "Lastname Firstname" with the case the DB stores, otherwise the
    # lock_salesman value won't match any rows.
    # NSW
    ("Robinson Peter",   "peter.robinson@hankooktyre.com.au",   "NSW", "BDE"),
    ("Buckley Paul",     "paul.buckley@hankooktyre.com.au",     "NSW", "SM"),
    ("Borghese Alessio", "alessio.borghese@hankooktyre.com.au", "NSW", "BDE"),
    ("Makris George",    "george.makris@hankooktyre.com.au",    "NSW", "BDE"),
    # QLD
    ("Marsh Aaron",      "aaron.marsh@hankooktyre.com.au",      "QLD", "SM"),
    ("Spires Steven",    "steven.spires@hankooktyre.com.au",    "QLD", "BDE"),
    ("Maclure Adam",     "adam.maclure@hankooktyre.com.au",     "QLD", "BDE"),
    ("Bovey Craig",      "craig.bovey@hankooktyre.com.au",      "QLD", "BDE"),
    # VIC / SA / TAS (one SM covers all three; the region filter
    # groups SA/TAS shops under VIC)
    ("Hobkirk Calvin",   "calvin.hobkirk@hankooktyre.com.au",   "VIC", "SM"),
    ("Bilston Kelley",   "kelley.bilston@hankooktyre.com.au",   "VIC", "BDE"),
    # Nicola Bellotto retired; Robert Ducie took over the VIC BDE seat.
    ("Ducie Robert",     "robert.ducie@hankooktyre.com.au",     "VIC", "BDE"),
    ("Gultjaeff Jason",  "jason.gultjaeff@hankooktyre.com.au",  "SA",  "BDE"),
    # WA
    ("Asim Qureshi",     "asim.qureshi@hankooktyre.com.au",     "WA",  "SM"),
    ("DAIS Jim",         "jim.dais@hankooktyre.com.au",         "WA",  "BDE"),
    # Leadership / dashboard ops — full scope
    ("Begbie Hayden",    "hayden.begbie@hankooktyre.com.au",    "NSW", "ALL"),
    ("Cho JunJong",      "junjong.cho@hankooktyre.com.au",      "NSW", "ALL"),
    ("Bhang Jayden",     "jayden.bhang@hankooktyre.com.au",     "NSW", "ALL"),
]
STATE_MANAGER_EMAIL = {
    "NSW": "paul.buckley@hankooktyre.com.au",
    "QLD": "aaron.marsh@hankooktyre.com.au",
    "VIC": "calvin.hobkirk@hankooktyre.com.au",
    "SA":  "calvin.hobkirk@hankooktyre.com.au",
    "TAS": "calvin.hobkirk@hankooktyre.com.au",
    "WA":  "asim.qureshi@hankooktyre.com.au",
}
ALWAYS_TO = ["hayden.begbie@hankooktyre.com.au",
             "junjong.cho@hankooktyre.com.au",
             "jayden.bhang@hankooktyre.com.au"]

def _bde_name_keys(full_name):
    """Generate uppercase lookup keys for both 'First Last' and 'Last First'."""
    if not full_name: return set()
    parts = full_name.strip().upper().split()
    if len(parts) < 2: return {full_name.strip().upper()}
    return {" ".join(parts), " ".join(reversed(parts))}

_BDE_EMAIL_MAP, _BDE_STATE_MAP, _BDE_ROLE_MAP = {}, {}, {}
_EMAIL_TO_DIR = {}     # email → (canonical_name, state, role)
for _entry in _BDE_DIRECTORY:
    _n, _e, _s, _r = _entry
    _EMAIL_TO_DIR.setdefault(_e.lower(), (_n, _s, _r))
    for _k in _bde_name_keys(_n):
        _BDE_EMAIL_MAP.setdefault(_k, _e)
        _BDE_STATE_MAP.setdefault(_k, _s)
        _BDE_ROLE_MAP.setdefault(_k, _r)

def _lookup_bde_email(name):
    return _BDE_EMAIL_MAP.get((name or "").strip().upper())

def _lookup_bde_state(name):
    return _BDE_STATE_MAP.get((name or "").strip().upper())

def _lookup_bde_role(name):
    return _BDE_ROLE_MAP.get((name or "").strip().upper())

# ── Microsoft Graph token cache ────────────────────────────────────
# Tokens live ~1h; we cache and refresh ~60s before expiry.
_GRAPH_TOKEN = {"access_token": None, "expires_at": 0.0}

def _get_graph_token():
    import time as _time, requests as _rq
    now = _time.time()
    cached = _GRAPH_TOKEN["access_token"]
    if cached and _GRAPH_TOKEN["expires_at"] > now + 60:
        return cached
    if not (GRAPH_TENANT_ID and GRAPH_CLIENT_ID and GRAPH_CLIENT_SECRET):
        raise RuntimeError("Graph creds not set: GRAPH_TENANT_ID / "
                           "GRAPH_CLIENT_ID / GRAPH_CLIENT_SECRET")
    r = _rq.post(
        f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}/oauth2/v2.0/token",
        data={
            "client_id":     GRAPH_CLIENT_ID,
            "client_secret": GRAPH_CLIENT_SECRET,
            "scope":         "https://graph.microsoft.com/.default",
            "grant_type":    "client_credentials",
        },
        timeout=15,
    )
    if r.status_code != 200:
        raise RuntimeError(f"token endpoint {r.status_code}: {r.text[:300]}")
    j = r.json()
    _GRAPH_TOKEN["access_token"] = j["access_token"]
    _GRAPH_TOKEN["expires_at"]   = now + int(j.get("expires_in", 3600))
    return _GRAPH_TOKEN["access_token"]

def _graph_send(to_list, cc_list, subject, html_body):
    """Send via Microsoft Graph /sendMail.  Raises on failure."""
    import requests as _rq
    token = _get_graph_token()
    payload = {
        "message": {
            "subject": subject,
            "body":    {"contentType": "HTML", "content": html_body},
            "from":    {"emailAddress": {"address": MAIL_FROM}},
            "toRecipients": [{"emailAddress": {"address": x}} for x in (to_list or [])],
            "ccRecipients": [{"emailAddress": {"address": x}} for x in (cc_list or [])],
        },
        "saveToSentItems": "true",
    }
    r = _rq.post(
        f"https://graph.microsoft.com/v1.0/users/{MAIL_FROM}/sendMail",
        headers={"Authorization": f"Bearer {token}",
                 "Content-Type":  "application/json"},
        json=payload, timeout=20,
    )
    if r.status_code not in (200, 202):
        raise RuntimeError(f"graph sendMail {r.status_code}: {r.text[:400]}")

def _smtp_send(to_list, cc_list, subject, html_body):
    """Legacy SMTP send.  Raises on failure."""
    import smtplib, re as _re
    from email.message import EmailMessage
    if not SMTP_PASSWORD:
        raise RuntimeError("SMTP_PASSWORD not set")
    msg = EmailMessage()
    msg["From"] = MAIL_FROM
    msg["To"]   = ", ".join(to_list or [])
    if cc_list: msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = subject
    txt = _re.sub(r"<[^>]+>", "", html_body)
    txt = txt.replace("&nbsp;", " ").replace("&amp;", "&")
    msg.set_content(txt)
    msg.add_alternative(html_body, subtype="html")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
        if SMTP_USE_TLS: s.starttls()
        if SMTP_USER and SMTP_PASSWORD:
            s.login(SMTP_USER, SMTP_PASSWORD)
        s.send_message(msg, from_addr=MAIL_FROM,
                       to_addrs=(to_list or []) + (cc_list or []))

def _send_mail_async(to_list, cc_list, subject, html_body):
    """Pick Graph if creds present, else SMTP, else just log.  Runs the
    actual delivery on a background thread so the API request returns
    immediately."""
    import threading
    if not (to_list or cc_list):
        return
    if GRAPH_CLIENT_ID and GRAPH_CLIENT_SECRET and GRAPH_TENANT_ID:
        path, sender = "graph", _graph_send
    elif SMTP_PASSWORD:
        path, sender = "smtp",  _smtp_send
    else:
        if MAIL_DEBUG:
            print(f"[mail] would send to {to_list} cc {cc_list}: {subject}")
        else:
            print(f"[mail] no Graph or SMTP creds set — skipping "
                  f"(to {to_list}, subject={subject!r})")
        return

    def _run():
        try:
            sender(to_list, cc_list, subject, html_body)
            if MAIL_DEBUG:
                print(f"[mail/{path}] sent: {subject} → {to_list} cc {cc_list}")
        except Exception as e:
            traceback.print_exc()
            print(f"[mail/{path}] send failed: {e}")
    threading.Thread(target=_run, daemon=True).start()

def _esc_html(s):
    return (str(s or "")
            .replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            .replace("\n","<br>"))

def _email_log_html(rid, bde_name, sold_to_name, ship_to, ship_to_name,
                    visit_date, met_person, notes, next_action, feedback,
                    visit_purpose="", prep_notes="", prep_history=None):
    rows = []
    def _row(lbl, val):
        return (f'<tr><td style="padding:4px 10px;color:#6b7280;font-size:11px;'
                f'text-transform:uppercase;letter-spacing:.4px;width:120px;'
                f'vertical-align:top;">{_esc_html(lbl)}</td>'
                f'<td style="padding:4px 10px;color:#111827;font-size:13px;'
                f'vertical-align:top;">{val}</td></tr>')
    rows.append(_row("BDE",         _esc_html(bde_name)))
    rows.append(_row("Visit date",  _esc_html(visit_date)))
    rows.append(_row("Sold-to",     _esc_html(sold_to_name)))
    rows.append(_row("Ship-to",     f"{_esc_html(ship_to)} — {_esc_html(ship_to_name)}"))
    if visit_purpose:
        rows.append(_row("Purpose",  _esc_html(visit_purpose)))
    if prep_notes:
        rows.append(_row("Prep",
                        f'<div style="background:#eff6ff;border-left:3px solid #2563eb;'
                        f'padding:6px 8px;color:#1e3a5f;white-space:pre-wrap;">'
                        f'{_esc_html(prep_notes)}</div>'))
    # Older prep notes for this same ship_to (from prior submissions
    # — typically the "Meeting Preparation" entries written before
    # the visit happened).  Shown chronologically so the reader has
    # the full prep context in one place.
    if prep_history:
        history_html = "".join(
            f'<div style="margin-top:4px;padding:6px 8px;background:#f8fafc;'
            f'border-left:3px solid #94a3b8;border-radius:3px;font-size:12px;'
            f'color:#334155;">'
            f'<div style="font-size:10px;color:#6b7280;margin-bottom:2px;">'
            f'<b>{_esc_html(h.get("bde_name") or "")}</b> · '
            f'{_esc_html(h.get("visit_date") or "")}</div>'
            f'<div style="white-space:pre-wrap;">{_esc_html(h.get("prep_notes") or "")}</div>'
            f'</div>'
            for h in prep_history
        )
        rows.append(_row("Prep history", history_html))
    if met_person:
        rows.append(_row("Met",      _esc_html(met_person)))
    if notes:
        rows.append(_row("Notes",       f'<div style="white-space:pre-wrap;">{_esc_html(notes)}</div>'))
    if next_action:
        rows.append(_row("Next step", _esc_html(next_action)))
    if feedback:
        rows.append(_row("Feedback",
                        f'<div style="background:#fef3c7;border-left:3px solid #f59e0b;'
                        f'padding:6px 8px;color:#78350f;white-space:pre-wrap;">'
                        f'{_esc_html(feedback)}</div>'))
    link        = f"{DASHBOARD_URL}/meeting"
    comment_url = f"{DASHBOARD_URL}/meeting#fb-{rid}"
    return (
        f'<div style="font-family:-apple-system,Segoe UI,sans-serif;font-size:13px;color:#111827;">'
        f'<div style="background:#1e3a5f;color:#fff;padding:10px 14px;border-radius:6px 6px 0 0;'
        f'font-weight:700;">BDE Visit Log #{rid}</div>'
        f'<table style="border-collapse:collapse;background:#fff;width:100%;'
        f'border:1px solid #e5e7eb;border-top:none;border-radius:0 0 6px 6px;">'
        f'{"".join(rows)}</table>'
        f'<div style="margin-top:12px;text-align:center;">'
        f'<a href="{comment_url}" '
        f'style="display:inline-block;background:#2563eb;color:#fff;'
        f'padding:9px 22px;border-radius:6px;text-decoration:none;'
        f'font-weight:600;font-size:13px;">💬 Make a comment</a>'
        f'</div>'
        f'<div style="margin-top:10px;font-size:11px;color:#6b7280;">'
        f'Opens the feedback box for this visit on the dashboard. '
        f'Full log list: <a href="{link}">{link}</a></div>'
        f'</div>'
    )

def _resolve_bde_state(bde_name):
    """Try the hardcoded directory first; if the BDE isn't there (new
    hire, name format drift, etc.) fall back to the customer master
    via salesman_name → bde_state.  Empty string when nothing matches."""
    s = _lookup_bde_state(bde_name)
    if s: return s
    if not bde_name: return ""
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute(
            "SELECT bde_state FROM customer "
            "WHERE UPPER(TRIM(salesman_name)) = %s "
            "AND bde_state IS NOT NULL AND TRIM(bde_state) <> '' "
            "ORDER BY (CASE WHEN bde_state IN ('NSW','QLD','VIC','SA','WA') "
            "          THEN 0 ELSE 1 END) "
            "LIMIT 1",
            (bde_name.strip().upper(),),
        )
        row = cur.fetchone()
        cur.close(); conn.close()
        return (row[0] or "").strip() if row else ""
    except Exception as e:
        print(f"[mail] _resolve_bde_state DB fallback failed: {e}")
        return ""

def _fetch_prep_history(ship_to, exclude_id=None):
    """Return every prior prep_notes entry for this ship_to so an
    actual-log notification can carry forward the prep context (the
    BDE writes the visit log into a fresh textarea on purpose; this
    keeps the email recipients seeing what was prepped beforehand)."""
    if not ship_to:
        return []
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        sql = (
            "SELECT id, visit_date, bde_name, prep_notes "
            "FROM meeting_log "
            "WHERE ship_to = %s "
            "  AND prep_notes IS NOT NULL AND TRIM(prep_notes) <> '' "
        )
        params = [ship_to]
        if exclude_id is not None:
            sql += "AND id <> %s "
            params.append(exclude_id)
        sql += "ORDER BY visit_date ASC, id ASC"
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
        for r in rows:
            if r.get("visit_date"):
                r["visit_date"] = r["visit_date"].strftime("%Y-%m-%d")
        cur.close(); conn.close()
        return rows
    except Exception as e:
        print(f"[mail] _fetch_prep_history failed: {e}")
        return []

def _notify_new_log(rid, bde_name, sold_to_name, ship_to, ship_to_name,
                    visit_date, met_person, notes, next_action,
                    visit_purpose="", plan_bde_email="", prep_notes=""):
    """Email Hayden + JJ + Jayden + the State Manager of the BDE's
    state.  CC the BDE author so they have a record of what was sent
    on their behalf.  Also CC the original scheduling BDE (plan_bde_email)
    if the log was created from someone else's calendar plan."""
    state    = _resolve_bde_state(bde_name)
    sm_email = STATE_MANAGER_EMAIL.get(state) if state else None
    author   = _lookup_bde_email(bde_name)
    to_list  = list(ALWAYS_TO)
    if sm_email and sm_email.lower() not in [x.lower() for x in to_list]:
        to_list.append(sm_email)
    cc_list  = [author] if author and author.lower() not in [x.lower() for x in to_list] else []
    if plan_bde_email and plan_bde_email.lower() != (author or "").lower() \
       and plan_bde_email.lower() not in [x.lower() for x in to_list] \
       and plan_bde_email.lower() not in [x.lower() for x in cc_list]:
        cc_list.append(plan_bde_email)
    subject  = (f"[BDE Visit] {bde_name} → {ship_to} "
                f"{ship_to_name or ''} ({visit_date})")
    # Surface the resolution so any "SM not getting mail" issue is
    # diagnosable from the console log without code changes.
    print(f"[mail] new_log #{rid} bde={bde_name!r} state={state!r} "
          f"sm={sm_email!r} plan_bde={plan_bde_email!r} → to={to_list} cc={cc_list}")
    prep_history = _fetch_prep_history(ship_to, exclude_id=rid)
    html = _email_log_html(rid, bde_name, sold_to_name, ship_to, ship_to_name,
                           visit_date, met_person, notes, next_action, None,
                           visit_purpose=visit_purpose,
                           prep_notes=prep_notes,
                           prep_history=prep_history)
    _send_mail_async(to_list, cc_list, subject, html)

def _notify_feedback_thread(rid, log_row, thread, current_author_email=""):
    """Email everyone in the feedback thread when a new comment lands.

    Recipients = original BDE + every prior commenter's email +
    Hayden/JJ + State Manager — minus the person who just typed
    (no point mailing yourself back).  Each thread participant sees
    the entire conversation chronologically in the message body.
    """
    if not log_row or not thread:
        return
    me = (current_author_email or "").strip().lower()

    # Gather participants
    participants = set()
    bde_email = (log_row.get("bde_email") or "").strip().lower() \
                or (_lookup_bde_email(log_row.get("bde_name") or "") or "").lower()
    if bde_email: participants.add(bde_email)
    for t in thread:
        em = (t.get("author_email") or "").strip().lower()
        if em: participants.add(em)
        # If we only have the name, resolve via the directory.
        if not em:
            resolved = _lookup_bde_email(t.get("author_name") or "")
            if resolved: participants.add(resolved.lower())

    # Always-notify recipients
    for em in ALWAYS_TO:
        participants.add(em.lower())
    state    = _lookup_bde_state(log_row.get("bde_name") or "")
    sm_email = STATE_MANAGER_EMAIL.get(state) if state else None
    if sm_email: participants.add(sm_email.lower())
    # Original scheduling BDE — if the log was created from a calendar
    # chip planned by someone else (e.g. a manager logged a visit on
    # behalf of the BDE who scheduled it), make sure that BDE is in the
    # loop on every comment.
    plan_em = (log_row.get("plan_bde_email") or "").strip().lower() \
              or (_lookup_bde_email(log_row.get("plan_bde_name") or "") or "").lower()
    if plan_em: participants.add(plan_em)

    # Don't mail the person who just commented.
    if me: participants.discard(me)
    if not participants:
        return

    last  = thread[-1]
    actor = (last.get("author_name") or last.get("author_email") or "Someone")
    subj  = (f"[BDE Visit] {actor} commented on {log_row.get('bde_name','')}'s "
             f"{log_row.get('visit_date','')} visit — {log_row.get('ship_to','')}")

    # Build the thread block in HTML so the reader gets the full
    # conversation, latest at the bottom (chronological reading order).
    thread_html_parts = []
    for t in thread:
        meta = (f"<div style='font-size:11px;color:#6b7280;margin-bottom:2px;'>"
                f"<b>{_esc_html(t.get('author_name') or t.get('author_email') or '—')}</b> "
                f"· {_esc_html(t.get('created_at') or '')}</div>")
        body_div = (f"<div style='white-space:pre-wrap;font-size:13px;color:#111827;"
                    f"padding:6px 10px;border-left:3px solid #f59e0b;"
                    f"background:#fef3c7;border-radius:4px;margin-bottom:6px;'>"
                    f"{_esc_html(t.get('text') or '')}</div>")
        thread_html_parts.append(meta + body_div)
    thread_html = "".join(thread_html_parts)

    # Log card + thread.  prep_history pulls every prior prep entry
    # for this ship_to so the reader of a feedback notification still
    # sees the full context, not just whatever was on this one row.
    prep_history = _fetch_prep_history(log_row.get("ship_to") or "",
                                       exclude_id=rid)
    card = _email_log_html(
        rid,
        log_row.get("bde_name") or "",
        log_row.get("sold_to_name") or log_row.get("sold_to") or "",
        log_row.get("ship_to") or "",
        log_row.get("ship_to_name") or "",
        log_row.get("visit_date") or "",
        log_row.get("met_person") or "",
        log_row.get("notes") or "",
        log_row.get("next_action") or "",
        None,           # feedback is shown as a thread below instead
        visit_purpose=log_row.get("visit_purpose") or "",
        prep_notes=log_row.get("prep_notes") or "",
        prep_history=prep_history,
    )
    link        = f"{DASHBOARD_URL}/meeting"
    comment_url = f"{DASHBOARD_URL}/meeting#fb-{rid}"
    full = (
        f"{card}"
        f"<div style='margin-top:14px;font-family:-apple-system,Segoe UI,sans-serif;"
        f"font-size:13px;color:#111827;'>"
        f"<div style='font-weight:700;color:#1e3a5f;margin-bottom:6px;'>Feedback thread</div>"
        f"{thread_html}"
        f"<div style='margin-top:12px;text-align:center;'>"
        f"<a href='{comment_url}' "
        f"style='display:inline-block;background:#2563eb;color:#fff;"
        f"padding:9px 22px;border-radius:6px;text-decoration:none;"
        f"font-weight:600;font-size:13px;'>💬 Make a comment</a>"
        f"</div>"
        f"<div style='margin-top:10px;font-size:11px;color:#6b7280;'>"
        f"Reply opens the feedback box for this visit on the dashboard. "
        f"<a href='{link}'>{link}</a></div>"
        f"</div>"
    )
    _send_mail_async(list(participants), [], subj, full)

def _ensure_meeting_log_table():
    """Create the meeting_log table on startup if it doesn't exist."""
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meeting_log (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                created_at   TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                visit_date   DATE NOT NULL,
                bde_email    VARCHAR(120) NOT NULL DEFAULT '',
                bde_name     VARCHAR(120) NOT NULL DEFAULT '',
                sold_to      VARCHAR(64)  NOT NULL DEFAULT '',
                sold_to_name VARCHAR(160) NOT NULL DEFAULT '',
                ship_to      VARCHAR(64)  NOT NULL DEFAULT '',
                met_person   VARCHAR(120) NOT NULL DEFAULT '',
                notes        TEXT,
                next_action  TEXT,
                photo_paths  TEXT,
                INDEX  idx_meeting_ship_to (ship_to),
                INDEX  idx_meeting_visit_date (visit_date),
                INDEX  idx_meeting_bde     (bde_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Threaded feedback — each row = one comment in the conversation
        # on a meeting_log entry.  Replaces the single TEXT column on
        # meeting_log (which is kept around as a denormalised "latest"
        # cache for older code paths).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meeting_feedback (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                meeting_id   INT NOT NULL,
                created_at   TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                author_email VARCHAR(120) NOT NULL DEFAULT '',
                author_name  VARCHAR(120) NOT NULL DEFAULT '',
                text         TEXT NOT NULL,
                INDEX idx_fb_meeting (meeting_id),
                INDEX idx_fb_created (created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Idempotent migration for older meeting_log tables.
        migrations = [
            "ALTER TABLE meeting_log MODIFY COLUMN sold_to VARCHAR(64) NOT NULL DEFAULT ''",
            "ALTER TABLE meeting_log MODIFY COLUMN ship_to VARCHAR(64) NOT NULL DEFAULT ''",
        ]
        for sql in migrations:
            try: cur.execute(sql)
            except Exception: pass
        # Add columns that may not exist on legacy tables.
        for col, defn, after in (
            ("sold_to_name",   "VARCHAR(160) NOT NULL DEFAULT ''", "sold_to"),
            ("visit_date",     "DATE NOT NULL DEFAULT '1970-01-01'", "created_at"),
            ("feedback",       "TEXT",                              "next_action"),
            ("visit_purpose",  "VARCHAR(40) NOT NULL DEFAULT ''",   "ship_to"),
            # Meeting Preparation note + the BDE who originally scheduled
            # the visit on the calendar (so feedback emails CC them).
            ("prep_notes",         "TEXT",                              "feedback"),
            ("plan_bde_email",     "VARCHAR(120) NOT NULL DEFAULT ''",  "bde_email"),
            ("plan_bde_name",      "VARCHAR(120) NOT NULL DEFAULT ''",  "plan_bde_email"),
            ("met_person_contact", "VARCHAR(80) NOT NULL DEFAULT ''",   "met_person"),
        ):
            try:
                cur.execute(f"ALTER TABLE meeting_log "
                            f"ADD COLUMN {col} {defn} AFTER {after}")
            except Exception:
                pass
        # One-time migration: if meeting_feedback is empty but some
        # meeting_log rows still carry single-string feedback, lift
        # those into the thread as a "(legacy)" first comment.
        try:
            cur.execute("SELECT COUNT(*) FROM meeting_feedback")
            (n,) = cur.fetchone()
            if n == 0:
                cur.execute("""
                    INSERT INTO meeting_feedback
                        (meeting_id, created_at, author_email, author_name, text)
                    SELECT id, created_at, '', '(legacy)', feedback
                    FROM meeting_log
                    WHERE feedback IS NOT NULL AND TRIM(feedback) <> ''
                """)
        except Exception as e:
            print(f"[meeting_feedback] legacy backfill skipped: {e}")
        # Backfill visit_date from created_at where it's still the sentinel.
        try:
            cur.execute("UPDATE meeting_log SET visit_date = DATE(created_at) "
                        "WHERE visit_date = '1970-01-01'")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE meeting_log "
                        "ADD INDEX idx_meeting_visit_date (visit_date)")
        except Exception:
            pass
        conn.commit()
        cur.close(); conn.close()
        os.makedirs(MEETING_PHOTO_DIR, exist_ok=True)
    except Exception as e:
        print(f"[meeting_log] schema init failed: {e}")

_ensure_meeting_log_table()

def _ensure_meeting_plan_table():
    """One-row-per-planned-visit table backing the drag-and-drop
    calendar on /meeting.  Independent from meeting_log: a plan is
    intent (\"visit this shop next Tuesday\"), a log is the memo
    written after the actual visit."""
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meeting_plan (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                plan_date    DATE NOT NULL,
                ship_to      VARCHAR(64) NOT NULL,
                sold_to      VARCHAR(64) NOT NULL DEFAULT '',
                ship_to_name VARCHAR(160) NOT NULL DEFAULT '',
                sold_to_name VARCHAR(160) NOT NULL DEFAULT '',
                bde_email    VARCHAR(120) NOT NULL DEFAULT '',
                bde_name     VARCHAR(120) NOT NULL DEFAULT '',
                created_at   TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_plan_date    (plan_date),
                INDEX idx_plan_ship_to (ship_to),
                INDEX idx_plan_bde     (bde_email),
                INDEX idx_plan_bde_nm  (bde_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # plan_group — label of the region/postcode card that dropped
        # this ship_to.  Empty for solo drags.  Rendering side groups
        # same-day, same-group, same-bde chips into one expandable
        # region chip.  Added idempotently for older deployments.
        cur.execute("SHOW COLUMNS FROM meeting_plan LIKE 'plan_group'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE meeting_plan "
                        "ADD COLUMN plan_group VARCHAR(60) NOT NULL DEFAULT ''")
            cur.execute("CREATE INDEX idx_plan_group ON meeting_plan (plan_date, plan_group)")
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"[meeting_plan] schema init failed: {e}")

_ensure_meeting_plan_table()
_ensure_request_log_table()
_ensure_submitted_orders_table()

def _bde_from_request():
    """Best-effort 'who is logged in'.  Cloudflare Access (if it's in
    front of the app) injects Cf-Access-Authenticated-User-Email; behind
    Tailscale / local dev that header is absent, so the form also sends
    bde_name as a fallback (manual pick)."""
    email = (request.headers.get("Cf-Access-Authenticated-User-Email")
             or request.headers.get("cf-access-authenticated-user-email")
             or "").strip().lower()
    return email

@app.get("/meeting")
def meeting_page():
    return send_from_directory("static", "meeting.html")

@app.get("/api/whoami")
def whoami():
    """Identify the current user from Cloudflare Access and return
    the scope locks the frontend should apply.

    BDE  → lock_salesman = own name (own territory + amber overlay
           markers for shops visited outside it on the map).
    SM   → lock_region   = own state.
    ALL  → no locks (leadership / Hayden / JJ / dashboard ops).

    No CF header (Tailscale / local dev) or unknown email → no locks
    so internal access from the office network keeps working."""
    email = _bde_from_request()
    out = {
        "email":          email,
        "name":           "",
        "role":           "ALL",
        "state":          "",
        "lock_salesman":  None,
        "lock_region":    None,
        "logged_in":      bool(email),
    }
    if not email:
        return jsonify(out)
    found = _EMAIL_TO_DIR.get(email.lower())
    if not found:
        # Recognised by CF but not in our directory — let them in with
        # full scope so we don't accidentally lock out a new hire.
        return jsonify(out)
    name, state, role = found
    out["name"]  = name
    out["state"] = state
    out["role"]  = role
    if role == "BDE":
        # BDEs see only their own data: lock the salesman dropdown to
        # their name AND the region buttons to their state, so they
        # can't accidentally widen the view by clicking another State.
        out["lock_salesman"] = name
        out["lock_region"]   = state
    elif role == "SM":
        out["lock_region"]   = state
    return jsonify(out)

@app.get("/api/mail_test")
def mail_test():
    """Diagnostic: try sending one test email synchronously so any error
    surfaces in the HTTP response.  Picks Graph if creds are set, else
    SMTP.  Use ?to=you@example.com."""
    to = (request.args.get("to") or "").strip()
    if not to:
        return jsonify({"error": "missing ?to=email"}), 400

    report = {
        "mail_from":          MAIL_FROM,
        "graph_tenant_set":   bool(GRAPH_TENANT_ID),
        "graph_client_set":   bool(GRAPH_CLIENT_ID),
        "graph_secret_set":   bool(GRAPH_CLIENT_SECRET),
        "smtp_host":          SMTP_HOST,
        "smtp_port":          SMTP_PORT,
        "smtp_user":          SMTP_USER,
        "smtp_password_set":  bool(SMTP_PASSWORD),
        "to":                 to,
    }
    subject = "[BDE Visit] Mail-path test from the dashboard"
    body    = ("<p>This is a test message from the dashboard.</p>"
               "<p>If you got it, mail delivery is wired up correctly.</p>")

    if GRAPH_TENANT_ID and GRAPH_CLIENT_ID and GRAPH_CLIENT_SECRET:
        report["path"] = "graph"
        try:
            _graph_send([to], [], subject, body)
            report["result"] = "SENT via Microsoft Graph — check inbox (and spam)"
            return jsonify(report)
        except Exception as e:
            report["result"] = f"FAILED — {type(e).__name__}: {e}"
            return jsonify(report), 500

    if SMTP_PASSWORD:
        report["path"] = "smtp"
        try:
            _smtp_send([to], [], subject, body)
            report["result"] = "SENT via SMTP — check inbox (and spam)"
            return jsonify(report)
        except Exception as e:
            report["result"] = f"FAILED — {type(e).__name__}: {e}"
            return jsonify(report), 500

    report["path"]   = "none"
    report["result"] = ("SKIPPED — set GRAPH_TENANT_ID / GRAPH_CLIENT_ID / "
                        "GRAPH_CLIENT_SECRET (recommended) or SMTP_PASSWORD in .env")
    return jsonify(report)

@app.get("/api/people")
def people_directory():
    """Names that can leave feedback — BDEs + State Managers + Hayden + JJ.
    Sourced from the hardcoded _BDE_DIRECTORY so it works without DB."""
    out = []
    seen = set()
    for nm, em, st, _role in _BDE_DIRECTORY:
        key = nm.upper()
        if key in seen: continue
        seen.add(key)
        out.append({"name": nm, "email": em, "state": st})
    return jsonify(out)

@app.get("/api/bdes_active")
def bdes_active():
    """Distinct salesman names from the customer master, dropdown source
    for the meeting form."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT DISTINCT TRIM(salesman_name) AS name
            FROM customer
            WHERE salesman_name IS NOT NULL
              AND TRIM(salesman_name) <> ''
              AND TRIM(salesman_name) <> '#N/A'
            ORDER BY name
        """)
        out = [r["name"] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _meeting_owner_or_403(cur, mid):
    """Return (ok, owner_email).  ok is False with the row missing/forbidden.
    Legacy rows may have a blank bde_email — match by bde_name in that case."""
    me_email = (_bde_from_request() or "").strip().lower()
    if not me_email:
        return False, None
    me_name = (_EMAIL_TO_DIR.get(me_email) or ("", "", ""))[0].strip().lower()
    cur.execute("SELECT bde_email, bde_name FROM meeting_log WHERE id=%s", (mid,))
    row = cur.fetchone()
    if not row:
        return False, None
    o_email = (row.get("bde_email") or "").strip().lower()
    o_name  = (row.get("bde_name")  or "").strip().lower()
    ok = (o_email and o_email == me_email) or (me_name and o_name == me_name)
    return bool(ok), o_email

@app.patch("/api/meeting/<int:mid>")
def meeting_patch(mid):
    """Edit one's own meeting_log row.  Only the original author (matched by
    bde_email) can update notes / next_action / prep_notes / met_person /
    visit_purpose / visit_date."""
    body = request.get_json(silent=True) or {}
    allowed = ("notes", "next_action", "prep_notes",
               "met_person", "met_person_contact",
               "visit_purpose", "visit_date")
    updates = {k: (body[k] or "").strip() if isinstance(body[k], str) else body[k]
               for k in allowed if k in body}
    if "visit_purpose" in updates and updates["visit_purpose"] and \
       updates["visit_purpose"] not in ("Promotion", "Product introduction",
                                        "Claim support", "Rebate follow-up",
                                        "Stock", "Other"):
        return jsonify({"error": "invalid visit_purpose"}), 400
    if not updates:
        return jsonify({"error": "no fields"}), 400
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        ok, _ = _meeting_owner_or_403(cur, mid)
        if not ok:
            return jsonify({"error": "forbidden"}), 403
        sets   = ", ".join(f"{k}=%s" for k in updates)
        params = list(updates.values()) + [mid]
        cur.execute(f"UPDATE meeting_log SET {sets} WHERE id=%s", params)
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close(); conn.close()

@app.delete("/api/meeting/<int:mid>")
def meeting_delete(mid):
    """Delete one's own meeting_log row (and its feedback thread)."""
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        ok, _ = _meeting_owner_or_403(cur, mid)
        if not ok:
            return jsonify({"error": "forbidden"}), 403
        cur.execute("DELETE FROM meeting_feedback WHERE meeting_id=%s", (mid,))
        cur.execute("DELETE FROM meeting_log WHERE id=%s", (mid,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close(); conn.close()

@app.patch("/api/meeting/feedback/<int:fid>")
def meeting_feedback_patch(fid):
    """Edit one's own feedback comment."""
    body = request.get_json(silent=True) or {}
    text = (body.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400
    me_email = (_bde_from_request() or "").strip().lower()
    if not me_email:
        return jsonify({"error": "auth required"}), 401
    me_name = (_EMAIL_TO_DIR.get(me_email) or ("", "", ""))[0].strip().lower()
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT author_email, author_name, meeting_id FROM meeting_feedback WHERE id=%s", (fid,))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        o_email = (row.get("author_email") or "").strip().lower()
        o_name  = (row.get("author_name")  or "").strip().lower()
        if not ((o_email and o_email == me_email) or (me_name and o_name == me_name)):
            return jsonify({"error": "forbidden"}), 403
        cur.execute("UPDATE meeting_feedback SET text=%s WHERE id=%s", (text, fid))
        # Refresh the denormalised latest-feedback cache on meeting_log
        cur.execute("""
            SELECT text FROM meeting_feedback
            WHERE meeting_id=%s
            ORDER BY created_at DESC, id DESC LIMIT 1
        """, (row["meeting_id"],))
        last = cur.fetchone()
        cur.execute("UPDATE meeting_log SET feedback=%s WHERE id=%s",
                    (last["text"] if last else None, row["meeting_id"]))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close(); conn.close()

@app.delete("/api/meeting/feedback/<int:fid>")
def meeting_feedback_delete(fid):
    """Delete one's own feedback comment."""
    me_email = (_bde_from_request() or "").strip().lower()
    if not me_email:
        return jsonify({"error": "auth required"}), 401
    me_name = (_EMAIL_TO_DIR.get(me_email) or ("", "", ""))[0].strip().lower()
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT author_email, author_name, meeting_id FROM meeting_feedback WHERE id=%s", (fid,))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        o_email = (row.get("author_email") or "").strip().lower()
        o_name  = (row.get("author_name")  or "").strip().lower()
        if not ((o_email and o_email == me_email) or (me_name and o_name == me_name)):
            return jsonify({"error": "forbidden"}), 403
        mid = row["meeting_id"]
        cur.execute("DELETE FROM meeting_feedback WHERE id=%s", (fid,))
        cur.execute("""
            SELECT text FROM meeting_feedback
            WHERE meeting_id=%s
            ORDER BY created_at DESC, id DESC LIMIT 1
        """, (mid,))
        last = cur.fetchone()
        cur.execute("UPDATE meeting_log SET feedback=%s WHERE id=%s",
                    (last["text"] if last else None, mid))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close(); conn.close()


@app.post("/api/meeting")
def meeting_post():
    """Insert one visit-log entry.  Accepts multipart/form-data so the
    optional photo uploads (up to 5 files) can come in the same request."""
    try:
        bde_email = _bde_from_request()
        bde_name  = (request.form.get("bde_name") or "").strip()
        sold_in   = (request.form.get("sold_to")  or "").strip()
        ship_to   = (request.form.get("ship_to")  or "").strip()
        met       = (request.form.get("met_person") or "").strip()
        met_contact = (request.form.get("met_person_contact") or "").strip()[:80]
        notes     = (request.form.get("notes") or "").strip()
        next_act  = (request.form.get("next_action") or "").strip()
        feedback  = (request.form.get("feedback") or "").strip()
        visit_in  = (request.form.get("visit_date") or "").strip()
        purpose   = (request.form.get("visit_purpose") or "").strip()
        prep      = (request.form.get("prep_notes") or "").strip()
        # If the user landed on the form by clicking a calendar chip,
        # the frontend passes the original scheduling BDE's name so we
        # can email them when feedback is later added on this log.
        plan_bde_name  = (request.form.get("plan_bde_name") or "").strip()
        plan_bde_email = _lookup_bde_email(plan_bde_name) or "" if plan_bde_name else ""

        if not ship_to:
            return jsonify({"error": "ship_to is required"}), 400
        # Prep-only submissions are allowed: a row that only carries
        # Meeting Preparation (pre-visit) is valid.  A row with notes
        # (a real visit log) still requires purpose.
        if not notes and not prep:
            return jsonify({"error": "notes or prep_notes is required"}), 400
        if notes and not purpose:
            return jsonify({"error": "visit_purpose is required when notes is provided"}), 400
        if purpose and purpose not in ("Promotion", "Product introduction",
                                       "Claim support", "Rebate follow-up",
                                       "Stock", "Other"):
            return jsonify({"error": "invalid visit_purpose"}), 400

        # Visit date: YYYY-MM-DD; default to today if blank/invalid.
        # Capped at today so BDE can backfill past visits but can't
        # invent future ones.
        from datetime import date as _date_cls
        visit_date_obj = _date_cls.today()
        if visit_in:
            try:
                visit_date_obj = datetime.strptime(visit_in, "%Y-%m-%d").date()
            except ValueError:
                pass
        if visit_date_obj > _date_cls.today():
            visit_date_obj = _date_cls.today()

        # The Sold-to picker submits the NAME (because /api/sold_to_names
        # only exposes names).  We resolve it to a sold-to CODE via the
        # customer master so meeting_log lines up with the rest of the
        # database.  If lookup fails, fall back to storing the raw input
        # truncated to fit the column.
        sold_to, sold_to_name = "", sold_in
        # Potential customers (POT-<id>) own their own sold_to_name in
        # the potential_customer row — use that directly so the log
        # carries the same labels the BDE originally typed.
        pot_id = _parse_potential_id(ship_to)
        if pot_id is not None:
            try:
                _conn = get_connection(); _cur = _conn.cursor(dictionary=True)
                _cur.execute("SELECT sold_to_name FROM potential_customer "
                             "WHERE id = %s", (pot_id,))
                r = _cur.fetchone()
                _cur.close(); _conn.close()
                if r and r.get("sold_to_name"):
                    sold_to_name = r["sold_to_name"]
            except Exception as e:
                print(f"[meeting] potential sold_to resolve failed: {e}")
        elif sold_in:
            try:
                _conn = get_connection()
                _cur  = _conn.cursor(dictionary=True)
                # If ship_to is known, prefer the sold_to that actually
                # owns this ship_to (handles the rare case of two sold_tos
                # sharing a sold_to_name).
                if ship_to:
                    _cur.execute(
                        "SELECT sold_to, sold_to_name FROM customer "
                        "WHERE ship_to = %s LIMIT 1", (ship_to,))
                    r = _cur.fetchone()
                    if r and r["sold_to"]:
                        sold_to      = str(r["sold_to"])
                        sold_to_name = (r["sold_to_name"] or sold_in).strip()
                if not sold_to:
                    _cur.execute(
                        "SELECT sold_to FROM customer "
                        "WHERE TRIM(sold_to_name) = %s "
                        "ORDER BY sold_to LIMIT 1", (sold_in,))
                    r = _cur.fetchone()
                    if r and r["sold_to"]:
                        sold_to = str(r["sold_to"])
                _cur.close(); _conn.close()
            except Exception as e:
                print(f"[meeting] sold_to resolve failed: {e}")
        # Belt-and-braces in case lookup failed and the raw input is huge:
        sold_to      = (sold_to or sold_in)[:64]
        sold_to_name = sold_to_name[:160]
        ship_to      = ship_to[:64]

        # Save uploaded photos under static/meeting_photos/YYYY/MM/...
        # using <timestamp>_<ship_to>_<idx>.<ext>.  Up to 5 files honoured;
        # everything else discarded.
        saved = []
        files = request.files.getlist("photos") or []
        if files:
            now    = datetime.now()
            subdir = os.path.join(MEETING_PHOTO_DIR,
                                  f"{now.year:04d}", f"{now.month:02d}")
            os.makedirs(subdir, exist_ok=True)
            ts = now.strftime("%Y%m%d_%H%M%S")
            for i, f in enumerate(files[:5]):
                if not f or not f.filename:
                    continue
                ext = os.path.splitext(f.filename)[1].lower()
                if ext not in (".jpg", ".jpeg", ".png", ".webp", ".heic"):
                    continue
                safe_ship = "".join(c for c in ship_to if c.isalnum()) or "X"
                fname = f"{ts}_{safe_ship}_{i}{ext}"
                path  = os.path.join(subdir, fname)
                f.save(path)
                # Store the URL relative to /static so the frontend can
                # link straight to it.
                saved.append(f"/static/meeting_photos/{now.year:04d}/{now.month:02d}/{fname}")

        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO meeting_log
                (visit_date, bde_email, bde_name,
                 plan_bde_email, plan_bde_name,
                 sold_to, sold_to_name,
                 ship_to, visit_purpose, met_person, met_person_contact,
                 notes, next_action,
                 feedback, prep_notes, photo_paths)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (visit_date_obj, bde_email, bde_name,
              plan_bde_email, plan_bde_name,
              sold_to, sold_to_name,
              ship_to, purpose, met, met_contact,
              notes, next_act,
              feedback if feedback else None,
              prep if prep else None,
              ",".join(saved) if saved else None))
        new_id = cur.lastrowid
        # Pull the resolved ship_to_name back out so the notification
        # email can show a friendly label without an extra round trip.
        ship_nm = ""
        try:
            cur.execute("SELECT MIN(NULLIF(TRIM(ship_to_name),'')) FROM customer "
                        "WHERE ship_to = %s", (ship_to,))
            r = cur.fetchone()
            if r and r[0]: ship_nm = r[0]
        except Exception:
            pass
        conn.commit()
        cur.close(); conn.close()

        # Fire-and-forget notification: Hayden + JJ + State Manager,
        # CC the BDE author.  Runs in a background thread so the form
        # POST returns immediately.
        try:
            _notify_new_log(new_id, bde_name, sold_to_name or sold_to,
                            ship_to, ship_nm,
                            visit_date_obj.strftime("%Y-%m-%d"),
                            met, notes, next_act,
                            visit_purpose=purpose,
                            plan_bde_email=plan_bde_email,
                            prep_notes=prep)
        except Exception as e:
            print(f"[meeting] notify_new_log failed: {e}")

        return jsonify({"ok": True, "id": new_id, "photos": saved})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/meeting/feedback")
def meeting_feedback():
    """Append a comment to a meeting_log's feedback thread.

    Body (JSON or form):  id, text, author_name
      • id           – meeting_log.id to comment on
      • text         – the comment body
      • author_name  – display name (resolved to email via the BDE
                       directory; falls back to the CF Access header
                       if running behind Cloudflare Access).

    The same payload key still accepts 'feedback' for backward compat
    with the older single-string call shape.
    """
    try:
        body  = request.get_json(silent=True) or {}
        rid   = body.get("id") or request.form.get("id")
        text  = (body.get("text") or body.get("feedback")
                 or request.form.get("text") or request.form.get("feedback")
                 or "").strip()
        a_nm  = (body.get("author_name") or request.form.get("author_name") or "").strip()
        try:
            rid = int(rid)
        except (TypeError, ValueError):
            return jsonify({"error": "invalid id"}), 400
        if not text:
            return jsonify({"error": "text is required"}), 400

        # Resolve author email: prefer CF Access header → fall back to
        # directory lookup by the picked name.
        a_email = _bde_from_request() or _lookup_bde_email(a_nm) or ""

        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("""
            INSERT INTO meeting_feedback
                (meeting_id, author_email, author_name, text)
            VALUES (%s, %s, %s, %s)
        """, (rid, a_email, a_nm, text))
        new_fb_id = cur.lastrowid

        # Keep meeting_log.feedback as a denormalised "latest comment"
        # cache so existing list queries that only read that column
        # still surface something useful.
        cur.execute("UPDATE meeting_log SET feedback=%s WHERE id=%s",
                    (text, rid))

        # Re-read the meeting_log row + everyone who has already
        # commented on this thread → those are the email recipients.
        cur.execute("""
            SELECT m.id, m.visit_date, m.bde_name, m.bde_email,
                   m.plan_bde_name, m.plan_bde_email,
                   m.sold_to, m.sold_to_name, m.ship_to,
                   m.visit_purpose, m.prep_notes,
                   m.met_person, m.notes, m.next_action,
                   COALESCE(NULLIF(TRIM(c.ship_to_name),''), m.ship_to) AS ship_to_name
            FROM meeting_log m
            LEFT JOIN (
                SELECT ship_to, MIN(NULLIF(TRIM(ship_to_name),'')) AS ship_to_name
                FROM customer GROUP BY ship_to
            ) c ON c.ship_to = m.ship_to
            WHERE m.id = %s
        """, (rid,))
        log_row = cur.fetchone()
        if log_row and log_row.get("visit_date"):
            log_row["visit_date"] = log_row["visit_date"].strftime("%Y-%m-%d")

        cur.execute("""
            SELECT id, created_at, author_email, author_name, text
            FROM meeting_feedback
            WHERE meeting_id = %s
            ORDER BY created_at ASC, id ASC
        """, (rid,))
        thread = cur.fetchall()
        for t in thread:
            if t.get("created_at"):
                t["created_at"] = t["created_at"].strftime("%Y-%m-%d %H:%M")

        conn.commit()
        cur.close(); conn.close()

        # Notify everyone touching this thread (original BDE +
        # everyone who has commented before + Hayden/JJ + State Manager)
        # except the person who just typed this comment.
        try:
            _notify_feedback_thread(rid, log_row, thread,
                                    current_author_email=a_email)
        except Exception as e:
            print(f"[meeting] notify_feedback_thread failed: {e}")

        return jsonify({"ok": True, "id": rid, "feedback_id": new_fb_id,
                        "thread": thread})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ── /meeting calendar planner ──────────────────────────────────────
@app.get("/api/meeting_plan")
def meeting_plan_list():
    """Return scheduled visits for a month (Y/M).  Scope filter mirrors
    the rest of the dashboard:
      • BDE → only their own plans.
      • SM  → plans by any BDE who belongs to their state.
      • ALL → everyone (Hayden / JJ / Jayden / unknown emails).
    """
    try:
        year  = int(request.args.get("year")  or datetime.now().year)
        month = int(request.args.get("month") or datetime.now().month)
    except Exception:
        return jsonify({"error": "year/month must be integers"}), 400

    email = _bde_from_request()
    me    = _EMAIL_TO_DIR.get(email.lower()) if email else None
    role  = me[2] if me else "ALL"
    bde   = me[0] if me else None
    state = me[1] if me else None

    wh = ["YEAR(plan_date) = %s", "MONTH(plan_date) = %s"]
    params = [year, month]
    if role == "BDE" and bde:
        wh.append("UPPER(bde_name) = %s")
        params.append(bde.upper())
    elif role == "SM" and state:
        # All BDEs whose home state matches this SM
        state_bdes = [n for (n, _e, s, _r) in _BDE_DIRECTORY if s == state]
        if state_bdes:
            placeholders = ",".join(["%s"] * len(state_bdes))
            wh.append(f"UPPER(bde_name) IN ({placeholders})")
            params.extend([n.upper() for n in state_bdes])

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT p.id, p.plan_date, p.ship_to, p.sold_to,
                   p.ship_to_name, p.sold_to_name,
                   p.bde_email, p.bde_name, p.plan_group,
                   EXISTS(
                     SELECT 1 FROM meeting_log m
                     WHERE m.ship_to = p.ship_to
                       AND YEAR(m.visit_date)  = YEAR(p.plan_date)
                       AND MONTH(m.visit_date) = MONTH(p.plan_date)
                       AND m.notes IS NOT NULL AND TRIM(m.notes) <> ''
                   ) AS logged,
                   (SELECT GROUP_CONCAT(DISTINCT m.visit_purpose
                                        ORDER BY m.visit_purpose SEPARATOR ',')
                    FROM meeting_log m
                    WHERE m.ship_to = p.ship_to
                      AND YEAR(m.visit_date)  = YEAR(p.plan_date)
                      AND MONTH(m.visit_date) = MONTH(p.plan_date)
                      AND m.visit_purpose IS NOT NULL
                      AND TRIM(m.visit_purpose) <> ''
                   ) AS purposes
            FROM meeting_plan p
            WHERE {' AND '.join(wh)}
            ORDER BY p.plan_date, p.id
        """, tuple(params))
        rows = cur.fetchall()
        for r in rows:
            if r.get("plan_date"):
                r["plan_date"] = r["plan_date"].strftime("%Y-%m-%d")
            r["logged"] = bool(r.get("logged"))
            r["purposes"] = (r.get("purposes") or "")
        cur.close(); conn.close()
        return jsonify(rows)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/meeting_plan")
def meeting_plan_create():
    """Schedule one ship_to on a date.  Body (JSON or form):
    {date: YYYY-MM-DD, ship_to: <code>, bde_name: <name> }
    bde_name is sent by the client (selected in the form) — for BDE
    role we override with their own name to prevent scheduling on
    behalf of someone else."""
    try:
        body  = request.get_json(silent=True) or {}
        date  = (body.get("date")     or request.form.get("date")     or "").strip()
        ship  = (body.get("ship_to")  or request.form.get("ship_to")  or "").strip()
        bdenm = (body.get("bde_name") or request.form.get("bde_name") or "").strip()
        pgrp  = (body.get("plan_group") or request.form.get("plan_group") or "").strip()[:60]
        if not date or not ship:
            return jsonify({"error": "date and ship_to are required"}), 400
        try:
            plan_d = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400

        # BDE-scope: force the planner to be the signed-in BDE.
        email = _bde_from_request()
        found = _EMAIL_TO_DIR.get(email.lower()) if email else None
        if found and found[2] == "BDE":
            bdenm = found[0]
            bde_email = email
        else:
            bde_email = email or ""
        # Without a BDE name the chip would be impossible to attribute
        # and the BDE-role calendar filter would hide it.  Return a
        # clear error instead of silently inserting an orphan row —
        # makes mis-configured directory entries obvious instead of
        # producing 'drag-drop does nothing' bug reports.
        if not bdenm.strip():
            return jsonify({"error":
                "Cannot find your BDE name.  "
                "Pick a BDE in the form first, or ask the admin to add "
                "your email to the directory."}), 400

        # Resolve sold_to + names from customer master so the calendar
        # chip can show a friendly label without a separate lookup.
        # Potential customers (POT-<id>) come from a different table.
        sold_to, sold_to_name, ship_to_name = "", "", ""
        try:
            conn = get_connection(); cur = conn.cursor(dictionary=True)
            pid = _parse_potential_id(ship)
            if pid is not None:
                cur.execute("SELECT name, sold_to_name FROM potential_customer "
                            "WHERE id = %s", (pid,))
                r = cur.fetchone()
                if r:
                    ship_to_name = r.get("name") or ""
                    sold_to_name = r.get("sold_to_name") or ""
            else:
                cur.execute("""
                    SELECT MIN(sold_to)      AS sold_to,
                           MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name,
                           MIN(NULLIF(TRIM(ship_to_name),'')) AS ship_to_name
                    FROM customer
                    WHERE ship_to = %s
                """, (ship,))
                r = cur.fetchone()
                if r:
                    sold_to      = str(r.get("sold_to") or "")
                    sold_to_name = r.get("sold_to_name") or ""
                    ship_to_name = r.get("ship_to_name") or ""
            cur.close(); conn.close()
        except Exception:
            pass

        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO meeting_plan
                (plan_date, ship_to, sold_to, ship_to_name, sold_to_name,
                 bde_email, bde_name, plan_group)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (plan_d, ship[:64], sold_to[:64], ship_to_name[:160], sold_to_name[:160],
              bde_email[:120], bdenm[:120], pgrp))
        new_id = cur.lastrowid
        conn.commit(); cur.close(); conn.close()
        return jsonify({"ok": True, "id": new_id,
                        "plan_date":   plan_d.strftime("%Y-%m-%d"),
                        "ship_to":     ship,
                        "ship_to_name": ship_to_name,
                        "sold_to":     sold_to,
                        "sold_to_name": sold_to_name,
                        "bde_name":    bdenm,
                        "bde_email":   bde_email,
                        "plan_group":  pgrp})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.delete("/api/meeting_plan/<int:plan_id>")
def meeting_plan_delete(plan_id):
    try:
        # If signed-in user is a BDE, restrict delete to their own rows.
        email = _bde_from_request()
        found = _EMAIL_TO_DIR.get(email.lower()) if email else None
        bde_scope = found[0] if (found and found[2] == "BDE") else None

        conn = get_connection(); cur = conn.cursor()
        if bde_scope:
            cur.execute("DELETE FROM meeting_plan "
                        "WHERE id = %s AND UPPER(bde_name) = %s",
                        (plan_id, bde_scope.upper()))
        else:
            cur.execute("DELETE FROM meeting_plan WHERE id = %s", (plan_id,))
        n = cur.rowcount
        conn.commit(); cur.close(); conn.close()
        return jsonify({"ok": True, "deleted": n})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/meeting_plan/shop_to_options")
def meeting_plan_shop_options():
    """Slim ship_to list for the drag-and-drop palette.  Returns one
    row per (ship_to, sold_to_name) combination so a ship_to that
    sits under multiple parents in the customer master shows up under
    each parent's Sold-to filter.  The frontend dedupes by ship_to
    for the unfiltered view.

    Also carries a postcode + region so the Region search can group
    ship_tos into region cards.  Postcode comes from customer.postcode
    when the column exists, otherwise the last 4-digit token in
    address_1 (matches the same fallback the fleet-demand endpoint
    uses).  Region falls back to bde_state → ship_to_state → 'COMMON'."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cust_cols = _list_columns(cur, "customer")
        has_pc      = "postcode" in cust_cols
        has_addr1   = "address_1" in cust_cols
        has_bde_st  = "bde_state" in cust_cols
        has_ship_st = "ship_to_state" in cust_cols
        has_city    = "city" in cust_cols
        pc_expr = "NULL"
        if has_pc:
            pc_expr = "NULLIF(TRIM(postcode),'')"
        elif has_addr1:
            # Grab the last 4-digit token from address_1 as postcode.
            pc_expr = ("NULLIF(TRIM(REGEXP_SUBSTR(address_1, '[0-9]{4}(?![0-9])')),'')"
                       if _mysql_supports_regexp(cur) else "NULL")
        # region: prefer bde_state, then ship_to_state, else 'COMMON'.
        region_expr = "'COMMON'"
        parts = []
        if has_bde_st:  parts.append("NULLIF(TRIM(bde_state),'')")
        if has_ship_st: parts.append("NULLIF(TRIM(ship_to_state),'')")
        if parts:
            region_expr = "COALESCE(" + ", ".join(parts) + ", 'COMMON')"
        # City fed straight through — the palette groups (city, postcode)
        # pairs so a city that spans multiple postcodes shows up as
        # multiple entries in the region-search suggest dropdown.
        city_expr = "NULLIF(TRIM(city),'')" if has_city else "NULL"
        cur.execute(f"""
            SELECT ship_to,
                   MIN(NULLIF(TRIM(ship_to_name),'')) AS ship_to_name,
                   sold_to,
                   NULLIF(TRIM(sold_to_name),'')      AS sold_to_name,
                   MIN({pc_expr})                     AS postcode,
                   MIN({region_expr})                 AS region,
                   MIN({city_expr})                   AS city
            FROM customer
            WHERE ship_to IS NOT NULL AND TRIM(ship_to_name) <> ''
            GROUP BY ship_to, sold_to, sold_to_name
            ORDER BY ship_to_name, sold_to_name
            LIMIT 8000
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify(rows)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _mysql_supports_regexp(cur):
    """MySQL 8+ supports REGEXP_SUBSTR; 5.7 does not.  Probe once."""
    try:
        cur.execute("SELECT REGEXP_SUBSTR('abc1234', '[0-9]{4}')")
        cur.fetchone()
        return True
    except Exception:
        return False

@app.get("/api/meeting/list")
def meeting_list():
    """Recent entries.  Optional ?ship_to=… / ?bde=… / ?sold_to=… /
    ?limit=N filters."""
    try:
        ship_to = (request.args.get("ship_to") or "").strip()
        bde     = (request.args.get("bde")     or "").strip()
        sold_to = (request.args.get("sold_to") or "").strip()
        limit   = max(1, min(500, int(request.args.get("limit", 100) or 100)))

        wh, params = [], []
        if ship_to:
            # Partial match against either the ship_to code OR the
            # shop name (from customer master via the LEFT JOIN below).
            like = f"%{ship_to}%"
            wh.append("(m.ship_to LIKE %s "
                      "  OR EXISTS (SELECT 1 FROM customer cc "
                      "             WHERE cc.ship_to = m.ship_to "
                      "             AND TRIM(cc.ship_to_name) LIKE %s))")
            params.extend([like, like])
        if bde:
            wh.append("m.bde_name = %s"); params.append(bde)
        if sold_to:
            # Partial match against any of: sold_to code, the
            # denormalised sold_to_name stored on the log row, and
            # the customer master's sold_to_name for the ship_to.
            like = f"%{sold_to}%"
            wh.append("(m.sold_to LIKE %s "
                      "  OR TRIM(m.sold_to_name) LIKE %s "
                      "  OR EXISTS (SELECT 1 FROM customer cc "
                      "             WHERE cc.ship_to = m.ship_to "
                      "             AND TRIM(cc.sold_to_name) LIKE %s))")
            params.extend([like, like, like])
        where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""

        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT m.id, m.created_at, m.visit_date,
                   m.bde_email, m.bde_name,
                   m.plan_bde_name, m.plan_bde_email,
                   m.sold_to, m.ship_to, m.visit_purpose,
                   m.met_person, m.met_person_contact,
                   m.notes, m.next_action, m.feedback, m.prep_notes, m.photo_paths,
                   COALESCE(NULLIF(TRIM(c.ship_to_name),''), m.ship_to) AS ship_to_name,
                   COALESCE(NULLIF(TRIM(c.sold_to_name),''),
                            NULLIF(TRIM(m.sold_to_name),''),
                            m.sold_to) AS sold_to_name,
                   c.contact_person AS contact_person,
                   c.phone          AS contact_phone,
                   -- Closest plan_date within the same month, for the same
                   -- ship_to, so the panel can show 'Planned vs Visited'.
                   (SELECT p.plan_date
                    FROM meeting_plan p
                    WHERE p.ship_to = m.ship_to
                      AND YEAR(p.plan_date)  = YEAR(m.visit_date)
                      AND MONTH(p.plan_date) = MONTH(m.visit_date)
                    ORDER BY ABS(DATEDIFF(p.plan_date, m.visit_date)), p.plan_date
                    LIMIT 1) AS plan_date
            FROM meeting_log m
            LEFT JOIN (
                SELECT ship_to,
                       MIN(NULLIF(TRIM(ship_to_name),'')) AS ship_to_name,
                       MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name,
                       CAST(NULL AS CHAR) AS contact_person,
                       CAST(NULL AS CHAR) AS phone
                FROM customer GROUP BY ship_to
                UNION ALL
                -- Potential customers come through as POT-<id> in ship_to;
                -- expose both names AND the contact info BDEs typed when
                -- creating the prospect so the panel can show how to
                -- re-contact them without a separate lookup.
                SELECT CONCAT('POT-', id) AS ship_to,
                       name              AS ship_to_name,
                       sold_to_name      AS sold_to_name,
                       contact_person    AS contact_person,
                       phone             AS phone
                FROM potential_customer
            ) c ON c.ship_to = m.ship_to
            {where_sql}
            ORDER BY m.bde_name ASC, m.ship_to ASC,
                     m.visit_purpose ASC, m.visit_date DESC, m.created_at DESC
            LIMIT {limit}
        """, tuple(params))
        rows = cur.fetchall()
        ids = [r["id"] for r in rows] if rows else []
        thread_by_meeting = {}
        if ids:
            placeholders = ",".join(["%s"] * len(ids))
            cur.execute(f"""
                SELECT id, meeting_id, created_at, author_email, author_name, text
                FROM meeting_feedback
                WHERE meeting_id IN ({placeholders})
                ORDER BY created_at ASC, id ASC
            """, tuple(ids))
            for t in cur.fetchall():
                if t.get("created_at"):
                    t["created_at"] = t["created_at"].strftime("%Y-%m-%d %H:%M")
                thread_by_meeting.setdefault(t["meeting_id"], []).append(t)
        for r in rows:
            r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else ""
            r["visit_date"] = r["visit_date"].strftime("%Y-%m-%d") if r["visit_date"] else ""
            if r.get("plan_date"):
                r["plan_date"] = r["plan_date"].strftime("%Y-%m-%d")
            r["photo_paths"] = r["photo_paths"].split(",") if r["photo_paths"] else []
            r["thread"]      = thread_by_meeting.get(r["id"], [])
            # Frontend State filter uses this — group SA/TAS under VIC
            # (same as the State Manager mapping) so the 4 main state
            # buttons cover every BDE.
            st = _resolve_bde_state(r.get("bde_name") or "") or ""
            r["bde_state"] = "VIC" if st in ("SA", "TAS") else st
        cur.close(); conn.close()
        return jsonify(rows)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────
# AI feature — Summary / Strategy generator for one meeting_log row.
# Self-contained block; remove everything between the START / END
# markers (and the matching frontend section in static/meeting.html) to
# fully revert.  When ANTHROPIC_API_KEY is unset, the feature stays
# dormant: GET /api/ai_status returns enabled=false and the frontend
# hides the button.
# ─────────────────────────────────────────────────────────────────────
# === AI FEATURE START ===
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
AI_MODEL          = os.getenv("AI_MODEL", "claude-haiku-4-5")
# Cap on rows the digest endpoint will accept in one call so token
# spend per click stays predictable.
AI_DIGEST_MAX_ROWS = int(os.getenv("AI_DIGEST_MAX_ROWS", "30"))

def _ai_fmt_row(r):
    bits = []
    if r.get("visit_date"):    bits.append(f"[{r['visit_date']}]")
    if r.get("bde_name"):      bits.append(f"BDE: {r['bde_name']}")
    if r.get("visit_purpose"): bits.append(f"Purpose: {r['visit_purpose']}")
    if r.get("met_person"):    bits.append(f"Met: {r['met_person']}")
    if r.get("ship_to"):       bits.append(f"Shop: {r['ship_to']}")
    if r.get("sold_to_name") or r.get("sold_to"):
        bits.append(f"Customer: {r.get('sold_to_name') or r.get('sold_to')}")
    lines = [" | ".join(bits)] if bits else []
    if r.get("prep_notes"):  lines.append(f"  Prep:  {r['prep_notes']}")
    if r.get("notes"):       lines.append(f"  Notes: {r['notes']}")
    if r.get("next_action"): lines.append(f"  Next:  {r['next_action']}")
    return "\n".join(lines)

def _ai_call_claude(prompt):
    if not ANTHROPIC_API_KEY:
        return None, "AI disabled — ANTHROPIC_API_KEY not configured on the server"
    import requests as _rq
    try:
        r = _rq.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": AI_MODEL,
                "max_tokens": 1200,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=60,
        )
        if r.status_code != 200:
            return None, f"Anthropic API {r.status_code}: {r.text[:300]}"
        data = r.json()
        text = "".join(b.get("text", "") for b in (data.get("content") or []))
        return text.strip(), None
    except Exception as e:
        return None, f"AI call error: {e}"

def _ai_parse(text):
    if not text: return "", ""
    import re as _re
    parts = _re.split(r"(?im)^\s*STRATEGY\s*:?\s*", text, maxsplit=1)
    if len(parts) == 2:
        summary = _re.sub(r"(?im)^\s*SUMMARY\s*:?\s*", "", parts[0]).strip()
        strategy = parts[1].strip()
    else:
        summary = _re.sub(r"(?im)^\s*SUMMARY\s*:?\s*", "", text).strip()
        strategy = ""
    return summary, strategy

@app.get("/api/ai_status")
def ai_status():
    return jsonify({"enabled": bool(ANTHROPIC_API_KEY), "model": AI_MODEL})

# ─── Visit Impact ────────────────────────────────────────────────────
# "Did this BDE visit actually move the needle on sales?" — answered
# by comparing the shop's sales in the month BEFORE a visit vs the
# month AFTER.  Strict pre/post, no in-month overlap, so causality is
# at least defensible (the model: "BDE visited mid-April; did May
# sales tick up vs March?").  We don't claim causation in the UI —
# this is correlation surfaced to the BDE/manager to spot patterns.
def _sales_table_for_ym(year, month):
    """Return the sales-table name that holds a given (year, month),
    or None if we don't have that month on disk."""
    try:
        now = datetime.now()
        if int(year) == now.year and int(month) == now.month:
            return "sales_thismonth"
        yymm = f"{int(year) % 100:02d}{int(month):02d}"
        table = f"sales_{yymm}"
        if table in REBATE_SALES_TABLES.values():
            return table
    except Exception:
        pass
    return None

def _sales_for_month(cur, ship_to, year, month):
    """SUM(qty), SUM(amt) for ship_to in (year, month).  Returns None
    when the table for that month doesn't exist on this server."""
    table = _sales_table_for_ym(year, month)
    if not table:
        return None
    try:
        cur.execute(
            "SELECT COALESCE(SUM(qty),0) AS qty, COALESCE(SUM(amt),0) AS amt "
            "FROM " + table + " "
            "WHERE ship_to = %s AND brand IN ('HK','LF')",
            (ship_to,)
        )
        r = cur.fetchone() or {}
        return {"qty": float(r.get("qty") or 0), "amt": float(r.get("amt") or 0)}
    except Exception:
        return None

@app.post("/api/meeting/impact")
def meeting_impact():
    """Batch pre/post sales delta for a list of meeting_log IDs.
    Body: { "ids": [int, ...] }
    Returns: { "impacts": { "<id>": {pre_year_month, post_year_month,
                                     pre, post, delta_qty, delta_amt,
                                     delta_qty_pct, delta_amt_pct,
                                     status}, ... } }
    status ∈ {ok, post_pending, pre_unavailable}.
    """
    body = request.get_json(silent=True) or {}
    raw_ids = body.get("ids") or []
    ids = []
    for x in raw_ids:
        try:
            ids.append(int(x))
        except Exception:
            pass
    if not ids:
        return jsonify({"impacts": {}})
    ids = ids[:200]   # cap so a chatty client can't tie up the DB

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        placeholders = ",".join(["%s"] * len(ids))
        cur.execute(
            f"SELECT id, ship_to, visit_date FROM meeting_log "
            f"WHERE id IN ({placeholders}) AND visit_date IS NOT NULL "
            f"  AND ship_to IS NOT NULL AND ship_to <> ''",
            tuple(ids)
        )
        visits = cur.fetchall()

        impacts = {}
        for v in visits:
            vd = v.get("visit_date")
            ship_to = v.get("ship_to")
            if not vd or not ship_to:
                continue
            year, month = vd.year, vd.month
            # Previous calendar month.
            py = year - 1 if month == 1 else year
            pm = 12 if month == 1 else month - 1
            # Next calendar month.
            ny = year + 1 if month == 12 else year
            nm = 1 if month == 12 else month + 1

            pre  = _sales_for_month(cur, ship_to, py, pm)
            post = _sales_for_month(cur, ship_to, ny, nm)

            entry = {
                "pre_year_month":  f"{py}-{pm:02d}",
                "post_year_month": f"{ny}-{nm:02d}",
            }
            if pre is None:
                entry["status"] = "pre_unavailable"
                impacts[str(v["id"])] = entry
                continue
            entry["pre"] = pre
            if post is None:
                entry["status"] = "post_pending"
                impacts[str(v["id"])] = entry
                continue
            entry["post"] = post
            entry["delta_qty"] = round(post["qty"] - pre["qty"], 2)
            entry["delta_amt"] = round(post["amt"] - pre["amt"], 2)
            entry["delta_qty_pct"] = round(((post["qty"] - pre["qty"]) / pre["qty"] * 100), 1) if pre["qty"] > 0 else None
            entry["delta_amt_pct"] = round(((post["amt"] - pre["amt"]) / pre["amt"] * 100), 1) if pre["amt"] > 0 else None
            entry["status"] = "ok"
            impacts[str(v["id"])] = entry

        cur.close(); conn.close()
        return jsonify({"impacts": impacts})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.get("/api/meeting/impact_summary")
def meeting_impact_summary():
    """Roll every meeting_log entry's pre/post sales delta up to the
    BDE that logged it.

      ?from=YYYY-MM-DD   default: 90 days ago
      ?to=YYYY-MM-DD     default: today
      ?bde=<name>        optional single-BDE drilldown

    Returns per-BDE totals: visits, up/flat/down counts, pre/post
    sums, delta $ and %, plus the top-3 winning and top-3 losing
    shop names inside the window.  Scope respects the same
    BDE / SM / ALL locks as /api/meeting_plan/list."""
    from datetime import date as _date, timedelta as _timedelta
    q_from = (request.args.get("from") or "").strip()
    q_to   = (request.args.get("to")   or "").strip()
    q_bde  = (request.args.get("bde")  or "").strip()
    try:
        d_from = datetime.strptime(q_from, "%Y-%m-%d").date() if q_from \
                 else (_date.today() - _timedelta(days=90))
        d_to   = datetime.strptime(q_to,   "%Y-%m-%d").date() if q_to   \
                 else _date.today()
    except Exception:
        return jsonify({"error": "from/to must be YYYY-MM-DD"}), 400
    if d_from > d_to:
        return jsonify({"error": "from must be <= to"}), 400

    # Scope — same as meeting_plan_list.
    email = _bde_from_request()
    me    = _EMAIL_TO_DIR.get(email.lower()) if email else None
    role  = me[2] if me else "ALL"
    scope_bde   = me[0] if me else None
    scope_state = me[1] if me else None

    # Build the WHERE with the `m.` alias baked in from the start so
    # the query below can reference the meeting_log fields safely.
    wh = ["m.visit_date BETWEEN %s AND %s",
          "m.ship_to IS NOT NULL", "m.ship_to <> ''"]
    params = [d_from, d_to]
    if role == "BDE" and scope_bde:
        wh.append("UPPER(m.bde_name) = %s"); params.append(scope_bde.upper())
    elif role == "SM" and scope_state:
        state_bdes = [n for (n, _e, s, _r) in _BDE_DIRECTORY if s == scope_state]
        if state_bdes:
            placeholders = ",".join(["%s"] * len(state_bdes))
            wh.append(f"UPPER(m.bde_name) IN ({placeholders})")
            params.extend([n.upper() for n in state_bdes])
    if q_bde:
        wh.append("UPPER(m.bde_name) = %s"); params.append(q_bde.upper())
    where_sql = "WHERE " + " AND ".join(wh)

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        # meeting_log carries ship_to code only; the readable name lives
        # on customer master.  Correlated subquery keeps rows whose
        # ship_to isn't in customer (potential customers, deleted rows).
        cur.execute(f"""
            SELECT m.id, m.bde_name, m.ship_to, m.visit_date,
                   (SELECT MIN(NULLIF(TRIM(c.ship_to_name),''))
                    FROM customer c WHERE c.ship_to = m.ship_to) AS ship_to_name
            FROM meeting_log m
            {where_sql}
            ORDER BY m.visit_date DESC
        """, tuple(params))
        visits = cur.fetchall()

        # Aggregate per BDE — reuse _sales_for_month for consistency.
        bdes = {}   # bde_name → dict
        for v in visits:
            bde = (v.get("bde_name") or "—").strip() or "—"
            g = bdes.setdefault(bde, {
                "bde_name": bde,
                "visits": 0, "up": 0, "flat": 0, "down": 0,
                "pre_qty": 0.0, "post_qty": 0.0,
                "pre_amt": 0.0, "post_amt": 0.0,
                "wins":   [],  # (delta_pct, ship_to_name)
                "losses": [],
                "pending": 0, "pre_missing": 0,
            })
            g["visits"] += 1
            vd = v.get("visit_date")
            if not vd: continue
            year, month = vd.year, vd.month
            py = year - 1 if month == 1 else year
            pm = 12 if month == 1 else month - 1
            ny = year + 1 if month == 12 else year
            nm = 1 if month == 12 else month + 1
            pre  = _sales_for_month(cur, v["ship_to"], py, pm)
            post = _sales_for_month(cur, v["ship_to"], ny, nm)
            if pre is None:
                g["pre_missing"] += 1
                continue
            if post is None:
                g["pending"] += 1
                continue
            g["pre_qty"]  += pre["qty"];  g["post_qty"] += post["qty"]
            g["pre_amt"]  += pre["amt"];  g["post_amt"] += post["amt"]
            dq  = post["qty"] - pre["qty"]
            dpct = ((post["qty"] - pre["qty"]) / pre["qty"] * 100) if pre["qty"] > 0 else None
            if dq > 0 and (dpct is None or dpct >= 1):
                g["up"] += 1
                if dpct is not None:
                    g["wins"].append((dpct, v.get("ship_to_name") or v["ship_to"]))
            elif dq < 0:
                g["down"] += 1
                if dpct is not None:
                    g["losses"].append((dpct, v.get("ship_to_name") or v["ship_to"]))
            else:
                g["flat"] += 1
        cur.close(); conn.close()

        # BDE-name → state lookup so state and grand totals can be
        # rolled up alongside the per-BDE rows.
        _bde_state_lc = {n.upper(): s for (n, _e, s, _r) in _BDE_DIRECTORY}
        # Finalise: round + take top 3 wins / bottom 3 losses.
        rows = []
        for g in bdes.values():
            g["state"] = _bde_state_lc.get(g["bde_name"].upper(), "")
            g["delta_qty"] = round(g["post_qty"] - g["pre_qty"], 1)
            g["delta_amt"] = round(g["post_amt"] - g["pre_amt"], 0)
            g["delta_pct"] = round((g["post_qty"] - g["pre_qty"]) / g["pre_qty"] * 100, 1) \
                             if g["pre_qty"] > 0 else None
            g["wins"].sort(key=lambda t: -t[0])
            g["losses"].sort(key=lambda t: t[0])
            g["top_wins"]   = [{"ship_to_name": n, "delta_pct": round(p, 1)}
                               for (p, n) in g["wins"][:3]]
            g["top_losses"] = [{"ship_to_name": n, "delta_pct": round(p, 1)}
                               for (p, n) in g["losses"][:3]]
            del g["wins"]; del g["losses"]
            for k in ("pre_qty", "post_qty"):  g[k] = round(g[k], 1)
            for k in ("pre_amt", "post_amt"):  g[k] = round(g[k], 0)
            rows.append(g)
        # Order by total visits desc so the busiest BDE lands on top.
        rows.sort(key=lambda r: -r["visits"])

        # Roll up state totals and one grand total from the per-BDE
        # rows.  Frontend renders these as separate table sections so
        # the reader can eyeball state-level performance without
        # having to sum in their head.
        def _agg(pool):
            if not pool:
                return None
            visits = sum(r["visits"] for r in pool)
            up     = sum(r["up"]     for r in pool)
            flat   = sum(r["flat"]   for r in pool)
            down   = sum(r["down"]   for r in pool)
            pending     = sum(r.get("pending", 0)     for r in pool)
            pre_missing = sum(r.get("pre_missing", 0) for r in pool)
            pre_qty  = round(sum(r["pre_qty"]  for r in pool), 1)
            post_qty = round(sum(r["post_qty"] for r in pool), 1)
            pre_amt  = round(sum(r["pre_amt"]  for r in pool), 0)
            post_amt = round(sum(r["post_amt"] for r in pool), 0)
            delta_qty = round(post_qty - pre_qty, 1)
            delta_amt = round(post_amt - pre_amt, 0)
            delta_pct = round((post_qty - pre_qty) / pre_qty * 100, 1) if pre_qty > 0 else None
            return {
                "visits": visits, "up": up, "flat": flat, "down": down,
                "pending": pending, "pre_missing": pre_missing,
                "pre_qty": pre_qty, "post_qty": post_qty,
                "pre_amt": pre_amt, "post_amt": post_amt,
                "delta_qty": delta_qty, "delta_amt": delta_amt,
                "delta_pct": delta_pct,
                "bde_count": len(pool),
            }
        by_state = {}
        for r in rows:
            s = r["state"] or "—"
            by_state.setdefault(s, []).append(r)
        states = []
        # NSW → QLD → VIC → WA → SA → TAS → NT → ACT → 기타 순.
        _STATE_ORDER = {"NSW":0,"QLD":1,"VIC":2,"WA":3,"SA":4,"TAS":5,"NT":6,"ACT":7}
        for st in sorted(by_state.keys(),
                         key=lambda k: (_STATE_ORDER.get(k, 99), k)):
            agg = _agg(by_state[st])
            if agg is not None:
                agg["state"] = st
                states.append(agg)
        total = _agg(rows) or {}
        if total:
            total["state"] = "TOTAL"
        return jsonify({
            "from": d_from.strftime("%Y-%m-%d"),
            "to":   d_to.strftime("%Y-%m-%d"),
            "bdes":   rows,
            "states": states,
            "total":  total,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── Shop Briefing Card ───────────────────────────────────────────────
# Per-shop summary view a BDE pulls up on the phone right before walking
# into a shop (or before calling its owner): last visit recap, recent
# contacts, 6-month sales trend, and a one-tap call/SMS/email row.  The
# rebate next-tier figures live on /api/rebate_data already and are
# fetched separately from the frontend so we don't duplicate that calc
# (and so the briefing renders even if rebate calc is slow).
@app.get("/shop/<ship_to>")
def shop_briefing_page(ship_to):
    return send_from_directory("static", "shop_briefing.html")

@app.get("/api/shop_briefing/<ship_to>")
def shop_briefing_data(ship_to):
    if not ship_to:
        return jsonify({"error": "ship_to required"}), 400
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        # 1. Shop master ----------------------------------------------------
        cur.execute(
            "SELECT ship_to, ship_to_name, sold_to, sold_to_name, "
            "       bde_state, salesman_name "
            "FROM customer WHERE ship_to = %s LIMIT 1",
            (ship_to,),
        )
        shop = cur.fetchone() or {"ship_to": ship_to}

        # 2. Last meeting log -----------------------------------------------
        cur.execute(
            "SELECT id, visit_date, bde_name, bde_email, visit_purpose, "
            "       met_person, met_person_contact, prep_notes, notes, "
            "       next_action, created_at "
            "FROM meeting_log "
            "WHERE ship_to = %s "
            "  AND notes IS NOT NULL AND TRIM(notes) <> '' "
            "ORDER BY visit_date DESC, id DESC "
            "LIMIT 1",
            (ship_to,),
        )
        last_meeting = cur.fetchone()
        if last_meeting:
            for k in ("visit_date", "created_at"):
                if last_meeting.get(k):
                    last_meeting[k] = last_meeting[k].strftime("%Y-%m-%d")

        # 3. Distinct met person + contact pairs (history, newest first) -----
        # Many BDEs talk to multiple people at the same shop over time —
        # the briefing surfaces all of them so a one-tap dial picks any.
        cur.execute(
            "SELECT met_person, met_person_contact, MAX(visit_date) AS last_seen "
            "FROM meeting_log "
            "WHERE ship_to = %s "
            "  AND ((met_person IS NOT NULL AND TRIM(met_person) <> '') "
            "    OR (met_person_contact IS NOT NULL AND TRIM(met_person_contact) <> '')) "
            "GROUP BY met_person, met_person_contact "
            "ORDER BY last_seen DESC "
            "LIMIT 10",
            (ship_to,),
        )
        contacts = []
        for r in cur.fetchall():
            person  = (r.get("met_person") or "").strip()
            contact = (r.get("met_person_contact") or "").strip()
            if not person and not contact:
                continue
            contacts.append({
                "name":      person,
                "contact":   contact,
                "is_email":  "@" in contact,
                "last_seen": r["last_seen"].strftime("%Y-%m-%d") if r.get("last_seen") else "",
            })

        # 4. 6-month sales trend --------------------------------------------
        # Loop through the per-month REBATE_SALES_TABLES + sales_thismonth
        # so the trend always reflects whatever ETL has produced so far —
        # don't hand-code month names.  Tables that don't exist yet (e.g.
        # before this fiscal-year refresh) silently contribute zero.
        trend_specs = [
            ("Jan", "sales_2601"),
            ("Feb", "sales_2602"),
            ("Mar", "sales_2603"),
            ("Apr", "sales_2604"),
            ("May", "sales_2605"),
            ("Jun", "sales_thismonth"),
        ]
        sales_trend = []
        for label, tbl in trend_specs:
            try:
                cur.execute(
                    "SELECT COALESCE(SUM(qty),0) AS qty, "
                    "       COALESCE(SUM(amt),0) AS amt "
                    "FROM " + tbl + " "
                    "WHERE ship_to = %s "
                    "  AND brand IN ('HK','LF')",
                    (ship_to,)
                )
                r = cur.fetchone() or {"qty": 0, "amt": 0}
                sales_trend.append({
                    "month": label,
                    "table": tbl,
                    "qty":   float(r.get("qty") or 0),
                    "amt":   float(r.get("amt") or 0),
                })
            except Exception:
                # Table missing for that month — show zero rather than 500.
                sales_trend.append({"month": label, "table": tbl, "qty": 0, "amt": 0})

        # 5. Meeting count (past 12 months) ----------------------------------
        cur.execute(
            "SELECT COUNT(*) AS n FROM meeting_log "
            "WHERE ship_to = %s AND visit_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)",
            (ship_to,)
        )
        rcnt = cur.fetchone() or {"n": 0}
        meeting_count_12m = int(rcnt.get("n") or 0)

        cur.close(); conn.close()

        return jsonify({
            "shop":              shop,
            "last_meeting":      last_meeting,
            "contacts":          contacts,
            "sales_trend":       sales_trend,
            "meeting_count_12m": meeting_count_12m,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── Admin: dashboard usage analytics ────────────────────────────────
# Read-only view onto request_log so we can quantify who's using the
# dashboard and which views matter.  Restricted to ALL-role users (the
# small admin set in _BDE_DIRECTORY — Hayden / JJ / Jayden).
def _is_admin_request():
    email = (_bde_from_request() or "").strip().lower()
    if not email:
        return False
    me = _EMAIL_TO_DIR.get(email)
    # role is the third tuple element; "ALL" = unscoped admin
    return bool(me and me[2] == "ALL")

@app.get("/admin/usage")
def admin_usage_page():
    if not _is_admin_request():
        return ("Forbidden — admin only.", 403)
    return send_from_directory("static", "admin_usage.html")

@app.get("/api/admin/usage")
def admin_usage_data():
    """Aggregate request_log over the requested window (default 30 days).
    Returns top paths, top users, daily totals, and per-user activity
    counts so the admin_usage.html page can render without further
    server round-trips."""
    if not _is_admin_request():
        return jsonify({"error": "admin only"}), 403
    try:
        days = int(request.args.get("days") or 30)
    except Exception:
        days = 30
    days = max(1, min(days, 365))

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        cur.execute(
            "SELECT COUNT(*) AS total, "
            "COUNT(DISTINCT user_email) AS uniq_users "
            "FROM request_log "
            "WHERE created_at >= NOW() - INTERVAL %s DAY",
            (days,)
        )
        summary = cur.fetchone() or {"total": 0, "uniq_users": 0}

        # Top paths — strip the cache-busting query string by indexing on
        # `path` only (already query-less since request.path).  Group / and
        # trailing-slash variants together by trimming a trailing slash.
        cur.execute(
            "SELECT path, COUNT(*) AS hits, "
            "COUNT(DISTINCT user_email) AS uniq_users "
            "FROM request_log "
            "WHERE created_at >= NOW() - INTERVAL %s DAY "
            "GROUP BY path "
            "ORDER BY hits DESC "
            "LIMIT 30",
            (days,)
        )
        top_paths = cur.fetchall()

        # Top users — include anonymous rows (empty email) too so the
        # unique-users KPI and this list always reconcile.  Subquery
        # picks the most common (country, IP) per user so the admin can
        # see "where they came from" without scanning every row.
        cur.execute(
            "SELECT r.user_email, COUNT(*) AS hits, "
            "       COUNT(DISTINCT DATE(r.created_at)) AS active_days, "
            "       MAX(r.created_at) AS last_seen, "
            "       (SELECT country FROM request_log r2 "
            "         WHERE r2.user_email = r.user_email "
            "           AND r2.created_at >= NOW() - INTERVAL %s DAY "
            "         GROUP BY country ORDER BY COUNT(*) DESC LIMIT 1) AS country, "
            "       (SELECT ip FROM request_log r3 "
            "         WHERE r3.user_email = r.user_email "
            "           AND r3.created_at >= NOW() - INTERVAL %s DAY "
            "         GROUP BY ip ORDER BY COUNT(*) DESC LIMIT 1) AS ip "
            "FROM request_log r "
            "WHERE r.created_at >= NOW() - INTERVAL %s DAY "
            "GROUP BY r.user_email "
            "ORDER BY hits DESC "
            "LIMIT 50",
            (days, days, days)
        )
        top_users = cur.fetchall()
        for r in top_users:
            if r.get("last_seen"):
                r["last_seen"] = r["last_seen"].strftime("%Y-%m-%d %H:%M")
            if not r.get("user_email"):
                # Empty-email rows = requests that didn't carry a
                # Cf-Access-Authenticated-User-Email header.  Most
                # commonly: the public /claim/* customer portal, or
                # health-check hits from Cloudflare itself.
                r["user_email"] = "(no auth — public /claim or health check)"
            r["country"] = r.get("country") or ""
            r["ip"] = r.get("ip") or ""

        cur.execute(
            "SELECT DATE(created_at) AS d, COUNT(*) AS hits, "
            "COUNT(DISTINCT user_email) AS uniq_users "
            "FROM request_log "
            "WHERE created_at >= NOW() - INTERVAL %s DAY "
            "GROUP BY DATE(created_at) "
            "ORDER BY d",
            (days,)
        )
        daily = cur.fetchall()
        for r in daily:
            if r.get("d"):
                r["d"] = r["d"].strftime("%Y-%m-%d")

        cur.close(); conn.close()
        return jsonify({
            "days": days,
            "total": int(summary.get("total") or 0),
            "uniq_users": int(summary.get("uniq_users") or 0),
            "top_paths": top_paths,
            "top_users": top_users,
            "daily": daily,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/admin/usage/user")
def admin_usage_user_detail():
    """Per-user drill-down: which paths did this user hit, when, and
    from where.  Drives the per-user activity panel on /admin/usage."""
    if not _is_admin_request():
        return jsonify({"error": "admin only"}), 403
    email = (request.args.get("email") or "").strip().lower()
    try:
        days = int(request.args.get("days") or 30)
    except Exception:
        days = 30
    days = max(1, min(days, 365))
    # The Top Users list shows "(no auth …)" for empty-email rows;
    # the frontend sends that label back verbatim — translate it into
    # the SQL-friendly empty string.
    if email.startswith("(no auth"):
        email = ""

    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)

        cur.execute(
            "SELECT path, COUNT(*) AS hits, MAX(created_at) AS last_seen "
            "FROM request_log "
            "WHERE user_email = %s "
            "  AND created_at >= NOW() - INTERVAL %s DAY "
            "GROUP BY path "
            "ORDER BY hits DESC "
            "LIMIT 30",
            (email, days)
        )
        paths = cur.fetchall()
        for r in paths:
            if r.get("last_seen"):
                r["last_seen"] = r["last_seen"].strftime("%Y-%m-%d %H:%M")

        cur.execute(
            "SELECT DATE(created_at) AS d, COUNT(*) AS hits "
            "FROM request_log "
            "WHERE user_email = %s "
            "  AND created_at >= NOW() - INTERVAL %s DAY "
            "GROUP BY DATE(created_at) ORDER BY d",
            (email, days)
        )
        daily = cur.fetchall()
        for r in daily:
            if r.get("d"):
                r["d"] = r["d"].strftime("%Y-%m-%d")

        cur.execute(
            "SELECT country, ip, COUNT(*) AS hits, MAX(created_at) AS last_seen "
            "FROM request_log "
            "WHERE user_email = %s "
            "  AND created_at >= NOW() - INTERVAL %s DAY "
            "GROUP BY country, ip "
            "ORDER BY hits DESC "
            "LIMIT 10",
            (email, days)
        )
        locations = cur.fetchall()
        for r in locations:
            if r.get("last_seen"):
                r["last_seen"] = r["last_seen"].strftime("%Y-%m-%d %H:%M")

        cur.close(); conn.close()
        return jsonify({
            "email": email or "(no auth)",
            "days": days,
            "paths": paths,
            "daily": daily,
            "locations": locations,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/meeting/voice_extract")
def meeting_voice_extract():
    """Convert a free-form voice transcript into structured meeting_log fields.
    Body: { "transcript": "<user's spoken recap>" }
    Returns: {
        notes, met_person, met_person_contact, next_action,
        visit_purpose ("" | "Promotion" | "Product introduction"
                       | "Claim support" | "Rebate follow-up" | "Stock" | "Other"),
        model
    }
    Each field is a string — empty when the model couldn't infer it.  The
    frontend only writes into empty form fields so user-typed values are
    never overwritten.
    """
    if not ANTHROPIC_API_KEY:
        return jsonify({"error": "AI feature disabled — ANTHROPIC_API_KEY is not set on the server"}), 503
    body = request.get_json(silent=True) or {}
    transcript = (body.get("transcript") or "").strip()
    if not transcript:
        return jsonify({"error": "transcript is empty"}), 400
    if len(transcript) > 4000:
        transcript = transcript[:4000]

    prompt = (
        "You extract structured fields from a BDE's spoken recap of a tyre-shop visit.\n"
        "Return ONLY a JSON object with these keys (all strings, empty when unknown):\n"
        '  - "notes": a clean written summary of what happened (1-3 sentences, in English).\n'
        '  - "met_person": the name or role of the person met (e.g. "John", "owner", "store manager"). Empty if not mentioned.\n'
        '  - "met_person_contact": phone or email if mentioned. Empty if not.\n'
        '  - "next_action": follow-up needed, if mentioned. Empty if not.\n'
        '  - "visit_purpose": MUST be exactly one of "Promotion", "Product introduction", '
        '"Claim support", "Rebate follow-up", "Stock", "Other", or empty string.\n'
        "    Pick the closest match based on what was discussed. Empty only if truly unclear.\n"
        "\n"
        "Transcript:\n"
        f"\"\"\"\n{transcript}\n\"\"\"\n"
        "\n"
        "Respond with ONLY the JSON object, no markdown, no commentary."
    )
    text, err = _ai_call_claude(prompt)
    if err:
        return jsonify({"error": err}), 502

    import json as _json, re as _re
    raw = (text or "").strip()
    m = _re.search(r"\{.*\}", raw, _re.DOTALL)
    if not m:
        return jsonify({"error": "AI returned no JSON", "raw": raw[:300]}), 502
    try:
        parsed = _json.loads(m.group(0))
    except Exception as e:
        return jsonify({"error": f"AI JSON parse failed: {e}", "raw": raw[:300]}), 502

    allowed_purposes = {"Promotion", "Product introduction", "Claim support",
                        "Rebate follow-up", "Stock", "Other", ""}
    out = {
        "notes":              str(parsed.get("notes") or "").strip(),
        "met_person":         str(parsed.get("met_person") or "").strip()[:120],
        "met_person_contact": str(parsed.get("met_person_contact") or "").strip()[:80],
        "next_action":        str(parsed.get("next_action") or "").strip(),
        "visit_purpose":      str(parsed.get("visit_purpose") or "").strip(),
        "model":              AI_MODEL,
    }
    if out["visit_purpose"] not in allowed_purposes:
        out["visit_purpose"] = ""
    return jsonify(out)

@app.post("/api/meeting/ai_digest")
def meeting_ai_digest():
    """Summarise an arbitrary set of meeting_log rows (the rows currently
    visible after the user's State/BDE/Purpose/Shop filters).  Body:
        { "ids": [int, int, ...] }
    Returns { summary, strategy, model, count }.
    """
    if not ANTHROPIC_API_KEY:
        return jsonify({"error": "AI feature disabled — ANTHROPIC_API_KEY is not set on the server"}), 503
    body = request.get_json(silent=True) or {}
    raw_ids = body.get("ids") or []
    try:
        ids = [int(x) for x in raw_ids if str(x).strip()]
    except (TypeError, ValueError):
        return jsonify({"error": "ids must be a list of integers"}), 400
    if not ids:
        return jsonify({"error": "No entries selected"}), 400
    if len(ids) > AI_DIGEST_MAX_ROWS:
        ids = ids[:AI_DIGEST_MAX_ROWS]
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        placeholders = ",".join(["%s"] * len(ids))
        cur.execute(f"""
            SELECT id, visit_date, bde_name,
                   sold_to, sold_to_name, ship_to, visit_purpose,
                   met_person, notes, next_action, prep_notes
            FROM meeting_log
            WHERE id IN ({placeholders})
            ORDER BY bde_name ASC, ship_to ASC,
                     visit_date DESC, created_at DESC
        """, tuple(ids))
        rows = cur.fetchall() or []
        cur.close(); conn.close()
        if not rows:
            return jsonify({"error": "No matching entries found"}), 404
        body_parts = [
            f"# {len(rows)} visit log entries to analyse "
            f"(the user has already narrowed the list with their filters)",
            "",
        ]
        for r in rows:
            body_parts.append(_ai_fmt_row(r))
            body_parts.append("")
        instructions = (
            "You are a sales-operations analyst for a tyre distributor in Australia.\n"
            "Treat the entries below as one cohort the user wants understood as a whole.\n"
            "Look for patterns across BDEs, customers, products, complaints, and outstanding next-actions.\n\n"
            "Reply in English with EXACTLY this format:\n\n"
            "SUMMARY:\n"
            "- 4 to 6 short bullets covering the cohort's main themes: what was discussed, "
            "recurring concerns, which customers/shops stand out, and any unresolved items.\n\n"
            "STRATEGY:\n"
            "- 4 to 6 concrete, prioritised next-step recommendations the team should take based on "
            "this cohort. Reference specific shops, BDEs, or product groups when relevant. "
            "Avoid generic advice.\n\n"
            "Keep each bullet under 30 words. Do not invent facts not present in the data."
        )
        prompt = instructions + "\n\n" + "\n".join(body_parts)
        text, err = _ai_call_claude(prompt)
        if err:
            return jsonify({"error": err}), 502
        summary, strategy = _ai_parse(text)
        return jsonify({
            "model":    AI_MODEL,
            "count":    len(rows),
            "summary":  summary,
            "strategy": strategy,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
# === AI FEATURE END ===


# ─────────────────────────────────────────────────────────────────────
# CLAIM portal — public form at /claim/<ship_to_code> + threaded back
# and forth at /claims for internal team.  Wrapped in START/END markers
# for easy revert.  No login on the customer side: the URL itself
# identifies the shop, supplemented by a honeypot field for bot spam.
# ─────────────────────────────────────────────────────────────────────
# === CLAIM FEATURE START ===
CLAIM_PHOTO_DIR = os.path.join(BASE_DIR, "static", "claim_photos")
CLAIM_TYPES = ("Road Hazard Warranty",)
CLAIM_STATUSES = ("New", "Reviewing", "Awaiting customer",
                  "Completed", "Resolved", "Rejected")

def _ensure_claim_tables():
    try:
        os.makedirs(CLAIM_PHOTO_DIR, exist_ok=True)
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS claim_submission (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                ship_to       VARCHAR(64)  NOT NULL,
                sold_to       VARCHAR(64)  NOT NULL DEFAULT '',
                ship_to_name  VARCHAR(160) NOT NULL DEFAULT '',
                sold_to_name  VARCHAR(160) NOT NULL DEFAULT '',
                contact_name  VARCHAR(120) NOT NULL DEFAULT '',
                contact_phone VARCHAR(40)  NOT NULL DEFAULT '',
                contact_email VARCHAR(160) NOT NULL DEFAULT '',
                claim_type    VARCHAR(40)  NOT NULL DEFAULT '',
                product_size  VARCHAR(80)  NOT NULL DEFAULT '',
                description   TEXT,
                photo_paths   TEXT,
                status        VARCHAR(40)  NOT NULL DEFAULT 'New',
                assigned_to   VARCHAR(120) NOT NULL DEFAULT '',
                created_at    TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP,
                resolved_at   TIMESTAMP    NULL,
                INDEX idx_ship    (ship_to),
                INDEX idx_status  (status),
                INDEX idx_created (created_at)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS claim_message (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                claim_id     INT NOT NULL,
                author_type  VARCHAR(20)  NOT NULL DEFAULT 'customer',
                author_name  VARCHAR(120) NOT NULL DEFAULT '',
                author_email VARCHAR(160) NOT NULL DEFAULT '',
                category     VARCHAR(40)  NOT NULL DEFAULT '',
                text         TEXT,
                photo_paths  TEXT,
                created_at   TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_claim   (claim_id),
                INDEX idx_created (created_at)
            )
        """)
        # Idempotent migration for tables that pre-date the category column.
        try:
            cur.execute("ALTER TABLE claim_message "
                        "ADD COLUMN category VARCHAR(40) NOT NULL DEFAULT '' "
                        "AFTER author_email")
        except Exception:
            pass
        cur.close(); conn.close()
    except Exception as e:
        print(f"[claim] schema init failed: {e}")
_ensure_claim_tables()

def _claim_lookup_shop(ship_to):
    """Resolve ship_to → (ship_to_name, sold_to, sold_to_name, bde_state, salesman_name)."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT ship_to, ship_to_name, sold_to, sold_to_name, "
            "       bde_state, salesman_name "
            "FROM customer WHERE ship_to = %s LIMIT 1",
            (ship_to,),
        )
        row = cur.fetchone()
        cur.close(); conn.close()
        return row
    except Exception as e:
        print(f"[claim] _claim_lookup_shop failed: {e}")
        return None

def _claim_save_photos(files, prefix, categories=None):
    """Save uploaded photo files under static/claim_photos/YYYY/MM/.
    Returns a list of public /static URLs.  When categories is given
    (parallel list of strings the size of files), the category is
    embedded in the filename so the viewer can label each photo
    without an extra DB column."""
    saved = []
    if not files: return saved
    now = datetime.now()
    subdir = os.path.join(CLAIM_PHOTO_DIR, f"{now.year:04d}", f"{now.month:02d}")
    os.makedirs(subdir, exist_ok=True)
    ts = now.strftime("%Y%m%d_%H%M%S")
    safe = "".join(c for c in (prefix or "X") if c.isalnum()) or "X"
    cats = list(categories or [])
    for i, f in enumerate(files[:10]):
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in (".jpg", ".jpeg", ".png", ".webp", ".heic", ".pdf"):
            continue
        cat = (cats[i] if i < len(cats) else "") or ""
        safe_cat = "".join(c for c in cat if c.isalnum() or c in "-_")[:24]
        suffix   = f"_{safe_cat}" if safe_cat else ""
        fname = f"{ts}_{safe}_{i}{suffix}{ext}"
        path  = os.path.join(subdir, fname)
        f.save(path)
        saved.append(f"/static/claim_photos/{now.year:04d}/{now.month:02d}/{fname}")
    return saved

# Fixed distribution list for claim notifications.  Every incoming
# claim + every customer reply on an open thread pings these three
# users.  Kept separate from ALWAYS_TO (meeting-log distribution) so
# a change on one channel doesn't leak into the other.  Order doesn't
# matter — the mail transport de-dupes.
CLAIM_NOTIFY_TO = [
    "elias@hankooktyre.com.au",
    "minku.lee@hankooktyre.com.au",
    "jayden.bhang@hankooktyre.com.au",
]


def _claim_notify_new(claim_id, shop, contact_name, claim_type, description):
    """Email the claim-handler team when a new claim is submitted."""
    to_list = list(CLAIM_NOTIFY_TO)
    subject = f"[Claim #{claim_id}] {shop.get('ship_to_name') if shop else ''} — {claim_type or 'New claim'}"
    link = f"{DASHBOARD_URL.rstrip('/')}/claims#{claim_id}"
    safe_desc = _esc_html((description or "")[:600])
    html = (
        f"<p>A new claim was submitted.</p>"
        f"<table style='border-collapse:collapse;font-family:Arial,sans-serif;font-size:13px;'>"
        f"<tr><td><b>Shop</b></td><td>{_esc_html(shop.get('ship_to_name') if shop else '')} "
        f"({_esc_html(shop.get('ship_to') if shop else '')})</td></tr>"
        f"<tr><td><b>Customer</b></td><td>{_esc_html(shop.get('sold_to_name') if shop else '')}</td></tr>"
        f"<tr><td><b>Contact</b></td><td>{_esc_html(contact_name or '')}</td></tr>"
        f"<tr><td><b>Type</b></td><td>{_esc_html(claim_type or '')}</td></tr>"
        f"<tr><td><b>Description</b></td><td><pre style='white-space:pre-wrap;'>{safe_desc}</pre></td></tr>"
        f"</table>"
        f"<p><a href='{link}'>Open claim #{claim_id} in the dashboard →</a></p>"
    )
    _send_mail_async(to_list, [], subject, html)

def _claim_notify_reply(claim_id, author_type, author_name, text):
    """Notify the other side when a new message lands on a thread."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT ship_to, ship_to_name, sold_to_name, contact_email, "
            "       contact_name, assigned_to "
            "FROM claim_submission WHERE id = %s", (claim_id,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row: return
        link = f"{DASHBOARD_URL.rstrip('/')}/claims#{claim_id}"
        cust_link = f"{CLAIM_PORTAL_URL.rstrip('/')}/claim/{row['ship_to']}"
        safe_text = _esc_html((text or "")[:600])
        if author_type == "customer":
            # Customer replied — ping the fixed handler team so someone
            # picks it up.  Same list the new-claim notification uses,
            # so a claim's whole life-cycle stays with the same three
            # people regardless of the shop's territory.
            to_list = list(CLAIM_NOTIFY_TO)
            subject = f"[Claim #{claim_id}] Customer replied — {row['ship_to_name']}"
            html = (f"<p><b>{_esc_html(author_name or row['contact_name'] or 'Customer')}</b> "
                    f"replied on claim #{claim_id} ({_esc_html(row['ship_to_name'])}):</p>"
                    f"<pre style='white-space:pre-wrap;font-family:Arial,sans-serif;'>{safe_text}</pre>"
                    f"<p><a href='{link}'>Open claim →</a></p>")
            _send_mail_async(to_list, [], subject, html)
        else:
            if row.get("contact_email"):
                subject = f"[Hankook Claim #{claim_id}] Reply from our team"
                html = (f"<p>Our team replied to your claim:</p>"
                        f"<pre style='white-space:pre-wrap;font-family:Arial,sans-serif;'>{safe_text}</pre>"
                        f"<p><a href='{cust_link}'>View and respond →</a></p>")
                _send_mail_async([row["contact_email"]], [], subject, html)
    except Exception as e:
        print(f"[claim] _claim_notify_reply failed: {e}")

# ── Customer-facing routes ───────────────────────────────────────────
@app.get("/claim/<ship_to>")
def claim_page(ship_to):
    return send_from_directory("static", "claim.html")

# Short-URL alias so QR codes (and the printable cards) can show
# claim.hankooktyre.com.au/735486 instead of …/claim/735486.  Restricted
# to ship-to-shaped paths so unrelated future routes (/meeting, /map,
# etc.) keep priority and random typos return 404 instead of leaking
# the claim page.
import re as _re
_SHIP_TO_SHORT_RE = _re.compile(r'^[A-Za-z0-9_-]{4,32}$')
_SHIP_TO_RESERVED = {
    "api", "static", "claim", "claims", "meeting", "map", "stock",
    "rebate", "price", "potentials", "favicon.ico", "robots.txt",
    "sitemap.xml", "health", "ping", "index.html",
}
@app.get("/<ship_to>")
def claim_short_url(ship_to):
    s = (ship_to or "").strip()
    if s.lower() in _SHIP_TO_RESERVED or not _SHIP_TO_SHORT_RE.match(s):
        from flask import abort
        abort(404)
    return send_from_directory("static", "claim.html")

@app.get("/api/claim/shop/<ship_to>")
def claim_shop_info(ship_to):
    shop = _claim_lookup_shop(ship_to)
    if not shop:
        return jsonify({"error": "Shop code not recognised", "ship_to": ship_to}), 404
    return jsonify({
        "ship_to":      shop["ship_to"],
        "ship_to_name": shop.get("ship_to_name") or "",
        "sold_to":      shop.get("sold_to") or "",
        "sold_to_name": shop.get("sold_to_name") or "",
        # Customer-facing base URL so the printable QR card embeds the
        # configured portal domain instead of whatever host happened to
        # serve the admin /claims page.
        "portal_url":   CLAIM_PORTAL_URL.rstrip("/"),
    })

@app.get("/api/claim/by_ship/<ship_to>")
def claim_list_for_shop(ship_to):
    """Customer-side: list this shop's prior claims + threads so they
    can check on status and continue any open conversation."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, claim_type, status, created_at, resolved_at,
                   contact_name, description
            FROM claim_submission
            WHERE ship_to = %s
            ORDER BY created_at DESC
            LIMIT 50
        """, (ship_to,))
        claims = cur.fetchall()
        ids = [c["id"] for c in claims]
        threads = {}
        if ids:
            ph = ",".join(["%s"] * len(ids))
            cur.execute(f"""
                SELECT id, claim_id, author_type, author_name, category,
                       text, photo_paths, created_at
                FROM claim_message
                WHERE claim_id IN ({ph})
                ORDER BY created_at ASC, id ASC
            """, tuple(ids))
            for m in cur.fetchall():
                if m.get("created_at"):
                    m["created_at"] = m["created_at"].strftime("%Y-%m-%d %H:%M")
                m["photo_paths"] = m["photo_paths"].split(",") if m["photo_paths"] else []
                threads.setdefault(m["claim_id"], []).append(m)
        for c in claims:
            if c.get("created_at"):
                c["created_at"] = c["created_at"].strftime("%Y-%m-%d %H:%M")
            if c.get("resolved_at"):
                c["resolved_at"] = c["resolved_at"].strftime("%Y-%m-%d %H:%M")
            c["thread"] = threads.get(c["id"], [])
        cur.close(); conn.close()
        return jsonify({"claims": claims})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/claim/by_ship/<ship_to>")
def claim_create(ship_to):
    # Honeypot — bots fill every field, real users don't see this one.
    if (request.form.get("website") or "").strip():
        return jsonify({"error": "spam"}), 400
    shop = _claim_lookup_shop(ship_to)
    if not shop:
        return jsonify({"error": "Shop code not recognised"}), 404
    contact_name  = (request.form.get("contact_name")  or "").strip()[:120]
    contact_phone = (request.form.get("contact_phone") or "").strip()[:40]
    contact_email = (request.form.get("contact_email") or "").strip()[:160]
    claim_type    = (request.form.get("claim_type")    or "").strip()[:40]
    product_size  = (request.form.get("product_size")  or "").strip()[:80]
    description   = (request.form.get("description")   or "").strip()
    if not contact_name or not contact_phone:
        return jsonify({"error": "Name and phone are required"}), 400
    if not description:
        return jsonify({"error": "Please describe the issue"}), 400
    # Only one claim type is currently exposed, so coerce any other
    # value to the default (frontend already restricts the dropdown).
    if claim_type not in CLAIM_TYPES:
        claim_type = CLAIM_TYPES[0]
    # Categories come as a comma-joined string in parallel to the
    # ordered photos files.  Frontend pushes one category per file
    # (multi-file slot like tread_measurement contributes N entries),
    # plus an empty-entry placeholder for empty slots, so the list
    # stays index-aligned with raw_files.
    cats = [c.strip() for c in (request.form.get("photo_categories") or "").split(",")]
    raw_files = request.files.getlist("photos") or []
    # Per-category minimums.  Six single-photo slots, tread measurement
    # requires at least 3 photos (one per groove), and proof of purchase
    # (invoice) is also required.  'others' is intentionally absent —
    # the slot exists for extras but doesn't block submission.
    REQUIRED_MIN = {
        "whole_tyre": 1, "tread": 1, "damaged": 1,
        "dot_code":   1, "dot_cut": 1, "serial_barcode": 1,
        "tread_measurement": 3,
        "invoice": 1,
    }
    got_by_cat = {}
    for i, f in enumerate(raw_files):
        if not (f and f.filename):
            continue
        cat = cats[i] if i < len(cats) else ""
        if cat in REQUIRED_MIN:
            got_by_cat[cat] = got_by_cat.get(cat, 0) + 1
    short = [c for c, need in REQUIRED_MIN.items() if got_by_cat.get(c, 0) < need]
    if short:
        return jsonify({"error":
            "Please attach all required product photos — missing: "
            + ", ".join(c.replace("_", " ") for c in short)}), 400
    photos = _claim_save_photos(raw_files, ship_to, categories=cats)
    total_required = sum(REQUIRED_MIN.values())
    if len(photos) < total_required:
        return jsonify({"error":
            f"Could not save all {total_required} required photos — "
            "please check file types (JPG / PNG / HEIC / PDF) and try again."}), 400
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO claim_submission
                (ship_to, sold_to, ship_to_name, sold_to_name,
                 contact_name, contact_phone, contact_email,
                 claim_type, product_size, description, photo_paths,
                 status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'New')
        """, (
            ship_to, shop.get("sold_to") or "",
            shop.get("ship_to_name") or "", shop.get("sold_to_name") or "",
            contact_name, contact_phone, contact_email,
            claim_type, product_size, description,
            ",".join(photos) if photos else None,
        ))
        cid = cur.lastrowid
        # First message: mirror the description so the thread reads as
        # a normal conversation from message #1.
        cur.execute("""
            INSERT INTO claim_message
                (claim_id, author_type, author_name, author_email, text, photo_paths)
            VALUES (%s, 'customer', %s, %s, %s, %s)
        """, (cid, contact_name, contact_email, description,
              ",".join(photos) if photos else None))
        cur.close(); conn.close()
        _claim_notify_new(cid, shop, contact_name, claim_type, description)
        return jsonify({"id": cid, "status": "New"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/claim/<int:cid>/customer_reply")
def claim_customer_reply(cid):
    """Customer-side reply.  Requires the same ship_to in the body so a
    leaked claim id alone can't be used to talk on someone else's thread."""
    if (request.form.get("website") or "").strip():
        return jsonify({"error": "spam"}), 400
    ship_to = (request.form.get("ship_to") or "").strip()
    text    = (request.form.get("text") or "").strip()
    name    = (request.form.get("contact_name") or "").strip()[:120]
    email   = (request.form.get("contact_email") or "").strip()[:160]
    if not ship_to or not text:
        return jsonify({"error": "ship_to and text are required"}), 400
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("SELECT ship_to FROM claim_submission WHERE id=%s", (cid,))
        row = cur.fetchone()
        if not row or row["ship_to"] != ship_to:
            cur.close(); conn.close()
            return jsonify({"error": "Claim not found for this shop"}), 404
        photos = _claim_save_photos(request.files.getlist("photos"), ship_to)
        cur.execute("""
            INSERT INTO claim_message
                (claim_id, author_type, author_name, author_email, text, photo_paths)
            VALUES (%s, 'customer', %s, %s, %s, %s)
        """, (cid, name, email, text, ",".join(photos) if photos else None))
        cur.close(); conn.close()
        _claim_notify_reply(cid, "customer", name, text)
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ── Internal routes ──────────────────────────────────────────────────
@app.get("/claims")
def claims_page():
    return send_from_directory("static", "claims.html")

@app.get("/api/claims")
def claims_list():
    """Internal: list claims with optional filters (status, state, BDE, search)."""
    status = (request.args.get("status") or "").strip()
    state  = (request.args.get("state")  or "").strip().upper()
    bde    = (request.args.get("bde")    or "").strip()
    q      = (request.args.get("q")      or "").strip()
    wh = []
    params = []
    if status:
        wh.append("c.status = %s"); params.append(status)
    if q:
        wh.append("(c.ship_to LIKE %s OR c.ship_to_name LIKE %s "
                  " OR c.sold_to_name LIKE %s OR c.contact_name LIKE %s)")
        like = f"%{q}%"
        params += [like, like, like, like]
    where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT c.id, c.ship_to, c.ship_to_name, c.sold_to_name,
                   c.contact_name, c.contact_phone, c.contact_email,
                   c.claim_type, c.status, c.assigned_to,
                   c.created_at, c.resolved_at,
                   c.description, c.product_size, c.photo_paths,
                   (SELECT COUNT(*) FROM claim_message m WHERE m.claim_id = c.id) AS msg_count
            FROM claim_submission c
            {where_sql}
            ORDER BY c.created_at DESC
            LIMIT 500
        """, tuple(params))
        rows = cur.fetchall()
        # State + BDE filters apply via the customer master so a single
        # join covers both — narrows the visible claims to those whose
        # shop matches the chosen scope.
        if (state and state in ("NSW","QLD","VIC","WA","SA","TAS")) or bde:
            ships = [r["ship_to"] for r in rows]
            keep = set()
            if ships:
                ph = ",".join(["%s"] * len(ships))
                cur.execute(f"""
                    SELECT DISTINCT ship_to, bde_state, salesman_name
                    FROM customer
                    WHERE ship_to IN ({ph})
                """, tuple(ships))
                cust_rows = cur.fetchall()
                for s in cust_rows:
                    if state:
                        st = STATE_REMAP.get(s.get("bde_state"), s.get("bde_state"))
                        if st != state: continue
                    if bde and (s.get("salesman_name") or "").strip().upper() != bde.upper():
                        continue
                    keep.add(s["ship_to"])
            rows = [r for r in rows if r["ship_to"] in keep]
        for r in rows:
            if r.get("created_at"):
                r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M")
            if r.get("resolved_at"):
                r["resolved_at"] = r["resolved_at"].strftime("%Y-%m-%d %H:%M")
            r["photo_paths"] = r["photo_paths"].split(",") if r["photo_paths"] else []
        # BDE list for the filter dropdown — narrowed by state when set,
        # so the picker can't include BDEs outside the current scope.
        bde_wh = ["salesman_name IS NOT NULL", "TRIM(salesman_name) <> ''"]
        bde_params = []
        if state and state in ("NSW","QLD","VIC","WA","SA","TAS"):
            bde_wh.append("bde_state = %s"); bde_params.append(state)
        cur.execute(f"""
            SELECT DISTINCT TRIM(salesman_name) AS name
            FROM customer
            WHERE {' AND '.join(bde_wh)}
            ORDER BY name
        """, tuple(bde_params))
        bdes = [r["name"] for r in cur.fetchall() if r.get("name")]
        cur.close(); conn.close()
        return jsonify({
            "claims":   rows,
            "statuses": list(CLAIM_STATUSES),
            "bdes":     bdes,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/claim/<int:cid>")
def claim_detail(cid):
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM claim_submission WHERE id=%s", (cid,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify({"error": "Claim not found"}), 404
        cur.execute("""
            SELECT id, claim_id, author_type, author_name, author_email,
                   category, text, photo_paths, created_at
            FROM claim_message
            WHERE claim_id = %s
            ORDER BY created_at ASC, id ASC
        """, (cid,))
        thread = cur.fetchall()
        for m in thread:
            if m.get("created_at"):
                m["created_at"] = m["created_at"].strftime("%Y-%m-%d %H:%M")
            m["photo_paths"] = m["photo_paths"].split(",") if m["photo_paths"] else []
        if row.get("created_at"):
            row["created_at"] = row["created_at"].strftime("%Y-%m-%d %H:%M")
        if row.get("resolved_at"):
            row["resolved_at"] = row["resolved_at"].strftime("%Y-%m-%d %H:%M")
        row["photo_paths"] = row["photo_paths"].split(",") if row["photo_paths"] else []
        cur.close(); conn.close()
        return jsonify({"claim": row, "thread": thread})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.patch("/api/claim/<int:cid>")
def claim_update(cid):
    body = request.get_json(silent=True) or {}
    fields = {}
    if "status" in body:
        s = (body["status"] or "").strip()
        if s and s not in CLAIM_STATUSES:
            return jsonify({"error": "invalid status"}), 400
        fields["status"] = s
        if s in ("Completed", "Resolved"):
            fields["resolved_at"] = datetime.now()
        else:
            fields["resolved_at"] = None
    if "assigned_to" in body:
        fields["assigned_to"] = (body["assigned_to"] or "").strip()[:120]
    if not fields:
        return jsonify({"error": "Nothing to update"}), 400
    try:
        conn = get_connection(); cur = conn.cursor()
        sets = ", ".join([f"{k} = %s" for k in fields])
        params = list(fields.values()) + [cid]
        cur.execute(f"UPDATE claim_submission SET {sets} WHERE id = %s", params)
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/claim/<int:cid>/internal_reply")
def claim_internal_reply(cid):
    email = _bde_from_request()
    me    = _EMAIL_TO_DIR.get(email.lower()) if email else None
    name  = (me[0] if me else "") or "Hankook team"
    text  = ""
    category = ""
    is_json = request.is_json
    if is_json:
        body = request.get_json(silent=True) or {}
        text = (body.get("text") or "").strip()
        category = (body.get("category") or "").strip()[:40]
    else:
        text = (request.form.get("text") or "").strip()
        category = (request.form.get("category") or "").strip()[:40]
    if not text:
        return jsonify({"error": "text is required"}), 400
    photos = _claim_save_photos(request.files.getlist("photos"), f"int{cid}") if not is_json else []
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO claim_message
                (claim_id, author_type, author_name, author_email,
                 category, text, photo_paths)
            VALUES (%s, 'internal', %s, %s, %s, %s, %s)
        """, (cid, name, email or "", category, text,
              ",".join(photos) if photos else None))
        # Auto-transition: a claim sitting in 'New' jumps to 'Reviewing'
        # the moment someone on our side replies, so the workflow
        # statuses match what's actually happening without the manager
        # having to flip the dropdown manually.
        cur.execute(
            "UPDATE claim_submission SET status = 'Reviewing' "
            "WHERE id = %s AND status = 'New'", (cid,))
        cur.close(); conn.close()
        _claim_notify_reply(cid, "internal", name, text)
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
# ── QR code generation (sticker per shop + bulk grid) ──────────────
@app.get("/api/claim/qr/<ship_to>")
def claim_qr_png(ship_to):
    """Render a PNG QR code that encodes /claim/<ship_to>.  The image
    is what the printable card/grid pages embed via <img src>."""
    try:
        import qrcode  # type: ignore
    except ImportError:
        return ("qrcode package not installed — pip install qrcode[pil]", 500)
    from io import BytesIO
    from flask import send_file
    url = f"{CLAIM_PORTAL_URL.rstrip('/')}/claim/{ship_to}"
    qr  = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10, border=2,
    )
    qr.add_data(url); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
    resp = send_file(buf, mimetype="image/png",
                     download_name=f"claim_qr_{ship_to}.png")
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp

@app.get("/claims/qr/<ship_to>")
def claim_qr_card(ship_to):
    """Single-shop printable A6 card."""
    return send_from_directory("static", "claim_qr_card.html")

@app.get("/claims/qr_bulk")
def claim_qr_bulk_page():
    """Bulk A4 grid of QR cards.  Frontend reads ?ships=A,B,C&state=NSW
    and renders one card per shop on the page."""
    return send_from_directory("static", "claim_qr_bulk.html")

@app.get("/api/claim/qr_shops")
def claim_qr_shops():
    """Helper for the bulk page — returns ship_tos matching the filters.
    Defaults to active shops (any in customer master)."""
    state   = (request.args.get("state") or "").strip().upper()
    bde     = (request.args.get("bde")   or "").strip()
    sold_to = (request.args.get("sold_to") or "").strip()
    ship_to_arg = (request.args.get("ship_to") or "").strip()
    # Frontend dropped its own cap so the page renders every match in
    # scope.  Server-side cap stays as a safety net (10 000 is well
    # above the largest customer master we've ever loaded).
    limit = min(int(request.args.get("limit") or "10000"), 10000)
    wh = ["ship_to IS NOT NULL", "TRIM(ship_to) <> ''"]
    params = []
    if state:
        # Map our region buttons back to the underlying bde_state values
        # so SA/TAS/ACT shops still print under the manager who serves them.
        wh.append("bde_state = %s")
        params.append(state)
    if bde:
        wh.append("UPPER(TRIM(salesman_name)) = UPPER(TRIM(%s))")
        params.append(bde)
    if sold_to:
        # Match either the sold-to code or its resolved name so the user
        # can type whichever they remember in the bulk-print search box.
        wh.append("("
                  "UPPER(TRIM(sold_to)) = UPPER(TRIM(%s)) "
                  "OR UPPER(TRIM(sold_to_name)) = UPPER(TRIM(%s))"
                  ")")
        params.extend([sold_to, sold_to])
    if ship_to_arg:
        # Same loose match as sold_to: accept either the ship_to code or
        # its name so the search box can take whatever the user typed.
        wh.append("("
                  "UPPER(TRIM(ship_to)) = UPPER(TRIM(%s)) "
                  "OR UPPER(TRIM(ship_to_name)) = UPPER(TRIM(%s))"
                  ")")
        params.extend([ship_to_arg, ship_to_arg])
    where_sql = "WHERE " + " AND ".join(wh)
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT ship_to,
                   MIN(NULLIF(TRIM(ship_to_name),'')) AS ship_to_name,
                   MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name,
                   MIN(NULLIF(TRIM(bde_state),''))   AS bde_state,
                   MIN(NULLIF(TRIM(salesman_name),'')) AS salesman_name
            FROM customer
            {where_sql}
            GROUP BY ship_to
            ORDER BY ship_to_name, ship_to
            LIMIT {limit}
        """, tuple(params))
        rows = cur.fetchall()
        # Also surface the available BDE list (for the bulk page dropdown)
        # narrowed by state if one is selected — keeps the picker scoped.
        bde_wh = ["salesman_name IS NOT NULL", "TRIM(salesman_name) <> ''"]
        bde_params = []
        if state:
            bde_wh.append("bde_state = %s"); bde_params.append(state)
        cur.execute(f"""
            SELECT DISTINCT TRIM(salesman_name) AS name
            FROM customer
            WHERE {' AND '.join(bde_wh)}
            ORDER BY name
        """, tuple(bde_params))
        bdes = [r["name"] for r in cur.fetchall() if r.get("name")]
        # Distinct sold-tos for the search box — narrowed by State + BDE
        # if either is set, so the suggestion list always matches the
        # currently-applied scope.  Restricted to sold_tos that actually
        # transact (sales_2526) so the list matches /api/sold_to_names
        # used by the graph view — keeps inactive / per-location master
        # rows out of the suggestion box.
        st_wh = ["sold_to IS NOT NULL", "TRIM(sold_to) <> ''",
                 "sold_to IN (SELECT DISTINCT sold_to FROM sales_2526 "
                 "WHERE sold_to IS NOT NULL)"]
        st_params = []
        if state:
            st_wh.append("bde_state = %s"); st_params.append(state)
        if bde:
            st_wh.append("UPPER(TRIM(salesman_name)) = UPPER(TRIM(%s))")
            st_params.append(bde)
        cur.execute(f"""
            SELECT sold_to,
                   MIN(NULLIF(TRIM(sold_to_name),'')) AS sold_to_name
            FROM customer
            WHERE {' AND '.join(st_wh)}
            GROUP BY sold_to
            ORDER BY sold_to_name, sold_to
        """, tuple(st_params))
        sold_tos = [{"sold_to": r["sold_to"], "name": r.get("sold_to_name") or r["sold_to"]}
                    for r in cur.fetchall() if r.get("sold_to")]
        cur.close(); conn.close()
        return jsonify({
            "shops":     rows,
            "bdes":      bdes,
            "sold_tos":  sold_tos,
            "base_url":  CLAIM_PORTAL_URL.rstrip("/"),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
# === CLAIM FEATURE END ===


# ─────────────────────────────────────────────────────────────────────
# POTENTIAL CUSTOMER feature — BDEs add prospect shops they're chasing
# so the planner / map can track excavation work alongside the real
# customer master.  Wrapped in START/END markers for easy revert.
# ─────────────────────────────────────────────────────────────────────
# === POTENTIAL CUSTOMER FEATURE START ===
POTENTIAL_STATUSES = ("Lead", "Visited", "Negotiating", "Converted", "Lost")

def _ensure_potential_table():
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS potential_customer (
                id             INT AUTO_INCREMENT PRIMARY KEY,
                name           VARCHAR(160) NOT NULL,
                sold_to_name   VARCHAR(160) NOT NULL DEFAULT '',
                address        VARCHAR(255) NOT NULL DEFAULT '',
                phone          VARCHAR(40)  NOT NULL DEFAULT '',
                contact_person VARCHAR(120) NOT NULL DEFAULT '',
                state          VARCHAR(8)   NOT NULL DEFAULT '',
                bde_name       VARCHAR(120) NOT NULL DEFAULT '',
                bde_email      VARCHAR(160) NOT NULL DEFAULT '',
                lat            DECIMAL(10,6) NULL,
                lon            DECIMAL(11,6) NULL,
                status         VARCHAR(40)  NOT NULL DEFAULT 'Lead',
                notes          TEXT,
                created_at     TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at     TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP
                               ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_state (state),
                INDEX idx_bde   (bde_name)
            )
        """)
        # Idempotent migration for rows created before sold_to_name existed.
        try:
            cur.execute("ALTER TABLE potential_customer "
                        "ADD COLUMN sold_to_name VARCHAR(160) NOT NULL DEFAULT '' "
                        "AFTER name")
        except Exception:
            pass
        cur.close(); conn.close()
    except Exception as e:
        print(f"[potential] schema init failed: {e}")
_ensure_potential_table()

def _potential_to_pseudo_ship(pid):
    """Pseudo ship_to code used in meeting_plan / meeting_log so the
    rest of the system can carry potential customers without a real
    customer-master row.  Format chosen to be obviously non-numeric."""
    return f"POT-{int(pid)}"

def _parse_potential_id(ship_to):
    """Reverse of the above.  Returns int id or None for real shops."""
    s = (ship_to or "").strip().upper()
    if not s.startswith("POT-"): return None
    try: return int(s[4:])
    except ValueError: return None

@app.get("/api/potential_customers")
def potential_customers_list():
    """List potential customers, optionally filtered by state / bde."""
    state = (request.args.get("state") or "").strip().upper()
    bde   = (request.args.get("bde")   or "").strip()
    wh, params = [], []
    if state and state in ("NSW","QLD","VIC","WA","SA","TAS"):
        wh.append("state = %s"); params.append(state)
    if bde:
        wh.append("UPPER(TRIM(bde_name)) = %s")
        params.append(bde.upper())
    where_sql = ("WHERE " + " AND ".join(wh)) if wh else ""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT id, name, sold_to_name,
                   address, phone, contact_person,
                   state, bde_name, bde_email, lat, lon, status, notes,
                   created_at, updated_at
            FROM potential_customer
            {where_sql}
            ORDER BY name ASC, id ASC
        """, tuple(params))
        rows = cur.fetchall()
        for r in rows:
            for k in ("created_at", "updated_at"):
                if r.get(k):
                    r[k] = r[k].strftime("%Y-%m-%d %H:%M")
            r["lat"] = float(r["lat"]) if r["lat"] is not None else None
            r["lon"] = float(r["lon"]) if r["lon"] is not None else None
            r["pseudo_ship_to"] = _potential_to_pseudo_ship(r["id"])
        cur.close(); conn.close()
        return jsonify({"potentials": rows, "statuses": list(POTENTIAL_STATUSES)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.post("/api/potential_customer")
def potential_customer_create():
    body = request.get_json(silent=True) or request.form
    name = (body.get("name") or "").strip()[:160]
    if not name:
        return jsonify({"error": "ship-to name is required"}), 400
    email = _bde_from_request()
    me    = _EMAIL_TO_DIR.get(email.lower()) if email else None
    sold_to_name = (body.get("sold_to_name") or "").strip()[:160]
    bde_name  = (body.get("bde_name") or (me[0] if me else "")).strip()[:120]
    state     = (body.get("state")    or (me[1] if me else "")).strip().upper()[:8]
    address   = (body.get("address")  or "").strip()[:255]
    phone     = (body.get("phone")    or "").strip()[:40]
    contact   = (body.get("contact_person") or "").strip()[:120]
    notes     = (body.get("notes")    or "").strip()
    status    = (body.get("status")   or "Lead").strip()
    if status not in POTENTIAL_STATUSES: status = "Lead"
    def _flt(v):
        try: return float(v)
        except (TypeError, ValueError): return None
    lat = _flt(body.get("lat"))
    lon = _flt(body.get("lon"))
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO potential_customer
                (name, sold_to_name, address, phone, contact_person,
                 state, bde_name, bde_email, lat, lon, status, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (name, sold_to_name, address, phone, contact,
              state, bde_name, email or "",
              lat, lon, status, notes))
        pid = cur.lastrowid
        cur.close(); conn.close()
        return jsonify({"id": pid, "pseudo_ship_to": _potential_to_pseudo_ship(pid)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.patch("/api/potential_customer/<int:pid>")
def potential_customer_update(pid):
    body = request.get_json(silent=True) or {}
    allowed = ("name", "sold_to_name", "address", "phone", "contact_person",
               "state", "bde_name", "status", "notes", "lat", "lon")
    sets, params = [], []
    for k in allowed:
        if k in body:
            v = body[k]
            if k in ("lat", "lon"):
                try: v = float(v) if v not in (None, "") else None
                except (TypeError, ValueError): v = None
            elif k == "status":
                v = (v or "").strip()
                if v not in POTENTIAL_STATUSES:
                    return jsonify({"error": "invalid status"}), 400
            elif k == "state":
                v = (v or "").strip().upper()[:8]
            else:
                v = (v or "").strip()
                if k == "name" and not v:
                    return jsonify({"error": "name cannot be empty"}), 400
            sets.append(f"{k} = %s"); params.append(v)
    if not sets:
        return jsonify({"error": "nothing to update"}), 400
    params.append(pid)
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute(f"UPDATE potential_customer SET {', '.join(sets)} "
                    f"WHERE id = %s", params)
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.delete("/api/potential_customer/<int:pid>")
def potential_customer_delete(pid):
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("DELETE FROM potential_customer WHERE id = %s", (pid,))
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/potential_customers/stats")
def potential_customers_stats():
    """Counts grouped by State and by Status for the management page header."""
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("SELECT COUNT(*) AS n FROM potential_customer")
        total = (cur.fetchone() or {}).get("n", 0)
        cur.execute("""
            SELECT state, COUNT(*) AS n
            FROM potential_customer
            GROUP BY state
            ORDER BY state
        """)
        by_state = {r["state"] or "—": r["n"] for r in cur.fetchall()}
        cur.execute("""
            SELECT status, COUNT(*) AS n
            FROM potential_customer
            GROUP BY status
            ORDER BY status
        """)
        by_status = {r["status"] or "Lead": r["n"] for r in cur.fetchall()}
        cur.execute("""
            SELECT bde_name, COUNT(*) AS n
            FROM potential_customer
            GROUP BY bde_name
            ORDER BY n DESC, bde_name ASC
            LIMIT 50
        """)
        by_bde = [{"bde_name": r["bde_name"] or "—", "n": r["n"]} for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({
            "total":    total,
            "by_state": by_state,
            "by_status": by_status,
            "by_bde":   by_bde,
            "statuses": list(POTENTIAL_STATUSES),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
# === POTENTIAL CUSTOMER FEATURE END ===


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))   # Cloudtype probes 5000
    from price_compare import price_dashboard, load_all_months, build_data
    app.add_url_rule('/price', 'price_dashboard', price_dashboard)

    @app.get("/api/price_debug")
    def price_debug():
        from flask import jsonify
        monthly = load_all_months()
        if not monthly:
            return jsonify({"error": "no monthly data"})
        data = build_data(monthly)
        return jsonify({
            "months": data["months"],
            "store_avg_tempe": data["store_avg"]["tempe"],
            "brand_size_data_keys": list(data["brand_size_data"].keys()),
            "sample": {abbr: list(sizes.keys())[:3] for abbr, sizes in list(data["brand_size_data"].items())[:3]},
        })

    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
